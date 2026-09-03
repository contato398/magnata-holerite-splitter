"""
magnata-holerite-splitter — app.py v2.95
Novidades vs v2.48:
  - Integração Secullum Ponto Web (Banco ID 149582) via blueprint isolado
    src/services/secullum_ponto.py (prefixo /secullum): autenticação Bearer,
    sincronização de funcionários (POST, CPF obrigatório) e varredura de
    Batidas Ímpares / Desvios de Carga Horária > 02:00, gerando alertas na
    tabela Pendências/Revisar do Airtable.

Histórico anterior (v2.6):
  - Novo endpoint /email/webhook (Fase 2 - Caixa de Entrada):
    recebe e-mails/anexos via Google Apps Script, registra em
    Emails Savian / Arquivos / Processar Arquivos, classifica o tipo de
    documento e cria pendências quando necessário. Suporta dry_run e
    é protegido por X-API-KEY.
"""

import os
import re
import sys
import io
import json
import gc
import time
import zipfile
import base64
import mimetypes
import hashlib
import secrets
import unicodedata
import logging
import tempfile
import requests
from collections import Counter, OrderedDict
from calendar import monthrange
from datetime import datetime, timezone, timedelta
from flask import Flask, request, jsonify, redirect
import pdfplumber
import uuid as _uuid
from pypdf import PdfReader, PdfWriter

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


def _mem_mb():
    try:
        with open('/proc/self/status') as f:
            for line in f:
                if line.startswith('VmRSS:'):
                    return int(line.split()[1]) // 1024
    except Exception:
        pass
    return -1


app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB máx upload

# ── Autenticação Administrativa Compartilhada V1 (wiring mínimo) ───────────────
# Só conecta a infraestrutura já implementada e testada no PR #118
# (magnata_os/autenticacao/) ao app Flask real -- nenhuma lógica nova aqui.
# configurar_sessao_segura lê MAGNATA_SESSION_SECRET_KEY do ambiente e falha
# de forma explícita (SegredoSessaoAusente) se ausente -- fail-closed, nunca
# um segredo gerado on-the-fly nem hardcoded.
from magnata_os.autenticacao.adapters.blueprint_login import auth_bp
from magnata_os.autenticacao.adapters.sessao import configurar_sessao_segura
configurar_sessao_segura(app)
app.register_blueprint(auth_bp)

# ── Integração Secullum Ponto Web (módulo isolado) ──────────────────────────────
from src.services.secullum_ponto import secullum_bp
app.register_blueprint(secullum_bp)
from src.sync_new_employees import sync_bp
app.register_blueprint(sync_bp)
from src.ingestao_secullum import ingestao_bp
app.register_blueprint(ingestao_bp)

# ── Módulo 01 (Ingestão) — Fase 0: observabilidade, sem efeito operacional ──────
from src.observability import observar_ingestao

@app.after_request
def _add_cors(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Requested-With'
    return response

# ── Airtable ──────────────────────────────────────────────────────────────────
AIRTABLE_API_KEY = os.environ.get('AIRTABLE_API_KEY', '')
BASE_ID     = 'appaCpIVj7Q97VhFy'
TABLE_FUNC  = 'tblNd8G66kjwos3eP'   # Funcionários
TABLE_HOL   = 'tblVaUgZeFfa5zRcH'   # Holerites
TABLE_CONT  = 'tblWITpkSbPg4SBAR'   # Contabilidade Mensal

F_HOL_NOME      = 'fldS42HdVbLhDRVOY'
F_HOL_STATUS    = 'fld0hQpNpQTVmDSeZ'
F_HOL_FUNC      = 'fldTXMjeHfgyDas9f'
F_HOL_DATA      = 'fld8hTVUDyDf5jfPE'
F_HOL_FOLHA     = 'fldqQZwNnMf8BGfyP'
F_HOL_PDF       = 'fldGXsgmuADtZIgtx'
F_HOL_MES_CONT  = 'fldUYB4uxkmBf7vDe'
F_HOL_VENCIM    = 'fldOal5gy1aqF5RPT'   # Total Vencimentos
F_HOL_DESCONTOS = 'fldRiNiS9Rqfjo5Hg'   # Total Descontos
F_HOL_LIQUIDO   = 'fldnOzg7FcXxes2dm'   # Valor Líquido
F_HOL_INSS      = 'fldvDY9x8dC0FCLrd'   # Descontos INSS

# Funcionários — campos usados no pré-cadastro (Fase 5C)
F_FUNC_STATUS    = 'fld5T04dlg1Yt6Xj8'
F_FUNC_CARGO     = 'fldK1lJS1L4tbZVmZ'
F_FUNC_CPF       = 'fld0Y3bXdArkSIJxo'
F_FUNC_NOME      = 'fld2fSiomk9AOLGDb'   # Nome Completo
F_FUNC_ADMISSAO  = 'fld5L1djmJugvLe8c'   # Data de Admissão
F_FUNC_PDF_FOLHA = 'fldgBhXpEFmy20yxd'   # PDF Folha Ponto (prontuário)
F_FUNC_RECIBOS   = 'fldFA03LEJ3wk3UmG'   # Recibos Pagamento (fatiados por colaborador) — v2.33
F_FUNC_RESUMO_PONTO = 'fldI6soHhol2CdQpP'  # Resumo Cartão Ponto (extraído)
F_FUNC_DOCS_ADMISSAO = 'fldV0KcIqK8ly8PKl'  # Documentos Admissionais (Kit Admissão agrupado)
F_FUNC_KIT_CONSOLIDADO = 'fld2aVtOjvJkOXn5P'  # Kit Admissão - PDF Consolidado (fundido)
F_FUNC_DOCUMENTOS = 'fldCWscgMwOT3Ej72'  # Documentos (campo genérico, visível na Interface por padrão)

# ── Tabelas/Campos Fase 2 — Caixa de Entrada ─────────────────────────────────
TABLE_EMAILS     = 'tblljRRrraXSipJd1'   # Emails Savian
TABLE_ARQUIVOS   = 'tblRsvhz8oOcUqhkv'   # Arquivos
TABLE_PROCESSAR  = 'tblXaLXvGJMyFOayc'   # Processar Arquivos
TABLE_PENDENCIAS = 'tblRkJBL6Wwf4fxVC'   # Pendências/Revisar

# ── Tabelas — Diagnóstico de Distribuição (read-only) ────────────────────────
TABLE_LOCAIS    = 'tblZy1WfzmGIeR8ZP'   # Locais
TABLE_CLIENTES  = 'tbl0znyuCEzoCHtCV'   # Clientes
TABLE_ENVIOS    = 'tblAu4wgdfTgLOoa4'   # Envios de Documentos

# Envios de Documentos — Fila de Envios / Recibo Digital de Leitura (Fase 3)
F_ENVIO_STATUS   = 'fldWm7mHYMwQpkQAr'   # Status (Preparando/Enviado/Concluído/Lido)
F_ENVIO_TIPO     = 'fld9PU6FeHnIk4XK5'   # Tipo (Relatório Mensal/Envio Manual)
F_ENVIO_CANAL    = 'fldVDCqA4oMzbZQRj'   # Canal (E-mail/WhatsApp)
F_ENVIO_DEST     = 'fldtQz5PhGwC8kxLC'   # Destinatário (texto livre: WhatsApp ou e-mail)
F_ENVIO_EMAIL    = 'flddvh7fZoWSnHXPW'   # Email
F_ENVIO_HOLERITES = 'fldpAcVyoVqDscvYr'  # Holerites (link)
F_ENVIO_CLIENTE  = 'flddPGn6vHiJw7vba'   # Cliente (link)
F_ENVIO_FUNC     = 'fldcm9bAj13phGQqS'   # Funcionário(s) Vinculado(s) (link)
F_ENVIO_HASH     = 'fldiu4OgOQolil57a'   # Hash Recibo
F_ENVIO_LIDO_EM  = 'fldwPmloi2rFZOL2a'   # Recibo Lido em
F_ENVIO_ARQUIVOS = 'fldiO4G7OO1FAjn5o'   # Arquivos (anexos genéricos — usado p/ Cartão Ponto)
F_ENVIO_DATA     = 'fldZm5SbVuFoxB6ge'   # Data envio
F_ENVIO_ERRO     = 'fldxaCcWELeEclZt7'   # Erro (texto do erro de disparo)
F_ENVIO_ENVIADO_EM = 'fldagEB513S7tPeJh' # Enviado em (timestamp exato BRT) — Protocolo de Entrega

TIPO_ENVIO_PONTO = 'Cartão Ponto Mensal'   # Tipo de envio espelhado p/ Folha de Ponto/Cartão Ponto
TIPO_ENVIO_COMBINADO = 'Holerite + Cartão Ponto Mensal'   # v2.25 — envio único c/ 2 PDFs (Holerite + Ponto)
TIPO_ENVIO_EMAIL_CLIENTE = 'Pacote Mensal Cliente'   # v2.29 — e-mail consolidado p/ cliente/escritório

# ── v2.29 — Distribuição por E-mail (SMTP) + fatiamento por cliente ──────────
# Remetente/SMTP: 100% por variável de ambiente (NUNCA hardcode de senha).
# Default = Gmail (contato@ é Gmail-hosted). Para Outlook/Hotmail, defina
# SMTP_HOST=smtp-mail.outlook.com no Render.
SMTP_HOST             = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT             = int(os.environ.get('SMTP_PORT', '587'))
EMAIL_SENDER          = os.environ.get('EMAIL_SENDER', '')
EMAIL_SENDER_PASSWORD = os.environ.get('EMAIL_SENDER_PASSWORD', '')
EMAIL_SENDER_NOME     = os.environ.get('EMAIL_SENDER_NOME', 'Magnata Portaria e Serviços')

TABLE_EXTRATO = 'tblJCUcFBVTH5W2kP'   # Extratos Mensais
TABLE_FGTS    = 'tbl8ehgLa00cE1U3s'   # FGTS Digital
TABLE_GUIAS   = 'tbl6FT1YzK1yqI77l'   # Guias e Comprovantes

# Assinaturas Digitais — v2.38, fluxo de assinatura nativa (IP/CPF/timestamp)
TABLE_ASSINATURAS    = 'tbl6xgW45637YJISv'
F_ASS_NOME           = 'fld0CM8g3rv2JxmL0'   # Nome (primário)
F_ASS_TIPO_DOC        = 'fldCKWzoTPeVarUmx'   # Tipo de Documento
F_ASS_STATUS          = 'fldW7KJrXuU6hHYVC'   # Status (Pendente/Assinado/Expirado)
F_ASS_HASH            = 'fldXgApUSbEJciu9r'   # Hash Token
F_ASS_CPF_INFORMADO   = 'fldGILBsjU2g8uRgb'   # CPF Informado
F_ASS_IP              = 'fldaObfgNMqszuGvb'   # IP Captura
F_ASS_USER_AGENT      = 'fldMyC9l4BODOjyYx'   # User Agent
F_ASS_DATA_ASSINATURA = 'fldOxZqGzXLtMs0xs'   # Data/Hora Assinatura
F_ASS_DATA_GERACAO    = 'fld8XKWxQk0IwZM2o'   # Data Geração
F_ASS_PROCESSAR_ID    = 'fldEw0a0UOEH6O4DZ'   # Processar Arquivos ID (texto — não link, ver nota abaixo)
F_ASS_FUNCIONARIO     = 'fldG4sjQ0QkFVakAu'   # Funcionário (link)
F_ASS_TENTATIVAS      = 'fldXk0fVbrBH4xJnf'   # Tentativas (proteção contra força bruta nos 4 dígitos)
F_ASS_EVIDENCIAS      = 'fldbfJnLF82YAhe96'   # Evidencias_Assinatura (texto consolidado: IP+Timestamp+User-Agent)
F_ASS_DOCUMENTO_PDF   = 'fldJXYMo9JFcfc2yt'   # Documento PDF (cópia do Kit Admissão consolidado)

# ✅ v3.6: 10 campos para identidade documental, idempotência e rastreamento
# Criar manualmente no Airtable: Assinaturas Digitais (tblEhCPB48RTj2xjx)

# Campos obrigatórios (identidade + idempotência) — v3.6 ATIVO
F_ASS_ARQUIVO_RECORD_ID = 'fldd5qwaEEubavIpR'         # [singleLineText] Record ID do arquivo
F_ASS_PDF_SHA256 = 'fldJHbPQ9fjmAaaCX'                # [singleLineText] SHA-256 do PDF
F_ASS_CHAVE_IDEMPOTENCIA = 'fldJawyGqeBRtUMHu'        # [singleLineText] Chave idempotência SHA256
F_ASS_REQUEST_ID = 'fldGoqOhYrHWMttJz'                # [singleLineText] ID da requisição HTTP

# Campos de contexto
F_ASS_FINALIDADE = 'fldXXXXXXXXXXXXXX'                 # [singleSelect] Tipo de fluxo: ASSINATURA|REENVIO|LOTE|KIT_ADMISSAO|FOLHA_PONTO|FICHA_EPI|RESCISAO (PLACEHOLDER)
F_ASS_VERSAO_FLUXO = 'fldXXXXXXXXXXXXXX'               # [singleLineText] Versão algo (v2.x, v3.6, etc) (PLACEHOLDER)

# Campos de rastreamento
F_ASS_STATUS_ENVIO = 'fldXXXXXXXXXXXXXX'               # [singleSelect] PREPARADO|AGUARDANDO_ENVIO|ENVIANDO|ENVIADO|FALHA_ENVIO|CANCELADO_DUPLICADO (PLACEHOLDER)
F_ASS_TENTATIVAS_ENVIO = 'fldXXXXXXXXXXXXXX'           # [number] Contador de tentativas (PLACEHOLDER)
F_ASS_ULTIMO_ERRO_ENVIO = 'fldXXXXXXXXXXXXXX'          # [multilineText] Mensagem de erro (PLACEHOLDER)

# Campo de reconciliação
F_ASS_REGISTRO_DUPLICADO_DE = 'fldXXXXXXXXXXXXXX'      # [singleLineText] Record ID do vencedor (se cancelado) (PLACEHOLDER)

# ✅ v3.6: Validação central dos Field IDs de assinatura
def _validar_configuracao_assinatura_v36():
    """
    Valida os 4 Field IDs v3.6 ATUALMENTE EM USO.

    Detecta:
    - valores vazios
    - placeholders (XXXX)
    - ausência de prefixo 'fld'
    - duplicação de IDs
    - constantes indefinidas

    Campos OBRIGATÓRIOS (lidos/escritos no fluxo atual):
    - F_ASS_ARQUIVO_RECORD_ID
    - F_ASS_PDF_SHA256
    - F_ASS_CHAVE_IDEMPOTENCIA
    - F_ASS_REQUEST_ID

    Campos PLANEJADOS (não bloqueantes até serem usados):
    - F_ASS_FINALIDADE
    - F_ASS_STATUS_ENVIO
    - F_ASS_TENTATIVAS_ENVIO
    - F_ASS_ULTIMO_ERRO_ENVIO
    - F_ASS_REGISTRO_DUPLICADO_DE
    - F_ASS_VERSAO_FLUXO

    Retorna: (bool, mensagem_detalhada)
    """
    # Validar apenas os 4 campos atualmente em uso
    campos_obrigatorios = {
        'F_ASS_ARQUIVO_RECORD_ID': F_ASS_ARQUIVO_RECORD_ID,
        'F_ASS_PDF_SHA256': F_ASS_PDF_SHA256,
        'F_ASS_CHAVE_IDEMPOTENCIA': F_ASS_CHAVE_IDEMPOTENCIA,
        'F_ASS_REQUEST_ID': F_ASS_REQUEST_ID,
    }

    campos_v36 = campos_obrigatorios

    erros = []
    valores_vistos = {}

    for nome_campo, valor_id in campos_v36.items():
        # 1. Verificar vazio
        if not valor_id or valor_id.strip() == '':
            erros.append(f'{nome_campo}: VAZIO')
            continue

        # 2. Verificar placeholder
        if 'XXXX' in valor_id:
            erros.append(f'{nome_campo}: PLACEHOLDER ({valor_id})')
            continue

        # 3. Verificar formato fld...
        if not valor_id.startswith('fld'):
            erros.append(f'{nome_campo}: sem prefixo fld ({valor_id})')
            continue

        # 4. Verificar formato: fld seguido de caracteres alfanuméricos, mínimo 10 chars total
        import re as _re_validation
        if not _re_validation.match(r'^fld[A-Za-z0-9]+$', valor_id):
            erros.append(f'{nome_campo}: formato inválido ({valor_id}) - deve ser fld[A-Za-z0-9]+')
            continue

        if len(valor_id) < 10:
            erros.append(f'{nome_campo}: muito curto ({valor_id}) - mínimo 10 caracteres')
            continue

        # 5. Verificar duplicação
        if valor_id in valores_vistos:
            erros.append(f'{nome_campo}: duplicado (também em {valores_vistos[valor_id]})')
            continue

        valores_vistos[valor_id] = nome_campo

    if erros:
        mensagem_erro = '\n  '.join(erros)
        logger.error(f'[ASSINATURA v3.6] Configuracao INVALIDA:\n  {mensagem_erro}')
        return False, f'Assinatura v3.6 INVALIDA ({len(erros)} erro em campos obrigatórios):\n  {mensagem_erro}'

    return True, 'Assinatura v3.6 OK (4/4 campos obrigatórios válidos)'


# ✅ FAIL-FAST REAL NO STARTUP COM VALIDAÇÃO RIGOROSA DE MODO (linhas 260-300)
# Validar configuração na inicialização
_config_ok, _config_msg = _validar_configuracao_assinatura_v36()

# Modo de validação: APENAS 'prod' (padrão) ou 'test' (bypass explícito)
_validation_mode_raw = os.environ.get('ASSINATURA_CONFIG_VALIDATION_MODE', 'prod')
_validation_mode = _validation_mode_raw.lower().strip()

# Validar modo: aceitar APENAS 'prod' ou 'test'
if _validation_mode not in ('prod', 'test'):
    msg_modo_invalido = f"""
╔════════════════════════════════════════════════════════════════════════════╗
║ [FATAL] MODO DE VALIDACAO INVALIDO                                        ║
║                                                                            ║
║ ASSINATURA_CONFIG_VALIDATION_MODE = '{_validation_mode_raw}'              ║
║ (normalizado: '{_validation_mode}')                                        ║
║                                                                            ║
║ Valores aceitos:                                                           ║
║   prod  → aplicação normal (padrão)                                        ║
║   test  → bypass para testes (config inválida permitida)                  ║
║                                                                            ║
║ Qualquer outro valor interrompe a inicialização.                           ║
╚════════════════════════════════════════════════════════════════════════════╝
"""
    logger.critical(msg_modo_invalido)
    sys.exit(1)

if not _config_ok:
    if _validation_mode == 'prod':
        # PRODUÇÃO: bloqueio total
        msg_fatal = f"""
╔════════════════════════════════════════════════════════════════════════════╗
║ [FATAL] ASSINATURA v3.6 INICIACAO FALHOU                                  ║
║                                                                            ║
║ Configuração dos 4 Field IDs obrigatórios está inválida.                   ║
║ A aplicação NÃO pode iniciar.                                              ║
║                                                                            ║
║ Erros detectados:                                                          ║
║ {_config_msg.replace(chr(10), chr(10) + '║ ')}║
║                                                                            ║
║ Ação necessária:                                                           ║
║ 1. Criar os 4 campos obrigatórios no Airtable:                            ║
║    - Arquivo Record ID                                                     ║
║    - PDF SHA-256                                                           ║
║    - Chave de Idempotência                                                 ║
║    - Request ID                                                            ║
║ 2. Obter os Field IDs (fldXXXXXXXXXXXXXX)                                  ║
║ 3. Substituir os placeholders em app.py (linhas 177-180)                  ║
║ 4. Reiniciar a aplicação                                                   ║
╚════════════════════════════════════════════════════════════════════════════╝
"""
        logger.critical(msg_fatal)
        sys.exit(1)  # Bloqueio total - não prosseguir
    elif _validation_mode == 'test':
        # TESTE: apenas warning
        logger.warning(f'[ASSINATURA v3.6] Bypass TEST permitido: {_config_msg}')
else:
    logger.info(f'[ASSINATURA v3.6] {_config_msg}')

# Tipos de documento canônicos (constantes, não strings livres)
TIPOS_DOCUMENTO_VALIDOS = {
    'KIT_ADMISSAO',
    'FOLHA_PONTO',
    'FICHA_EPI',
    'RESCISAO',
    'CONTRATO_EXPERIENCIA',
    'CONTRATO_TRABALHO',
    'HOLERITE_FOLHA_PONTO',
}

# Nomes de exibição (mapa de canônicos para português)
NOMES_DOCUMENTOS = {
    'KIT_ADMISSAO': 'Kit de Admissão',
    'FOLHA_PONTO': 'Folha de Ponto',
    'FICHA_EPI': 'Ficha de EPI',
    'RESCISAO': 'Rescisão',
    'CONTRATO_EXPERIENCIA': 'Contrato de Experiência',
    'CONTRATO_TRABALHO': 'Contrato de Trabalho',
    'HOLERITE_FOLHA_PONTO': 'Holerite + Folha de Ponto',
}

# ─── PACOTE HOLERITE + FOLHA DE PONTO (extensão mínima, decisão arquitetural) ──
#
# Decisão (documentada, não presumida): o Holerite passa a ser assinável
# SOMENTE dentro de um pacote atômico junto com a Folha de Ponto da MESMA
# competência — nunca isolado. Isto evita reabrir os 65 registros
# "Holerite" órfãos encontrados na auditoria (tipo pré-v3.6, sem mais
# suporte) com o mesmo problema estrutural (Holerite sem vínculo formal
# de competência a outro documento). O pacote reaproveita 100% da tabela
# "Assinaturas Digitais" e os 4 campos v3.6 já reais (nenhum campo novo no
# Airtable, nenhuma rota nova):
#   - F_ASS_ARQUIVO_RECORD_ID passa a conter "<rec_holerite>|<rec_ponto>"
#   - F_ASS_PDF_SHA256        passa a conter "<sha_holerite>|<sha_ponto>"
# (convenção documentada aqui e no código; delimitador "|" nunca aparece
# em um Record ID do Airtable nem em um hex SHA-256, então é seguro).
TIPO_PACOTE_HOLERITE_PONTO = 'HOLERITE_FOLHA_PONTO'
_SEPARADOR_PACOTE = '|'

# Estados de assinatura (máquina de estados)
ESTADOS_ASSINATURA = {
    'PREPARADO': 'Record criado, aguardando envio',
    'AGUARDANDO_ENVIO': 'Pronto para enviar WhatsApp',
    'ENVIANDO': 'Tentativa de envio em andamento',
    'ENVIADO_AGUARDANDO_ASSINATURA': 'Link enviado, aguardando assinatura',
    'ASSINADO': 'Documento assinado e comprovado',
    'FALHA_ENVIO': 'Falha no envio WhatsApp (permite retry)',
    'EXPIRADO': 'Prazo de assinatura expirado',
    'CANCELADO': 'Assinatura cancelada',
}

MAX_TENTATIVAS_ASSINATURA = 5   # após isso, marca Status="Expirado"

CNPJ_MAGNATA = '17987187000161'   # CNPJ do empregador — nunca é cliente-destino

# Extratos Mensais
F_EXT_STATUS  = 'fldozFq8FJAOZMqns'
F_EXT_FOLHA   = 'fldDqvqtFqN0U5UBD'
F_EXT_CLIENTE = 'fldKtdZpZ4fd7XpAX'
F_EXT_PDF     = 'fldznv1E24rfbZt34'
F_EXT_DATA    = 'fldZ3q923bEAqUekZ'
# FGTS Digital
F_FGTS_STATUS  = 'fldNb2ueeOHjyRi65'
F_FGTS_CLIENTE = 'fldGFwcySH5TXBjDB'
F_FGTS_FOLHA   = 'fld22SuzevUvtaMkg'
F_FGTS_PDF     = 'fldYZS5KB9yKK4lMH'
F_FGTS_DATA    = 'fldoFNd8CgL5ZZ63C'
# Guias e Comprovantes (documentos COMUNS — broadcast)
F_GUIA_STATUS  = 'fldwOQfBzyH7rVx2O'
F_GUIA_TIPO    = 'fldZc4A6stiQPI8qt'
F_GUIA_PDF     = 'fldmC813USDUA3eVl'   # PDF GUIA
F_GUIA_COMPROV = 'fldBSFBpAUvJFMzbH'   # PDF COMPROVANTE (comprovantes de pagamento, VR/VA…)
F_GUIA_NOME    = 'fldOLOkac6twvlfZR'   # Nome documento
# Benefícios (Clientes) — v3.12: Horas Extras/Assiduidade (Sicoob, fatiado
# por colaborador e fundido por cliente) e VR/VA (lote via Excel+comprovante
# agregado, ou PIX individual de colaborador novo sem cartão ainda).
F_CLI_NOME          = 'fldx6AShzJCfQeBMa'
F_CLI_HORAS_EXTRAS  = 'fldmnmpHN2toqaQkr'
F_CLI_ASSIDUIDADE   = 'fldIW9QdiKDYeDAEZ'
F_CLI_VRVA          = 'fldk9aUnVOXlWUHDr'
F_CLI_ALMOCO_JANTA  = 'fldUyDDKQqytLgCkD'
F_CLI_DIARIAS       = 'fldjcLYeyOg9CvAe3'
# Envios — campos de e-mail (link p/ os documentos do pacote)
F_ENVIO_TEXTO    = 'fldGeaadeNYWaa4Og'   # Texto do Email
F_ENVIO_EXTRATO  = 'fldITxIfLG8TKntBu'   # Extrato Mensal (link)
F_ENVIO_FGTS     = 'fldnsrSDCpK3YAXbm'   # FGTS Digital (link)
F_ENVIO_GUIAS    = 'fld0Pm44oScZX7yRx'   # Guias e Comprovantes (link)
F_ENVIO_CERTIDOES = 'fldJAcWLvXyTkkruw'  # Certidões (link) — v2.31

# URL pública (p/ o link de Recibo Digital de Leitura no WhatsApp — comprovação)
RECIBO_BASE_URL = os.environ.get('PUBLIC_BASE_URL', 'https://magnata-holerite-splitter.onrender.com').rstrip('/')

# Nomes (não ids) dos campos lookup de anexos na Envios — usados p/ montar o
# e-mail. Aceitam qualquer tipo de arquivo (PDF, XLS, etc.). "PDF COMPROVANTES"
# cobre comprovantes de pagamento (FGTS, DCTFWeb, VR/VA — broadcast).
ENVIO_LOOKUP_PDFS = [
    'PDF HOLERITES', 'PDF EXTRATO MENSAL', 'PDF FGTS Digital', 'PDF GUIAS',
    'PDF COMPROVANTES', 'PDF CERTIDÕES', 'PDF FERIAS', 'PDF CONTRATAÇÃO - RESCISÃO',
    'Arquivos',
]

# Emails Savian
F_EMAIL_NAME     = 'fldwKHHiVVEKySwx4'
F_EMAIL_STATUS   = 'fld44SrJN9Va8avMX'
F_EMAIL_ASSUNTO  = 'fld66diI0hksJE5PS'
F_EMAIL_CONTEUDO = 'fldzi2kWBoT2kfEhL'
F_EMAIL_MSGID    = 'fldCCdUEMF3hlTngA'

# Arquivos
F_ARQ_NOME     = 'fldjVGYri7DZDJuee'   # Arquivo (campo primário)
F_ARQ_STATUS   = 'fld9yxb30SlLJxqoM'
F_ARQ_DATA     = 'fldKcfvu5Anec54xa'
F_ARQ_ATTACH   = 'fldm6S1xnp8S6sKFE'
F_ARQ_EMAILS   = 'fld2yYAHWe0smV5Bb'
F_ARQ_NOME_ARQ = 'fldsOySQRfZ8rPGDw'   # Nome do Arquivo
F_ARQ_HASH     = 'fldOB09YlKDEqKSFO'  # Hash do Anexo

# Processar Arquivos
F_PROC_NAME      = 'fldmrG1ZTHHU4QYQK'
F_PROC_STATUS    = 'fldvN9T5MiuKZGDi0'
F_PROC_DATA      = 'flddNzmqp1Im1D02m'
F_PROC_ARQUIVOS  = 'fldQtevv6jAwKVdEN'  # Campo "Arquivos" (multipleAttachments) — localizado via API 2026-07-20
F_PROC_ARQUIVOS2 = 'fldLWSmK81i8jbtCG'  # Vínculo (não usar para upload)
F_PROC_TIPO_DOC  = 'fldvkOVlwCMywGTES'
F_PROC_ERROR_CODE = 'fldYYYYYYYYYYYYYY'  # ⚠️ OPCIONAL: campo para error_code (se existir)

# Pendências/Revisar
F_PEND_NOME   = 'fldovcs6bySCshoXI'
F_PEND_STATUS = 'fldf1an8HCV2DxEwk'
F_PEND_TIPO   = 'fldyZgyB5F5fv6kUX'
F_PEND_ORIGEM = 'fldUk3hr2mCkfu1Wb'
F_PEND_OBS    = 'fld2bqGLlotCVRBn5'
F_PEND_DATA   = 'fldRolmP0rSbJevUZ'

EMAIL_WEBHOOK_KEY = os.environ.get('EMAIL_WEBHOOK_KEY', '')

# Caixa fiscal (v2.96) — Recibos/Guias/Extrato mensal de competência tributária.
# Remetente adicionado à lista de confiança do Apps Script (REMETENTES_CONFIAVEIS)
# sem alterar o e-mail já existente do Departamento Pessoal. Documentos vindos
# deste remetente são roteados no /email/webhook para as tabelas de Extrato/
# FGTS/Guias já usadas na distribuição a clientes, em vez do fluxo de
# Processar Arquivos do DP (Kit Admissão etc.), que não se aplica a eles.
REMETENTE_FISCAL = 'dpfiscal.contabilidade2@hotmail.com'

# Tipos fiscais/mensais roteados via _processar_anexo_fiscal — o roteamento
# NÃO depende só do remetente ser REMETENTE_FISCAL: achado real em produção
# (10/07/2026) mostrou uma Guia de FGTS chegando por outro remetente
# confiável (não o fiscal dedicado), que teria caído no fluxo de Processar
# Arquivos/Kit Admissão do DP se dependesse só do remetente. Documento
# classificado como um destes tipos vai para a tabela certa não importa
# de qual caixa monitorada ele veio.
TIPOS_FISCAIS = {
    'Extrato da Folha de Pagamento', 'FGTS',
    'DCTFWeb - Recibo de Entrega', 'DCTFWeb - Declaração',
}

# Evolution API (gateway WhatsApp) — v2.28. URL e instância não são segredo
# (default embutido); a API KEY é secreta e DEVE vir por variável de ambiente.
EVOLUTION_API_URL  = os.environ.get('EVOLUTION_API_URL', 'http://143.95.214.239:8080').rstrip('/')
EVOLUTION_INSTANCE = os.environ.get('EVOLUTION_INSTANCE', 'magnata')
EVOLUTION_API_KEY  = os.environ.get('EVOLUTION_API_KEY', '')

# Limite estrito para o conteudo binario aceito pela rota dedicada de video.
# Mantido abaixo do limite global do Flask para evitar payloads excessivos na
# memoria e para produzir arquivos adequados ao transporte pelo WhatsApp.
WHATSAPP_VIDEO_MAX_BYTES = 15 * 1024 * 1024
# Mesmo raciocínio para a rota dedicada de PDF direto (sem Airtable/assinatura).
WHATSAPP_DOCUMENTO_MAX_BYTES = 15 * 1024 * 1024

# Regras de classificação de documento (Fase 2)
# Lista de (tipo_documento, [regex de palavras-chave]) — primeira que casar vence
TIPO_DOC_REGRAS = [
    # Rescisão — precisa vir ANTES de Holerite e de Contrato de Trabalho/
    # Experiência. Achado real em produção (05/07/2026): um TERMO DE
    # RESCISÃO DO CONTRATO DE TRABALHO (TRCT) genuíno contém "Valor
    # Líquido" (discriminação das verbas rescisórias), que também bate com
    # o padrão genérico de Holerite. Como Holerite vinha ANTES nesta lista
    # e o classificador para no primeiro tipo com hit, o TRCT foi
    # classificado como Holerite — mesmo tendo "TRCT" e "Termo de Rescisão"
    # no texto, esses padrões nunca chegaram a ser testados. Rescisão
    # também precisa vir antes de Contrato de Trabalho/Experiência pelo
    # mesmo motivo (o texto de uma rescisão quase sempre cita "contrato de
    # trabalho").
    # NOTA: 'Líquido rescisão' e 'Férias Rescisão' foram REMOVIDOS destes
    # padrões (estavam aqui como "variações reais observadas em produção").
    # Achado real em produção (05/07/2026): são apenas rubricas de linha
    # dentro de holerites (ex.: "1/3 FERIAS RESCISAO", "LIQUIDO RESCISAO"),
    # que aparecem normalmente em lotes de holerite mensal de colaboradores
    # desligados no período — um lote de 110 páginas de holerites reais foi
    # classificado erroneamente como Rescisão por causa desses 2 padrões.
    # Os padrões restantes (TRCT, Termo/Aviso/Homologação/Cálculo de
    # Rescisão, Motivo/Data demissão) são específicos o suficiente de um
    # documento de rescisão de verdade e não geram esse falso positivo.
    ('Rescisão', [r'Termo\s+de\s+Rescis[ãa]o', r'Aviso\s+de\s+Rescis[ãa]o',
                   r'\bTRCT\b', r'Rescis[ãa]o\s+(?:do\s+)?Contrato\s+de\s+Trabalho',
                   r'Homologa[çc][ãa]o\s+(?:de\s+)?Rescis[ãa]o',
                   r'C[áa]lculo\s+de\s+Rescis[ãa]o',
                   r'Motivo\s+demiss[ãa]o', r'Data\s+demiss[ãa]o',
                   r'Data\s+de\s+demiss[ãa]o']),
    # Extrato da Folha de Pagamento (caixa fiscal, v2.96) — precisa vir ANTES
    # de Holerite: é um relatório-resumo da folha (não o holerite individual),
    # mas costuma citar "Total de Vencimentos"/"Valor Líquido" também.
    # "Extrato Mensal" foi adicionado depois de um caso real (10/07/2026) em
    # que o contador manda o documento com esse nome, sem a frase completa
    # "Extrato da Folha de Pagamento" — caiu como "Outro"/Pendência.
    ('Extrato da Folha de Pagamento',
        [r'Extrato\s+(?:da\s+)?Folha\s+de\s+Pagamento', r'Extrato\s+Mensal\b']),
    # FGTS/DCTFWeb (caixa fiscal, v2.96) — precisam vir ANTES de Holerite,
    # EPI, Ficha de Registro, Contrato de Trabalho/Experiência e Férias.
    # Achado real em produção (10/07/2026): a "Guia Emitida" do FGTS Digital
    # (98 trabalhadores) trouxe no rodapé o glossário de categorias FGTS, que
    # cita literalmente "Contrato de trabalho Verde e Amarelo" várias vezes —
    # como "Contrato de Trabalho" vinha ANTES na lista, o documento (que é
    # 100% FGTS: "Detalhe da Guia Emitida", "FGTS Mensal na Guia", "Qtd.
    # Trabalhadores FGTS") foi classificado errado como Contrato de Trabalho,
    # e as regras de FGTS (que exigiam "FGTS Digital"/"Guia do FGTS"/"GFD" —
    # nenhuma dessas frases aparece nesse documento real) nunca chegaram a
    # ser testadas. Recibo de Entrega do DCTFWeb precisa vir antes da
    # Declaração genérica, senão o \bDCTFWeb\b da Declaração sempre vence
    # primeiro e o recibo nunca é distinguido.
    ('DCTFWeb - Recibo de Entrega',
        [r'Recibo\s+de\s+Entrega.{0,60}DCTFWeb', r'DCTFWeb.{0,60}Recibo\s+de\s+Entrega']),
    ('Guia DCTFWeb/DARF',
        [r'Guia\s+(?:de\s+Recolhimento\s+)?(?:da\s+)?DCTFWeb\b',
         r'DCTFWeb\s*[-–—:/]?\s*DARF\b',
         r'DARF\s*[-–—:/]?(?:\s+da)?\s+DCTFWeb\b']),
    ('DCTFWeb - Declaração', [r'\bDCTFWeb\b']),
    ('FGTS', [r'FGTS\s+Digital', r'Guia\s+do\s+FGTS', r'\bGFD\b',
              r'Detalhe\s+da\s+Guia\s+Emitida', r'FGTS\s+Mensal\s+na\s+Guia',
              r'Qtd\.?\s*Trabalhadores\s+FGTS', r'Total\s+FGTS\b']),
    ('Holerite', [r'Recibo\s+de\s+Pagamento', r'Total\s+de\s+Vencimentos', r'Valor\s+L[íi]quido']),
    ('Folha de Ponto', [r'Folha\s+de\s+Ponto', r'Espelho\s+de\s+Ponto',
                         r'Cart[ãa]o\s+(?:de\s+)?Ponto', r'Secullum',
                         r'Ponto\s+Web']),
    # EPI — checado antes dos tipos de contrato; padrão de Ficha de EPI não
    # colide com "Ficha de Registro de Empregado" (regras distintas abaixo).
    ('EPI', [r'Ficha\s+de\s+(?:Controle\s+de\s+)?EPI\b',
              r'Equipamento\s+de\s+Prote[çc][ãa]o\s+Individual',
              r'Ficha\s+de\s+EPI[\'’]?s', r'\bEPI\b',
              r'Recebi\s+(?:o|os|do|dos)\s+(?:seguintes\s+)?equipamentos?']),
    # Termo de Prorrogação de Contrato de Experiência — precisa vir ANTES de
    # "Contrato de Experiência" pela mesma razão (o termo de prorrogação
    # sempre menciona "contrato de experiência" no corpo do texto).
    ('Termo de Prorrogação de Contrato de Experiência',
        [r'Prorroga[çc][ãa]o\s+(?:do\s+)?Contrato\s+de\s+Experi[êe]ncia',
          r'Termo\s+(?:Aditivo\s+)?de\s+Prorroga[çc][ãa]o']),
    # Ficha de Registro de Empregado — precisa vir ANTES de "Contrato de
    # Trabalho", pois a ficha de registro cita o número da CTPS (que também
    # casaria com a regra \bCTPS\b de Contrato de Trabalho abaixo).
    ('Ficha de Registro de Empregado',
        [r'Ficha\s+de\s+Registro\s+de\s+Empregados?', r'Registro\s+de\s+Empregados?\b.{0,40}Ficha',
          r'Livro\s+(?:de\s+)?Registro\s+de\s+Empregados?',
          # Variação real observada em produção: título "REGISTRO DE
          # EMPREGADO" isolado (sem "Ficha de" antes), formulário de eSocial.
          r'Registro\s+de\s+Empregados?\b', r'Matr[íi]cula\s+eSocial']),
    ('Contrato de Experiência', [r'Contrato\s+de\s+Experi[êe]ncia']),
    # "Contrato de Trabalho" e "Kit de Admissão" NÃO se misturam com
    # fechamentos mensais de folha/extratos/FGTS: este tipo é exclusivamente
    # o documento de admissão que o contador envia (nome dele mesmo,
    # "Contrato de Trabalho"); "Kit de Admissão" é o nome interno do pacote
    # consolidado (Ficha + Contrato + Termo de VT) — nenhum dos dois deve
    # capturar documentos fiscais/mensais (ver nota de FGTS acima).
    ('Contrato de Trabalho', [r'Contrato\s+de\s+Trabalho', r'\bCTPS\b']),
    ('Férias', [r'Aviso\s+de\s+F[ée]rias', r'Recibo\s+de\s+F[ée]rias', r'Per[íi]odo\s+de\s+Gozo']),
    ('Guia', [r'Guia\s+de\s+Recolhimento', r'\bGPS\b', r'\bDARF\b']),
    ('Boleto', [r'\d{5}\.\d{5}\s+\d{5}\.\d{6}\s+\d{5}\.\d{6}\s+\d\s+\d{14}', r'Linha\s+Digit[áa]vel']),
    ('Nota Fiscal', [r'NFS-?e', r'Nota\s+Fiscal\s+de\s+Servi[çc]o', r'DANFE']),
]


def classificar_documento(texto: str):
    """Retorna (tipo_documento, confianca) com base em palavras-chave no texto extraído."""
    for tipo, padroes in TIPO_DOC_REGRAS:
        hits = sum(1 for p in padroes if re.search(p, texto, re.IGNORECASE))
        if hits > 0:
            return tipo, hits
    return 'Outro', 0


MESES_PT = [
    'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
]

# ── Rate limiter Airtable ─────────────────────────────────────────────────────
_last_at_call = 0.0


def _at_throttle():
    global _last_at_call
    now = time.monotonic()
    gap = now - _last_at_call
    if gap < 0.23:
        time.sleep(0.23 - gap)
    _last_at_call = time.monotonic()


def _at_headers():
    return {
        'Authorization': f'Bearer {AIRTABLE_API_KEY}',
        'Content-Type': 'application/json'
    }


# ── Erros Airtable (Fase 6 — diagnóstico/sanitização) ────────────────────────

class AirtableError(Exception):
    """
    Erro de gravação/leitura no Airtable que preserva o motivo exato
    retornado pela API (status HTTP + corpo da resposta), em vez de deixar
    apenas a mensagem genérica de requests.raise_for_status() ("422 Client
    Error: Unprocessable Entity for url: ...").
    """

    def __init__(self, status_code: int, body_text: str, context: str = ''):
        self.status_code = status_code
        self.body_text = body_text or ''
        self.context = context
        try:
            self.error_json = json.loads(body_text) if body_text else None
        except ValueError:
            self.error_json = None
        super().__init__(self._mensagem())

    def _mensagem(self) -> str:
        detalhe = None
        if isinstance(self.error_json, dict):
            err = self.error_json.get('error')
            if isinstance(err, dict):
                detalhe = f"{err.get('type', '')}: {err.get('message', '')}".strip(': ')
            elif isinstance(err, str):
                detalhe = err
        if not detalhe:
            detalhe = self.body_text[:300]
        prefixo = f'[{self.context}] ' if self.context else ''
        return f'{prefixo}Airtable HTTP {self.status_code}: {detalhe}'


def _raise_for_airtable(r, context: str = ''):
    """Loga e levanta AirtableError com o erro exato retornado pela API, se a resposta não for 2xx."""
    if not r.ok:
        logger.error(f'[AT] {context} HTTP {r.status_code}: {r.text[:1000]}')
        raise AirtableError(r.status_code, r.text, context)


# ── Helpers PDF ───────────────────────────────────────────────────────────────

def extrair_cpf(texto: str):
    if not texto:
        return None
    linhas = texto.split('\n')
    for i, linha in enumerate(linhas):
        if 'CPF' in linha:
            m = re.search(r'\d{3}\.\d{3}\.\d{3}-\d{2}', linha)
            if m:
                return m.group()
            if i + 1 < len(linhas):
                m = re.search(r'\d{3}\.\d{3}\.\d{3}-\d{2}', linhas[i + 1])
                if m:
                    return m.group()
    m = re.search(r'\d{3}\.\d{3}\.\d{3}-\d{2}', texto)
    return m.group() if m else None


def _mascarar_cpf(cpf: str) -> str:
    """Blindagem LGPD para log/exibição: nunca imprimir o CPF completo.
    Mantém só o último grupo (3 dígitos) + dígitos verificadores, no
    formato ***.***.123-45 -- suficiente para correlacionar um caso
    específico num log sem expor o documento inteiro. Entrada fora do
    formato esperado (vazia, incompleta) nunca é ecoada como está --
    volta um placeholder igualmente seguro."""
    digitos = re.sub(r'\D', '', cpf or '')
    if len(digitos) != 11:
        return '***.***.***-**'
    return f'***.***.{digitos[6:9]}-{digitos[9:11]}'


def extrair_nome_funcionario(texto: str):
    if not texto:
        return 'Desconhecido'
    linhas = texto.split('\n')
    for i, linha in enumerate(linhas):
        if 'Nome do Funcionário' in linha or 'Nome do Funcionario' in linha:
            if i + 1 < len(linhas):
                proxima = linhas[i + 1].strip()
                partes = proxima.split()
                if partes and partes[0].isdigit():
                    nome_partes = []
                    for p in partes[1:]:
                        if re.match(r'^\d{5,6}$', p):
                            break
                        nome_partes.append(p)
                    if nome_partes:
                        return ' '.join(nome_partes)
    # Formato B (Folhas Manuais de Ponto, v2.99): sem CPF impresso — só
    # "Colaborador: NOME" no cabeçalho. Usado como fallback quando o padrão
    # "Nome do Funcionário" (holerites/Secullum) não bate.
    for linha in linhas:
        m = re.search(r'Colaborador:\s*(.+)', linha, re.IGNORECASE)
        if m:
            nome = m.group(1).strip()
            if nome:
                return nome
    return 'Desconhecido'


# Linha de dia do "Cartão Ponto" (Secullum Ponto Web), ex.:
#   "29/04/26 - Qua - C1 18:56 01:00 01:53 09:05"
#   "28/04/26 - Ter - C2 FOLGA FOLGA FOLGA FOLGA FOLGA FOLGA"
_LINHA_CARTAO_PONTO_RE = re.compile(
    r'^(\d{2}/\d{2}/\d{2})\s*-\s*([A-Za-zÀ-ú]{3})\s*-\s*(C\d)\s+(.*)$'
)
_HORA_RE = re.compile(r'\d{2}:\d{2}')
_PERIODO_CARTAO_PONTO_RE = re.compile(
    r'Per[íi]odo:\s*(\d{2}/\d{2}/\d{4})\s*at[ée]\s*(\d{2}/\d{2}/\d{4})'
)


def extrair_cartao_ponto(texto: str) -> dict:
    """
    Fase 2 — Folha de Ponto / "Cartão Ponto" (ainda NÃO usado em produção).

    Extrai os dias de um relatório "Cartão Ponto" (formato Secullum Ponto
    Web, identificado a partir de PDF real). Cada dia aparece numa linha como:
      "29/04/26 - Qua - C1 18:56 01:00 01:53 09:05"
      "28/04/26 - Ter - C2 FOLGA FOLGA FOLGA FOLGA FOLGA FOLGA"

    Retorna:
      {
        'cpf': '123.456.789-00' | None,
        'periodo_inicio': 'dd/mm/aaaa' | None,
        'periodo_fim': 'dd/mm/aaaa' | None,
        'dias': [
          {'data': 'dd/mm/aaaa', 'dia_semana': 'Qua', 'ciclo': 'C1',
           'folga': False,
           'entrada_1': 'HH:MM'|None, 'saida_1': 'HH:MM'|None,
           'entrada_2': 'HH:MM'|None, 'saida_2': 'HH:MM'|None,
           'entrada_3': 'HH:MM'|None, 'saida_3': 'HH:MM'|None},
          ...
        ],
      }

    Esta função é independente do restante do pipeline — precisa ser
    revisada com mais exemplos (ex.: linhas com (*) batida manual, ('¨')
    abono parcial, ('^') pré-assinalado) antes de ser usada em
    _processar_folha_ponto.
    """
    cpf = extrair_cpf(texto)

    periodo_inicio = periodo_fim = None
    m = _PERIODO_CARTAO_PONTO_RE.search(texto)
    if m:
        periodo_inicio, periodo_fim = m.group(1), m.group(2)

    dias = []
    for linha in texto.split('\n'):
        linha = linha.strip()
        m = _LINHA_CARTAO_PONTO_RE.match(linha)
        if not m:
            continue
        data_2dig, dia_semana, ciclo, resto = m.groups()
        dia, mes, ano2 = data_2dig.split('/')
        data = f'{dia}/{mes}/20{ano2}'

        folga = 'FOLGA' in resto.upper()
        horarios = [] if folga else _HORA_RE.findall(resto)
        horarios = (horarios + [None] * 6)[:6]

        dias.append({
            'data': data,
            'dia_semana': dia_semana,
            'ciclo': ciclo,
            'folga': folga,
            'entrada_1': horarios[0],
            'saida_1':   horarios[1],
            'entrada_2': horarios[2],
            'saida_2':   horarios[3],
            'entrada_3': horarios[4],
            'saida_3':   horarios[5],
        })

    return {
        'cpf': cpf,
        'periodo_inicio': periodo_inicio,
        'periodo_fim': periodo_fim,
        'dias': dias,
    }


def parse_br_float(s: str):
    """Converte string no formato brasileiro '1.234,56' para float."""
    try:
        return float(s.replace('.', '').replace(',', '.'))
    except Exception:
        return None


def extrair_valores_holerite(texto: str) -> dict:
    """
    Extrai valores financeiros do texto de um holerite.

    Formato esperado no PDF:
      ...
      998 INSS M2  7,94  181,79
      Total de Vencimentos  Total de Descontos
      181,79

      Valor Líquido  2.108,34
      ...

    Retorna dict com chaves (todas opcionais):
      total_descontos, valor_liquido, total_vencimentos, inss
    """
    if not texto:
        return {}

    resultado = {}

    # 1. Valor Líquido — "Valor Líquido 2.108,34" (pode ser 0,00 para afastados INSS)
    m = re.search(
        r'Valor\s+L[íi]quido\s+([\d.]+,\d{2})',
        texto, re.IGNORECASE
    )
    if m:
        v = parse_br_float(m.group(1))
        if v is not None:
            resultado['valor_liquido'] = v

    # 2. Total Descontos — primeiro número após a linha dupla de totais
    #    "Total de Vencimentos  Total de Descontos\n182,55"
    m = re.search(
        r'Total\s+de\s+Vencimentos\s+Total\s+de\s+Descontos\s*\n\s*([\d.]+,\d{2})',
        texto, re.IGNORECASE
    )
    if m:
        v = parse_br_float(m.group(1))
        if v is not None:
            resultado['total_descontos'] = v
    else:
        # fallback: procura "Total de Descontos" seguido do valor na mesma ou próxima linha
        m = re.search(
            r'Total\s+de\s+Descontos\s+([\d.]+,\d{2})',
            texto, re.IGNORECASE
        )
        if m:
            v = parse_br_float(m.group(1))
            if v is not None:
                resultado['total_descontos'] = v

    # 3. Total Vencimentos = Total Descontos + Valor Líquido
    if 'total_descontos' in resultado and 'valor_liquido' in resultado:
        resultado['total_vencimentos'] = round(
            resultado['total_descontos'] + resultado['valor_liquido'], 2
        )

    # 4. INSS — linha com código INSS (geralmente 998 INSS M2 <ref> <valor>)
    #    pega o ÚLTIMO número monetário da linha que contém "INSS"
    for linha in texto.split('\n'):
        if re.search(r'\bINSS\b', linha, re.IGNORECASE):
            # só considera linhas de item (começam com dígito = código)
            if re.match(r'^\d', linha.strip()):
                numeros = re.findall(r'[\d.]+,\d{2}', linha)
                if numeros:
                    v = parse_br_float(numeros[-1])
                    if v is not None and v > 0:
                        resultado['inss'] = v
                    break

    return resultado


def _cpf_digitos_validos(cpf_num: str) -> bool:
    """Valida CPF (11 dígitos) pelo algoritmo dos dígitos verificadores."""
    if len(cpf_num) != 11 or cpf_num == cpf_num[0] * 11:
        return False
    for i in (9, 10):
        soma = sum(int(cpf_num[n]) * (i + 1 - n) for n in range(i))
        digito = (soma * 10 % 11) % 10
        if digito != int(cpf_num[i]):
            return False
    return True


# Padrões que indicam que o posto de trabalho é a própria empresa/sede,
# não um cliente/posto externo.
PADROES_LOCAL_INTERNO = [
    r'pr[óo]pria\s+empresa',
    r'sede\s+da\s+(?:empresa|contratante|magnata)',
    r'(?:o\s+)?mesmo\s+da\s+empresa',
    r'mesmo\s+endere[çc]o\s+da\s+(?:empresa|contratante)',
    r'nas\s+depend[êe]ncias\s+da\s+(?:empresa|contratante)',
    r'na\s+sede\s+d[aeo]\s+(?:empresa|contratante|magnata)',
    r'instala[çc][õo]es\s+da\s+(?:empresa|contratante)',
]

# Trechos genéricos/sem sentido isolado — se forem o único conteúdo
# extraído para o local, tratamos como "desconhecido" em vez de um posto real.
FILLERS_LOCAL_DESCONHECIDO = {
    'situa-se', 'o mesmo', 'a mesma', 'conforme', 'acima', 'abaixo',
    'mencionado', 'descrito', 'indicado', 'especificado',
}


def _resolver_local_posto(texto: str):
    """
    Determina local_posto / tipo_local / avisos para o posto de trabalho.

    "Magnata / Sede" é um local VÁLIDO (comum para folguistas e colaboradores
    sem posto fixo) — não é tratado como erro nem reduz confiança.

    Retorna (local_posto, tipo_local, avisos_qualidade):
      - ("Magnata / Sede", "interno_empresa", [])      -> posto é a própria empresa/sede (válido)
      - ("Edifício Sky", "externo", [])                -> posto externo identificado
      - (None, "desconhecido", [aviso])                -> truncado, ambíguo, conflito ou indeterminado
      - (None, "desconhecido", [])                     -> nada encontrado no texto
    """
    avisos = []

    interno = any(re.search(p, texto, re.IGNORECASE) for p in PADROES_LOCAL_INTERNO)

    # Tenta extrair um candidato a posto externo
    m = re.search(
        r'(?:local de trabalho|prestar[áa]?\s+(?:os\s+)?servi[çc]os?\s+(?:em|no|na))\s*:?\s*'
        r'([A-Za-zÀ-ú0-9\s,./-]{2,80}?)\s*[,\.\n]',
        texto, re.IGNORECASE
    )
    candidato = None
    if m:
        bruto = re.sub(r'\s+', ' ', m.group(1)).strip(' .,-')
        # Remove prefixos de localização que não fazem parte do nome do posto
        bruto = re.sub(
            r'^(?:que\s+)?(?:situa-se|localiza-se|estabelecido\s+em|situado\s+em)\s*',
            '', bruto, flags=re.IGNORECASE
        ).strip(' .,-')
        candidato_valido = (
            len(bruto) >= 4
            and bruto.lower() not in FILLERS_LOCAL_DESCONHECIDO
            and re.search(r'[A-Za-zÀ-ú]{3,}', bruto)
            and not any(re.search(p, bruto, re.IGNORECASE) for p in PADROES_LOCAL_INTERNO)
        )
        if candidato_valido:
            candidato = bruto
        elif not interno:
            # Trecho encontrado, mas truncado/genérico demais para ser um posto real
            avisos.append(
                f'Local de trabalho extraído de forma truncada/confusa: "{bruto}" — revisar manualmente.'
            )

    if interno and candidato:
        # Contrato menciona "própria empresa/sede" E um posto externo nomeado —
        # o sistema não consegue determinar qual prevalece.
        avisos.append(
            f'Conflito entre indicação de posto externo ("{candidato}") e indicação de '
            'sede/própria empresa no contrato — revisar manualmente.'
        )
        return None, 'desconhecido', avisos

    if interno:
        return 'Magnata / Sede', 'interno_empresa', avisos

    if candidato:
        return candidato, 'externo', avisos

    return None, 'desconhecido', avisos


def extrair_dados_contrato(texto: str):
    """
    Extrai dados de um Contrato de Experiência / Contrato de Trabalho (Fase 5B).

    Heurísticas por regex, tolerantes — qualquer campo não encontrado retorna None.
    Não levanta exceção: na ausência de casamento, simplesmente omite o valor.

    Retorna (dados, avisos_qualidade):
      dados: dict com chaves (todas opcionais)
        nome_funcionario, cpf, data_admissao, cargo_funcao, local_posto,
        tipo_local, salario, jornada_escala, cnpj_empresa
      avisos_qualidade: lista de strings com alertas sobre extrações
        confusas/truncadas (não bloqueiam o resultado, apenas sinalizam).
    """
    resultado = {
        'nome_funcionario': None,
        'cpf': None,
        'data_admissao': None,
        'cargo_funcao': None,
        'local_posto': None,
        'tipo_local': 'desconhecido',
        'salario': None,
        'jornada_escala': None,
        'cnpj_empresa': None,
    }
    avisos_qualidade = []
    if not texto:
        return resultado, avisos_qualidade

    # CPF — reaproveita o extrator já validado para holerites
    resultado['cpf'] = extrair_cpf(texto)

    # CNPJ da empresa contratante
    m = re.search(r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}', texto)
    if m:
        resultado['cnpj_empresa'] = m.group()

    # Nome do funcionário/empregado — padrões comuns em contratos.
    # Captura só letras/espaços (sem quebra de linha) e para antes de termos
    # como "portador", "carteira de trabalho", "CTPS", "CPF", "RG", etc.
    _stop_nome = (
        r'(?:\n|,|\.|portador|portadora|carteira\s+de\s+trabalho|CTPS|'
        r'CPF|RG|residente|doravante)'
    )
    padroes_nome = [
        rf'(?:EMPREGAD[OA]|CONTRATAD[OA])\s*:?\s*\n?\s*([A-ZÀ-Ú][A-ZÀ-Ú ]{{3,59}}?)(?=\s*{_stop_nome}|$)',
        rf'Sr\.?\(?a?\)?\.?\s+([A-ZÀ-Ú][A-ZÀ-Ú ]{{3,59}}?),\s+(?:portador|inscrit[oa]|CPF)',
        rf'(?:nome|funcion[áa]rio)\s*:?\s*([A-ZÀ-Ú][A-ZÀ-Ú ]{{3,59}}?)(?=\s*{_stop_nome}|$)',
    ]
    for p in padroes_nome:
        m = re.search(p, texto, re.IGNORECASE)
        if m:
            nome = m.group(1).strip()
            if len(nome.split()) >= 2:
                resultado['nome_funcionario'] = nome
                break

    # Data de admissão
    padroes_data = [
        r'(?:admiss[ãa]o|admitid[oa]|a partir de)\D{0,20}(\d{2}/\d{2}/\d{4})',
        r'Admiss[ãa]o:\s*(\d{2}/\d{2}/\d{4})',
    ]
    for p in padroes_data:
        m = re.search(p, texto, re.IGNORECASE)
        if m:
            resultado['data_admissao'] = m.group(1)
            break

    # Cargo/função — para na pontuação OU em conectores/sufixos que indicam
    # que o texto deixou de falar do cargo (ex.: "..., doravante denominado...")
    m = re.search(
        r'fun[çc][ãa]o\s+(?:de|do cargo de)?\s*:?\s*([A-Za-zÀ-ú0-9\s/\-]{3,60}?)'
        r'(?:\s*[,\.\n]'
        r'|\s+(?:doravante|conforme|de acordo com|nos termos|na forma|'
        r'e mais|contratad[oa]|jornada|sal[áa]rio|remunera[çc][ãa]o))',
        texto, re.IGNORECASE
    )
    if m:
        cargo = re.sub(r'\s+', ' ', m.group(1)).strip(' .,-')
        cargo = re.sub(
            r'\s+(?:e|conforme|de acordo com|nos termos|na forma)\s*$',
            '', cargo, flags=re.IGNORECASE
        )
        resultado['cargo_funcao'] = cargo or None

    # Local/posto de trabalho — normaliza "própria empresa/sede" vs. posto externo
    local_posto, tipo_local, avisos_local = _resolver_local_posto(texto)
    resultado['local_posto'] = local_posto
    resultado['tipo_local'] = tipo_local
    avisos_qualidade.extend(avisos_local)

    # Salário — só aceita valor claramente associado a salário/remuneração/
    # ordenado/vencimento/R$. Aceita "1.234,56" (com centavos) ou valores
    # inteiros com 3+ dígitos (rejeita números isolados/incompatíveis como "8.2").
    m = re.search(
        r'(?:sal[áa]rio|remunera[çc][ãa]o|ordenado|vencimento)[^\d]{0,30}R?\$?\s*'
        r'([\d]{1,3}(?:\.\d{3})*,\d{2}|\d{3,}(?:\.\d{3})*)',
        texto, re.IGNORECASE
    )
    if m:
        valor_str = m.group(1)
        if ',' in valor_str:
            valor = parse_br_float(valor_str)
        else:
            try:
                valor = float(valor_str.replace('.', ''))
            except ValueError:
                valor = None
        resultado['salario'] = valor if valor and valor >= 100 else None

    # Jornada/escala
    m = re.search(
        r'(?:jornada|escala)\D{0,30}((?:\d{1,2}\s*(?:horas|h)\b[^\n,.]{0,30})|(?:\d{1,2}\s*x\s*\d{1,2}))',
        texto, re.IGNORECASE
    )
    if m:
        resultado['jornada_escala'] = m.group(1).strip()

    return resultado, avisos_qualidade


def extrair_nome_documento_dp(texto: str):
    """
    Extrator genérico (leve) de nome de funcionário para documentos de DP
    que não são holerite/contrato — Rescisão, EPI, Ficha de Registro de
    Empregado, Termo de Prorrogação. Reaproveita os mesmos padrões de rótulo
    ("Empregado:", "Funcionário:", "Nome:") usados em extrair_dados_contrato,
    sem repetir toda a extração de cargo/salário/local (irrelevantes aqui).
    Retorna None se não encontrar.
    """
    if not texto:
        return None
    _stop_nome = (
        r'(?:\n|,|\.|portador|portadora|carteira\s+de\s+trabalho|CTPS|'
        r'CPF|RG|residente|doravante)'
    )
    padroes_nome = [
        rf'(?:EMPREGAD[OA]|CONTRATAD[OA]|FUNCION[ÁA]RI[OA]|COLABORADOR[A]?)\s*:?\s*\n?\s*'
        rf'([A-ZÀ-Ú][A-ZÀ-Ú ]{{3,59}}?)(?=\s*{_stop_nome}|$)',
        rf'Nome\s*:?\s*([A-ZÀ-Ú][A-ZÀ-Ú ]{{3,59}}?)(?=\s*{_stop_nome}|$)',
    ]
    for p in padroes_nome:
        m = re.search(p, texto, re.IGNORECASE)
        if m:
            nome = m.group(1).strip()
            if len(nome.split()) >= 2:
                return nome
    return None


def extrair_dados_rescisao(texto: str):
    """
    Extrai dados de um Termo/Aviso de Rescisão de Contrato de Trabalho (TRCT).

    Heurísticas por regex, tolerantes — qualquer campo não encontrado retorna
    None. Não levanta exceção.

    Retorna dict com chaves (todas opcionais):
      nome_funcionario, cpf, data_rescisao, tipo_aviso, motivo_rescisao
    """
    resultado = {
        'nome_funcionario': None,
        'cpf': None,
        'data_rescisao': None,
        'tipo_aviso': None,
        'motivo_rescisao': None,
    }
    if not texto:
        return resultado

    resultado['cpf'] = extrair_cpf(texto)

    # Nome — TRCT oficial do eSocial (layout numerado "10 PIS/PASEP 11 Nome")
    # tentado primeiro: o documento tem DOIS campos "Nome" — o primeiro é do
    # EMPREGADOR ("Razão Social/Nome"), só o segundo (após "PIS/PASEP") é do
    # trabalhador. O extrator genérico extrair_nome_documento_dp pega o
    # primeiro "Nome" que encontrar (o do empregador, errado) — por isso
    # este padrão específico tem prioridade aqui.
    m_nome_trct = re.search(
        r'PIS/PASEP.{0,20}?Nome\s*\n\s*([A-ZÀ-Ú][A-ZÀ-Ú ]{3,59}?)(?=\s*\n|$)',
        texto,
    )
    if m_nome_trct and len(m_nome_trct.group(1).split()) >= 2:
        resultado['nome_funcionario'] = m_nome_trct.group(1).strip()
    else:
        resultado['nome_funcionario'] = extrair_nome_documento_dp(texto)

    # Data de rescisão / desligamento / saída
    # TRCT oficial do eSocial: rótulos numa linha ("... Data de Admissão ...
    # Data de Afastamento ...") e os valores na linha seguinte — a data de
    # afastamento (a que nos interessa) é a ÚLTIMA data da linha de valores.
    m_datas_trct = re.search(r'Data\s+de\s+Admiss[ãa]o.*\n(.*)\n', texto, re.IGNORECASE)
    if m_datas_trct:
        datas = re.findall(r'\d{2}/\d{2}/\d{4}', m_datas_trct.group(1))
        if datas:
            resultado['data_rescisao'] = datas[-1]

    if not resultado['data_rescisao']:
        padroes_data = [
            r'(?:data\s+(?:do\s+)?(?:afastamento|desligamento|rescis[ãa]o|aviso))\D{0,20}(\d{2}/\d{2}/\d{4})',
            r'(?:rescis[ãa]o|desligamento)\D{0,20}(?:em|de)?\s*(\d{2}/\d{2}/\d{4})',
            r'Data\s*:?\s*(\d{2}/\d{2}/\d{4})',
        ]
        for p in padroes_data:
            m = re.search(p, texto, re.IGNORECASE)
            if m:
                resultado['data_rescisao'] = m.group(1)
                break

    # Tipo de aviso — indenizado vs. trabalhado, ou pedido de demissão
    if re.search(r'aviso\s+pr[ée]vio\s+indenizado', texto, re.IGNORECASE):
        resultado['tipo_aviso'] = 'Aviso Prévio Indenizado'
    elif re.search(r'aviso\s+pr[ée]vio\s+trabalhado', texto, re.IGNORECASE):
        resultado['tipo_aviso'] = 'Aviso Prévio Trabalhado'
    elif re.search(r'pedido\s+de\s+demiss[ãa]o', texto, re.IGNORECASE):
        resultado['tipo_aviso'] = 'Pedido de Demissão'

    # Motivo — sem justa causa / com justa causa / acordo (art. 484-A CLT)
    if re.search(r'sem\s+justa\s+causa', texto, re.IGNORECASE):
        resultado['motivo_rescisao'] = 'Dispensa sem justa causa'
    elif re.search(r'com\s+justa\s+causa', texto, re.IGNORECASE):
        resultado['motivo_rescisao'] = 'Dispensa com justa causa'
    elif re.search(r'acordo\s+(?:entre\s+as\s+partes|mútuo|m[úu]tuo)|484-?A', texto, re.IGNORECASE):
        resultado['motivo_rescisao'] = 'Acordo (Art. 484-A CLT)'

    return resultado


def construir_mapa_cpf(caminho_pdf: str) -> tuple[dict, int]:
    """
    Pass 1 — LEVE.
    Extrai CPF, nome, índice de página E valores financeiros por colaborador.
    Retorna: ({cpf: {'nome': str, 'paginas': [int], 'valores': dict}}, total_paginas)
    """
    mapa: dict = {}
    ultima_chave = None  # última chave tocada no mapa (CPF real OU sem_cpf_N) — p/ continuação
    logger.info(f'[P1] Iniciando varredura | RAM: {_mem_mb()} MB')

    with pdfplumber.open(caminho_pdf) as pdf:
        total = len(pdf.pages)
        logger.info(f'[P1] Total de páginas: {total}')

        for i in range(total):
            page = pdf.pages[i]
            try:
                texto = page.extract_text() or ''
                cpf   = extrair_cpf(texto)
                nome  = extrair_nome_funcionario(texto)
                vals  = extrair_valores_holerite(texto)

                if not cpf and nome in (None, '', 'Desconhecido') and ultima_chave is not None:
                    # Página de continuação (sem CPF impresso e sem "Colaborador:
                    # NOME"/"Nome do Funcionário") — tanto Folhas Manuais (Formato
                    # B: tabela + bloco de assinatura) quanto exports da Secullum
                    # (Formato A: página de legenda/observações após os dados)
                    # podem ter mais de 1 página por colaborador. Sem essa
                    # checagem, essa página vira um "colaborador fantasma"
                    # isolado e a página anterior fica sem o restante do conteúdo.
                    mapa[ultima_chave]['paginas'].append(i)
                    for k, v in vals.items():
                        if v is not None and k not in mapa[ultima_chave]['valores']:
                            mapa[ultima_chave]['valores'][k] = v
                    del texto
                    logger.info(f'[P1] Pág {i+1:03d}/{total}: continuação de {ultima_chave}')
                    if (i + 1) % 15 == 0:
                        gc.collect()
                    continue

                if not cpf:
                    cpf = f'sem_cpf_{i}'

                if cpf not in mapa:
                    mapa[cpf] = {'nome': nome, 'paginas': [], 'valores': {}}
                mapa[cpf]['paginas'].append(i)
                ultima_chave = cpf

                # Atualiza valores: só substitui se o novo for não-nulo (aceita zero)
                for k, v in vals.items():
                    if v is not None and k not in mapa[cpf]['valores']:
                        mapa[cpf]['valores'][k] = v

                del texto
                logger.info(
                    f'[P1] Pág {i+1:03d}/{total}: CPF={cpf} | Nome={nome} | Vals={vals}'
                )

            except Exception as exc:
                logger.warning(f'[P1] Pág {i+1}/{total}: erro → {exc}')
                mapa[f'sem_cpf_{i}'] = {
                    'nome': 'Erro extração', 'paginas': [i], 'valores': {}
                }

            if (i + 1) % 15 == 0:
                gc.collect()
                logger.info(f'[P1] GC executado | RAM: {_mem_mb()} MB')

    gc.collect()
    logger.info(f'[P1] Concluído: {len(mapa)} colaboradores | RAM: {_mem_mb()} MB')
    return mapa, total


def extrair_pdf_colaborador(caminho_pdf: str, indices: list) -> bytes:
    reader = PdfReader(caminho_pdf)
    writer = PdfWriter()
    for idx in indices:
        writer.add_page(reader.pages[idx])
    buf = io.BytesIO()
    writer.write(buf)
    resultado = buf.getvalue()
    del reader, writer, buf
    return resultado


def extrair_pdf_do_request(campo: str = 'pdf') -> str:
    """campo (v3.00): nome do campo multipart a buscar. Com o padrão 'pdf',
    mantém o comportamento antigo (cai para qualquer arquivo enviado, se
    'pdf' não existir) — usado por quem já chama sem argumento. Com um
    nome explícito diferente (ex.: 'pdf_manual'), busca só aquele campo,
    sem fallback, para não pegar o arquivo errado quando há 2 PDFs no
    mesmo request."""
    ct = request.content_type or ''
    if 'multipart/form-data' in ct:
        arq = request.files.get(campo)
        if not arq and campo == 'pdf':
            arq = next(iter(request.files.values()), None) if request.files else None
        if arq:
            caminho = os.path.join(tempfile.gettempdir(), f'holerite_{_uuid.uuid4().hex}.pdf')
            arq.save(caminho)
            tamanho = os.path.getsize(caminho)
            logger.info(f'[PDF] Salvo em disco: {caminho} | {tamanho // 1024} KB')
            return caminho
    elif 'application/json' in ct and campo == 'pdf':
        data = request.get_json(force=True, silent=True) or {}
        if 'pdf_base64' in data:
            try:
                caminho = os.path.join(tempfile.gettempdir(), f'holerite_{_uuid.uuid4().hex}.pdf')
                with open(caminho, 'wb') as f:
                    f.write(base64.b64decode(data['pdf_base64']))
                return caminho
            except Exception:
                pass
    return None


# ── Helpers Airtable ──────────────────────────────────────────────────────────

def mes_anterior_info():
    hoje = datetime.now()
    if hoje.month == 1:
        return MESES_PT[11], hoje.year - 1, 12
    return MESES_PT[hoje.month - 2], hoje.year, hoje.month - 1


def extrair_competencia_holerite(texto: str):
    """Lê a competência impressa no holerite ("Competência: MM/AAAA") e
    devolve (folha_mensal, data_str), ex.: ('Abril 2026', '2026-04-01').

    Usado para arquivar cada holerite no mês contábil CORRETO em vez de
    confiar no default "mês anterior" — essencial para o backlog retroativo
    de 15/06, que traz holerites de competências antigas importados de uma vez.

    Retorna (None, None) se não encontrar a competência no texto.
    """
    if not texto:
        return None, None

    # Formato 1 — holerite: mês por extenso ("Mensalista Abril de 2026",
    # "Março de 2026"). É o formato real dos holerites desta folha.
    m = re.search(
        r'\b(Janeiro|Fevereiro|Mar[çc]o|Abril|Maio|Junho|Julho|Agosto|'
        r'Setembro|Outubro|Novembro|Dezembro)\s+de\s+(\d{4})',
        texto, re.IGNORECASE,
    )
    if m:
        nome_norm = m.group(1).lower().replace('ç', 'c')
        nomes = ['janeiro', 'fevereiro', 'marco', 'abril', 'maio', 'junho',
                 'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']
        if nome_norm in nomes:
            mes = nomes.index(nome_norm) + 1
            ano = int(m.group(2))
            return f'{MESES_PT[mes - 1]} {ano}', f'{ano}-{mes:02d}-01'

    # Formato 2 — fallback numérico ("Competência: 04/2026"), usado por
    # extratos e alguns layouts.
    m = re.search(r'Compet[êe]ncia[:\s]+(\d{2})\s*/\s*(\d{4})', texto, re.IGNORECASE)
    if m:
        mes, ano = int(m.group(1)), int(m.group(2))
        if 1 <= mes <= 12:
            return f'{MESES_PT[mes - 1]} {ano}', f'{ano}-{mes:02d}-01'

    return None, None


def _extrair_texto_pdf_bytes(pdf_bytes: bytes) -> str:
    """Extrai o texto de todas as páginas de um PDF a partir dos bytes
    (mesmo padrão pdfplumber já usado em outros pontos do arquivo — não
    duplica lógica nova, só nomeia o padrão existente para reúso no
    pacote Holerite+Folha de Ponto)."""
    if not pdf_bytes:
        return ''
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            return '\n'.join((p.extract_text() or '') for p in pdf.pages)
    except Exception:
        return ''


def _extrair_competencia_folha_ponto(texto: str):
    """Lê a competência de uma Folha/Cartão de Ponto. Tenta primeiro os
    mesmos 2 formatos de `extrair_competencia_holerite` (mês por extenso
    "de AAAA" e "Competência: MM/AAAA") — sem duplicar a função já
    validada na Macro 5, só delegando — e, se nenhum bater, tenta um
    formato adicional específico de cartão de ponto: intervalo de datas
    "DD/MM/AAAA a DD/MM/AAAA" (usa o mês da data final do período).

    Retorna (folha_mensal, data_str) ou (None, None) se indeterminável.
    Nunca adivinha por proximidade temporal ou nome de arquivo — se não
    achar o texto, retorna None (comportamento exigido: falha explícita,
    nunca "competência divergente" presumida por omissão).
    """
    folha_mensal, data_str = extrair_competencia_holerite(texto)
    if folha_mensal:
        return folha_mensal, data_str

    if not texto:
        return None, None

    m = re.search(
        r'(\d{2})/(\d{2})/(\d{4})\s*(?:a|até|-)\s*(\d{2})/(\d{2})/(\d{4})',
        texto, re.IGNORECASE,
    )
    if m:
        mes, ano = int(m.group(5)), int(m.group(6))
        if 1 <= mes <= 12:
            return f'{MESES_PT[mes - 1]} {ano}', f'{ano}-{mes:02d}-01'

    return None, None


def buscar_mes_contabilidade_atual():
    hoje = datetime.now()
    nome = f'{MESES_PT[hoje.month - 1]} {hoje.year}'
    _at_throttle()
    r = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_CONT}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        params={
            'filterByFormula': f'{{Mês - Contabilidade}}="{nome}"',
            'maxRecords': 1,
            'fields[]': ['Mês - Contabilidade'],
        },
        timeout=30,
    )
    if r.ok:
        records = r.json().get('records', [])
        if records:
            return records[0]['id'], nome
    return None, nome


def buscar_funcionario_por_cpf(cpf: str):
    headers = {'Authorization': f'Bearer {AIRTABLE_API_KEY}'}
    cpf_num = re.sub(r'\D', '', cpf)
    formulas = [
        f'{{CPF}}="{cpf}"',
        f'{{CPF}}="{cpf_num}"',
        f'{{CPF}}={cpf_num}',
    ]
    for formula in formulas:
        _at_throttle()
        try:
            r = requests.get(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}',
                headers=headers,
                params={
                    'filterByFormula': formula,
                    'maxRecords': 1,
                    'fields[]': ['Nome Completo', 'CPF'],
                },
                timeout=30,
            )
            if r.ok:
                records = r.json().get('records', [])
                if records:
                    nome = records[0].get('fields', {}).get('Nome Completo', '')
                    logger.info(f'[AT] Funcionário encontrado com fórmula: {formula}')
                    return records[0]['id'], nome
        except requests.exceptions.Timeout:
            logger.warning(f'[AT] Timeout CPF {_mascarar_cpf(cpf)} (fórmula: {formula})')
            continue
        except Exception as exc:
            logger.warning(f'[AT] Erro CPF {_mascarar_cpf(cpf)}: {exc}')
            continue
    return None, None


def _normalizar_nome(s: str) -> str:
    """Maiúsculas, sem acento, espaços colapsados — para casar nomes 'sujos'
    (cadastro tem acentos/espaços duplos/faltantes inconsistentes)."""
    if not s:
        return ''
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'\s+', ' ', s).strip()
    return s.upper()


def buscar_funcionario_por_nome(nome: str):
    """
    Fallback para layouts de holerite que não imprimem CPF na página (ex.:
    layout "Mensalista" identificado em 05/07/2026 — só tem Nome do
    Funcionário, sem CPF). Casa por igualdade exata (case-insensitive) com
    Nome Completo no Airtable e, se não achar, por igualdade normalizada
    (sem acento, espaços colapsados) — o cadastro tem inconsistências desse
    tipo (ex.: "Celso Conceição Lima " com espaço sobrando, "GABRIEL
    GONÇALVES..." vs. PDF sem cedilha).
    """
    if not nome:
        return None, None
    headers = {'Authorization': f'Bearer {AIRTABLE_API_KEY}'}
    nome_escapado = nome.replace('"', '\\"')
    formula = f'UPPER({{Nome Completo}})=UPPER("{nome_escapado}")'
    _at_throttle()
    try:
        r = requests.get(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}',
            headers=headers,
            params={
                'filterByFormula': formula,
                'maxRecords': 1,
                'fields[]': ['Nome Completo', 'CPF'],
            },
            timeout=30,
        )
        if r.ok:
            records = r.json().get('records', [])
            if records:
                nome_at = records[0].get('fields', {}).get('Nome Completo', '')
                logger.info(f'[AT] Funcionário encontrado por nome: "{nome}"')
                return records[0]['id'], nome_at
    except Exception as exc:
        logger.warning(f'[AT] Erro busca por nome "{nome}": {exc}')

    # Fallback — igualdade normalizada contra a tabela inteira. Não dá para
    # restringir via SEARCH()/filterByFormula porque a própria inconsistência
    # de acento (ex.: "JOAO" no PDF vs. "JOÃO" no cadastro) faz o SEARCH do
    # Airtable falhar — a formula não remove acento, então a pré-seleção
    # excluiria justamente os casos que este fallback existe para resolver.
    alvo_norm = _normalizar_nome(nome)
    offset = None
    try:
        while True:
            params = {
                'fields[]': ['Nome Completo', 'CPF'],
                'pageSize': 100,
            }
            if offset:
                params['offset'] = offset
            _at_throttle()
            r = requests.get(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}',
                headers=headers, params=params, timeout=30,
            )
            if not r.ok:
                break
            data = r.json()
            for rec in data.get('records', []):
                nome_at = rec.get('fields', {}).get('Nome Completo', '')
                if _normalizar_nome(nome_at) == alvo_norm:
                    logger.info(f'[AT] Funcionário encontrado por nome normalizado: "{nome}" -> "{nome_at}"')
                    return rec['id'], nome_at
            offset = data.get('offset')
            if not offset:
                break
    except Exception as exc:
        logger.warning(f'[AT] Erro busca por nome normalizado "{nome}": {exc}')

    return None, None


def criar_registro_holerite(nome, func_id, folha_mensal, data_str, mes_cont_id, valores=None, status='Concluído'):
    """Cria registro de holerite no Airtable, incluindo valores financeiros se disponíveis.

    'status' permite criar o registro como 'Concluído' (fluxo normal) ou
    'Revisão Manual' (sanitização — usado quando algum dado não pôde ser
    gravado/extraído com confiança). 'typecast=True' permite ao Airtable
    criar a opção "Revisão Manual" automaticamente no campo Status, se ainda
    não existir como opção do singleSelect.
    """
    campos = {
        F_HOL_NOME:     f'{nome} - {folha_mensal}',
        F_HOL_STATUS:   status,
        F_HOL_FUNC:     [func_id],
        F_HOL_DATA:     data_str,
        F_HOL_FOLHA:    folha_mensal,
    }
    # mes_cont_id pode vir None (mês de contabilidade não encontrado) — um
    # link [None] é rejeitado pelo Airtable (422), então omitimos o campo
    # nesse caso em vez de quebrar a criação do holerite.
    if mes_cont_id:
        campos[F_HOL_MES_CONT] = [mes_cont_id]
    if valores:
        if valores.get('total_vencimentos') is not None:
            campos[F_HOL_VENCIM]    = valores['total_vencimentos']
        if valores.get('total_descontos') is not None:
            campos[F_HOL_DESCONTOS] = valores['total_descontos']
        if valores.get('valor_liquido') is not None:
            campos[F_HOL_LIQUIDO]   = valores['valor_liquido']
        if valores.get('inss') is not None:
            campos[F_HOL_INSS]      = valores['inss']

    _at_throttle()
    r = requests.post(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_HOL}',
        headers=_at_headers(),
        json={'fields': campos, 'typecast': True},
        timeout=15,
    )
    _raise_for_airtable(r, f'criar holerite ({nome})')
    return r.json()['id']


def atualizar_valores_holerite(record_id: str, valores: dict):
    """Atualiza apenas os campos financeiros de um holerite existente."""
    campos = {}
    if valores.get('total_vencimentos') is not None:
        campos[F_HOL_VENCIM]    = valores['total_vencimentos']
    if valores.get('total_descontos') is not None:
        campos[F_HOL_DESCONTOS] = valores['total_descontos']
    if valores.get('valor_liquido') is not None:
        campos[F_HOL_LIQUIDO]   = valores['valor_liquido']
    if valores.get('inss') is not None:
        campos[F_HOL_INSS]      = valores['inss']

    if not campos:
        return None

    _at_throttle()
    r = requests.patch(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_HOL}/{record_id}',
        headers=_at_headers(),
        json={'fields': campos, 'typecast': True},
        timeout=15,
    )
    _raise_for_airtable(r, f'atualizar valores holerite {record_id}')
    return r.json()


def anexar_pdf_holerite(record_id, pdf_bytes, filename):
    _at_throttle()
    url = (
        f'https://content.airtable.com/v0/{BASE_ID}'
        f'/{record_id}/{F_HOL_PDF}/uploadAttachment'
    )
    r = requests.post(
        url,
        headers={
            'Authorization': f'Bearer {AIRTABLE_API_KEY}',
            'Content-Type': 'application/json',
        },
        json={
            'contentType': 'application/pdf',
            'filename': filename,
            'file': base64.b64encode(pdf_bytes).decode('utf-8'),
        },
        timeout=60,
    )
    _raise_for_airtable(r, f'anexar PDF holerite {record_id}')
    return r.json()


# ── Helpers genéricos Airtable (Fase 2) ──────────────────────────────────────

def _criar_registro(table_id: str, fields: dict) -> str:
    _at_throttle()
    r = requests.post(
        f'https://api.airtable.com/v0/{BASE_ID}/{table_id}',
        headers=_at_headers(),
        json={'fields': fields, 'typecast': True},
        timeout=30,
    )
    _raise_for_airtable(r, f'create {table_id}')
    return r.json()['id']


def _buscar_por_campo(table_id: str, campo_nome: str, valor: str):
    """Busca o 1º registro onde {campo_nome} = valor. Retorna o registro ou None."""
    _at_throttle()
    valor_escapado = valor.replace('"', '\\"')
    formula = f'{{{campo_nome}}}="{valor_escapado}"'
    r = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{table_id}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        params={'filterByFormula': formula, 'maxRecords': 1},
        timeout=30,
    )
    r.raise_for_status()
    records = r.json().get('records', [])
    return records[0] if records else None


def _anexar_attachment(table_id: str, record_id: str, field_id: str,
                        conteudo_bytes: bytes, filename: str,
                        content_type: str = 'application/pdf'):
    _at_throttle()
    url = f'https://content.airtable.com/v0/{BASE_ID}/{record_id}/{field_id}/uploadAttachment'
    r = requests.post(
        url,
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}', 'Content-Type': 'application/json'},
        json={
            'contentType': content_type,
            'filename': filename,
            'file': base64.b64encode(conteudo_bytes).decode('utf-8'),
        },
        timeout=60,
    )
    _raise_for_airtable(r, f'attach {table_id}/{record_id}')
    return r.json()


def _substituir_attachment(table_id: str, record_id: str, field_id: str,
                            conteudo_bytes: bytes, filename: str,
                            content_type: str = 'application/pdf'):
    """Sobe um novo anexo e SUBSTITUI todo o conteúdo do campo por ele —
    remove qualquer anexo anterior, em vez de acumular (comportamento padrão
    de _anexar_attachment, que só adiciona). Usado no campo "PDF Folha
    Ponto" (v3.00): a versão assinada substitui a provisória, nunca as duas
    juntas no mesmo campo."""
    resultado_upload = _anexar_attachment(table_id, record_id, field_id, conteudo_bytes, filename, content_type)
    anexos_apos_upload = (resultado_upload.get('fields', {}) or {}).get(field_id) or []
    if not anexos_apos_upload:
        return resultado_upload
    novo = anexos_apos_upload[-1]
    _at_throttle()
    r = requests.patch(
        f'https://api.airtable.com/v0/{BASE_ID}/{table_id}/{record_id}',
        headers=_at_headers(),
        json={'fields': {field_id: [{'id': novo['id']}]}, 'typecast': True},
        timeout=30,
    )
    _raise_for_airtable(r, f'substituir attachment {table_id}/{record_id}/{field_id}')
    return r.json()


def _formatar_resumo_cartao_ponto(dados: dict) -> str:
    """Formata o retorno de extrair_cartao_ponto em texto legível para o
    campo "Resumo Cartão Ponto (extraído)" em Funcionários."""
    linhas = []
    if dados['periodo_inicio'] and dados['periodo_fim']:
        linhas.append(f'Período: {dados["periodo_inicio"]} até {dados["periodo_fim"]}')
    for dia in dados['dias']:
        cabecalho = f'{dia["data"]} ({dia["dia_semana"]}, {dia["ciclo"]})'
        if dia['folga']:
            linhas.append(f'{cabecalho}: FOLGA')
            continue
        horarios = ' '.join(
            h for h in (dia['entrada_1'], dia['saida_1'], dia['entrada_2'],
                         dia['saida_2'], dia['entrada_3'], dia['saida_3'])
            if h
        )
        linhas.append(f'{cabecalho}: {horarios}' if horarios else f'{cabecalho}: (sem batidas)')
    return '\n'.join(linhas)


def _atualizar_resumo_ponto_funcionario(func_id: str, resumo_texto: str):
    _at_throttle()
    r = requests.patch(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}/{func_id}',
        headers=_at_headers(),
        json={'fields': {F_FUNC_RESUMO_PONTO: resumo_texto}, 'typecast': True},
        timeout=15,
    )
    _raise_for_airtable(r, f'atualizar resumo cartão ponto {func_id}')
    return r.json()


# ── Helpers Fase 4 — /processar-fila ─────────────────────────────────────────

def _atualizar_status_processar(record_id: str, status: str):
    _at_throttle()
    r = requests.patch(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_PROCESSAR}/{record_id}',
        headers=_at_headers(),
        json={'fields': {F_PROC_STATUS: status}, 'typecast': True},
        timeout=15,
    )
    _raise_for_airtable(r, f'update status {record_id}')
    return r.json()


def _criar_pendencia(arquivo_id: str, tipo_problema: str, observacao: str, nome: str = None):
    """`nome` (opcional, Macro 6A — idempotência das Pendências do Kit de
    Admissão): permite substituir o rótulo default (f'{tipo_problema}:
    {arquivo_id}') por uma chave determinística própria de quem chama —
    usado por _criar_ou_reaproveitar_pendencia_kit_identidade para poder
    buscar essa mesma Pendência depois (F_PEND_NOME é texto livre, sem
    campo dedicado de chave, e não há alteração de schema nesta fase).
    Todos os chamadores existentes continuam com o comportamento default
    (sem passar `nome`) — mudança aditiva, não quebra nada."""
    return _criar_registro(TABLE_PENDENCIAS, {
        F_PEND_NOME:   nome if nome else f'{tipo_problema}: {arquivo_id}',
        F_PEND_STATUS: 'Pendente',
        F_PEND_TIPO:   tipo_problema,
        F_PEND_ORIGEM: [arquivo_id],
        F_PEND_OBS:    observacao[:500] if observacao else '',
        F_PEND_DATA:   datetime.now().isoformat(),
    })


# Sufixos societários/comuns que não ajudam na identificação de uma pessoa ou
# empresa e que variam entre o PDF e o cadastro (ex.: "MAGNATA SERVIÇOS LTDA"
# vs "MAGNATA SERVIÇOS LTDA ME").
_SUFIXOS_NOME_RE = re.compile(
    r'\b(LTDA|ME|EPP|EIRELI|S\s*/?\s*A|SA)\b\.?'
)


def _normalizar_telefone_br(numero: str):
    """Normaliza um número de telefone brasileiro para o formato esperado
    pelo WhatsApp/Evolution API: DDI 55 + DDD + número, somente dígitos
    (ex.: "5511999998888"). Retorna None se `numero` já estiver correto,
    estiver vazio ou não for possível normalizar com confiança."""
    if not numero:
        return None
    digitos = re.sub(r'\D', '', str(numero))
    if not digitos:
        return None
    # Remove "0" de discagem interurbana (ex.: "0 11 99999-8888")
    if digitos.startswith('0') and len(digitos) in (12, 13):
        digitos = digitos[1:]
    if digitos.startswith('55') and len(digitos) in (12, 13):
        return None  # já normalizado
    if len(digitos) in (10, 11):
        return '55' + digitos
    return None


def _normalizar_nome_busca(texto: str) -> str:
    """Normaliza nomes para comparação "fuzzy": maiúsculas, sem acentos,
    sem pontuação, espaços colapsados e sem sufixos societários comuns
    (LTDA, ME, EPP, EIRELI, S/A) — necessário pois o nome extraído do PDF
    raramente é idêntico, byte a byte, ao cadastrado no Airtable."""
    if not texto:
        return ''
    sem_acentos = unicodedata.normalize('NFKD', texto)
    sem_acentos = ''.join(c for c in sem_acentos if not unicodedata.combining(c))
    upper = sem_acentos.upper()
    upper = re.sub(r'[^\w\s]', ' ', upper)
    upper = _SUFIXOS_NOME_RE.sub(' ', upper)
    return re.sub(r'\s+', ' ', upper).strip()


def _buscar_funcionario_por_nome(nome: str):
    """Fallback de busca por Nome Completo quando não há CPF.

    Faz comparação fuzzy (ver _normalizar_nome_busca): ignora maiúsculas/
    minúsculas, acentos, pontuação, espaços extras e sufixos societários.
    """
    nome_normalizado = _normalizar_nome_busca(nome)
    if not nome_normalizado:
        return None, None

    for rec in _at_listar_todos(TABLE_FUNC, ['Nome Completo']):
        nome_cadastrado = rec['fields'].get('Nome Completo', '')
        if _normalizar_nome_busca(nome_cadastrado) == nome_normalizado:
            return rec['id'], nome_cadastrado
    return None, None


_CONECTORES_NOME_RE = re.compile(r'\b(?:DA|DE|DO|DAS|DOS|E)\b')


def _primeiro_ultimo_nome(nome_normalizado: str) -> tuple:
    """Extrai (primeiro, último) token de um nome já normalizado (maiúsculas,
    sem acento), ignorando conectores comuns (DA/DE/DO/DAS/DOS/E) — usado
    quando a folha manual grafa o nome sem os conectores do cadastro."""
    sem_conectores = _CONECTORES_NOME_RE.sub(' ', nome_normalizado)
    tokens = [t for t in sem_conectores.split() if t]
    if not tokens:
        return '', ''
    return tokens[0], tokens[-1]


def _buscar_funcionario_por_nome_parcial(nome: str):
    """Fallback (v3.00) quando a busca exata por nome (_buscar_funcionario_
    por_nome) não encontra nada — ex.: folha manual grafa o nome sem os
    conectores ("da"/"de"/"dos") presentes no cadastro. Casa por Primeiro
    Nome + Último Nome, ambos normalizados.

    Se mais de um Funcionário bater (homônimos), NÃO escolhe sozinho —
    devolve a lista de candidatos para o chamador criar uma Pendência
    explícita em vez de arriscar vincular à pessoa errada.

    Retorna (func_id, nome_at, candidatos):
      - achou exatamente 1: (id, nome, None)
      - homônimos (2+):     (None, None, [{'id':..., 'nome':...}, ...])
      - não achou:          (None, None, None)
    """
    nome_norm = _normalizar_nome_busca(nome)
    primeiro, ultimo = _primeiro_ultimo_nome(nome_norm)
    if not primeiro or not ultimo:
        return None, None, None

    candidatos = []
    for rec in _at_listar_todos(TABLE_FUNC, ['Nome Completo']):
        nome_cadastrado = rec['fields'].get('Nome Completo', '')
        p_cad, u_cad = _primeiro_ultimo_nome(_normalizar_nome_busca(nome_cadastrado))
        if p_cad == primeiro and u_cad == ultimo:
            candidatos.append({'id': rec['id'], 'nome': nome_cadastrado})

    if len(candidatos) == 1:
        return candidatos[0]['id'], candidatos[0]['nome'], None
    if len(candidatos) > 1:
        return None, None, candidatos
    return None, None, None


def _buscar_cpf_por_funcionario_id(func_id: str) -> str:
    """Busca o CPF cadastrado de um Funcionário pelo record id — usado no
    Formato B da Folha de Ponto (v2.99), que não tem CPF impresso no PDF."""
    _at_throttle()
    r = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}/{func_id}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        timeout=15,
    )
    if r.ok:
        cpf = r.json().get('fields', {}).get('CPF', '') or ''
        if not cpf:
            logger.warning(f'[PONTO] CPF vazio para func_id={func_id} — resposta Airtable sem o campo CPF preenchido.')
        return cpf
    logger.error(f'[PONTO] falha ao buscar CPF de func_id={func_id}: HTTP {r.status_code} — {r.text[:300]}')
    return ''


def _buscar_contabilidade_mensal_por_nome(nome: str):
    _at_throttle()
    nome_escapado = nome.replace('"', '\\"')
    r = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_CONT}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        params={
            'filterByFormula': f'{{Mês - Contabilidade}}="{nome_escapado}"',
            'maxRecords': 1,
            'fields[]': ['Mês - Contabilidade'],
        },
        timeout=30,
    )
    if r.ok:
        records = r.json().get('records', [])
        if records:
            return records[0]['id']
    return None


def _buscar_holerite_existente(func_id: str, folha_mensal: str):
    """Procura holerite já criado para este funcionário nesta folha mensal."""
    _at_throttle()
    folha_escapada = folha_mensal.replace('"', '\\"')
    r = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_HOL}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        params={
            'filterByFormula': f'{{Folha Mensal}}="{folha_escapada}"',
            'returnFieldsByFieldId': 'true',
        },
        timeout=30,
    )
    r.raise_for_status()
    for rec in r.json().get('records', []):
        links = rec.get('fields', {}).get(F_HOL_FUNC) or []
        link_ids = [l['id'] if isinstance(l, dict) else l for l in links]
        if func_id in link_ids:
            return rec
    return None


def _verificar_anexo_em_lista(attachments: list, nome_arquivo: str, pdf_hash: str) -> str:
    """
    Verifica se já existe, numa lista de anexos (multipleAttachments), um
    arquivo com o mesmo nome.

    Retorna:
      'identico'   — já existe anexo com mesmo nome e mesmo hash (não duplicar)
      'conflito'   — já existe anexo com mesmo nome mas conteúdo diferente
                      (substituição precisa de autorização — não anexa)
      'novo'       — não há anexo com esse nome (pode anexar normalmente)
    """
    for att in attachments:
        if att.get('filename') != nome_arquivo:
            continue
        try:
            existente_bytes = requests.get(att['url'], timeout=60).content
            existente_hash = hashlib.sha256(existente_bytes).hexdigest()
        except Exception as exc:
            logger.warning(f'[FILA] Falha ao baixar anexo existente para comparar hash: {exc}')
            return 'conflito'
        return 'identico' if existente_hash == pdf_hash else 'conflito'
    return 'novo'


def _verificar_anexo_holerite(holerite_fields: dict, nome_arquivo: str, pdf_hash: str) -> str:
    """Verifica duplicidade de anexo no campo PDF Holerite (ver _verificar_anexo_em_lista)."""
    attachments = holerite_fields.get(F_HOL_PDF) or []
    return _verificar_anexo_em_lista(attachments, nome_arquivo, pdf_hash)


def _processar_holerite(ctx: dict, dry_run: bool) -> dict:
    """
    Handler de Holerite para /processar-fila.

    ctx: {proc_id, arquivo_id, pdf_bytes, pdf_hash, nome_arquivo, texto,
          folha_mensal, mes_cont_id, data_holerite}

    Retorno padronizado:
      {"acao": str, "status_final": "Concluído"|"Erro",
       "detalhes": dict, "pendencia": {"tipo":..., "observacao":...} | None}
    """
    texto = ctx['texto']
    folha_mensal = ctx['folha_mensal']
    mes_cont_id = ctx['mes_cont_id']
    data_holerite = ctx['data_holerite']

    # 0. Competência manda. O default folha_mensal vem do "mês anterior"
    #    (Maio, hoje), mas o backlog de 15/06 traz holerites de competências
    #    antigas (Jan-Abr/2026). Quando o PDF traz "Competência: MM/AAAA",
    #    usamos ela para arquivar no mês contábil correto (e não bagunçar Maio).
    comp_folha, comp_data = extrair_competencia_holerite(texto)
    competencia_pdf_usada = False
    if comp_folha and comp_folha != folha_mensal:
        folha_mensal = comp_folha
        data_holerite = comp_data
        mes_cont_id = _buscar_contabilidade_mensal_por_nome(comp_folha) or mes_cont_id
        competencia_pdf_usada = True

    # 0b. Guarda de segurança: holerite SEMPRE traz a competência (mês por
    #     extenso). Se não detectamos NENHUMA, este PDF provavelmente não é um
    #     holerite de fato — no backlog de 15/06 há rescisões mal classificadas
    #     como "Holerite". Em vez de criar um holerite-fantasma no mês errado
    #     (default Maio), sinalizamos para revisão e NÃO gravamos nada.
    if not comp_folha:
        return {
            'acao': 'competencia_nao_detectada',
            'status_final': 'Erro',
            'detalhes': {
                'folha_mensal_default': ctx['folha_mensal'],
                'nome_arquivo': ctx.get('nome_arquivo'),
            },
            'pendencia': {
                'tipo': 'Holerite sem competência detectada',
                'observacao': (
                    'Não foi possível ler a competência (mês/ano) no PDF. Pode ser '
                    'documento mal classificado como Holerite (ex.: Rescisão) ou PDF '
                    'ilegível. NÃO arquivado, para não cair no mês errado — revisar '
                    'classificação/legibilidade manualmente.'
                ),
            },
        }

    # 1. Localizar funcionário por CPF, com fallback por nome extraído do PDF
    cpf = extrair_cpf(texto)
    func_id, nome = (None, None)
    if cpf:
        func_id, nome = buscar_funcionario_por_cpf(cpf)

    nome_pdf = extrair_nome_funcionario(texto)
    if not func_id and nome_pdf:
        func_id, nome = _buscar_funcionario_por_nome(nome_pdf)

    if not func_id:
        return {
            'acao': 'funcionario_nao_encontrado',
            'status_final': 'Erro',
            'detalhes': {
                'cpf_extraido': cpf,
                'nome_extraido': nome_pdf,
                'folha_mensal': folha_mensal,
            },
            'pendencia': {
                'tipo': 'Funcionário não encontrado',
                'observacao': (
                    f'CPF extraído: {cpf or "(nenhum)"} | '
                    f'Nome extraído: {nome_pdf or "(nenhum)"} | '
                    f'Folha: {folha_mensal}'
                ),
            },
        }

    # 2. Extrair valores financeiros do holerite
    valores = extrair_valores_holerite(texto)

    # 2b. Trava de segurança — valores financeiros inconsistentes
    v = valores.get('total_vencimentos')
    d = valores.get('total_descontos')
    liq = valores.get('valor_liquido')
    TOLERANCIA_CENTAVOS = 0.02
    valores_inconsistentes = False
    motivo_inconsistencia = None

    if v is not None and d is not None and liq is not None:
        if liq > 0 and v == d:
            valores_inconsistentes = True
            motivo_inconsistencia = (
                f'Total Vencimentos ({v}) igual a Total Descontos ({d}) com Valor Líquido ({liq}) > 0.'
            )
        elif v < liq:
            valores_inconsistentes = True
            motivo_inconsistencia = (
                f'Total Vencimentos ({v}) menor que Valor Líquido ({liq}).'
            )
        elif abs((v - d) - liq) > TOLERANCIA_CENTAVOS:
            valores_inconsistentes = True
            motivo_inconsistencia = (
                f'Total Vencimentos ({v}) - Total Descontos ({d}) = {v - d}, '
                f'diferente de Valor Líquido ({liq}) além da tolerância de R$ {TOLERANCIA_CENTAVOS}.'
            )

    if valores_inconsistentes:
        return {
            'acao': 'valores_inconsistentes',
            'status_final': 'Erro',
            'detalhes': {
                'funcionario_id': func_id,
                'funcionario_nome': nome,
                'cpf_extraido': cpf,
                'folha_mensal': folha_mensal,
                'valores': valores,
            },
            'pendencia': {
                'tipo': 'Valores inconsistentes',
                'observacao': (
                    f'{motivo_inconsistencia} '
                    f'Holerite NÃO atualizado, requer revisão manual.'
                ),
            },
        }

    # 3. Evitar duplicidade — checar se já existe holerite deste funcionário nesta folha
    existente = _buscar_holerite_existente(func_id, folha_mensal)

    if dry_run:
        pdf_ja_existia = False
        pdf_duplicado_evitado = False
        status_anexo = None

        if existente:
            status_anexo = _verificar_anexo_holerite(
                existente.get('fields', {}), ctx['nome_arquivo'], ctx['pdf_hash'],
            )
            if status_anexo == 'identico':
                pdf_ja_existia = True
                pdf_duplicado_evitado = True
                acao = 'atualizaria_holerite_existente_sem_duplicar_pdf'
            elif status_anexo == 'conflito':
                acao = 'atualizaria_holerite_existente_anexo_precisaria_autorizacao'
            else:  # 'novo'
                acao = 'atualizaria_holerite_existente_e_anexaria_pdf'
        else:
            acao = 'criaria_holerite'

        return {
            'acao': acao,
            'status_final': 'Concluído',
            'detalhes': {
                'funcionario_id': func_id,
                'funcionario_nome': nome,
                'cpf_extraido': cpf,
                'folha_mensal': folha_mensal,
                'competencia_pdf_usada': competencia_pdf_usada,
                'valores': valores,
                'holerite_existente_id': existente['id'] if existente else None,
                'pdf_anexado': False,
                'pdf_ja_existia': pdf_ja_existia,
                'pdf_duplicado_evitado': pdf_duplicado_evitado,
                'status_anexo': status_anexo,
            },
            'pendencia': None,
        }

    pendencia = None
    status_final = 'Concluído'
    avisos_sanitizacao = []

    if existente:
        holerite_id = existente['id']

        try:
            atualizar_valores_holerite(holerite_id, valores)
        except AirtableError as exc:
            logger.error(f'[FILA] {ctx["proc_id"]}: erro ao atualizar valores do holerite {holerite_id}: {exc}')
            avisos_sanitizacao.append(f'Valores financeiros NÃO atualizados: {exc}')
            status_final = 'Erro'

        status_anexo = _verificar_anexo_holerite(
            existente.get('fields', {}), ctx['nome_arquivo'], ctx['pdf_hash'],
        )
        if status_anexo == 'novo':
            try:
                anexar_pdf_holerite(holerite_id, ctx['pdf_bytes'], ctx['nome_arquivo'])
                acao = 'holerite_atualizado' if not avisos_sanitizacao else 'holerite_atualizado_parcial'
            except AirtableError as exc:
                logger.error(f'[FILA] {ctx["proc_id"]}: erro ao anexar PDF no holerite {holerite_id}: {exc}')
                avisos_sanitizacao.append(f'PDF NÃO anexado: {exc}')
                acao = 'holerite_atualizado_parcial'
                status_final = 'Erro'
        elif status_anexo == 'identico':
            acao = 'holerite_atualizado_anexo_ja_existia' if not avisos_sanitizacao else 'holerite_atualizado_parcial'
        else:  # 'conflito'
            acao = 'holerite_atualizado_anexo_nao_substituido'
            avisos_sanitizacao.append(
                f'Holerite {holerite_id}: já existe anexo "{ctx["nome_arquivo"]}" com '
                f'conteúdo diferente do PDF processado. Anexo NÃO foi substituído — '
                f'requer autorização manual.'
            )

        if avisos_sanitizacao:
            pendencia = {
                'tipo': 'Holerite atualizado com pendências (revisão manual)',
                'observacao': f'Holerite {holerite_id}: ' + ' | '.join(avisos_sanitizacao),
            }
    else:
        try:
            holerite_id = criar_registro_holerite(
                nome, func_id, folha_mensal, data_holerite, mes_cont_id, valores=valores,
            )
            anexar_pdf_holerite(holerite_id, ctx['pdf_bytes'], ctx['nome_arquivo'])
            acao = 'holerite_criado'
        except AirtableError as exc:
            # Sanitização: o payload completo (com valores financeiros) foi
            # rejeitado pelo Airtable. Em vez de perder o documento, tenta
            # novamente com um payload mínimo seguro (sem valores) e status
            # "Revisão Manual", preservando o PDF e registrando o erro exato
            # numa Pendência para correção manual posterior.
            logger.error(f'[FILA] {ctx["proc_id"]}: Airtable rejeitou criação do holerite (payload completo): {exc}')
            try:
                holerite_id = criar_registro_holerite(
                    nome, func_id, folha_mensal, data_holerite, mes_cont_id,
                    valores=None, status='Revisão Manual',
                )
                anexar_pdf_holerite(holerite_id, ctx['pdf_bytes'], ctx['nome_arquivo'])
            except AirtableError as exc2:
                logger.error(f'[FILA] {ctx["proc_id"]}: Airtable rejeitou também o payload mínimo: {exc2}')
                return {
                    'acao': 'holerite_falhou',
                    'status_final': 'Erro',
                    'detalhes': {
                        'funcionario_id': func_id,
                        'funcionario_nome': nome,
                        'cpf_extraido': cpf,
                        'folha_mensal': folha_mensal,
                        'valores': valores,
                    },
                    'pendencia': {
                        'tipo': 'Erro Airtable ao criar Holerite',
                        'observacao': (
                            f'Falha ao criar holerite para {nome} ({folha_mensal}) mesmo com '
                            f'payload mínimo. Erro completo: {exc} | Erro do payload mínimo: {exc2}'
                        ),
                    },
                }
            acao = 'holerite_criado_revisao_manual'
            status_final = 'Concluído'
            pendencia = {
                'tipo': 'Holerite criado em Revisão Manual',
                'observacao': (
                    f'Holerite {holerite_id} ({nome} - {folha_mensal}) criado com status '
                    f'"Revisão Manual" — valores financeiros não foram gravados devido a '
                    f'erro do Airtable: {exc}. PDF anexado normalmente; preencher os '
                    f'valores manualmente.'
                ),
            }

    return {
        'acao': acao,
        'status_final': status_final,
        'detalhes': {
            'holerite_id': holerite_id,
            'funcionario_id': func_id,
            'funcionario_nome': nome,
            'cpf_extraido': cpf,
            'folha_mensal': folha_mensal,
            'valores': valores,
        },
        'pendencia': pendencia,
    }


CAMPOS_CONTRATO_OBRIGATORIOS = ['nome_funcionario', 'cpf', 'data_admissao']

# 'tipo_local' é metadado de classificação (sempre preenchido, mesmo que
# "desconhecido") e não entra no cálculo de confiança da extração.
CAMPOS_PARA_CONFIANCA = [
    'nome_funcionario', 'cpf', 'data_admissao', 'cargo_funcao',
    'local_posto', 'salario', 'jornada_escala', 'cnpj_empresa',
]

# ── Padrão genérico de decisão (reaproveitável para outros documentos) ──────
#
# Toda decisão automática de um handler de documento se encaixa em uma destas
# 5 categorias (ver função decidir_acao_documento):
#   - executar_automaticamente          -> ação principal sem revisão humana
#   - executar_com_status_intermediario -> registro/ação em estado de transição
#   - funcionario_ou_registro_ja_existente
#                                        -> já existe registro equivalente, nada a fazer
#   - enviar_para_revisao               -> Pendência, humano decide
#   - ignorar_documento                 -> documento não aplicável a este fluxo
#
# Automação é a regra; revisão manual é a exceção (erros, divergências,
# duplicidades, baixa confiança ou dados essenciais faltantes).
#
# Para contratos (Fase 5C) esse padrão se traduz em 'decisao_5c':
#   executar_automaticamente              -> cadastrar_ativo_automaticamente
#   executar_com_status_intermediario     -> criar_pre_cadastro_seguro
#   funcionario_ou_registro_ja_existente  -> funcionario_ja_existente
#   enviar_para_revisao                   -> enviar_para_revisao
#   ignorar_documento                     -> não usado nesta fase (reservado)
#
# ── Tipos de documento futuros (apenas previstos, SEM processamento real) ───
#
# Os tipos abaixo já existem como possíveis 'Tipo de Documento' em Processar
# Arquivos, mas continuam mapeados para None em PROCESSADORES_DOCUMENTO (ver
# registro de handlers mais abaixo) — /processar-fila retorna 400 "ainda não
# implementado" para eles, sem qualquer leitura/escrita adicional. Quando
# forem implementados, cada um receberá um handler próprio que também deverá
# produzir uma das 5 categorias genéricas acima, por exemplo:
#   - Folha de Ponto        -> tipicamente executar_automaticamente (registrar
#                              ponto) ou enviar_para_revisao (inconsistências
#                              de horário/falta).
#   - Certidões             -> executar_automaticamente (arquivar/atualizar
#                              validade) ou enviar_para_revisao (certidão
#                              vencida/divergente).
#   - Extrato da Folha      -> executar_com_status_intermediario (conciliação
#                              pendente de confirmação) ou enviar_para_revisao.
#   - FGTS Digital          -> executar_automaticamente (registrar guia/valor)
#                              ou enviar_para_revisao (divergência de valores).
#   - Guias/Boletos         -> executar_automaticamente (registrar pagamento)
#                              ou enviar_para_revisao (vencimento/valor
#                              inconsistente).
#   - Notas Fiscais         -> executar_automaticamente (registrar nota) ou
#                              enviar_para_revisao (dados fiscais incompletos).
#   - Outros docs por e-mail -> ignorar_documento (não aplicável) ou
#                              enviar_para_revisao (encaminhar para triagem).
# Nenhum desses handlers é implementado nesta etapa; nenhuma extração real,
# leitura adicional ou escrita é feita para esses tipos.

# Fase 5C — pré-cadastro/cadastro automático seguro em Funcionários
CONFIANCA_MINIMA_PRE_CADASTRO = 0.70   # abaixo disso -> enviar_para_revisao
CONFIANCA_ATIVO_AUTOMATICO = 0.85      # acima/igual -> cadastrar_ativo_automaticamente

# Campos do pré-cadastro: nome legível -> field ID em Funcionários.
# Somente campos de escrita simples e já existentes no schema (nenhum campo novo).
CAMPOS_FUNC_PRE_CADASTRO = {
    'Nome Completo':    F_FUNC_NOME,
    'CPF':              F_FUNC_CPF,
    'Status':           F_FUNC_STATUS,
    'Cargo':            F_FUNC_CARGO,
    'Data de Admissão': F_FUNC_ADMISSAO,
    'WhatsApp':         'WhatsApp',
}

# Termos cujo aparecimento no nome extraído indicam resíduo de texto do
# contrato (regex capturou além do nome) — usados por _nome_suspeito().
_TERMOS_NOME_SUSPEITO = (
    'PORTADOR', 'PORTADORA', 'CARTEIRA', 'CTPS', 'RESIDENTE',
    'DORAVANTE', 'CONTRATO', 'CONTRATAD', 'EMPREGAD', 'EMPRESA',
)


def _nome_suspeito(nome):
    """
    Heurística para 'nome suspeito': vazio, muito curto, contendo dígitos
    ou termos residuais de contrato (sinal de captura incorreta pela regex).
    """
    if not nome:
        return True
    nome = nome.strip()
    if len(nome) < 5:
        return True
    if re.search(r'\d', nome):
        return True
    nome_upper = nome.upper()
    return any(termo in nome_upper for termo in _TERMOS_NOME_SUSPEITO)


def _data_br_para_iso(data_br):
    """Converte 'DD/MM/AAAA' para 'AAAA-MM-DD' (formato aceito pelo campo date). None se inválido."""
    if not data_br:
        return None
    try:
        return datetime.strptime(data_br, '%d/%m/%Y').strftime('%Y-%m-%d')
    except ValueError:
        return None


def decidir_acao_documento(*, erros_bloqueantes, avisos_qualidade, confianca,
                            confianca_minima, confianca_ativo,
                            registro_existente, divergencia):
    """
    Decisão genérica reaproveitável (ver comentário acima de
    CONFIANCA_MINIMA_PRE_CADASTRO) — produz uma das 5 categorias:
      - 'enviar_para_revisao'
      - 'funcionario_ou_registro_ja_existente'
      - 'executar_automaticamente'
      - 'executar_com_status_intermediario'
      ('ignorar_documento' não é produzido aqui — reservado para handlers
      futuros que precisem descartar um documento sem ação nenhuma).

    erros_bloqueantes: lista de motivos que, se não vazia, força revisão
        (dados essenciais inválidos/ausentes, local não identificado etc.).
    avisos_qualidade: lista de avisos da extração — também força revisão.
    registro_existente: True se já existe um registro equivalente (ex.: CPF
        já cadastrado em Funcionários).
    divergencia: motivo de divergência com o registro existente, ou None.
    """
    if erros_bloqueantes:
        return 'enviar_para_revisao', '; '.join(erros_bloqueantes)
    # Aprovação por exceção (v2.66): avisos de qualidade e confiança baixa NÃO
    # retêm mais o documento — o cadastro é criado imediatamente com status
    # intermediário ("Validação Pendente") para o humano revisar/ajustar depois.
    # Só erros verdadeiramente bloqueantes (CPF inválido, nome ausente) e
    # divergência com registro já existente continuam indo para revisão sem criar.
    if avisos_qualidade:
        return 'executar_com_status_intermediario', 'Avisos de qualidade na extração: ' + '; '.join(avisos_qualidade)
    if confianca < confianca_minima:
        return 'executar_com_status_intermediario', (
            f'Confiança da extração ({confianca}) abaixo do mínimo '
            f'({confianca_minima}) — criado como Validação Pendente (aprovação por exceção).'
        )
    if registro_existente:
        if divergencia:
            return 'enviar_para_revisao', divergencia
        return 'funcionario_ou_registro_ja_existente', (
            'Registro equivalente já existe e dados compatíveis — nenhuma ação necessária.'
        )
    if confianca >= confianca_ativo:
        return 'executar_automaticamente', (
            f'Dados essenciais válidos, sem avisos de qualidade e '
            f'confiança {confianca} >= {confianca_ativo} — apto para execução automática.'
        )
    return 'executar_com_status_intermediario', (
        f'Dados essenciais válidos, mas confiança {confianca} entre '
        f'{confianca_minima} e {confianca_ativo} — execução em estado intermediário de segurança.'
    )


def _montar_campos_pre_cadastro(dados: dict, status: str) -> dict:
    """
    Monta os campos (nome legível -> valor) do cadastro em Funcionários.

    'status' é "Ativo" (cadastrar_ativo_automaticamente) ou "Pré-cadastro"
    (criar_pre_cadastro_seguro), conforme decidido em _processar_contrato_stub.

    Inclui apenas campos mapeados e de escrita simples (Nome Completo, CPF,
    Status, Cargo, Data de Admissão). "Locais de trabalho" é um vínculo para
    a tabela Locais e "Origem do cadastro/Contrato de origem" não tem campo
    equivalente em Funcionários — nenhum dos dois é incluído aqui
    (ver 'observacoes' no retorno do handler).
    """
    campos = {
        'Nome Completo': dados['nome_funcionario'],
        'CPF': dados['cpf'],
        'Status': status,
    }
    if dados.get('cargo_funcao'):
        campos['Cargo'] = dados['cargo_funcao']
    data_iso = _data_br_para_iso(dados.get('data_admissao'))
    if data_iso:
        campos['Data de Admissão'] = data_iso
    if dados.get('telefone'):
        whatsapp_normalizado = _normalizar_numero_evolution(dados['telefone'])
        if whatsapp_normalizado:
            campos['WhatsApp'] = whatsapp_normalizado
    return campos


def _criar_pre_cadastro_funcionario(campos_legiveis: dict) -> str:
    """Cria registro em Funcionários a partir dos campos mapeados (Status="Ativo" ou "Pré-cadastro")."""
    campos_at = {
        CAMPOS_FUNC_PRE_CADASTRO[nome]: valor
        for nome, valor in campos_legiveis.items()
        if nome in CAMPOS_FUNC_PRE_CADASTRO
    }
    return _criar_registro(TABLE_FUNC, campos_at)


# ── Kit de Admissão — agrupamento de Contrato + Ficha de Registro + Vale-Transporte ──
#
# Regra de negócio (definida pelo usuário em 2026-06-22): a contabilidade
# sempre envia esses 3 documentos JUNTOS, no mesmo e-mail, para o mesmo
# colaborador. Decisões:
#   1. Ficha de Registro de Empregado é a FONTE PRIMÁRIA de dados (mais
#      completa: nome, CPF, nascimento, admissão, cargo, telefone) — o
#      Contrato de Experiência/Trabalho NUNCA cria o cadastro se sua Ficha
#      de Registro irmã ainda não foi processada (ver _processar_contrato_stub).
#   2. Contrato e Vale-Transporte (tipo "Outro") são vinculados ao MESMO
#      funcionário — anexados ao campo "Documentos Admissionais" — em vez
#      de ficarem soltos sem ligação com o cadastro.
#
# Detecção de "documento irmão" (mesmo Kit de Admissão): não há campo de
# vínculo direto entre Processar Arquivos de attachments diferentes do
# mesmo e-mail acessível por filterByFormula (linked record), então usamos
# uma janela de tempo curta em "Data Processo" — na prática, o robô de
# e-mail (Apps Script) processa todos os anexos de um mesmo e-mail em
# sequência, poucos segundos um do outro (verificado empiricamente em
# produção: gaps de 2-8s entre anexos do mesmo e-mail vs. minutos entre
# e-mails diferentes). 20s é uma janela folgada e segura para esse padrão.
JANELA_KIT_ADMISSAO_SEGUNDOS = 20
TIPOS_KIT_ADMISSAO = ('Contrato de Experiência', 'Contrato de Trabalho',
                       'Ficha de Registro de Empregado', 'Outro')


def _buscar_documentos_irmaos_kit_admissao(ctx: dict):
    """
    Busca outros registros em "Processar Arquivos" cuja "Data Processo"
    caia dentro de ±JANELA_KIT_ADMISSAO_SEGUNDOS do registro atual (ctx) e
    cujo tipo esteja em TIPOS_KIT_ADMISSAO — heurística para "mesmo e-mail,
    mesmo Kit de Admissão" (ver comentário acima).

    Correção final (Macro 6A — revisão independente do patch, achado
    material): o vínculo com a mensagem de origem deixou de ser "reforço
    quando disponível" e passou a ser OBRIGATÓRIO. A janela de tempo
    sozinha NUNCA prova pertencimento — só restringe o universo de busca.
    Sem `ctx['email_savian_id']` (origem determinística ausente), não há
    como confirmar que QUALQUER candidato pertence à mesma mensagem: a
    função devolve lista vazia, em vez de cair de volta para a janela de
    tempo desprotegida como na versão anterior desta correção. Chamadores
    que não fornecem `email_savian_id` (antigos ou de teste) passam a
    nunca agrupar nada — comportamento seguro por padrão ("origem de
    e-mail ausente → rejeita"), nunca mais permissivo por omissão.

    Mesma disciplina para cada candidato: sem Arquivo vinculado, ou sem
    conseguir confirmar o e-mail de origem dele (falha de leitura, campo
    vazio, ou apontando para OUTRA mensagem), o candidato nunca entra na
    lista — silenciosamente descartado AQUI só significa "não é irmão
    confirmado desta mensagem", nunca "documento perdido": ele segue seu
    próprio processamento normal via /processar-fila como documento
    independente.

    Retorna lista de dicts: {proc_id, tipo_documento, status, arquivo_id, nome_arquivo}.
    Nunca inclui o próprio ctx['proc_id'].
    """
    email_savian_id_atual = ctx.get('email_savian_id')
    if not email_savian_id_atual:
        return []
    _at_throttle()
    r = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_PROCESSAR}/{ctx["proc_id"]}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        timeout=15,
    )
    if not r.ok:
        return []
    data_processo_iso = r.json().get('fields', {}).get(F_PROC_DATA)
    if not data_processo_iso:
        return []
    try:
        centro = datetime.fromisoformat(data_processo_iso.replace('Z', '+00:00'))
    except ValueError:
        return []
    inicio = (centro - timedelta(seconds=JANELA_KIT_ADMISSAO_SEGUNDOS)).isoformat()
    fim = (centro + timedelta(seconds=JANELA_KIT_ADMISSAO_SEGUNDOS)).isoformat()

    formula = (
        f'AND(IS_AFTER({{Data Processo}}, "{inicio}"), '
        f'IS_BEFORE({{Data Processo}}, "{fim}"))'
    )
    _at_throttle()
    r2 = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_PROCESSAR}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        params={'filterByFormula': formula, 'maxRecords': 20},
        timeout=30,
    )
    if not r2.ok:
        return []

    irmaos = []
    for rec in r2.json().get('records', []):
        if rec['id'] == ctx['proc_id']:
            continue
        fields = rec.get('fields', {})
        tipo = fields.get('Tipo de Documento')
        if tipo not in TIPOS_KIT_ADMISSAO:
            continue
        arquivos_link = fields.get('Arquivos 2') or []
        arquivo_id = arquivos_link[0]['id'] if arquivos_link and isinstance(arquivos_link[0], dict) else (arquivos_link[0] if arquivos_link else None)
        if not arquivo_id:
            continue  # sem Arquivo vinculado -- nunca confirma origem, nunca inclui

        _at_throttle()
        r_arq_irmao = requests.get(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ARQUIVOS}/{arquivo_id}',
            headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
            params={'returnFieldsByFieldId': 'true'},
            timeout=15,
        )
        if not r_arq_irmao.ok:
            continue  # sem confirmação de origem -- nunca inclui por omissão
        email_links_irmao = r_arq_irmao.json().get('fields', {}).get(F_ARQ_EMAILS) or []
        email_id_irmao = (
            email_links_irmao[0]['id'] if email_links_irmao and isinstance(email_links_irmao[0], dict)
            else (email_links_irmao[0] if email_links_irmao else None)
        )
        if email_id_irmao != email_savian_id_atual:
            continue  # candidato de OUTRA mensagem de origem -- nunca agrupa

        irmaos.append({
            'proc_id': rec['id'],
            'tipo_documento': tipo,
            'status': fields.get('Status'),
            'arquivo_id': arquivo_id,
        })
    return irmaos


def _vincular_documento_ao_funcionario(func_id: str, arquivo_id: str):
    """
    Anexa o PDF do Arquivo informado ao campo "Documentos Admissionais" do
    Funcionário — usado para agrupar Contrato/Vale-Transporte ao mesmo
    cadastro criado a partir da Ficha de Registro (ou vice-versa). Evita
    duplicar o anexo se ele já estiver lá (compara por nome do arquivo).
    """
    if not arquivo_id:
        return False
    _at_throttle()
    r_arq = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ARQUIVOS}/{arquivo_id}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        params={'returnFieldsByFieldId': 'true'},
        timeout=15,
    )
    if not r_arq.ok:
        return False
    arquivo_fields = r_arq.json().get('fields', {})
    attachments = arquivo_fields.get(F_ARQ_ATTACH) or []
    if not attachments:
        return False
    pdf_info = attachments[0]
    nome_arquivo = pdf_info.get('filename', 'documento.pdf')

    _at_throttle()
    r_func = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}/{func_id}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        params={'returnFieldsByFieldId': 'true'},
        timeout=15,
    )
    if r_func.ok:
        ja_anexados = r_func.json().get('fields', {}).get(F_FUNC_DOCS_ADMISSAO) or []
        if any(a.get('filename') == nome_arquivo for a in ja_anexados):
            return False  # já vinculado — não duplica

    pdf_bytes = requests.get(pdf_info['url'], timeout=60).content
    _anexar_attachment(TABLE_FUNC, func_id, F_FUNC_DOCS_ADMISSAO, pdf_bytes, nome_arquivo)
    return True


def _baixar_pdf_arquivo(arquivo_id: str):
    """Retorna (conteudo_bytes, filename) do PDF anexado a um registro de
    Arquivos, ou (None, None) se não existir/falhar."""
    if not arquivo_id:
        return None, None
    _at_throttle()
    r = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ARQUIVOS}/{arquivo_id}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        params={'returnFieldsByFieldId': 'true'},
        timeout=15,
    )
    if not r.ok:
        return None, None
    attachments = r.json().get('fields', {}).get(F_ARQ_ATTACH) or []
    if not attachments:
        return None, None
    pdf_info = attachments[0]
    conteudo = requests.get(pdf_info['url'], timeout=60).content
    return conteudo, pdf_info.get('filename', 'documento.pdf')


def _vale_transporte_preenchido(arquivo_id: str, cpf_funcionario: str) -> bool:
    """
    Verifica se um documento de Vale-Transporte (tipo "Outro") já vem
    preenchido com o CPF do funcionário — se sim, é o documento real da
    contabilidade e deve ser usado como está. Se vier em branco/sem CPF
    correspondente (ou não existir), o robô gera a declaração nativamente
    (ver _gerar_pdf_declaracao_renuncia_vt).
    """
    conteudo, _ = _baixar_pdf_arquivo(arquivo_id)
    if not conteudo:
        return False
    try:
        with pdfplumber.open(io.BytesIO(conteudo)) as pdf:
            texto = '\n'.join((p.extract_text() or '') for p in pdf.pages)
    except Exception:
        return False
    cpf_doc = extrair_cpf(texto)
    cpf_doc_num = re.sub(r'\D', '', cpf_doc or '')
    cpf_func_num = re.sub(r'\D', '', cpf_funcionario or '')
    return bool(cpf_doc_num) and cpf_doc_num == cpf_func_num


def _gerar_pdf_declaracao_renuncia_vt(nome: str, cpf: str, data_admissao_br: str = None) -> bytes:
    """
    Gera nativamente (fpdf2 — já é dependência do projeto, sem custo extra)
    a Declaração de Renúncia de Vale-Transporte, no mesmo padrão de texto
    usado pela contabilidade (ver "Modelo 1.pdf" real, calibrado em
    2026-06-22), preenchendo Nome e CPF extraídos da Ficha de Registro.
    """
    from fpdf import FPDF

    _TZ = timezone(timedelta(hours=-3))
    data_admissao_br = data_admissao_br or datetime.now(tz=_TZ).strftime('%d/%m/%Y')
    agora = datetime.now(tz=_TZ)
    data_extenso = agora.strftime('%d de %m de %Y')

    pdf = FPDF()
    pdf.add_page()
    largura = pdf.w - pdf.l_margin - pdf.r_margin

    pdf.set_font('Helvetica', 'B', 12)
    pdf.cell(largura, 8, 'DECLARACAO DE RENUNCIA DO VALE TRANSPORTE', align='C', ln=True)
    pdf.ln(4)

    pdf.set_font('Helvetica', '', 11)
    pdf.cell(largura, 6, 'A', ln=True)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(largura, 6, 'MAGNATA PORTARIA E SERVICOS LTDA', ln=True)
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(largura, 6, 'RUA R FERMINO JOSE DE ARAUJO, VILA NOVA, ITAPETININGA - SP', ln=True)
    pdf.cell(largura, 6, 'CNPJ: 17.987.187/0001-61', ln=True)
    pdf.ln(4)

    texto_declaracao = (
        f'Eu, {nome}, portador(a) do CPF no {cpf}, empregado(a) de MAGNATA PORTARIA '
        f'E SERVICOS LTDA, admitido(a) em {data_admissao_br}, declaro que nao vou '
        f'utilizar o beneficio do "Vale Transporte", desde ja isentando esta empresa '
        f'do pagamento deste beneficio.'
    )
    pdf.multi_cell(largura, 6, texto_declaracao)
    pdf.ln(8)
    pdf.cell(largura, 6, f'ITAPETININGA, {data_extenso}.', ln=True)
    pdf.ln(10)
    pdf.cell(largura, 6, '_' * 50, ln=True)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(largura, 6, nome, ln=True)
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(largura, 6, f'CPF: {cpf}', ln=True)
    pdf.ln(6)
    pdf.set_font('Helvetica', 'I', 8)
    pdf.multi_cell(
        largura, 5,
        'Documento gerado automaticamente pelo sistema de processamento de admissoes '
        '(robo Magnata) por ausencia de declaracao preenchida da contabilidade.'
    )

    return bytes(pdf.output())


def _montar_pdf_consolidado_kit_admissao(pdf_ficha: bytes, pdf_contrato: bytes = None,
                                          pdf_vt: bytes = None) -> bytes:
    """
    Funde, na ordem Ficha de Registro + Contrato + Termo de Renúncia de VT,
    os PDFs do Kit de Admissão num único arquivo (via pypdf — nativo, sem
    custo). Documentos ausentes (None) são simplesmente pulados, mantendo
    a ordem dos demais.
    """
    writer = PdfWriter()
    for conteudo in (pdf_ficha, pdf_contrato, pdf_vt):
        if not conteudo:
            continue
        reader = PdfReader(io.BytesIO(conteudo))
        for page in reader.pages:
            writer.add_page(page)
    buffer = io.BytesIO()
    writer.write(buffer)
    return buffer.getvalue()


def _classificar_papel_kit_por_conteudo(texto: str) -> str:
    """
    Determina o PAPEL de um documento dentro do Kit de Admissão (ficha /
    contrato / vt) pelo CONTEÚDO real do PDF — nunca pelo campo "Tipo de
    Documento" salvo em Processar Arquivos, que pode estar desatualizado
    (achado real, 22/06/2026: 3 Fichas de Registro foram classificadas
    como "Contrato de Trabalho" por uma corrida de deploy do classificador
    geral, e como o agrupamento do Kit confiava nesse campo, o Kit nunca
    foi montado). Ao reclassificar aqui, o agrupamento fica imune a esse
    tipo de erro passado ou futuro, seja qual for a causa.
    """
    if not texto:
        return 'contrato'  # fallback conservador — sobra só esse papel
    if (re.search(r'Ficha\s+de\s+Registro\s+de\s+Empregados?', texto, re.IGNORECASE)
            or re.search(r'Registro\s+de\s+Empregados?\b', texto, re.IGNORECASE)
            or re.search(r'Matr[íi]cula\s+eSocial', texto, re.IGNORECASE)
            or re.search(r'Livro\s+(?:de\s+)?Registro\s+de\s+Empregados?', texto, re.IGNORECASE)):
        return 'ficha'
    if (re.search(r'Vale[\s-]+Transporte', texto, re.IGNORECASE)
            and re.search(r'Ren[úu]ncia', texto, re.IGNORECASE)):
        return 'vt'
    return 'contrato'


def _baixar_e_classificar_papel_kit(arquivo_id: str):
    """Baixa o PDF de um Arquivo e devolve (papel, conteudo_bytes, texto) —
    papel via _classificar_papel_kit_por_conteudo. `texto` é devolvido para
    permitir checagem de identidade (CPF, ver _cpf_compativel_com_kit) por
    quem chama, sem baixar/extrair o PDF de novo. (None, None, '') se
    falhar."""
    conteudo, _ = _baixar_pdf_arquivo(arquivo_id)
    if not conteudo:
        return None, None, ''
    try:
        with pdfplumber.open(io.BytesIO(conteudo)) as pdf:
            texto = '\n'.join((p.extract_text() or '') for p in pdf.pages)
    except Exception:
        texto = ''
    return _classificar_papel_kit_por_conteudo(texto), conteudo, texto


def _cpf_compativel_com_kit(cpf_extraido, cpf_esperado):
    """Identidade forte e OBRIGATÓRIA para o agrupamento do Kit de Admissão.

    Correção final (Macro 6A — revisão independente do patch, achado
    material): a versão anterior devolvia True ("compatível, pode
    vincular") quando qualquer um dos dois CPFs estava ausente — sob a
    lógica de "sem como provar divergência". Isso contradiz a regra de
    identidade forte aprovada: ausência de contradição NUNCA é confirmação
    positiva. A regra agora exige confirmação POSITIVA dos dois lados —
    só devolve True quando `cpf_extraido` E `cpf_esperado` estão presentes
    E são iguais (comparação só de dígitos, formato ignorado). Qualquer
    ausência (candidato sem CPF legível, ou funcionário-alvo sem CPF
    cadastrado) ou divergência devolve False — o candidato nunca é
    vinculado nem incorporado ao Kit nesses casos, mas também nunca é
    descartado silenciosamente: quem chama (_montar_e_disparar_kit_admissao)
    encaminha para revisão humana com motivo sanitizado. CPF é a única
    base de decisão aqui — nunca nome aproximado."""
    if not cpf_extraido or not cpf_esperado:
        return False
    return re.sub(r'\D', '', cpf_extraido) == re.sub(r'\D', '', cpf_esperado)


def _chave_pendencia_kit_identidade(func_id: str, proc_id: str, arquivo_id: str, motivo_sanitizado: str) -> str:
    """Chave determinística para idempotência da Pendência de rejeição de
    identidade do Kit de Admissão (Macro 6A — último fechamento).

    Formada só por IDs técnicos + a categoria sanitizada da rejeição —
    NUNCA CPF, nome ou trecho de PDF. A MESMA combinação (funcionário-alvo,
    processo/registro de origem, arquivo rejeitado, categoria) produz
    SEMPRE a mesma chave. Reaproveitada como valor de F_PEND_NOME — não há
    campo dedicado de "chave de idempotência" em Pendências, e criar um
    exigiria alterar o schema do Airtable (fora do escopo desta fase);
    F_PEND_NOME já é texto livre e serve aos dois propósitos: rótulo
    legível por humano E identidade determinística para busca."""
    return (
        f'Kit de Admissão — identidade não confirmada [{motivo_sanitizado}] '
        f'funcionario={func_id} origem={proc_id} arquivo={arquivo_id}'
    )


def _buscar_pendencia_kit_identidade_existente(chave: str):
    """Busca (somente leitura) uma Pendência já existente com a MESMA
    chave determinística — base da idempotência: nunca deixa criar uma
    segunda Pendência para a mesma rejeição.

    Não usa filterByFormula: o nome de EXIBIÇÃO real do campo F_PEND_NOME
    nesta base não é conhecido com segurança nesta sessão (sem conector
    Airtable habilitado para confirmar o schema), e uma fórmula com o nome
    errado falharia silenciosamente ou de forma imprevisível. Em vez
    disso, pagina a tabela usando Field IDs (`returnFieldsByFieldId=true`,
    já conhecidos e estáveis — os mesmos usados em toda a leitura desta
    base) e compara a chave no lado do Python.

    Se houver mais de uma Pendência com a mesma chave (ex.: sobra de uma
    corrida concorrente antiga, antes desta correção), devolve a PRIMEIRA
    encontrada de forma determinística — nunca tenta mesclar as duas nem
    associa a rejeição à Pendência errada.

    Falha de leitura (rede, HTTP, JSON) devolve None — o chamador trata
    como "não encontrada" e segue para criar uma nova. Prioridade
    deliberada: nunca deixar a rejeição sem NENHUM registro por causa de
    uma falha nesta busca (ver _criar_ou_reaproveitar_pendencia_kit_identidade)."""
    offset = None
    while True:
        params = {'returnFieldsByFieldId': 'true'}
        if offset:
            params['offset'] = offset
        _at_throttle()
        try:
            r = requests.get(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_PENDENCIAS}',
                headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
                params=params, timeout=30,
            )
        except Exception:
            return None
        if not r.ok:
            return None
        try:
            data = r.json()
        except Exception:
            return None
        for rec in data.get('records', []):
            if rec.get('fields', {}).get(F_PEND_NOME) == chave:
                return rec['id']
        offset = data.get('offset')
        if not offset:
            return None


def _marcar_pendencia_kit_identidade_vista_novamente(pendencia_id: str) -> bool:
    """Atualização NÃO destrutiva de uma Pendência já existente,
    reaproveitada por idempotência: só o carimbo de data/hora (F_PEND_DATA)
    é tocado — Status, Tipo, Nome e Observação continuam exatamente como
    estavam. Nunca reabre nem sobrescreve uma decisão humana já registrada
    em Status (ex.: um revisor já pode ter marcado como resolvida). Serve
    só para sinalizar "esta rejeição ainda está acontecendo" sem duplicar
    o registro. Não-bloqueante: qualquer falha devolve False, nunca levanta."""
    try:
        _at_throttle()
        r = requests.patch(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_PENDENCIAS}/{pendencia_id}',
            headers=_at_headers(),
            json={'fields': {F_PEND_DATA: datetime.now().isoformat()}, 'typecast': True},
            timeout=15,
        )
        return r.ok
    except Exception:
        return False


def _criar_ou_reaproveitar_pendencia_kit_identidade(func_id: str, proc_id: str, arquivo_id: str,
                                                      motivo_sanitizado: str) -> dict:
    """Idempotência da Pendência de rejeição de identidade do Kit de
    Admissão (Macro 6A — último fechamento, achado do Ultrareview:
    reprocessar o mesmo documento rejeitado podia criar Pendências
    duplicadas).

    Antes de criar: procura uma Pendência já existente com a MESMA chave
    determinística (funcionário-alvo + processo de origem + arquivo
    rejeitado + categoria — nunca CPF/nome/texto do PDF). Se existir,
    REAPROVEITA o mesmo pendencia_id (só atualiza o carimbo de data, de
    forma não destrutiva) — nunca cria uma segunda. Se não existir, cria
    uma nova.

    Qualquer falha (busca ou criação) é não-bloqueante: log sanitizado
    (só IDs técnicos) e devolve pendencia_id=None — nunca derruba o
    processamento dos demais documentos do Kit (mesma disciplina já
    usada em todo o resto desta função).

    Devolve {'pendencia_id': str | None, 'reaproveitada': bool}."""
    chave = _chave_pendencia_kit_identidade(func_id, proc_id, arquivo_id, motivo_sanitizado)

    pendencia_existente_id = _buscar_pendencia_kit_identidade_existente(chave)
    if pendencia_existente_id:
        _marcar_pendencia_kit_identidade_vista_novamente(pendencia_existente_id)
        return {'pendencia_id': pendencia_existente_id, 'reaproveitada': True}

    try:
        pendencia_id = _criar_pendencia(
            arquivo_id,
            'Kit de Admissão — identidade não confirmada',
            (
                f'Documento com o mesmo e-mail de origem do Kit de {func_id} '
                f'(Processar: {proc_id}, Arquivo: {arquivo_id}), mas a identidade '
                f'por CPF não pôde ser confirmada (motivo: {motivo_sanitizado}). '
                'NÃO vinculado nem incorporado ao Kit — revisar manualmente e '
                'vincular/anexar com decisão humana.'
            ),
            nome=chave,
        )
        return {'pendencia_id': pendencia_id, 'reaproveitada': False}
    except Exception as exc:
        logger.warning(
            f'[KIT] Falha ao criar Pendência de identidade para '
            f'{proc_id}/{arquivo_id}: {exc} — não bloqueante.'
        )
        return {'pendencia_id': None, 'reaproveitada': False}


def _montar_e_disparar_kit_admissao(ctx: dict, func_id: str, dry_run: bool,
                                     dados_para_vt: dict = None) -> dict:
    """
    Automação real (v3.3, 06/07/2026) de montagem + disparo do Kit de
    Admissão — substitui o processo que antes exigia auditoria manual
    periódica para achar Kits parados.

    Reúne o documento atual + os irmãos (mesmo e-mail, ver
    _buscar_documentos_irmaos_kit_admissao), reclassifica CADA UM pelo
    conteúdo real (não pelo "Tipo de Documento" salvo — ver
    _classificar_papel_kit_por_conteudo), monta o PDF consolidado assim
    que a Ficha de Registro é identificada (gera a Declaração de
    Renúncia de VT nativamente se não houver uma irmã preenchida) e
    dispara a Assinatura Nativa via WhatsApp automaticamente.

    Idempotente: a MONTAGEM/DISPARO do Kit nunca repete se o Funcionário já
    tiver Kit Admissão Consolidado — mas o documento atual É SEMPRE
    vinculado ao funcionário (mesmo nesse caso, ver
    'documento_atual_vinculado_tardiamente' no retorno), para nunca deixar
    um anexo atrasado órfão (Macro 6A — fechamento autônomo, §6). Chamado
    tanto por _processar_ficha_registro_stub quanto por
    _processar_contrato_stub — qualquer um dos documentos do Kit que for
    processado primeiro pelo /processar-fila completa o trabalho.
    """
    resultado = {'kit_montado': False, 'assinatura_disparada': False, 'motivo': None}
    if dry_run or not func_id:
        resultado['motivo'] = 'dry_run ou funcionario_id ausente'
        return resultado

    # ✅ v3.6: DEFESA PROFUNDA — Validar configuração antes de qualquer operação
    config_ok, config_msg = _validar_configuracao_assinatura_v36()
    if not config_ok:
        logger.error(f'[KIT ADMISSAO] Configuracao invalida: {config_msg}')
        resultado['motivo'] = f'Configuracao v3.6 incompleta — Kit não pode prosseguir'
        return resultado

    _at_throttle()
    r_func = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}/{func_id}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        params={'returnFieldsByFieldId': 'true'}, timeout=15,
    )

    # Vincula o documento ATUAL ao funcionário incondicionalmente, antes de
    # checar se o Kit já foi concluído (Macro 6A — fechamento autônomo,
    # revisão adversarial §6, achado "anexos atrasados"): antes desta
    # correção, quando o Kit já estava concluído (ex.: Ficha assemblou o
    # Kit sozinha, dentro da janela, e o Contrato real chegou depois,
    # fora da janela ou simplesmente processado depois), o documento
    # atrasado NUNCA era vinculado ao funcionário — ficava silenciosamente
    # sem nenhum rastro no cadastro correto, mesmo sendo o documento real
    # (ex.: o Contrato assinado de verdade, quando o Kit já foi montado só
    # com Ficha + Declaração de VT gerada automaticamente). Vincular aqui é
    # uma operação separada, idempotente por nome de arquivo (ver
    # _vincular_documento_ao_funcionario) e não depende de o Kit ainda
    # estar em aberto — sempre seguro, mesmo que o Kit já tenha sido
    # concluído ou disparado.
    _vincular_documento_ao_funcionario(func_id, ctx['arquivo_id'])

    if r_func.ok and (r_func.json().get('fields', {}).get(F_FUNC_KIT_CONSOLIDADO) or []):
        resultado['motivo'] = (
            'Kit já existente — nada a fazer na montagem/disparo, mas o '
            'documento atual foi vinculado ao funcionário para não ficar '
            'órfão (chegada tardia).'
        )
        resultado['documento_atual_vinculado_tardiamente'] = True
        return resultado

    # CPF do funcionário-alvo, já disponível na mesma leitura acima — base
    # de comparação para _cpf_compativel_com_kit (Macro 6A, correção de
    # auditoria: nunca vincula documento de OUTRA pessoa por coincidência
    # de janela de tempo).
    cpf_esperado = r_func.json().get('fields', {}).get(F_FUNC_CPF) if r_func.ok else None

    docs = {}
    papel_atual = _classificar_papel_kit_por_conteudo(ctx.get('texto', ''))
    if ctx.get('pdf_bytes'):
        docs[papel_atual] = (ctx['arquivo_id'], ctx['pdf_bytes'], ctx['proc_id'])

    irmaos_rejeitados_por_cpf = []
    pendencias_identidade_kit = []
    for irmao in _buscar_documentos_irmaos_kit_admissao(ctx):
        if not irmao.get('arquivo_id'):
            continue
        papel, conteudo, texto_irmao = _baixar_e_classificar_papel_kit(irmao['arquivo_id'])
        if not (papel and conteudo):
            continue
        cpf_irmao = extrair_cpf(texto_irmao)
        if not _cpf_compativel_com_kit(cpf_irmao, cpf_esperado):
            # Correção final (Macro 6A — revisão independente do patch,
            # achado material): candidato com o MESMO e-mail de origem
            # confirmado, mas identidade por CPF não comprovada (CPF do
            # candidato ausente, CPF do funcionário-alvo ausente, ou os
            # dois divergentes) — nunca vinculado, nunca incorporado ao
            # Kit, e NUNCA descartado silenciosamente: encaminhado para
            # revisão humana com motivo sanitizado (classificação
            # genérica, nunca o valor do CPF nem trecho do PDF) e
            # referência aos IDs técnicos (proc_id/arquivo_id/func_id)
            # para rastreabilidade.
            if not cpf_irmao:
                motivo_sanitizado = 'cpf_candidato_ausente'
            elif not cpf_esperado:
                motivo_sanitizado = 'cpf_esperado_ausente'
            else:
                motivo_sanitizado = 'cpf_divergente'
            logger.warning(
                f'[KIT] Documento irmão {irmao["proc_id"]}/{irmao["arquivo_id"]} '
                f'rejeitado do Kit de {func_id}: identidade não confirmada '
                f'({motivo_sanitizado}).'
            )
            irmaos_rejeitados_por_cpf.append(irmao['proc_id'])
            # Idempotência (Macro 6A — último fechamento, achado do
            # Ultrareview): reprocessar o mesmo documento/rejeição nunca
            # cria uma segunda Pendência — ver
            # _criar_ou_reaproveitar_pendencia_kit_identidade.
            pendencia_resultado = _criar_ou_reaproveitar_pendencia_kit_identidade(
                func_id, irmao['proc_id'], irmao['arquivo_id'], motivo_sanitizado,
            )
            pendencias_identidade_kit.append({
                'proc_id': irmao['proc_id'], 'arquivo_id': irmao['arquivo_id'],
                'pendencia_id': pendencia_resultado['pendencia_id'],
                'reaproveitada': pendencia_resultado['reaproveitada'],
                'motivo': motivo_sanitizado,
            })
            continue
        _vincular_documento_ao_funcionario(func_id, irmao['arquivo_id'])
        if papel not in docs:
            docs[papel] = (irmao['arquivo_id'], conteudo, irmao['proc_id'])
    if irmaos_rejeitados_por_cpf:
        resultado['irmaos_rejeitados_por_cpf'] = irmaos_rejeitados_por_cpf
    if pendencias_identidade_kit:
        resultado['pendencias_identidade_kit'] = pendencias_identidade_kit

    if 'ficha' not in docs:
        resultado['motivo'] = (
            'Ficha de Registro de Empregado ainda não identificada entre os '
            'documentos disponíveis — Kit aguarda.'
        )
        return resultado

    pdf_ficha_bytes = docs['ficha'][1]
    pdf_contrato_bytes = docs['contrato'][1] if 'contrato' in docs else None

    if 'vt' in docs:
        pdf_vt_bytes = docs['vt'][1]
    else:
        dados_para_vt = dados_para_vt or {}
        pdf_vt_bytes = _gerar_pdf_declaracao_renuncia_vt(
            dados_para_vt.get('nome_funcionario'), dados_para_vt.get('cpf'),
            dados_para_vt.get('data_admissao'),
        )

    pdf_consolidado = _montar_pdf_consolidado_kit_admissao(pdf_ficha_bytes, pdf_contrato_bytes, pdf_vt_bytes)
    nome_pessoa = (dados_para_vt or {}).get('nome_funcionario') or func_id
    nome_kit = f'Kit Admissao - {nome_pessoa}.pdf'
    _anexar_attachment(TABLE_FUNC, func_id, F_FUNC_KIT_CONSOLIDADO, pdf_consolidado, nome_kit)
    resultado['kit_montado'] = True

    # O registro que disparou esta chamada tem seu Status atualizado pelo
    # laço genérico de /processar-fila (via 'status_final' no retorno do
    # handler) — aqui cuidamos apenas dos OUTROS documentos envolvidos.
    for papel, (_arquivo_id, _conteudo, proc_id) in docs.items():
        if proc_id and proc_id != ctx['proc_id']:
            try:
                _atualizar_status_processar(proc_id, 'Concluído')
            except Exception:
                logger.warning(f'[KIT] Falha ao marcar {proc_id} ({papel}) como Concluído — não bloqueante.')

    # ✅ v3.6: Criar registro em Arquivos para PDF consolidado, obter Record ID
    arquivo_record_id = None
    try:
        _at_throttle()
        r_arquivo_novo = requests.post(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ARQUIVOS}',  # ✅ corrigido: tblbgJqXh0u1Cjq1s não existe na base
            headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'} ,
            json={
                'fields': {
                    F_ARQ_NOME: nome_kit,
                    F_ARQ_ATTACH: [{  # deve já ter sido criado acima
                        'url': '',  # Será atualizado
                        'filename': nome_kit,
                    }],
                }
            },
            timeout=15,
        )
        if r_arquivo_novo.ok:
            arquivo_record_id = r_arquivo_novo.json().get('id')
    except Exception as exc:
        logger.warning(f'[KIT] Falha ao criar Record em Arquivos ({func_id}): {exc}')

    try:
        _gerar_assinatura_core(
            funcionario_id=func_id,
            tipo_documento='KIT_ADMISSAO',  # ✅ Tipo canônico (corrigido de 'Kit de Admissão')
            arquivo_record_id=arquivo_record_id,  # ✅ Record ID obrigatório (corrigido de ausente)
            nome_documento=nome_kit,
            disparar_whatsapp=not dry_run,
            dry_run=dry_run,
        )
        resultado['assinatura_disparada'] = not dry_run
    except Exception as exc:
        logger.error(f'[KIT] Kit montado mas falha ao disparar assinatura ({func_id}): {exc}')
        resultado['motivo'] = f'Kit montado mas assinatura falhou: {exc}'

    return resultado


CAMPOS_FICHA_REGISTRO_OBRIGATORIOS = ['nome_funcionario', 'cpf', 'data_admissao']
CAMPOS_FICHA_REGISTRO_PARA_CONFIANCA = [
    'nome_funcionario', 'cpf', 'data_admissao', 'cargo_funcao', 'data_nascimento',
]


def extrair_dados_ficha_registro(texto: str):
    """
    Extrai dados de uma Ficha/Registro de Empregado (formulário eSocial —
    "REGISTRO DE EMPREGADO"). Esse documento costuma trazer mais campos que
    o Contrato de Experiência (data de nascimento, endereço, telefone), por
    isso ganha extrator próprio em vez de reaproveitar extrair_dados_contrato.

    Retorna (dados, avisos_qualidade) no mesmo formato de extrair_dados_contrato,
    para reaproveitar decidir_acao_documento / _montar_campos_pre_cadastro.
    """
    resultado = {
        'nome_funcionario': None, 'cpf': None, 'data_nascimento': None,
        'data_admissao': None, 'cargo_funcao': None, 'telefone': None,
    }
    avisos_qualidade = []
    if not texto:
        return resultado, avisos_qualidade

    resultado['cpf'] = extrair_cpf(texto)

    m = re.search(
        r'Empregado\s+Benefici[áa]rios\s*\n\s*([A-ZÀ-Ú][A-ZÀ-Ú ]{3,59})',
        texto, re.IGNORECASE
    )
    if m:
        nome = m.group(1).strip()
        if len(nome.split()) >= 2:
            resultado['nome_funcionario'] = nome

    m = re.search(
        r'Data\s+de\s+nascimento\s+Local.*?\n\s*(\d{2}/\d{2}/\d{4})',
        texto, re.IGNORECASE
    )
    if m:
        resultado['data_nascimento'] = m.group(1)

    m = re.search(
        r'Data\s+de\s+Admiss[ãa]o[^\n]*\n\s*(\d{2}/\d{2}/\d{4})',
        texto, re.IGNORECASE
    )
    if m:
        resultado['data_admissao'] = m.group(1)

    m = re.search(
        r'Cargo\s+Fun[çc][ãa]o\s+C\.B\.O\.\s*\n\s*([A-ZÀ-Ú][A-ZÀ-Ú /]+?)\s+\d{5,7}',
        texto, re.IGNORECASE
    )
    if m:
        resultado['cargo_funcao'] = re.sub(r'\s+', ' ', m.group(1)).strip()

    # Telefone — quando a ficha vem com o campo preenchido (nem sempre vem).
    m = re.search(
        r'Telefone\s+Celular\D{0,10}(\(?\d{2}\)?\s?9?\s?\d{4}-?\s?\d{4})',
        texto, re.IGNORECASE
    )
    if m:
        resultado['telefone'] = m.group(1)

    return resultado, avisos_qualidade


def _processar_ficha_registro_stub(ctx, dry_run):
    """
    Handler de Ficha de Registro de Empregado (Fase 5C — mesmo vocabulário de
    decisão usado em Contrato de Experiência/Trabalho).

    Extrai dados, consulta (somente leitura) se o CPF já existe em
    Funcionários, e decide entre cadastrar_ativo_automaticamente /
    criar_pre_cadastro_seguro / funcionario_ja_existente / enviar_para_revisao.
    Se a ficha trouxer telefone preenchido, ele já entra como WhatsApp no
    pré-cadastro (ver _montar_campos_pre_cadastro) — economiza o passo manual
    de preencher o telefone depois.
    """
    dados, avisos_qualidade = extrair_dados_ficha_registro(ctx['texto'])

    cpf_num = re.sub(r'\D', '', dados['cpf'] or '')

    campos_faltantes = [
        campo for campo in CAMPOS_FICHA_REGISTRO_OBRIGATORIOS if not dados.get(campo)
    ]

    total_campos = len(CAMPOS_FICHA_REGISTRO_PARA_CONFIANCA)
    preenchidos = sum(1 for c in CAMPOS_FICHA_REGISTRO_PARA_CONFIANCA if dados.get(c) is not None)
    confianca = round(preenchidos / total_campos, 2) if total_campos else 0.0
    confianca = round(max(0.0, confianca - 0.1 * len(avisos_qualidade)), 2)

    funcionario_existe = None
    nome_existente = None
    func_id = None
    if dados['cpf']:
        func_id, nome_existente = buscar_funcionario_por_cpf(dados['cpf'])
        funcionario_existe = func_id is not None

    divergencia = None
    if funcionario_existe and dados['nome_funcionario'] and nome_existente:
        if dados['nome_funcionario'].strip().upper() != nome_existente.strip().upper():
            divergencia = (
                f'Nome da ficha ("{dados["nome_funcionario"]}") difere do '
                f'cadastro existente em Funcionários ("{nome_existente}") para o mesmo CPF.'
            )

    nome_suspeito = _nome_suspeito(dados['nome_funcionario'])

    erros_bloqueantes = []
    if campos_faltantes:
        erros_bloqueantes.append(f'Campos obrigatórios faltando: {", ".join(campos_faltantes)}.')
    if not (bool(cpf_num) and _cpf_digitos_validos(cpf_num)):
        erros_bloqueantes.append('CPF inválido (dígitos verificadores não conferem).')
    if nome_suspeito:
        erros_bloqueantes.append(f'Nome do funcionário ausente ou suspeito: "{dados["nome_funcionario"]}".')

    categoria_generica, motivo_decisao = decidir_acao_documento(
        erros_bloqueantes=erros_bloqueantes,
        avisos_qualidade=avisos_qualidade,
        confianca=confianca,
        confianca_minima=CONFIANCA_MINIMA_PRE_CADASTRO,
        confianca_ativo=CONFIANCA_ATIVO_AUTOMATICO,
        registro_existente=funcionario_existe,
        divergencia=divergencia,
    )

    _MAPA_CATEGORIA_PARA_DECISAO_5C = {
        'executar_automaticamente': 'cadastrar_ativo_automaticamente',
        'executar_com_status_intermediario': 'criar_pre_cadastro_seguro',
        'funcionario_ou_registro_ja_existente': 'funcionario_ja_existente',
        'enviar_para_revisao': 'enviar_para_revisao',
    }
    decisao_5c = _MAPA_CATEGORIA_PARA_DECISAO_5C[categoria_generica]

    pre_cadastro_simulado = None
    pre_cadastro_id = None
    pendencia_simulada = None

    if decisao_5c in ('cadastrar_ativo_automaticamente', 'criar_pre_cadastro_seguro'):
        status_destino = 'Validação Pendente'  # aprovação por exceção (v2.66): sempre entra pendente de validação manual
        campos_pre_cadastro = _montar_campos_pre_cadastro(dados, status_destino)
        pre_cadastro_simulado = {'campos': campos_pre_cadastro}
        if not dry_run:
            func_id_recheck, _ = buscar_funcionario_por_cpf(dados['cpf'])
            if func_id_recheck:
                decisao_5c = 'funcionario_ja_existente'
                motivo_decisao = (
                    f'CPF já existe em Funcionários (ID={func_id_recheck}), detectado na '
                    f'rechecagem imediatamente antes da criação — criação abortada.'
                )
                pre_cadastro_simulado = None
                func_id = func_id_recheck
            else:
                pre_cadastro_id = _criar_pre_cadastro_funcionario(campos_pre_cadastro)
                func_id = pre_cadastro_id
    elif decisao_5c == 'enviar_para_revisao':
        pendencia_simulada = {
            'tipo': 'Ficha de Registro — revisão Fase 5C',
            'observacao': f'{motivo_decisao} (Arquivo: {ctx["nome_arquivo"]}, Processar: {ctx["proc_id"]}).',
        }

    # Kit de Admissão — monta (Ficha + Contrato + Termo de VT, reclassificando
    # cada irmão pelo CONTEÚDO real, não pelo "Tipo de Documento" salvo — ver
    # _montar_e_disparar_kit_admissao) e dispara a Assinatura Nativa via
    # WhatsApp automaticamente assim que a Ficha estiver disponível.
    kit_resultado = {'kit_montado': False, 'assinatura_disparada': False, 'motivo': None}
    if not dry_run and func_id:
        kit_resultado = _montar_e_disparar_kit_admissao(ctx, func_id, dry_run, dados_para_vt=dados)

    detalhes = {
        'tipo_documento': ctx['tipo_documento'],
        'arquivo_id': ctx['arquivo_id'],
        'nome_arquivo': ctx['nome_arquivo'],
        'dados_extraidos': dados,
        'confianca': confianca,
        'funcionario_existe_em_funcionarios': funcionario_existe,
        'decisao_5c': decisao_5c,
        'pre_cadastro_simulado': pre_cadastro_simulado,
        'pre_cadastro_id': pre_cadastro_id,
        'funcionario_id': func_id,
        'kit_admissao': kit_resultado,
        'pendencia_simulada': pendencia_simulada,
        'proxima_acao_sugerida': motivo_decisao,
    }

    pendencia_retorno = pendencia_simulada if decisao_5c == 'enviar_para_revisao' else None

    return {
        'acao': 'ficha_registro_extraida_sem_gravar',
        'status_final': 'Concluído' if kit_resultado['kit_montado'] else ctx['status_atual'],
        'detalhes': detalhes,
        'pendencia': pendencia_retorno,
    }


def _processar_contrato_stub(ctx, dry_run):
    """
    Handler de Contrato de Experiência / Contrato de Trabalho (Fase 5B + 5C).

    Fase 5B: extrai dados do contrato e consulta (somente leitura) se o CPF já
    existe em Funcionários.

    Fase 5C: decisao_5c (padrão genérico reaproveitável — ver comentário acima de
    CONFIANCA_MINIMA_PRE_CADASTRO) ∈ {cadastrar_ativo_automaticamente,
    criar_pre_cadastro_seguro, funcionario_ja_existente, enviar_para_revisao}.
    Em dry_run=true apenas simula (nada gravado). Em dry_run=false executa a ação
    decidida — cria em Funcionários com Status="Ativo" ou "Pré-cadastro", ou cria
    Pendência. Nunca altera Status/cargo/local/etc. de um funcionário já existente
    — nesse caso apenas compara e reporta.
    """
    dados, avisos_qualidade = extrair_dados_contrato(ctx['texto'])

    cpf_num = re.sub(r'\D', '', dados['cpf'] or '')

    validacao = {
        'cpf_valido': bool(cpf_num) and _cpf_digitos_validos(cpf_num),
        'nome_presente': bool(dados['nome_funcionario']),
        'data_admissao_valida': bool(dados['data_admissao']),
    }

    campos_faltantes = [
        campo for campo in CAMPOS_CONTRATO_OBRIGATORIOS if not dados.get(campo)
    ]

    total_campos = len(CAMPOS_PARA_CONFIANCA)
    preenchidos = sum(1 for campo in CAMPOS_PARA_CONFIANCA if dados.get(campo) is not None)
    confianca = round(preenchidos / total_campos, 2) if total_campos else 0.0

    # Cada aviso de qualidade (local truncado/ambíguo/conflitante) reduz a
    # confiança — mas "Magnata / Sede" (interno_empresa) não gera aviso e,
    # portanto, não penaliza.
    confianca = round(max(0.0, confianca - 0.1 * len(avisos_qualidade)), 2)

    funcionario_existe = None
    nome_existente = None
    func_id = None
    if dados['cpf']:
        func_id, nome_existente = buscar_funcionario_por_cpf(dados['cpf'])
        funcionario_existe = func_id is not None

    # Kit de Admissão — Ficha de Registro de Empregado é a fonte primária de
    # dados (mais completa). Se o funcionário ainda não existe E há uma Ficha
    # de Registro irmã (mesmo e-mail, ver _buscar_documentos_irmaos_kit_admissao)
    # que ainda não foi processada, o Contrato CEDE a vez — não cria o cadastro
    # a partir dos seus dados mais pobres. Cria Pendência orientando processar
    # a Ficha primeiro; quando isso acontecer, o Contrato será vinculado ao
    # funcionário criado pela Ficha (ver bloco de agrupamento mais abaixo).
    if not funcionario_existe and not dry_run:
        irmaos_aguardando_ficha = [
            i for i in _buscar_documentos_irmaos_kit_admissao(ctx)
            if i['tipo_documento'] == 'Ficha de Registro de Empregado'
            and i['status'] in ('Pendente', 'Processando')
        ]
        if irmaos_aguardando_ficha:
            return {
                'acao': 'cedido_a_ficha_de_registro',
                'status_final': 'Processando',
                'detalhes': {
                    'tipo_documento': ctx['tipo_documento'],
                    'arquivo_id': ctx['arquivo_id'],
                    'nome_arquivo': ctx['nome_arquivo'],
                    'motivo': (
                        'Kit de Admissão detectado: existe uma Ficha de Registro de '
                        'Empregado irmã (mesmo e-mail) ainda não processada. Por ela ser '
                        'a fonte primária de dados, o Contrato aguarda — processe a Ficha '
                        'de Registro primeiro (Processar Arquivos: '
                        f'{irmaos_aguardando_ficha[0]["proc_id"]}).'
                    ),
                },
                'pendencia': {
                    'tipo': 'Kit de Admissão — aguardando Ficha de Registro',
                    'observacao': (
                        f'Contrato "{ctx["nome_arquivo"]}" (Processar: {ctx["proc_id"]}) aguarda '
                        f'o processamento da Ficha de Registro de Empregado irmã '
                        f'(Processar: {irmaos_aguardando_ficha[0]["proc_id"]}), que é a fonte '
                        f'primária de dados do Kit de Admissão.'
                    ),
                },
            }

    # ── Fase 5C — decisão de pré-cadastro / revisão ──────────────────────────
    divergencia = None
    if funcionario_existe and dados['nome_funcionario'] and nome_existente:
        if dados['nome_funcionario'].strip().upper() != nome_existente.strip().upper():
            divergencia = (
                f'Nome do contrato ("{dados["nome_funcionario"]}") difere do '
                f'cadastro existente em Funcionários ("{nome_existente}") para o mesmo CPF.'
            )

    nome_suspeito = _nome_suspeito(dados['nome_funcionario'])

    erros_bloqueantes = []
    if campos_faltantes:
        erros_bloqueantes.append(f'Campos obrigatórios faltando: {", ".join(campos_faltantes)}.')
    if not validacao['cpf_valido']:
        erros_bloqueantes.append('CPF inválido (dígitos verificadores não conferem).')
    if nome_suspeito:
        erros_bloqueantes.append(f'Nome do funcionário ausente ou suspeito: "{dados["nome_funcionario"]}".')
    if dados['tipo_local'] == 'desconhecido':
        erros_bloqueantes.append('Local de trabalho não identificado (tipo_local="desconhecido").')

    categoria_generica, motivo_decisao = decidir_acao_documento(
        erros_bloqueantes=erros_bloqueantes,
        avisos_qualidade=avisos_qualidade,
        confianca=confianca,
        confianca_minima=CONFIANCA_MINIMA_PRE_CADASTRO,
        confianca_ativo=CONFIANCA_ATIVO_AUTOMATICO,
        registro_existente=funcionario_existe,
        divergencia=divergencia,
    )

    # Mapeamento da categoria genérica para o vocabulário específico de
    # contratos (Fase 5C) — ver comentário acima de CONFIANCA_MINIMA_PRE_CADASTRO.
    _MAPA_CATEGORIA_PARA_DECISAO_5C = {
        'executar_automaticamente': 'cadastrar_ativo_automaticamente',
        'executar_com_status_intermediario': 'criar_pre_cadastro_seguro',
        'funcionario_ou_registro_ja_existente': 'funcionario_ja_existente',
        'enviar_para_revisao': 'enviar_para_revisao',
    }
    decisao_5c = _MAPA_CATEGORIA_PARA_DECISAO_5C[categoria_generica]

    origem_contrato = (
        f'Contrato: {ctx["tipo_documento"]}, arquivo "{ctx["nome_arquivo"]}" '
        f'(Processar Arquivos: {ctx["proc_id"]}, Arquivo: {ctx["arquivo_id"]}).'
    )

    pre_cadastro_simulado = None
    pre_cadastro_id = None
    pendencia_simulada = None

    if decisao_5c in ('cadastrar_ativo_automaticamente', 'criar_pre_cadastro_seguro'):
        status_destino = 'Validação Pendente'  # aprovação por exceção (v2.66): sempre entra pendente de validação manual
        campos_pre_cadastro = _montar_campos_pre_cadastro(dados, status_destino)
        observacoes = []
        if dados['local_posto']:
            observacoes.append(
                f'Local/Posto sugerido: "{dados["local_posto"]}" '
                f'(tipo_local={dados["tipo_local"]}) — campo "Locais de trabalho" é um '
                'vínculo para a tabela Locais e não é preenchido automaticamente; '
                'requer associação manual.'
            )
        observacoes.append(
            origem_contrato + ' Não há campo "Origem do cadastro"/"Contrato de origem" '
            'mapeado em Funcionários — referência mantida apenas neste relatório.'
        )
        pre_cadastro_simulado = {
            'campos': campos_pre_cadastro,
            'observacoes': observacoes,
        }
        if not dry_run:
            # Regra de segurança: rechecar CPF em Funcionários imediatamente
            # antes de criar, para nunca gerar duplicidade (ex.: requisição
            # duplicada criou o registro entre a decisão e a gravação).
            func_id_recheck, _ = buscar_funcionario_por_cpf(dados['cpf'])
            if func_id_recheck:
                decisao_5c = 'funcionario_ja_existente'
                motivo_decisao = (
                    f'CPF já existe em Funcionários (ID={func_id_recheck}), '
                    'detectado na rechecagem imediatamente antes da criação — '
                    'criação abortada para evitar duplicidade.'
                )
                pre_cadastro_simulado = None
                func_id = func_id_recheck
            else:
                pre_cadastro_id = _criar_pre_cadastro_funcionario(campos_pre_cadastro)
                func_id = pre_cadastro_id

    elif decisao_5c == 'enviar_para_revisao':
        partes_obs = [motivo_decisao]
        if dados['local_posto']:
            partes_obs.append(
                f'Local/Posto identificado: "{dados["local_posto"]}" (tipo_local={dados["tipo_local"]}).'
            )
        partes_obs.append(origem_contrato)
        pendencia_simulada = {
            'tipo': 'Contrato — revisão Fase 5C',
            'observacao': ' '.join(partes_obs),
        }

    proxima_acao = motivo_decisao

    # Kit de Admissão — mesma automação da Ficha de Registro (função
    # compartilhada _montar_e_disparar_kit_admissao): se, reclassificando
    # pelo conteúdo, os documentos disponíveis já incluem uma Ficha, monta
    # o Kit e dispara a assinatura aqui mesmo — não depende de qual dos
    # documentos irmãos o /processar-fila processa primeiro.
    kit_resultado = {'kit_montado': False, 'assinatura_disparada': False, 'motivo': None}
    if not dry_run and func_id:
        kit_resultado = _montar_e_disparar_kit_admissao(ctx, func_id, dry_run, dados_para_vt=dados)

    detalhes = {
        'tipo_documento': ctx['tipo_documento'],
        'record_id': ctx['proc_id'],
        'arquivo_id': ctx['arquivo_id'],
        'nome_arquivo': ctx['nome_arquivo'],
        'dados_extraidos': dados,
        'local_posto': dados['local_posto'],
        'tipo_local': dados['tipo_local'],
        'avisos_qualidade': avisos_qualidade,
        'validacao': validacao,
        'campos_faltantes': campos_faltantes,
        'confianca': confianca,
        'funcionario_existe_em_funcionarios': funcionario_existe,
        'decisao_5c': decisao_5c,
        'pre_cadastro_simulado': pre_cadastro_simulado,
        'pre_cadastro_id': pre_cadastro_id,
        'funcionario_id': func_id,
        'kit_admissao': kit_resultado,
        'pendencia_simulada': pendencia_simulada,
        'proxima_acao_sugerida': proxima_acao,
        'observacao': (
            'Fase 5B/5C. Em dry_run=true nada é gravado — "pre_cadastro_simulado" e '
            '"pendencia_simulada" mostram o que seria feito. Em dry_run=false, '
            'conforme "decisao_5c": cadastrar_ativo_automaticamente cria em '
            'Funcionários com Status="Ativo"; criar_pre_cadastro_seguro cria com '
            'Status="Pré-cadastro"; enviar_para_revisao cria Pendência. '
            'Funcionário existente nunca é sobrescrito.'
        ),
        'log_operacional': {
            'disponivel': False,
            'motivo': (
                'Tabela "Processar Arquivos" não possui campo de texto/observação '
                'para registrar decisão, ID do funcionário criado/encontrado ou erro '
                '(campos atuais: Name, Status, Tipo, Arquivos, Data Processo, '
                'PROCESSO IDA, Arquivos 2, Tipo de Documento). Nenhum campo novo foi '
                'criado; esta informação é retornada apenas neste JSON.'
            ),
        },
    }

    pendencia_retorno = pendencia_simulada if decisao_5c == 'enviar_para_revisao' else None

    return {
        'acao': 'contrato_extraido_sem_gravar',
        'status_final': 'Concluído' if kit_resultado['kit_montado'] else ctx['status_atual'],
        'detalhes': detalhes,
        'pendencia': pendencia_retorno,
    }


def _processar_rescisao_stub(ctx: dict, dry_run: bool) -> dict:
    """
    Handler de Termo/Aviso de Rescisão de Contrato de Trabalho (TRCT).

    Extrai nome, CPF, data de rescisão, tipo de aviso e motivo. Consulta
    (somente leitura) se o CPF corresponde a um funcionário cadastrado.

    Por segurança, este handler NUNCA altera o Status do funcionário
    automaticamente (não existe campo "Data de Desligamento" em
    Funcionários, e inativar alguém por engano tem custo alto — folha de
    pagamento, FGTS, benefícios). Em vez disso, sempre cria uma Pendência
    rica (com todos os dados extraídos prontos para 1 clique de confirmação
    humana), igual ao padrão de segurança já usado em Contrato (Fase 5C):
    "funcionário existente nunca é sobrescrito automaticamente".

    Mesmo em dry_run=true, nada é gravado — apenas simula o que seria criado.
    """
    dados = extrair_dados_rescisao(ctx['texto'])

    cpf_num = re.sub(r'\D', '', dados['cpf'] or '')
    cpf_valido = bool(cpf_num) and _cpf_digitos_validos(cpf_num)

    func_id, nome_cadastrado = (None, None)
    if dados['cpf']:
        func_id, nome_cadastrado = buscar_funcionario_por_cpf(dados['cpf'])

    campos_faltantes = [
        campo for campo in ('nome_funcionario', 'cpf', 'data_rescisao')
        if not dados.get(campo)
    ]

    partes_obs = [
        f'RESCISÃO — dados extraídos automaticamente do documento '
        f'"{ctx["nome_arquivo"]}" (Processar Arquivos: {ctx["proc_id"]}, '
        f'Arquivo: {ctx["arquivo_id"]}).',
        f'Nome: {dados["nome_funcionario"] or "(não identificado)"}',
        f'CPF: {dados["cpf"] or "(não identificado)"}'
        + (' [inválido]' if dados['cpf'] and not cpf_valido else ''),
        f'Data da rescisão: {dados["data_rescisao"] or "(não identificada)"}',
        f'Tipo de aviso: {dados["tipo_aviso"] or "(não identificado)"}',
        f'Motivo: {dados["motivo_rescisao"] or "(não identificado)"}',
    ]
    if func_id:
        partes_obs.append(
            f'Funcionário ENCONTRADO em Funcionários: "{nome_cadastrado}" (ID={func_id}). '
            f'AÇÃO SUGERIDA: confirmar dados e alterar Status para "Inativo" manualmente '
            f'(este robô não altera Status automaticamente por segurança).'
        )
    elif dados['cpf']:
        partes_obs.append(
            'CPF extraído NÃO encontrado em Funcionários — confirmar se é '
            'erro de leitura do PDF ou se o colaborador nunca foi cadastrado.'
        )
    if campos_faltantes:
        partes_obs.append(f'Campos não identificados automaticamente: {", ".join(campos_faltantes)}.')

    observacao = ' '.join(partes_obs)

    detalhes = {
        'tipo_documento': ctx['tipo_documento'],
        'record_id': ctx['proc_id'],
        'arquivo_id': ctx['arquivo_id'],
        'nome_arquivo': ctx['nome_arquivo'],
        'dados_extraidos': dados,
        'cpf_valido': cpf_valido,
        'funcionario_encontrado': func_id is not None,
        'funcionario_id': func_id,
        'funcionario_nome_cadastrado': nome_cadastrado,
        'campos_faltantes': campos_faltantes,
        'observacao': (
            'Fase 6 — Rescisão. Nunca altera Status de Funcionários '
            'automaticamente; sempre cria Pendência com os dados extraídos '
            'para confirmação humana de 1 clique.'
        ),
    }

    pendencia = {
        'tipo': 'Rescisão — confirmar e inativar funcionário',
        'observacao': observacao,
    }

    if dry_run:
        return {
            'acao': 'rescisao_extraida_simulada',
            'status_final': ctx['status_atual'],
            'detalhes': detalhes,
            'pendencia': None,
        }

    return {
        'acao': 'rescisao_extraida_pendencia_criada',
        'status_final': 'Revisão Manual',
        'detalhes': detalhes,
        'pendencia': pendencia,
    }


def _processar_folha_ponto(ctx: dict, dry_run: bool) -> dict:
    """
    Handler de Folha de Ponto para /processar-fila (Fase 1 — anexação ao
    prontuário do funcionário).

    Ainda NÃO extrai horários/batidas do PDF (formato não mapeado). O que
    este handler faz é: identificar o funcionário pelo CPF (ou, na falta
    dele, pelo nome) presente no PDF e anexar o arquivo ao campo
    "PDF Folha Ponto" (prontuário, em Funcionários), evitando duplicar o
    mesmo arquivo. Itens sem funcionário identificável vão para "Revisão
    Manual" com Pendência, em vez de ficarem invisíveis.
    """
    # ✅ v3.6: DEFESA PROFUNDA — Validar configuração antes de qualquer operação
    config_ok, config_msg = _validar_configuracao_assinatura_v36()
    if not config_ok:
        logger.error(f'[FOLHA PONTO] Configuracao invalida: {config_msg}')
        return {
            'acao': 'folha_ponto_bloqueada',
            'status_final': 'Revisão Manual',
            'detalhes': f'Configuracao v3.6 incompleta — folha não pode prosseguir',
        }

    texto = ctx['texto']

    cpf = extrair_cpf(texto)
    func_id, nome = (None, None)
    if cpf:
        func_id, nome = buscar_funcionario_por_cpf(cpf)

    nome_pdf = extrair_nome_funcionario(texto)
    if not func_id and nome_pdf and nome_pdf != 'Desconhecido':
        func_id, nome = _buscar_funcionario_por_nome(nome_pdf)

    if not func_id:
        detalhes = {
            'cpf_extraido': cpf,
            'nome_extraido': nome_pdf,
            'nome_arquivo': ctx['nome_arquivo'],
        }
        observacao = (
            f'Folha de Ponto "{ctx["nome_arquivo"]}" não pôde ser vinculada a um '
            f'funcionário. CPF extraído: {cpf or "(nenhum)"} | '
            f'Nome extraído: {nome_pdf or "(nenhum)"} '
            f'(Processar Arquivos: {ctx["proc_id"]}, Arquivo: {ctx["arquivo_id"]}).'
        )
        if dry_run:
            return {
                'acao': 'funcionario_nao_encontrado',
                'status_final': 'Revisão Manual',
                'detalhes': detalhes,
                'pendencia': None,
            }
        return {
            'acao': 'funcionario_nao_encontrado',
            'status_final': 'Revisão Manual',
            'detalhes': detalhes,
            'pendencia': {
                'tipo': 'Funcionário não encontrado (Folha de Ponto)',
                'observacao': observacao,
            },
        }

    # Verificar se este arquivo já está anexado ao prontuário (evitar duplicar)
    _at_throttle()
    r = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}/{func_id}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        params={'returnFieldsByFieldId': 'true'},
        timeout=30,
    )
    r.raise_for_status()
    func_fields = r.json().get('fields', {})
    attachments = func_fields.get(F_FUNC_PDF_FOLHA) or []
    status_anexo = _verificar_anexo_em_lista(attachments, ctx['nome_arquivo'], ctx['pdf_hash'])

    detalhes = {
        'funcionario_id': func_id,
        'funcionario_nome': nome,
        'cpf_extraido': cpf,
        'status_anexo': status_anexo,
    }

    # Fase 2 — extrai período/dias do Cartão Ponto e atualiza o resumo no
    # prontuário do funcionário (best-effort, não afeta status_final).
    dados_cartao = extrair_cartao_ponto(texto)
    detalhes['cartao_ponto_periodo_inicio'] = dados_cartao['periodo_inicio']
    detalhes['cartao_ponto_periodo_fim'] = dados_cartao['periodo_fim']
    detalhes['cartao_ponto_dias_extraidos'] = len(dados_cartao['dias'])

    if not dry_run and dados_cartao['dias']:
        try:
            _atualizar_resumo_ponto_funcionario(
                func_id, _formatar_resumo_cartao_ponto(dados_cartao),
            )
        except AirtableError as exc:
            logger.warning(
                f'[FILA] {ctx["proc_id"]}: falha ao atualizar resumo do Cartão '
                f'Ponto de {func_id}: {exc}'
            )

    if dry_run:
        acao = {
            'novo':     'anexaria_folha_ponto_ao_prontuario',
            'identico': 'folha_ponto_ja_anexada',
            'conflito': 'folha_ponto_conflito_precisaria_autorizacao',
        }[status_anexo]
        return {'acao': acao, 'status_final': 'Concluído', 'detalhes': detalhes, 'pendencia': None}

    if status_anexo == 'identico':
        return {'acao': 'folha_ponto_ja_anexada', 'status_final': 'Concluído', 'detalhes': detalhes, 'pendencia': None}

    if status_anexo == 'conflito':
        return {
            'acao': 'folha_ponto_conflito_anexo_nao_substituido',
            'status_final': 'Revisão Manual',
            'detalhes': detalhes,
            'pendencia': {
                'tipo': 'Conflito de anexo (Folha de Ponto)',
                'observacao': (
                    f'Já existe um anexo "{ctx["nome_arquivo"]}" no prontuário do funcionário '
                    f'{nome} ({func_id}) com conteúdo diferente. Anexo NÃO foi substituído — '
                    f'requer revisão manual.'
                ),
            },
        }

    # status_anexo == 'novo'
    try:
        _anexar_attachment(TABLE_FUNC, func_id, F_FUNC_PDF_FOLHA, ctx['pdf_bytes'], ctx['nome_arquivo'])
        return {'acao': 'folha_ponto_anexada_ao_prontuario', 'status_final': 'Concluído', 'detalhes': detalhes, 'pendencia': None}
    except AirtableError as exc:
        logger.error(f'[FILA] {ctx["proc_id"]}: erro ao anexar Folha de Ponto ao prontuário de {func_id}: {exc}')
        return {
            'acao': 'folha_ponto_falhou_anexo',
            'status_final': 'Revisão Manual',
            'detalhes': detalhes,
            'pendencia': {
                'tipo': 'Erro ao anexar Folha de Ponto',
                'observacao': f'Funcionário {func_id} ({nome}): {exc}',
            },
        }


def _processar_documento_sem_automacao(ctx: dict, dry_run: bool) -> dict:
    """
    Handler genérico (Fase 6 — sanitização) para tipos de documento sem
    processamento automático implementado ainda (Folha de Ponto, Férias,
    FGTS, Guia, Boleto, Nota Fiscal, Outro).

    Antes desta mudança, esses tipos ficavam mapeados para None em
    PROCESSADORES_DOCUMENTO e /processar-fila retornava 400 "ainda não
    implementado" para a requisição inteira — os itens correspondentes
    ficavam parados em "Pendente" em Processar Arquivos, sem nenhuma
    Pendência/visibilidade (ex.: e-mails com FGTS Digital "desapareciam").

    Agora cada item é marcado individualmente com status_final="Revisão
    Manual" e uma Pendência é criada apontando o Arquivo/PDF original, para
    que um humano trate o documento — em vez de o item ficar invisível
    indefinidamente.

    Também usado para EPI, Ficha de Registro de Empregado e Termo de
    Prorrogação de Contrato de Experiência (Fase 6b) — para esses três tipos
    não existe campo dedicado em Funcionários para anexar automaticamente
    (ex.: não há campo "Ficha de EPI"), então a automação segura possível é
    extrair nome/CPF e deixá-los prontos na Pendência, sem gravar nada em
    Funcionários.
    """
    tipo = ctx['tipo_documento']
    nome_extraido = extrair_nome_documento_dp(ctx['texto'])
    cpf_extraido = extrair_cpf(ctx['texto'])

    func_id, nome_cadastrado = (None, None)
    if cpf_extraido:
        func_id, nome_cadastrado = buscar_funcionario_por_cpf(cpf_extraido)

    partes_obs = [
        f'Processamento automático de "{tipo}" ainda não implementado. '
        f'Arquivo "{ctx["nome_arquivo"]}" requer tratamento manual '
        f'(Processar Arquivos: {ctx["proc_id"]}, Arquivo: {ctx["arquivo_id"]}).',
        f'Nome identificado: {nome_extraido or "(não identificado)"}',
        f'CPF identificado: {cpf_extraido or "(não identificado)"}',
    ]
    if func_id:
        partes_obs.append(f'Funcionário correspondente em Funcionários: "{nome_cadastrado}" (ID={func_id}).')
    elif cpf_extraido:
        partes_obs.append('CPF identificado não corresponde a nenhum funcionário cadastrado.')

    observacao = ' '.join(partes_obs)

    detalhes = {
        'tipo_documento': tipo,
        'arquivo_id': ctx['arquivo_id'],
        'nome_arquivo': ctx['nome_arquivo'],
        'motivo': observacao,
        'nome_extraido': nome_extraido,
        'cpf_extraido': cpf_extraido,
        'funcionario_id': func_id,
    }

    if dry_run:
        return {
            'acao': 'marcaria_para_revisao_manual',
            'status_final': 'Revisão Manual',
            'detalhes': detalhes,
            'pendencia': None,
        }

    return {
        'acao': 'marcado_para_revisao_manual',
        'status_final': 'Revisão Manual',
        'detalhes': detalhes,
        'pendencia': {
            'tipo': f'Sem automação: {tipo}',
            'observacao': observacao,
        },
    }


# Registro de handlers por tipo de documento (Fase 4 / Fase 6).
# Tipos sem handler dedicado usam _processar_documento_sem_automacao, que
# marca o item como "Revisão Manual" + Pendência em vez de deixá-lo invisível
# (status_final="Revisão Manual" é criado automaticamente como nova opção do
# singleSelect "Status" via typecast=True, tanto em Processar Arquivos quanto
# em Holerites).
PROCESSADORES_DOCUMENTO = {
    'Holerite': _processar_holerite,
    'Folha de Ponto': _processar_folha_ponto,
    'Contrato de Experiência': _processar_contrato_stub,
    'Contrato de Trabalho': _processar_contrato_stub,
    'Rescisão': _processar_rescisao_stub,
    'EPI': _processar_documento_sem_automacao,
    'Ficha de Registro de Empregado': _processar_ficha_registro_stub,
    'Termo de Prorrogação de Contrato de Experiência': _processar_documento_sem_automacao,
    'Férias': _processar_documento_sem_automacao,
    'FGTS': _processar_documento_sem_automacao,
    'Guia': _processar_documento_sem_automacao,
    'Boleto': _processar_documento_sem_automacao,
    'Nota Fiscal': _processar_documento_sem_automacao,
    'Outro': _processar_documento_sem_automacao,
}

# Tipos de documento para os quais o vocabulário de decisão 'decisao_5c'
# (Fase 5C) se aplica e que, portanto, exigem as restrições extras de
# dry_run=false abaixo (record_id específico + limit=1).
TIPOS_CONTRATO_5C = ('Contrato de Experiência', 'Contrato de Trabalho', 'Ficha de Registro de Empregado')


# v3.11 — Metadados de fatiamento/destino por tipo de documento, usados só
# para compor o texto de /status-documentos. A LISTA de tipos em si vem de
# PROCESSADORES_DOCUMENTO + TIPOS_FISCAIS (fonte real do código) — isto aqui
# é só a descrição. Um tipo novo em qualquer um dos dois já aparece no
# comando automaticamente, mesmo sem entrada aqui (cai no texto genérico
# "ver código-fonte" até alguém descrever o destino real).
CAPACIDADES_DOCUMENTO = {
    'Holerite': ('Individual por CPF', 'Holerites'),
    'Contrato de Experiência': ('Pré-cadastro por exceção', 'Funcionários (Status "Validação Pendente")'),
    'Contrato de Trabalho': ('Pré-cadastro por exceção', 'Funcionários (Status "Validação Pendente")'),
    'Ficha de Registro de Empregado': ('Pré-cadastro por exceção', 'Funcionários (Status "Validação Pendente")'),
    'Rescisão': ('Extração de dados + Pendência para confirmação manual', 'Pendências/Revisar'),
    'Extrato da Folha de Pagamento': ('Por cliente (CNPJ/Nome)', 'Extratos Mensais'),
    'FGTS': ('Por cliente (CNPJ/Nome)', 'FGTS Digital'),
    'DCTFWeb - Recibo de Entrega': ('Broadcast (mesmo documento p/ todos os clientes ativos)', 'Guias e Comprovantes'),
    'DCTFWeb - Declaração': ('Broadcast (mesmo documento p/ todos os clientes ativos)', 'Guias e Comprovantes'),
}

# 'Folha de Ponto' tem 1 entrada só em PROCESSADORES_DOCUMENTO (a fila
# genérica /processar-fila), mas o motor real de produção é a rota dedicada
# /processar-folha-ponto (v3.00+), com 2 formatos de entrada distintos e
# Regra de Prevalência entre eles — por isso é desdobrado manualmente em 2
# linhas no relatório, em vez de vir só da chave 'Folha de Ponto'.
CAPACIDADES_FOLHA_PONTO = [
    ('Folha de Ponto — Secullum (mestre)', 'Individual por CPF', 'PDF Folha Ponto (Funcionários) / Assinaturas Digitais'),
    ('Folha de Ponto — Manual', 'Individual por colaborador, com Prevalência sobre o Secullum', 'PDF Folha Ponto (Funcionários) / Assinaturas Digitais'),
]

# v3.12 — Benefícios: rotas dedicadas (/processar-beneficios e
# /processar-vr-va), fora de PROCESSADORES_DOCUMENTO/TIPO_DOC_REGRAS (mesmo
# motivo da Folha de Ponto acima — lógica própria demais pra fila genérica).
CAPACIDADES_BENEFICIOS = [
    ('Extrato Bancário Consolidado (PIX)', 'Classifica por palavra-chave na descrição (VR/VA, Almoço/Jantar, Horas Extras, Diárias) e ignora o resto; nome por match fuzzy (CPF vem mascarado na origem)', 'Clientes — colunas "VR/VA Benefícios" / "Hora Almoço e Jantar" / "Horas Extras" / "Diárias"'),
    ('Horas Extras — mestre avulso (Sicoob, pré-extrato)', 'Individual por CPF, fundido por Cliente/posto', 'Clientes — coluna "Horas Extras"'),
    ('Assiduidade (Sicoob, trava R$ 110,00)', 'Individual por CPF, fundido por Cliente/posto', 'Clientes — coluna "Assiduidade"'),
    ('VR/VA — Lote (planilha .xls/.xlsx + comprovante agregado)', 'Por linha da planilha, via CPF→Local→Cliente (rota /processar-vr-va, v2.34)', 'Extratos Mensais (relatório gerado) + Guias e Comprovantes (comprovante)'),
    ('VR/VA — PIX individual avulso (fora do extrato)', 'Individual por CPF, fundido por Cliente/posto — caso raro, uso normal passou a ser via Extrato Bancário', 'Clientes — coluna "VR/VA Benefícios"'),
]


def _listar_capacidades_documentos() -> str:
    """v3.11 — Monta o texto de /status-documentos a partir de
    PROCESSADORES_DOCUMENTO e TIPOS_FISCAIS (fonte real do código), não de
    uma lista escrita à mão — um tipo novo em qualquer um dos dois aparece
    aqui sozinho no próximo deploy."""
    automatizados = []
    sem_automacao = []
    for tipo, handler in PROCESSADORES_DOCUMENTO.items():
        if tipo == 'Outro':
            continue
        if tipo == 'Folha de Ponto':
            continue  # desdobrado via CAPACIDADES_FOLHA_PONTO abaixo
        if tipo in TIPOS_FISCAIS:
            continue  # tratado no loop de TIPOS_FISCAIS abaixo — tem rota
            # própria (_processar_anexo_fiscal) com automação real, mesmo
            # aparecendo aqui mapeado para _processar_documento_sem_automacao
            # (esse handler só vale para o mesmo tipo chegando pela fila do
            # Departamento Pessoal, não pela caixa fiscal).
        if handler is _processar_documento_sem_automacao:
            sem_automacao.append(tipo)
        else:
            fatiamento, destino = CAPACIDADES_DOCUMENTO.get(tipo, ('ver código-fonte', 'ver código-fonte'))
            automatizados.append((tipo, fatiamento, destino))
    for tipo in TIPOS_FISCAIS:
        fatiamento, destino = CAPACIDADES_DOCUMENTO.get(tipo, ('ver código-fonte', 'ver código-fonte'))
        automatizados.append((tipo, fatiamento, destino))

    linhas = ['🤖 LISTA DE CAPACIDADES DO ROBÔ (DOCUMENTOS ATIVOS):']
    idx = 1
    for nome, fatiamento, destino in CAPACIDADES_FOLHA_PONTO:
        linhas.append(f'{idx}. {nome} (Fatiamento: {fatiamento} → Tabela: {destino})')
        idx += 1
    for nome, fatiamento, destino in CAPACIDADES_BENEFICIOS:
        linhas.append(f'{idx}. {nome} (Fatiamento: {fatiamento} → Tabela: {destino})')
        idx += 1
    for tipo, fatiamento, destino in sorted(automatizados):
        linhas.append(f'{idx}. {tipo} (Fatiamento: {fatiamento} → Tabela: {destino})')
        idx += 1

    if sem_automacao:
        linhas.append('')
        linhas.append('📋 CLASSIFICADOS, MAS SEM AUTOMAÇÃO (viram Pendência para revisão manual):')
        linhas.append(', '.join(sorted(sem_automacao)))

    return '\n'.join(linhas)


# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.route('/health', methods=['GET'])
def health():
    at_ok = False
    at_status = None
    try:
        r = requests.get(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}',
            headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
            params={'maxRecords': 1, 'fields[]': ['CPF']},
            timeout=10,
        )
        at_status = r.status_code
        at_ok = r.ok
    except Exception as exc:
        at_status = str(exc)
    return jsonify({
        'status': 'ok',
        'servico': 'magnata-holerite-splitter',
        'versao': '3.17',
        'ram_mb': _mem_mb(),
        'airtable_ok': at_ok,
        'airtable_status': at_status,
    })


# ── Códigos canônicos de erro para /separar ────────────────────────────────────
ERRO_CODIGOS_SEPARAR = {
    'PDF_PASSWORD_INCORRECT': 'PDF protegido por senha',
    'PDF_INVALID': 'Dados não são um PDF válido',
    'PDF_CORRUPTED': 'PDF corrompido ou inválido',
    'PDF_EMPTY': 'PDF vazio ou muito pequeno',
    'PDF_NOT_SUPPORTED': 'Formato ou versão de PDF não suportada',
    'ATTACHMENT_NOT_FOUND': 'Anexo PDF não recebido',
    'EMPLOYEE_NOT_IDENTIFIED': 'Nenhum funcionário/CPF encontrado no PDF',
    'DOCUMENT_TYPE_NOT_IDENTIFIED': 'Tipo de documento não identificado',
    'AIRTABLE_RECORD_NOT_FOUND': 'Registro não encontrado no Airtable',
    'PROCESSING_ERROR': 'Erro ao processar PDF',
}


def _error_response_separar(
    error_code: str,
    record_id: str | None = None,
    message: str | None = None,
    log_detail: str | None = None,
) -> tuple[dict, int]:
    """
    Padroniza resposta de erro do /separar.
    Retorna HTTP 200 com success=false para erros de negócio previsíveis.
    Log_detail é registrado apenas nos logs internos (nunca na resposta).
    """
    canonical_msg = ERRO_CODIGOS_SEPARAR.get(error_code, 'Erro desconhecido')
    final_msg = message or canonical_msg

    if log_detail:
        logger.warning(f'[/separar] {error_code}: {log_detail}')
    else:
        logger.warning(f'[/separar] {error_code}: {final_msg}')

    resp = {
        'success': False,
        'status': 'erro',
        'error_code': error_code,
        'message': final_msg,
    }
    if record_id:
        resp['processar_arquivo_record_id'] = record_id

    return jsonify(resp), 200


def _success_response_separar(
    data: dict,
    record_id: str | None = None,
) -> tuple[dict, int]:
    """
    Padroniza resposta de sucesso do /separar.
    Retorna HTTP 200 com success=true e os dados do resultado.
    """
    resp = {
        'success': True,
        'status': 'concluido',
        'error_code': None,
        'message': 'PDF processado com sucesso',
        'resultado': data,
    }
    if record_id:
        resp['processar_arquivo_record_id'] = record_id

    return jsonify(resp), 200


@app.route('/keep-alive', methods=['GET', 'POST'])
def keep_alive():
    """Endpoint leve para evitar cold start do Render (plano free).
    Deve ser chamado a cada 5-10 minutos via Make.com ou similar.
    Retorna 200 se o serviço está aquecido."""
    return jsonify({'status': 'ok', 'timestamp': datetime.now().isoformat()}), 200


@app.route('/status-documentos', methods=['GET'])
def status_documentos():
    """v3.11 — Diagnóstico de capacidades: lista os tipos de documento que o
    robô sabe classificar e processar hoje, com fatiamento e tabela de
    destino. Gerado a partir de PROCESSADORES_DOCUMENTO/TIPOS_FISCAIS (ver
    _listar_capacidades_documentos) — nunca escrito à mão, então não fica
    desatualizado quando um tipo novo entra no código."""
    texto = _listar_capacidades_documentos()
    if request.args.get('formato') == 'json':
        return jsonify({'status': 'ok', 'texto': texto})
    return texto, 200, {'Content-Type': 'text/plain; charset=utf-8'}


# ── FUNÇÕES DE SUPORTE PARA /separar ASSÍNCRONO ────────────────────────────────

def _fazer_upload_pdf_airtable(record_id: str, pdf_bytes: bytes, filename: str = 'holerite.pdf') -> bool:
    """
    Faz upload do PDF para o campo 'Arquivos' no Airtable.

    Args:
        record_id: ID do registro (começa com 'rec')
        pdf_bytes: Bytes do PDF
        filename: Nome do arquivo

    Returns:
        True se sucesso, False se erro
    """
    try:
        # Validar Field ID
        if not F_PROC_ARQUIVOS or F_PROC_ARQUIVOS == 'fldXXXXXXXXXXXXXX':
            logger.error('[UPLOAD] Field ID do campo Arquivos não configurado')
            return False

        # Preparar requisição (usar base64)
        pdf_b64 = base64.b64encode(pdf_bytes).decode('utf-8')

        # Upload via Airtable attachments endpoint
        upload_url = f'https://content.airtable.com/v0/{BASE_ID}/{record_id}/{F_PROC_ARQUIVOS}/uploadAttachment'

        r = requests.post(
            upload_url,
            headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
            json={
                'contentType': 'application/pdf',
                'file': pdf_b64,
                'filename': filename,
            },
            timeout=60,
        )

        if r.status_code in (200, 201):
            logger.info(f'[UPLOAD] PDF enviado para Airtable: {record_id} ({len(pdf_bytes)} bytes)')
            return True
        else:
            logger.error(f'[UPLOAD] Erro {r.status_code}: {r.text[:200]}')
            return False

    except Exception as e:
        logger.error(f'[UPLOAD] Erro ao fazer upload: {type(e).__name__}: {str(e)[:200]}')
        return False


def _enfileirar_processamento(record_id: str, pdf_hash: str, pdf_url: str | None = None) -> tuple[bool, str]:
    """
    Enfileira tarefa Celery para processar PDF.

    Args:
        record_id: ID do registro Airtable
        pdf_hash: SHA256 do PDF
        pdf_url: URL do anexo (para worker baixar)

    Returns:
        (sucesso: bool, task_id: str)
    """
    try:
        from tarefas_processar_pdf import processar_pdf_task, gerar_idempotency_key

        # Gerar chave de idempotência
        idempotency_key = gerar_idempotency_key(record_id, pdf_hash)

        # Enfileirar tarefa
        task = processar_pdf_task.apply_async(
            args=(record_id, idempotency_key, pdf_url),
        )

        logger.info(f'[FILA] Tarefa enfileirada: {record_id} | task_id={task.id}')
        return True, task.id

    except Exception as e:
        logger.error(f'[FILA] Erro ao enfileirar: {type(e).__name__}: {str(e)[:200]}')
        return False, ''


@app.route('/separar', methods=['POST'])
@observar_ingestao('/separar')
def separar():
    """
    Versão ASSÍNCRONA do /separar (v4.0) — SEM dependência de /tmp.
    Recebe PDF em memória, valida, faz upload para Airtable,
    enfileira para Celery, retorna HTTP 202 imediatamente.
    """
    try:
        # Validar Record ID (PRIMEIRO, antes de validar PDF)
        record_id = request.form.get('processar_arquivo_record_id') or request.args.get('record_id')
        app.logger.info(f"FORM={request.form}")
        app.logger.info(f"RECORD_ID={record_id!r}")
        if not record_id or not record_id.startswith('rec'):
            logger.warning(f'[/separar] Record ID inválido: {record_id}')
            return jsonify({
                'success': False,
                'status': 'erro',
                'error_code': 'INVALID_RECORD_ID',
                'message': 'Record ID obrigatório (começa com "rec")',
            }), 400

        # Obter PDF diretamente de request.files (EM MEMÓRIA)
        if 'pdf' not in request.files:
            return jsonify({
                'success': False,
                'status': 'erro',
                'error_code': 'ATTACHMENT_NOT_FOUND',
                'message': 'PDF não recebido.',
                'processar_arquivo_record_id': record_id,
            }), 200

        pdf_file = request.files['pdf']

        # Validar nome do arquivo e extensão
        if not pdf_file.filename:
            return jsonify({
                'success': False,
                'status': 'erro',
                'error_code': 'PDF_NO_FILENAME',
                'message': 'Arquivo sem nome.',
                'processar_arquivo_record_id': record_id,
            }), 200

        if not pdf_file.filename.lower().endswith('.pdf'):
            return jsonify({
                'success': False,
                'status': 'erro',
                'error_code': 'PDF_INVALID_EXTENSION',
                'message': 'Arquivo deve ter extensão .pdf',
                'processar_arquivo_record_id': record_id,
            }), 200

        # Validar content-type
        content_type = pdf_file.content_type or ''
        if content_type and not content_type.startswith('application/pdf'):
            logger.warning(f'[/separar] Content-Type não é PDF: {content_type}')
            # Permitir se não foi fornecido, mas avisar se foi e está errado

        # Ler PDF_BYTES em memória (NÃO em disco)
        pdf_bytes = pdf_file.read()

        # Validar tamanho mínimo
        if len(pdf_bytes) < 100:
            return jsonify({
                'success': False,
                'status': 'erro',
                'error_code': 'PDF_EMPTY',
                'message': 'PDF muito pequeno (< 100 bytes).',
                'processar_arquivo_record_id': record_id,
            }), 200

        # Validar tamanho máximo (já protegido por MAX_CONTENT_LENGTH, mas checar mesmo assim)
        max_size = 50 * 1024 * 1024  # 50 MB
        if len(pdf_bytes) > max_size:
            return jsonify({
                'success': False,
                'status': 'erro',
                'error_code': 'PDF_TOO_LARGE',
                'message': f'PDF muito grande (> {max_size/1024/1024:.0f}MB).',
                'processar_arquivo_record_id': record_id,
            }), 200

        # Validar assinatura %PDF
        if not pdf_bytes.startswith(b'%PDF'):
            return jsonify({
                'success': False,
                'status': 'erro',
                'error_code': 'PDF_INVALID',
                'message': 'Dados não são um PDF válido.',
                'processar_arquivo_record_id': record_id,
            }), 200

        # Verificar registro no Airtable
        logger.info(f'[/separar] Validando: {record_id}')
        try:
            r = requests.get(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_PROCESSAR}/{record_id}',
                headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
                timeout=10,
            )
            if r.status_code != 200:
                logger.error(f'[/separar] Registro não encontrado: {record_id}')
                return jsonify({
                    'success': False,
                    'status': 'erro',
                    'error_code': 'AIRTABLE_RECORD_NOT_FOUND',
                    'message': 'Registro não encontrado.',
                    'processar_arquivo_record_id': record_id,
                }), 200
        except Exception as e:
            logger.error(f'[/separar] Erro ao verificar: {type(e).__name__}')
            return jsonify({
                'success': False,
                'status': 'erro',
                'error_code': 'AIRTABLE_CHECK_FAILED',
                'message': 'Erro ao validar registro.',
                'processar_arquivo_record_id': record_id,
            }), 200

        # Calcular SHA256 sobre pdf_bytes (em memória)
        pdf_hash = hashlib.sha256(pdf_bytes).hexdigest()

        # Upload do PDF para Airtable (passando os bytes diretos)
        logger.info(f'[/separar] Upload: {record_id} ({len(pdf_bytes)} bytes)')
        success_upload = _fazer_upload_pdf_airtable(
            record_id,
            pdf_bytes,
            filename=pdf_file.filename or 'holerite.pdf'
        )

        if not success_upload:
            logger.error(f'[/separar] Falha upload: {record_id}')
            try:
                app.logger.info(f"PATCH URL: https://api.airtable.com/v0/{BASE_ID}/{TABLE_PROCESSAR}/{record_id}")
                app.logger.info(f"TABLE_PROCESSAR={TABLE_PROCESSAR}")
                app.logger.info(f"record_id={record_id!r}")
                requests.patch(
                    f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_PROCESSAR}/{record_id}',
                    headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
                    json={'fields': {F_PROC_STATUS: 'Erro', F_PROC_TIPO_DOC: 'UPLOAD_FAILED'}},
                    timeout=10,
                )
            except Exception:
                pass
            return jsonify({
                'success': False,
                'status': 'erro',
                'error_code': 'PDF_UPLOAD_FAILED',
                'message': 'Falha ao enviar PDF.',
                'processar_arquivo_record_id': record_id,
            }), 200

        logger.info(f'[/separar] Enfileirando: {record_id}')

        # Obter URL do anexo para worker — com retry e validação robusta
        pdf_url = None
        max_retries = 5
        for tentativa in range(1, max_retries + 1):
            try:
                r = requests.get(
                    f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_PROCESSAR}/{record_id}?returnFieldsByFieldId=true',
                    headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
                    timeout=10,
                )
                if r.status_code != 200:
                    logger.warning(f'[/separar] Tentativa {tentativa}: status {r.status_code}')
                    if tentativa < max_retries:
                        time.sleep(0.5)
                    continue

                fields = r.json().get('fields', {})
                attachments = fields.get(F_PROC_ARQUIVOS, []) or []

                if attachments and isinstance(attachments, list) and len(attachments) > 0:
                    candidate_url = attachments[0].get('url')
                    if candidate_url and candidate_url.startswith('http'):
                        pdf_url = candidate_url
                        logger.info(f'[/separar] URL obtida na tentativa {tentativa}: {pdf_url[:60]}...')
                        break
                    else:
                        logger.warning(f'[/separar] Tentativa {tentativa}: URL inválida')
                else:
                    logger.warning(f'[/separar] Tentativa {tentativa}: campo vazio')

                if tentativa < max_retries:
                    time.sleep(0.5)

            except Exception as e:
                logger.warning(f'[/separar] Tentativa {tentativa}: {type(e).__name__}: {str(e)[:100]}')
                if tentativa < max_retries:
                    time.sleep(0.5)

        # Validar que pdf_url foi obtida
        if not pdf_url:
            logger.error(f'[/separar] Falha: URL do PDF não disponível após {max_retries} tentativas')
            return jsonify({
                'success': False,
                'status': 'erro',
                'error_code': 'PDF_ATTACHMENT_URL_NOT_AVAILABLE',
                'message': 'URL do anexo não disponível no Airtable.',
                'processar_arquivo_record_id': record_id,
            }), 200

        # Enfileirar tarefa Celery
        success_enfileira, task_id = _enfileirar_processamento(record_id, pdf_hash, pdf_url)

        if not success_enfileira:
            logger.error(f'[/separar] Falha enfileirar: {record_id}')
            return jsonify({
                'success': False,
                'status': 'erro',
                'error_code': 'QUEUE_FAILED',
                'message': 'Falha ao enfileirar.',
                'processar_arquivo_record_id': record_id,
            }), 200

        # Responder HTTP 202 (sucesso assíncrono)
        logger.info(f'[/separar] 202: {record_id} | task_id={task_id}')
        return jsonify({
            'success': True,
            'status': 'enfileirado',
            'message': 'PDF armazenado e enviado para processamento',
            'error_code': None,
            'processar_arquivo_record_id': record_id,
            'task_id': task_id,
        }), 202

    except Exception as exc:
        logger.exception(f'ERRO INESPERADO em /separar: {type(exc).__name__}')
        return jsonify({
            'error': f'{type(exc).__name__}: erro interno',
            'detail': 'Entre em contato com suporte',
        }), 500


@app.route('/tarefas/<task_id>', methods=['GET'])
def consultar_tarefa(task_id):
    """
    Consulta status de uma tarefa Celery.

    Retorna:
    {
      "task_id": "...",
      "state": "PENDING|STARTED|SUCCESS|FAILURE",
      "processar_arquivo_record_id": "rec..." (se disponível)
    }
    """
    try:
        from tarefas_processar_pdf import processar_pdf_task
        from celery_app import celery_app

        # Obter resultado da tarefa
        result = celery_app.AsyncResult(task_id)

        resp = {
            'task_id': task_id,
            'state': result.state,
        }

        # Se tiver metadata, incluir
        if result.info:
            if isinstance(result.info, dict):
                if 'processar_arquivo_record_id' in result.info:
                    resp['processar_arquivo_record_id'] = result.info.get('processar_arquivo_record_id')

        logger.info(f'[TAREFAS] Consulta: {task_id} → {result.state}')
        return jsonify(resp), 200

    except Exception as e:
        logger.error(f'[TAREFAS] Erro ao consultar {task_id}: {type(e).__name__}')
        return jsonify({
            'error': 'Erro ao consultar tarefa',
            'task_id': task_id,
        }), 500


@app.route('/separar/zip', methods=['POST'])
def separar_zip():
    record_id = request.args.get('record_id') or (request.get_json(silent=True) or {}).get('record_id')
    caminho_pdf = None
    try:
        caminho_pdf = extrair_pdf_do_request()
        if not caminho_pdf or not os.path.exists(caminho_pdf):
            return _error_response_separar(
                'ATTACHMENT_NOT_FOUND',
                record_id=record_id,
                message='PDF não recebido.'
            )

        tamanho = os.path.getsize(caminho_pdf)
        if tamanho < 100:
            if caminho_pdf and os.path.exists(caminho_pdf):
                os.unlink(caminho_pdf)
            return _error_response_separar(
                'PDF_EMPTY',
                record_id=record_id,
                message='PDF muito pequeno (< 100 bytes).'
            )

        try:
            mapa, _ = construir_mapa_cpf(caminho_pdf)
        except Exception as exc:
            exc_type = type(exc).__name__
            if 'password' in str(exc).lower() or exc_type in ('PdfReadError',):
                return _error_response_separar(
                    'PDF_PASSWORD_INCORRECT',
                    record_id=record_id,
                    log_detail=f'{exc_type}: {str(exc)[:200]}'
                )
            elif 'corrupt' in str(exc).lower() or exc_type in ('PdfSyntaxError',):
                return _error_response_separar(
                    'PDF_CORRUPTED',
                    record_id=record_id,
                    log_detail=f'{exc_type}: {str(exc)[:200]}'
                )
            else:
                return _error_response_separar(
                    'PROCESSING_ERROR',
                    record_id=record_id,
                    log_detail=f'{exc_type}: {str(exc)[:200]}'
                )

        if not mapa:
            return _error_response_separar(
                'EMPLOYEE_NOT_IDENTIFIED',
                record_id=record_id
            )

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for cpf, dados in mapa.items():
                pdf_ind = extrair_pdf_colaborador(caminho_pdf, dados['paginas'])
                nome_arq = (
                    f"holerite_{cpf.replace('.','').replace('-','')}"
                    f"_{dados['nome'].replace(' ','_')[:30]}.pdf"
                )
                zf.writestr(nome_arq, pdf_ind)
                del pdf_ind
                gc.collect()

        zip_b64 = base64.b64encode(zip_buf.getvalue()).decode('utf-8')
        return _success_response_separar(
            {
                'total_funcionarios': len(mapa),
                'zip_base64': zip_b64,
                'tamanho_zip_bytes': len(zip_buf.getvalue()),
            },
            record_id=record_id
        )

    except Exception as exc:
        logger.exception(f'ERRO INESPERADO em /separar/zip: {type(exc).__name__}: {str(exc)[:200]}')
        if caminho_pdf and os.path.exists(caminho_pdf):
            try:
                os.unlink(caminho_pdf)
            except Exception as e:
                logger.warning(f'Erro ao limpar {caminho_pdf}: {e}')
        return jsonify({
            'error': f'{type(exc).__name__}: erro interno no servidor',
            'detail': 'Entre em contato com suporte'
        }), 500
    finally:
        if caminho_pdf and os.path.exists(caminho_pdf):
            try:
                os.unlink(caminho_pdf)
            except Exception as e:
                logger.warning(f'Erro ao limpar {caminho_pdf}: {e}')


@app.route('/processar-holerites', methods=['POST', 'OPTIONS'])
def processar_holerites():
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    etapa = 'init'
    caminho_pdf = None
    try:
        if not AIRTABLE_API_KEY:
            return jsonify({
                'status': 'erro',
                'erro': 'AIRTABLE_API_KEY não configurada no servidor.',
                'etapa': etapa,
            }), 500

        etapa = 'receber_pdf'
        caminho_pdf = extrair_pdf_do_request()
        if not caminho_pdf or not os.path.exists(caminho_pdf) or os.path.getsize(caminho_pdf) < 100:
            return jsonify({'status': 'erro', 'erro': 'PDF não recebido.', 'etapa': etapa}), 400
        with open(caminho_pdf, 'rb') as _f:
            _header = _f.read(4)
        if _header != b'%PDF':
            os.unlink(caminho_pdf)
            caminho_pdf = None
            return jsonify({'status': 'erro', 'erro': 'Dados não são PDF válido.', 'etapa': etapa}), 400

        pdf_kb = os.path.getsize(caminho_pdf) // 1024
        logger.info(f'[INICIO] PDF salvo em disco: {pdf_kb} KB | RAM: {_mem_mb()} MB')

        etapa = 'parametros_mes'
        folha_mensal = (
            request.form.get('folha_mensal')
            or request.args.get('folha_mensal')
            or (request.get_json(silent=True) or {}).get('folha_mensal')
        )
        mes_cont_id = (
            request.form.get('mes_cont_id')
            or request.args.get('mes_cont_id')
            or (request.get_json(silent=True) or {}).get('mes_cont_id')
        )

        nome_mes, ano, mes_num = mes_anterior_info()
        if not folha_mensal:
            folha_mensal = f'{nome_mes} {ano}'

        etapa = 'buscar_mes_contabilidade'
        if not mes_cont_id:
            mes_cont_id, nome_mes_cont = buscar_mes_contabilidade_atual()
            if not mes_cont_id:
                return jsonify({
                    'status': 'erro',
                    'erro': f'Contabilidade Mensal "{nome_mes_cont}" não encontrada. '
                            'Passe mes_cont_id manualmente.',
                    'etapa': etapa,
                }), 500

        ultimo_dia    = monthrange(ano, mes_num)[1]
        data_holerite = f'{ano}-{mes_num:02d}-{ultimo_dia:02d}'

        # casamento_manual — override pontual para casos que passaram por
        # revisão humana (divergência de grafia entre PDF e cadastro que não
        # é seguro resolver automaticamente, ex.: nome incompleto/typo real).
        # Formato: {"<índice da página no PDF enviado>": "<record_id do Funcionário>"}
        casamento_manual_raw = (
            request.form.get('casamento_manual')
            or request.args.get('casamento_manual')
            or (request.get_json(silent=True) or {}).get('casamento_manual')
        )
        casamento_manual = {}
        if casamento_manual_raw:
            try:
                casamento_manual = {
                    int(k): v for k, v in json.loads(casamento_manual_raw).items()
                }
            except Exception:
                logger.warning(f'[CONFIG] casamento_manual inválido, ignorando: {casamento_manual_raw}')

        logger.info(
            f'[CONFIG] folha_mensal={folha_mensal} | data={data_holerite} | mes_cont_id={mes_cont_id} | '
            f'casamento_manual={casamento_manual}'
        )

        etapa = 'pass1_mapear_cpf'
        mapa, total_paginas = construir_mapa_cpf(caminho_pdf)

        if not mapa:
            return jsonify({
                'status': 'erro', 'erro': 'Nenhum CPF encontrado no PDF.', 'etapa': etapa,
            }), 422

        total_colab = len(mapa)
        logger.info(
            f'[P2] Iniciando criação | {total_colab} colaboradores | '
            f'{total_paginas} páginas | RAM: {_mem_mb()} MB'
        )

        criados: list = []
        erros:   list = []
        contador = 0

        for cpf, dados in mapa.items():
            contador += 1
            nome_pdf = dados['nome']
            paginas  = dados['paginas']
            valores  = dados.get('valores', {})
            tag      = f'[P2 {contador:02d}/{total_colab}]'

            if cpf.startswith('sem_cpf'):
                # Layout sem CPF impresso na página (ex.: "Mensalista" 06/2026) —
                # cai para casamento por nome, já extraído em construir_mapa_cpf.
                etapa = f'buscar_funcionario_por_nome:{nome_pdf}'
                func_id, nome_at = buscar_funcionario_por_nome(nome_pdf)
                if not func_id and paginas[0] in casamento_manual:
                    func_id = casamento_manual[paginas[0]]
                    nome_at = nome_pdf
                    logger.info(f'{tag} Casado por override manual (pág {paginas[0]}) -> {func_id}')
                if not func_id:
                    logger.warning(f'{tag} CPF não extraído e nome não casou | nome: {nome_pdf} | págs: {paginas}')
                    erros.append({'cpf': 'N/A', 'nome': nome_pdf,
                                  'motivo': 'CPF não extraído e nome não encontrado no Airtable',
                                  'paginas': paginas})
                    continue
                logger.info(f'{tag} Casado por nome: "{nome_pdf}" -> {nome_at}')
            else:
                etapa = f'buscar_funcionario:{cpf}'
                func_id, nome_at = buscar_funcionario_por_cpf(cpf)
                if not func_id:
                    logger.warning(f'{tag} Funcionário não encontrado | CPF: {_mascarar_cpf(cpf)}')
                    erros.append({'cpf': cpf, 'nome': nome_pdf,
                                  'motivo': 'Funcionário não encontrado no Airtable'})
                    continue

            nome_final = nome_at or nome_pdf
            filename   = f'Holerite {folha_mensal} - {nome_final}.pdf'
            logger.info(
                f'{tag} {nome_final} (CPF: {cpf}) | págs: {paginas} | vals: {valores}'
            )

            pdf_ind = None
            try:
                etapa   = f'extrair_pdf:{cpf}'
                pdf_ind = extrair_pdf_colaborador(caminho_pdf, paginas)
                logger.info(f'{tag} PDF: {len(pdf_ind)//1024} KB | RAM: {_mem_mb()} MB')

                etapa     = f'criar_registro:{cpf}'
                record_id = criar_registro_holerite(
                    nome_final, func_id, folha_mensal, data_holerite, mes_cont_id, valores
                )
                logger.info(f'{tag} Registro criado: {record_id} | valores: {valores}')

                etapa = f'anexar_pdf:{cpf}'
                anexar_pdf_holerite(record_id, pdf_ind, filename)
                logger.info(f'{tag} PDF anexado ✓')

                criados.append({
                    'cpf': cpf, 'nome': nome_final,
                    'record_id': record_id, 'arquivo': filename,
                    'paginas': paginas, 'valores': valores,
                })

            except requests.HTTPError as exc:
                motivo = f'HTTP {exc.response.status_code}: {exc.response.text[:300]}'
                logger.error(f'{tag} ERRO Airtable: {motivo}')
                erros.append({'cpf': cpf, 'nome': nome_final,
                              'motivo': motivo, 'etapa': etapa})
            except Exception as exc:
                logger.exception(f'{tag} ERRO inesperado: {exc}')
                erros.append({'cpf': cpf, 'nome': nome_final,
                              'motivo': str(exc), 'etapa': etapa})
            finally:
                if pdf_ind is not None:
                    del pdf_ind
                gc.collect()

        if caminho_pdf and os.path.exists(caminho_pdf):
            os.unlink(caminho_pdf)
        gc.collect()

        logger.info(
            f'[FIM] páginas={total_paginas} | colaboradores={total_colab} | '
            f'criados={len(criados)} | erros={len(erros)} | RAM: {_mem_mb()} MB'
        )

        return jsonify({
            'status': 'concluido',
            'folha_mensal': folha_mensal,
            'mes_contabilidade_id': mes_cont_id,
            'total_paginas': total_paginas,
            'total_colaboradores': total_colab,
            'total_criados': len(criados),
            'total_erros': len(erros),
            'criados': criados,
            'erros': erros,
        })

    except Exception as exc:
        logger.exception(f'[ERRO CRÍTICO] etapa={etapa}')
        return jsonify({
            'status': 'erro_critico',
            'erro': str(exc),
            'etapa': etapa,
        }), 500
    finally:
        # Blindagem LGPD: garante a remoção do PDF de origem (lote inteiro
        # de holerites, com CPF/valores de vários colaboradores) do /tmp
        # mesmo quando uma exceção não prevista interrompe o processamento
        # antes do unlink explícito do caminho feliz, acima. Mesmo padrão
        # já usado em separar_zip().
        if caminho_pdf and os.path.exists(caminho_pdf):
            try:
                os.unlink(caminho_pdf)
            except Exception as e:
                logger.warning(f'Erro ao limpar {caminho_pdf}: {e}')


def _processar_folha_ponto_arquivo(caminho_pdf, folha_mensal, disparar_assinatura, cpfs_excluir):
    """v3.00 — Processa UM arquivo de Folha de Ponto (master Secullum OU
    Folha Manual): fatia por CPF/nome, anexa o cartão individual no campo
    "PDF Folha Ponto" e (opcional) dispara a Assinatura Nativa.

    Compartilhado pelas duas fases do pipeline (Folhas Manuais processadas
    PRIMEIRO, master da Secullum depois) — `cpfs_excluir` é a Regra de
    Prevalência: qualquer CPF nesse conjunto é IGNORADO por completo (sem
    fatiar, sem salvar, sem disparar WhatsApp).

    Formato A (CPF explícito) e Formato B ("Colaborador: NOME", sem CPF) são
    tratados juntos aqui: Formato B busca o Funcionário por nome exato
    normalizado e, se não achar, por Primeiro+Último Nome (ignora conectores
    da/de/do/dos) — homônimos nesse fallback parcial viram Pendência
    explícita em vez de vincular por adivinhação.
    """
    mapa, total_paginas = construir_mapa_cpf(caminho_pdf)
    anexados, erros, ignorados = [], [], []
    cpfs_processados = set()

    for idx, (cpf, dados) in enumerate(mapa.items()):
        nome_pdf = dados.get('nome')
        paginas = dados.get('paginas')
        cpf_formato = 'A'  # A = CPF explícito no PDF; B = só nome ("Colaborador:")

        if cpf.startswith('sem_cpf'):
            # Formato B (Folhas Manuais): sem CPF impresso — o único dado
            # disponível é o nome extraído via "Colaborador: NOME" (ver
            # extrair_nome_funcionario). Busca exata por nome primeiro.
            func_id, nome_at = _buscar_funcionario_por_nome(nome_pdf)
            if not func_id:
                # Fallback (v3.00): Primeiro Nome + Último Nome, ignorando
                # conectores — cobre o caso do PDF grafar o nome sem "da"/
                # "de"/"dos" presentes no cadastro. Homônimos NÃO são
                # resolvidos automaticamente.
                func_id, nome_at, homonimos = _buscar_funcionario_por_nome_parcial(nome_pdf)
                if homonimos:
                    pendencia_id = _criar_registro(TABLE_PENDENCIAS, {
                        F_PEND_NOME: f'Homônimos na Folha de Ponto: "{nome_pdf}"',
                        F_PEND_STATUS: 'Pendente',
                        F_PEND_TIPO: 'Homônimos - Folha de Ponto (Formato B)',
                        F_PEND_OBS: 'Candidatos encontrados: ' + '; '.join(
                            f"{c['nome']} ({c['id']})" for c in homonimos),
                        F_PEND_DATA: datetime.now().isoformat(),
                    })
                    erros.append({'cpf': 'N/A', 'nome': nome_pdf,
                                  'motivo': 'Homônimos encontrados — Pendência criada para revisão manual',
                                  'pendencia_record_id': pendencia_id, 'candidatos': homonimos,
                                  'paginas': paginas})
                    continue
            if not func_id:
                erros.append({'cpf': 'N/A', 'nome': nome_pdf,
                              'motivo': 'Sem CPF no PDF e nome não encontrado no Airtable (Formato B)',
                              'paginas': paginas})
                continue
            cpf_formato = 'B'
            cpf_real = _buscar_cpf_por_funcionario_id(func_id)
            cpf_digitos = re.sub(r'\D', '', cpf_real or '')
        else:
            func_id, nome_at = buscar_funcionario_por_cpf(cpf)
            if not func_id:
                erros.append({'cpf': cpf, 'nome': nome_pdf,
                              'motivo': 'Funcionário não encontrado no Airtable'})
                continue
            cpf_digitos = re.sub(r'\D', '', cpf)

        # Regra de Prevalência (v3.00): pula quem já recebeu Folha Manual —
        # não anexa, não dispara assinatura, não reprocessa o cartão
        # zerado/com faltas do master da Secullum por cima.
        if cpf_digitos and cpf_digitos in cpfs_excluir:
            ignorados.append({'cpf': cpf_digitos, 'nome': nome_at, 'funcionario_id': func_id,
                              'motivo': 'ja_recebeu_folha_manual_nesta_competencia'})
            continue

        nome_final = nome_at or nome_pdf
        filename = f'Folha de Ponto {folha_mensal} - {nome_final}.pdf'
        try:
            pdf_ind = extrair_pdf_colaborador(caminho_pdf, paginas)
            _anexar_attachment(TABLE_FUNC, func_id, F_FUNC_PDF_FOLHA, pdf_ind, filename)
            item = {
                'cpf': cpf_digitos or cpf, 'nome': nome_final, 'funcionario_id': func_id,
                'paginas': paginas, 'arquivo': filename, 'formato': cpf_formato,
            }

            if disparar_assinatura:
                try:
                    _at_throttle()
                    r_check = requests.get(
                        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}/{func_id}',
                        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
                        params={'returnFieldsByFieldId': 'true'}, timeout=15,
                    )
                    anexo_url = None
                    arquivo_record_id = None
                    anexo_url = None
                    if r_check.ok:
                        anexos_ponto = r_check.json().get('fields', {}).get(F_FUNC_PDF_FOLHA) or []
                        if anexos_ponto:
                            anexo_url = anexos_ponto[-1].get('url')
                            # ✅ v3.6: Criar Record em Arquivos para garantir identidade documental
                            try:
                                _at_throttle()
                                r_arquivo = requests.post(
                                    f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ARQUIVOS}',  # ✅ corrigido: tblbgJqXh0u1Cjq1s não existe na base
                                    headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
                                    json={
                                        'fields': {
                                            F_ARQ_NOME: filename,
                                            F_ARQ_ATTACH: [{'url': anexo_url, 'filename': filename}],
                                        }
                                    },
                                    timeout=15,
                                )
                                if r_arquivo.ok:
                                    arquivo_record_id = r_arquivo.json().get('id')
                            except Exception as exc:
                                logger.warning(f'[PONTO] Falha ao criar Record em Arquivos ({func_id}): {exc}')

                    if arquivo_record_id:
                        resultado_ass, _ = _gerar_assinatura_core(
                            funcionario_id=func_id,
                            tipo_documento='FOLHA_PONTO',  # ✅ Tipo canônico (corrigido de 'Folha de Ponto')
                            arquivo_record_id=arquivo_record_id,  # ✅ Record ID obrigatório (corrigido de ausente)
                            nome_documento=f'Folha de Ponto {folha_mensal} - {nome_final}',
                            mensagem_extra='Por favor, revise os dias e horários abaixo e confirme.',
                            disparar_whatsapp=True,
                        )
                        item['assinatura'] = resultado_ass
                    else:
                        item['assinatura'] = {'status': 'erro', 'erro': 'arquivo_record_id_nao_criado'}
                except Exception as exc:
                    logger.error(f'[PONTO] falha ao disparar assinatura {_mascarar_cpf(cpf)}: {exc}')
                    item['assinatura'] = {'status': 'erro', 'erro': str(exc)}

            anexados.append(item)
            if cpf_digitos:
                cpfs_processados.add(cpf_digitos)
            del pdf_ind
        except Exception as exc:
            logger.error(f'[PONTO] erro ao anexar Folha de Ponto {_mascarar_cpf(cpf)}: {exc}')
            erros.append({'cpf': cpf, 'nome': nome_final, 'motivo': str(exc)})

        # Otimização de memória (v3.00, anti-timeout): libera RAM a cada 15
        # colaboradores processados neste loop de anexo/disparo, além do
        # gc.collect() que a Pass 1 (construir_mapa_cpf) já faz na varredura.
        if (idx + 1) % 15 == 0:
            gc.collect()

    return {
        'total_colaboradores': len(mapa),
        'total_paginas': total_paginas,
        'anexados': anexados,
        'ignorados': ignorados,
        'erros': erros,
        'cpfs_processados': cpfs_processados,
    }


@app.route('/processar-folha-ponto', methods=['POST', 'OPTIONS'])
def processar_folha_ponto_master():
    """
    v3.00 — Pipeline de fechamento de Folha de Ponto: master Secullum
    (Formato A, CPF explícito) + Folhas Manuais da Magnata (Formato B,
    "Colaborador: NOME", sem CPF), com Regra de Prevalência (Manual sempre
    vence) e disparo de Assinatura Nativa. Escopo estritamente isolado a
    este endpoint e ao campo "PDF Folha Ponto" de Funcionários — não toca em
    nenhuma rota de Admissão/Rescisão/cadastro nem no motor genérico de
    WhatsApp usado por elas.

    Body multipart:
      - "pdf": PDF master da Secullum [obrigatório]
      - "pdf_manual": PDF das Folhas Manuais [opcional] — se enviado, é
        processado PRIMEIRO; todo CPF resolvido nele entra automaticamente
        na lista de exclusão (cpfs_manuais_processados, em memória) antes
        de processar o master — nenhuma página da Secullum desses CPFs é
        fatiada, salva ou disparada.
      - "folha_mensal": ex. "Junho 2026" [opcional, default = mês anterior]
      - "disparar_assinatura": true/false [opcional, default true]
      - "cpfs_ja_processados": JSON/CSV de CPFs extras a excluir (soma-se
        aos vindos de pdf_manual, se ambos forem enviados) [opcional]

    Fase Definitiva (assinatura confirmada): tratada à parte, no webhook
    /assinatura/<hash_token> — ver bloco isolado "Folha de Ponto" lá, que
    substitui o PDF provisório pelo assinado (nunca acumula os dois).

    Não exige X-API-KEY (espelha /processar-holerites); usa a AIRTABLE_API_KEY
    do servidor.
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    caminho_master = extrair_pdf_do_request('pdf')
    if not caminho_master or not os.path.exists(caminho_master) or os.path.getsize(caminho_master) < 100:
        return jsonify({'status': 'erro', 'erro': 'PDF (master) não recebido.'}), 400
    with open(caminho_master, 'rb') as _f:
        if _f.read(4) != b'%PDF':
            os.unlink(caminho_master)
            return jsonify({'status': 'erro', 'erro': 'Dados do master não são PDF válido.'}), 400

    caminho_manual = extrair_pdf_do_request('pdf_manual')
    if caminho_manual and (not os.path.exists(caminho_manual) or os.path.getsize(caminho_manual) < 100):
        try:
            os.unlink(caminho_manual)
        except OSError:
            pass
        caminho_manual = None
    if caminho_manual:
        with open(caminho_manual, 'rb') as _f:
            if _f.read(4) != b'%PDF':
                os.unlink(caminho_manual)
                caminho_manual = None

    folha_mensal = (
        request.form.get('folha_mensal')
        or request.args.get('folha_mensal')
        or (request.get_json(silent=True) or {}).get('folha_mensal')
    )
    if not folha_mensal:
        nome_mes, ano, _ = mes_anterior_info()
        folha_mensal = f'{nome_mes} {ano}'

    disparar_assinatura = str(
        request.form.get('disparar_assinatura',
                          request.args.get('disparar_assinatura', 'true'))
    ).strip().lower() in ('1', 'true', 'yes', 'sim')

    cpfs_ja_processados_raw = (
        request.form.get('cpfs_ja_processados')
        or request.args.get('cpfs_ja_processados')
        or (request.get_json(silent=True) or {}).get('cpfs_ja_processados')
    )
    cpfs_ja_processados = set()
    if cpfs_ja_processados_raw:
        if isinstance(cpfs_ja_processados_raw, str):
            try:
                lista = json.loads(cpfs_ja_processados_raw)
            except ValueError:
                lista = cpfs_ja_processados_raw.split(',')
        else:
            lista = cpfs_ja_processados_raw
        cpfs_ja_processados = {re.sub(r'\D', '', str(c)) for c in lista if str(c).strip()}

    # Regra de Prevalência (v3.00): array em memória — CPFs resolvidos nas
    # Folhas Manuais (processadas primeiro) somam-se aos excluídos explícitos
    # antes de processar o master da Secullum.
    cpfs_manuais_processados = set(cpfs_ja_processados)
    resultado_manual = None

    # Blindagem LGPD: try/finally garante a remoção dos 2 PDFs de origem
    # (master Secullum + manual, ambos com cartão de ponto de vários
    # colaboradores) mesmo se _processar_folha_ponto_arquivo levantar uma
    # exceção não prevista -- antes, o unlink de cada um só acontecia na
    # linha seguinte à chamada, então uma exceção ali deixava o PDF preso
    # no /tmp indefinidamente.
    try:
        if caminho_manual:
            r_manual = _processar_folha_ponto_arquivo(
                caminho_manual, folha_mensal, disparar_assinatura, cpfs_excluir=cpfs_ja_processados)
            cpfs_manuais_processados |= r_manual['cpfs_processados']
            resultado_manual = {
                'total_colaboradores': r_manual['total_colaboradores'],
                'total_paginas': r_manual['total_paginas'],
                'total_anexados': len(r_manual['anexados']),
                'total_ignorados': len(r_manual['ignorados']),
                'total_erros': len(r_manual['erros']),
                'anexados': r_manual['anexados'],
                'ignorados': r_manual['ignorados'],
                'erros': r_manual['erros'],
            }

        r_master = _processar_folha_ponto_arquivo(
            caminho_master, folha_mensal, disparar_assinatura, cpfs_excluir=cpfs_manuais_processados)

        if not resultado_manual and r_master['total_colaboradores'] == 0:
            return jsonify({'status': 'erro', 'erro': 'Nenhum CPF/nome encontrado no PDF.'}), 422

        todos_anexados = r_master['anexados'] + (resultado_manual['anexados'] if resultado_manual else [])
        total_assinaturas_enviadas = sum(
            1 for a in todos_anexados
            if disparar_assinatura and (a.get('assinatura') or {}).get('whatsapp_disparo') == 'enviado'
        )
        total_assinaturas_falha = sum(
            1 for a in todos_anexados
            if disparar_assinatura and (a.get('assinatura') or {}).get('status') != 'ok'
        ) if disparar_assinatura else 0

        return jsonify({
            'status': 'concluido',
            'folha_mensal': folha_mensal,
            'disparar_assinatura': disparar_assinatura,
            'total_assinaturas_enviadas': total_assinaturas_enviadas,
            'total_assinaturas_falha': total_assinaturas_falha,
            'folhas_manuais': resultado_manual,  # None se "pdf_manual" não foi enviado
            'master_secullum': {
                'total_colaboradores': r_master['total_colaboradores'],
                'total_paginas': r_master['total_paginas'],
                'total_anexados': len(r_master['anexados']),
                'total_ignorados_prevalencia_manual': len(r_master['ignorados']),
                'total_erros': len(r_master['erros']),
                'anexados': r_master['anexados'],
                'ignorados': r_master['ignorados'],
                'erros': r_master['erros'],
            },
            # CPFs processados nesta chamada (manual + master) — útil para
            # repassar em cpfs_ja_processados numa reexecução futura, se preciso.
            'cpfs_processados': sorted(cpfs_manuais_processados | r_master['cpfs_processados']),
        })
    finally:
        if caminho_manual and os.path.exists(caminho_manual):
            try:
                os.unlink(caminho_manual)
            except OSError:
                pass
        if caminho_master and os.path.exists(caminho_master):
            try:
                os.unlink(caminho_master)
            except OSError:
                pass


@app.route('/corrigir-valores', methods=['POST', 'OPTIONS'])
def corrigir_valores():
    """
    Atualiza valores financeiros de holerites já criados mas sem valores
    (ou com valores zerados), SEM criar registros novos — apenas PATCH.

    Fluxo:
      1. Busca holerites para folha_mensal + mes_contabilidade especificados,
         que estejam sem Total Vencimentos (BLANK ou 0)
      2. Para cada registro, baixa o PDF já anexado no Airtable
      3. Extrai valores financeiros via pdfplumber
      4. Atualiza (PATCH) apenas os campos financeiros do registro existente

    Parâmetros JSON:
      folha_mensal       — ex. "Maio 2026"   [obrigatório]
      mes_contabilidade  — ex. "Junho 2026"  [opcional, recomendado]
      dry_run            — true: apenas simula, NÃO grava nada no Airtable
      limit              — int: processa no máximo N registros (teste controlado)
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    data = request.get_json(force=True, silent=True) or {}

    def _param(nome):
        return data.get(nome) if data.get(nome) is not None else request.args.get(nome)

    folha_mensal = _param('folha_mensal')
    mes_contabilidade = _param('mes_contabilidade')
    if not folha_mensal:
        return jsonify({'status': 'erro', 'erro': 'folha_mensal é obrigatório'}), 400

    dry_run_raw = _param('dry_run')
    dry_run = str(dry_run_raw).strip().lower() in ('1', 'true', 'yes', 'sim')

    limit = None
    limit_raw = _param('limit')
    if limit_raw is not None:
        try:
            limit = int(limit_raw)
            if limit <= 0:
                limit = None
        except (TypeError, ValueError):
            limit = None

    logger.info(
        f'[CORR] Iniciando correção | folha_mensal={folha_mensal} | '
        f'mes_contabilidade={mes_contabilidade or "(qualquer)"} | '
        f'dry_run={dry_run} | limit={limit if limit is not None else "(sem limite)"}'
    )

    # 1. Buscar holerites da folha/mês sem Total Vencimentos (BLANK ou 0)
    condicoes = [
        f'{{Folha Mensal}}="{folha_mensal}"',
        'OR({Total Vencimentos}=BLANK(), {Total Vencimentos}=0)',
    ]
    if mes_contabilidade:
        condicoes.append(f'FIND("{mes_contabilidade}", ARRAYJOIN({{Mês Contabilidade}}))')
    formula = 'AND(' + ', '.join(condicoes) + ')'

    todos: list = []
    offset = None
    pagina = 0
    while True:
        pagina += 1
        _at_throttle()
        params = {
            'filterByFormula': formula,
            'fields[]': [
                'Holerite', 'PDF HOLERITE', 'Folha Mensal',
                'Mês Contabilidade', 'Total Vencimentos',
            ],
            'pageSize': 100,
        }
        if offset:
            params['offset'] = offset
        r = requests.get(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_HOL}',
            headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
            params=params,
            timeout=30,
        )
        if not r.ok:
            return jsonify({
                'status': 'erro',
                'erro': f'Airtable list error: {r.status_code} {r.text[:300]}'
            }), 500

        body = r.json()
        todos.extend(body.get('records', []))
        offset = body.get('offset')
        logger.info(f'[CORR] Página {pagina}: {len(body.get("records", []))} registros')
        if not offset:
            break

    logger.info(f'[CORR] Total sem valores: {len(todos)}')
    if not todos:
        return jsonify({
            'status': 'concluido',
            'folha_mensal': folha_mensal,
            'mes_contabilidade': mes_contabilidade,
            'dry_run': dry_run,
            'total_sem_valores': 0,
            'atualizados': 0,
            'erros': 0,
            'detalhe': [],
        })

    total_encontrados = len(todos)
    if limit is not None:
        todos = todos[:limit]
        logger.info(f'[CORR] limit={limit} aplicado | processando {len(todos)} de {total_encontrados}')

    atualizados = 0
    erros_list: list = []
    processados: list = []

    for rec in todos:
        record_id = rec['id']
        nome_hol  = rec.get('fields', {}).get('Holerite', record_id)
        attachments = rec.get('fields', {}).get('PDF HOLERITE', [])

        if not attachments:
            logger.warning(f'[CORR] {nome_hol}: sem PDF anexado')
            erros_list.append({'record_id': record_id, 'nome': nome_hol,
                               'motivo': 'sem PDF anexado'})
            continue

        pdf_url = attachments[0].get('url')
        if not pdf_url:
            erros_list.append({'record_id': record_id, 'nome': nome_hol,
                               'motivo': 'URL do PDF não disponível'})
            continue

        # Baixar o PDF do Airtable CDN
        tmp_path = None
        try:
            resp = requests.get(pdf_url, timeout=30)
            resp.raise_for_status()
            tmp_path = os.path.join(tempfile.gettempdir(), f'corr_{_uuid.uuid4().hex}.pdf')
            with open(tmp_path, 'wb') as f:
                f.write(resp.content)

            # Extrair texto e valores
            texto_completo = ''
            with pdfplumber.open(tmp_path) as pdf_doc:
                for pg in pdf_doc.pages:
                    texto_completo += (pg.extract_text() or '') + '\n'

            valores = extrair_valores_holerite(texto_completo)
            logger.info(f'[CORR] {nome_hol}: valores extraídos = {valores}')

            if not valores:
                erros_list.append({'record_id': record_id, 'nome': nome_hol,
                                   'motivo': 'valores não encontrados no PDF'})
                continue

            if dry_run:
                logger.info(f'[CORR] {nome_hol}: [DRY RUN] não gravado | valores = {valores}')
                processados.append({'record_id': record_id, 'nome': nome_hol,
                                     'valores': valores, 'gravado': False})
            else:
                # Atualizar no Airtable
                atualizar_valores_holerite(record_id, valores)
                logger.info(f'[CORR] {nome_hol}: atualizado ✓ | valores = {valores}')
                processados.append({'record_id': record_id, 'nome': nome_hol,
                                     'valores': valores, 'gravado': True})

            atualizados += 1

        except requests.HTTPError as exc:
            motivo = f'HTTP {exc.response.status_code} ao baixar PDF'
            logger.error(f'[CORR] {nome_hol}: {motivo}')
            erros_list.append({'record_id': record_id, 'nome': nome_hol, 'motivo': motivo})
        except Exception as exc:
            logger.exception(f'[CORR] {nome_hol}: erro inesperado')
            erros_list.append({'record_id': record_id, 'nome': nome_hol, 'motivo': str(exc)})
        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.unlink(tmp_path)
            gc.collect()

    logger.info(
        f'[CORR] Concluído | dry_run={dry_run} | total_encontrados={total_encontrados} | '
        f'processados={len(todos)} | atualizados={atualizados} | erros={len(erros_list)}'
    )
    return jsonify({
        'status': 'concluido',
        'folha_mensal': folha_mensal,
        'mes_contabilidade': mes_contabilidade,
        'dry_run': dry_run,
        'limit': limit,
        'total_sem_valores': total_encontrados,
        'total_processados': len(todos),
        'atualizados': atualizados,
        'erros': len(erros_list),
        'detalhe_erros': erros_list,
        'detalhe_processados': processados,
    })


def _detectar_competencia_fiscal(texto: str):
    """v2.96 (Macro 6A — correção de auditoria, achado #2) — Extrai a
    competência (mês/ano) de um documento fiscal a partir de marcadores
    semânticos explícitos: 'Competência MM/AAAA', 'Período de Apuração
    MM/AAAA' / 'Período de Referência MM/AAAA', ou a mesma família por
    extenso ('Competência: Maio de 2026' — mesmo padrão já reconhecido em
    extrair_competencia_holerite, aqui restrito à proximidade do marcador
    para nunca casar com um mês solto no meio do texto).

    NUNCA usa vencimento/pagamento/emissão/recebimento como substituto de
    competência — são datas diferentes e podem apontar para outro mês.
    NUNCA usa nome do arquivo, data do e-mail ou mês corrente como
    confirmação isolada.

    Antes desta correção, a ausência de marcador explícito fazia esta
    função assumir silenciosamente 'mês anterior ao recebimento' — a
    competência então gravada em TODOS os registros de cliente fatiados a
    partir do mesmo PDF (ver _processar_doc_cliente_master) podia ficar
    errada sem nenhum sinal de revisão. Agora, na ausência de marcador
    comprovado, devolve None — quem chama (_processar_anexo_fiscal) DEVE
    tratar isso como competência não comprovada e rotear para revisão
    manual, nunca gravar um mês adivinhado.

    Ambiguidade (Macro 6A — fechamento autônomo, revisão adversarial):
    quando o texto tem MAIS DE UM marcador de competência com valores
    DIVERGENTES (ex.: um PDF consolidado com "Competência: 05/2026" numa
    página e "Competência: 08/2026" noutra), a versão anterior desta
    correção usava `re.search` (só o primeiro achado) e escolheria um dos
    dois silenciosamente — mesmo risco de "adivinhar" que a correção
    original eliminou para o caso de ausência total de marcador. Agora
    todos os achados (numérico + por extenso) são coletados; se houver mais
    de um valor (mês, ano) DISTINTO entre eles, é tratado como ambíguo —
    devolve None e registra um aviso sanitizado (sem nenhum trecho do PDF)
    para auditoria. Repetições do MESMO valor (ex.: cabeçalho + rodapé) não
    são ambíguas."""
    if not texto:
        return None

    marcador = r'(?:Compet[êe]ncia|Per[íi]odo\s+de\s+(?:Apura[çc][ãa]o|Refer[êe]ncia))'
    nomes_meses = ['janeiro', 'fevereiro', 'marco', 'abril', 'maio', 'junho',
                   'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']

    valores_encontrados = []  # lista de (mes_num, ano_str), na ordem em que aparecem

    # Formato numérico — "Competência: 05/2026" / "Período de Apuração 05/2026"
    for m in re.finditer(marcador + r'\D{0,10}(\d{2})/(\d{4})', texto, re.IGNORECASE):
        mes_num, ano = int(m.group(1)), m.group(2)
        if 1 <= mes_num <= 12:
            valores_encontrados.append((mes_num, ano))

    # Formato por extenso — "Competência: Maio de 2026" / "Período de
    # Referência Maio/2026" — mesma lista de meses de extrair_competencia_holerite.
    for m in re.finditer(
        marcador + r'\D{0,15}(Janeiro|Fevereiro|Mar[çc]o|Abril|Maio|Junho|Julho|Agosto|'
        r'Setembro|Outubro|Novembro|Dezembro)(?:\s+de)?\s+(\d{4})',
        texto, re.IGNORECASE,
    ):
        nome_norm = m.group(1).lower().replace('ç', 'c')
        if nome_norm in nomes_meses:
            valores_encontrados.append((nomes_meses.index(nome_norm) + 1, m.group(2)))

    if not valores_encontrados:
        # Nenhum marcador semântico de competência encontrado — não adivinha.
        return None

    distintos = sorted(set(valores_encontrados))
    if len(distintos) > 1:
        logger.warning(
            '[FISCAL] Competência ambígua: %d valores distintos encontrados no '
            'mesmo documento (%s) — não é possível determinar qual prevalece, '
            'tratado como não comprovada.',
            len(distintos),
            ', '.join(f'{mes:02d}/{ano}' for mes, ano in distintos),
        )
        return None

    mes_num, ano = distintos[0]
    return f'{MESES_PT[mes_num - 1]} {ano}'


def _processar_anexo_fiscal(conteudo: bytes, nome_arquivo: str, tipo_doc: str, texto: str) -> dict:
    """v2.96 — Roteia um anexo vindo da caixa fiscal (REMETENTE_FISCAL) para a
    tabela correta, reaproveitando a mesma infraestrutura já usada para
    Extrato/FGTS de clientes (fatiamento por CNPJ/Nome) e Guias broadcast
    (DCTFWeb), em vez do fluxo de Processar Arquivos do Departamento Pessoal
    (Kit Admissão etc.), que não se aplica a documentos fiscais da empresa."""
    folha_mensal = _detectar_competencia_fiscal(texto)

    # Competência não comprovada (Macro 6A, achado #2): antes, a ausência de
    # marcador explícito era coberta com um "mês anterior" adivinhado, que
    # seria gravado em TODOS os registros de cliente fatiados a partir deste
    # PDF (Extrato/FGTS) ou no comprovante de Guia — sem nenhum sinal de
    # revisão. Agora, sem competência comprovada, o documento NÃO é
    # processado automaticamente; vai para Pendência, igual ao tratamento já
    # existente para "tipo não reconhecido" logo abaixo.
    if not folha_mensal:
        pendencia_id = _criar_registro(TABLE_PENDENCIAS, {
            F_PEND_NOME:   f'Competência fiscal não detectada: {nome_arquivo}',
            F_PEND_STATUS: 'Pendente',
            F_PEND_TIPO:   'Competência fiscal não detectada',
            F_PEND_OBS: (
                f'Tipo: {tipo_doc}. Não foi encontrado marcador explícito de '
                'competência ("Competência MM/AAAA" ou "Período de Apuração/'
                'Referência MM/AAAA", numérico ou por extenso) no texto '
                'extraído do PDF. NÃO processado automaticamente, para não '
                'gravar competência errada em registros de cliente/guia — '
                'revisar manualmente e reprocessar.'
            ),
            F_PEND_DATA:   datetime.now().isoformat(),
        })
        return {'acao': 'competencia_fiscal_nao_detectada', 'tipo_documento': tipo_doc,
                'pendencia_record_id': pendencia_id}

    if tipo_doc in ('Extrato da Folha de Pagamento', 'FGTS'):
        tipo_fatiador = 'extrato' if tipo_doc == 'Extrato da Folha de Pagamento' else 'fgts'
        tmp_path = os.path.join(tempfile.gettempdir(), f'fiscal_{_uuid.uuid4().hex}.pdf')
        try:
            with open(tmp_path, 'wb') as f:
                f.write(conteudo)
            resultado_fatiamento = _processar_doc_cliente_master(tmp_path, tipo_fatiador, folha_mensal)
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
        return {'acao': 'fatiado_por_cliente', 'tipo_documento': tipo_doc,
                'folha_mensal': folha_mensal, 'resultado_fatiamento': resultado_fatiamento}

    if tipo_doc in ('DCTFWeb - Recibo de Entrega', 'DCTFWeb - Declaração'):
        rec_id = _criar_registro(TABLE_GUIAS, {
            F_GUIA_STATUS: 'Recebido', F_GUIA_TIPO: tipo_doc,
            F_GUIA_NOME: f'{tipo_doc} - {folha_mensal}',
        })
        _anexar_attachment(TABLE_GUIAS, rec_id, F_GUIA_PDF, conteudo, nome_arquivo)
        return {'acao': 'arquivado_guia_comprovante', 'tipo_documento': tipo_doc,
                'folha_mensal': folha_mensal, 'guia_record_id': rec_id}

    # Tipo não reconhecido — cria Pendência para revisão manual em vez de
    # descartar (documento fiscal sem classificação clara é risco de
    # compliance, não deve ficar perdido silenciosamente).
    pendencia_id = _criar_registro(TABLE_PENDENCIAS, {
        F_PEND_NOME:   f'Documento fiscal não reconhecido: {nome_arquivo}',
        F_PEND_STATUS: 'Pendente',
        F_PEND_TIPO:   'Documento fiscal não reconhecido',
        F_PEND_OBS:    texto[:500] if texto.strip() else '(sem texto extraído do PDF)',
        F_PEND_DATA:   datetime.now().isoformat(),
    })
    return {'acao': 'pendencia_documento_nao_reconhecido', 'tipo_documento': tipo_doc,
            'pendencia_record_id': pendencia_id}


@app.route('/email/webhook', methods=['POST', 'OPTIONS'])
def email_webhook():
    """
    Fase 2 — Caixa de Entrada.

    Recebe e-mail + anexos (JSON) vindos do Google Apps Script, registra em
    Emails Savian / Arquivos / Processar Arquivos, classifica o tipo de
    documento e cria pendências quando necessário.

    Protegido por header X-API-KEY (variável de ambiente EMAIL_WEBHOOK_KEY).

    Payload esperado (JSON):
      {
        "message_id": "<...@mail.gmail.com>",   [obrigatório]
        "assunto": "...",
        "remetente": "...",
        "corpo": "...",
        "anexos": [{"nome_arquivo": "x.pdf", "conteudo_base64": "..."}],
        "dry_run": true
      }
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401

    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    data = request.get_json(force=True, silent=True) or {}

    message_id = data.get('message_id')
    if not message_id:
        return jsonify({'status': 'erro', 'erro': 'message_id é obrigatório'}), 400

    assunto   = data.get('assunto', '') or ''
    remetente = data.get('remetente', '') or ''
    corpo     = data.get('corpo', '') or ''
    anexos    = data.get('anexos', []) or []
    dry_run   = str(data.get('dry_run', False)).strip().lower() in ('1', 'true', 'yes', 'sim')

    logger.info(
        f'[EMAIL] message_id={message_id} | assunto={assunto!r} | '
        f'remetente={remetente!r} | anexos={len(anexos)} | dry_run={dry_run}'
    )

    resultado = {
        'status': 'ok',
        'dry_run': dry_run,
        'message_id': message_id,
        'anexos_processados': [],
    }

    # 1. Checar Message-ID duplicado — se já existe, não processa nada
    existente = _buscar_por_campo(TABLE_EMAILS, 'MESSAGE ID', message_id)
    if existente:
        resultado['email_savian'] = {
            'acao': 'duplicado_message_id', 'record_id': existente['id'], 'gravado': False,
        }
        logger.info(f'[EMAIL] Message-ID já existe em Emails Savian: {existente["id"]} → ignorando')
        return jsonify(resultado)

    email_savian_id = None
    if dry_run:
        resultado['email_savian'] = {'acao': 'criaria_novo', 'gravado': False}
    else:
        email_savian_id = _criar_registro(TABLE_EMAILS, {
            F_EMAIL_NAME:     f'{assunto or "(sem assunto)"} — {message_id[:40]}',
            F_EMAIL_STATUS:   'Recebido',
            F_EMAIL_ASSUNTO:  assunto,
            F_EMAIL_CONTEUDO: corpo,
            F_EMAIL_MSGID:    message_id,
        })
        resultado['email_savian'] = {'acao': 'criado', 'gravado': True, 'record_id': email_savian_id}
        logger.info(f'[EMAIL] Emails Savian criado: {email_savian_id}')

    # 2. Processar cada anexo
    for anexo in anexos:
        nome_arquivo = anexo.get('nome_arquivo', 'arquivo.pdf')
        b64 = anexo.get('conteudo_base64', '')
        item = {'nome_arquivo': nome_arquivo}

        try:
            conteudo = base64.b64decode(b64)
        except Exception as exc:
            item['erro'] = f'base64 inválido: {exc}'
            resultado['anexos_processados'].append(item)
            continue

        hash_anexo = hashlib.sha256(conteudo).hexdigest()
        item['hash'] = hash_anexo
        item['tamanho_bytes'] = len(conteudo)

        # Checar duplicidade pelo hash do anexo
        dup = _buscar_por_campo(TABLE_ARQUIVOS, 'Hash do Anexo', hash_anexo)
        if dup:
            item['duplicado'] = True
            item['arquivo_record_id'] = dup['id']
            resultado['anexos_processados'].append(item)
            logger.info(f'[EMAIL] Anexo {nome_arquivo} duplicado (hash já existe): {dup["id"]}')
            continue
        item['duplicado'] = False

        # Extrair texto (só PDFs) e classificar
        texto = ''
        if nome_arquivo.lower().endswith('.pdf'):
            tmp_path = os.path.join(tempfile.gettempdir(), f'email_{_uuid.uuid4().hex}.pdf')
            try:
                with open(tmp_path, 'wb') as f:
                    f.write(conteudo)
                with pdfplumber.open(tmp_path) as pdf_doc:
                    for pg in pdf_doc.pages:
                        texto += (pg.extract_text() or '') + '\n'
            except Exception as exc:
                item['erro_extracao'] = str(exc)
            finally:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)

        if not texto.strip():
            tipo_doc, confianca = 'Não Identificado', 0
            tipo_problema = 'PDF ilegível'
        else:
            tipo_doc, confianca = classificar_documento(texto)
            tipo_problema = 'Documento não reconhecido' if tipo_doc == 'Outro' else None

        item['tipo_documento'] = tipo_doc
        item['confianca'] = confianca
        item['trecho_texto'] = texto[:500]
        item['pendencia'] = tipo_problema

        eh_fiscal = (REMETENTE_FISCAL.lower() in remetente.lower()) or (tipo_doc in TIPOS_FISCAIS)

        if dry_run:
            if eh_fiscal:
                item['acao'] = f'roteamento_fiscal({tipo_doc})'
            else:
                item['acao'] = 'criaria_arquivo_e_processar' + (f'_e_pendencia({tipo_problema})' if tipo_problema else '')
            resultado['anexos_processados'].append(item)
            continue

        # Caixa fiscal (v2.96) — não usa o fluxo de Processar Arquivos do DP
        # (Kit Admissão etc.); vai direto para Extrato/FGTS/Guias. Aplica-se
        # tanto ao remetente fiscal dedicado quanto a qualquer documento já
        # classificado como um tipo fiscal, venha de onde vier (ver nota em
        # TIPOS_FISCAIS acima).
        if eh_fiscal:
            item.update(_processar_anexo_fiscal(conteudo, nome_arquivo, tipo_doc, texto))
            resultado['anexos_processados'].append(item)
            continue

        # Criar registro em Arquivos + anexar PDF
        arquivo_fields = {
            F_ARQ_NOME:     nome_arquivo,
            F_ARQ_NOME_ARQ: nome_arquivo,
            F_ARQ_HASH:     hash_anexo,
            F_ARQ_STATUS:   'Recebido',
            F_ARQ_DATA:     datetime.now().isoformat(),
        }
        if email_savian_id:
            arquivo_fields[F_ARQ_EMAILS] = [email_savian_id]
        arquivo_id = _criar_registro(TABLE_ARQUIVOS, arquivo_fields)
        _anexar_attachment(TABLE_ARQUIVOS, arquivo_id, F_ARQ_ATTACH, conteudo, nome_arquivo)
        item['arquivo_record_id'] = arquivo_id
        logger.info(f'[EMAIL] Arquivo criado: {arquivo_id} ({nome_arquivo})')

        # Criar registro em Processar Arquivos
        processar_id = _criar_registro(TABLE_PROCESSAR, {
            F_PROC_NAME:      nome_arquivo,
            F_PROC_STATUS:    'Pendente',
            F_PROC_TIPO_DOC:  tipo_doc,
            F_PROC_DATA:      datetime.now().isoformat(),
            F_PROC_ARQUIVOS2: [arquivo_id],
        })
        item['processar_record_id'] = processar_id
        logger.info(f'[EMAIL] Processar Arquivos criado: {processar_id} | tipo={tipo_doc}')

        # Criar pendência, se necessário
        if tipo_problema:
            pendencia_id = _criar_registro(TABLE_PENDENCIAS, {
                F_PEND_NOME:   f'{tipo_problema}: {nome_arquivo}',
                F_PEND_STATUS: 'Pendente',
                F_PEND_TIPO:   tipo_problema,
                F_PEND_ORIGEM: [arquivo_id],
                F_PEND_OBS:    texto[:500] if texto.strip() else '(sem texto extraído do PDF)',
                F_PEND_DATA:   datetime.now().isoformat(),
            })
            item['pendencia_record_id'] = pendencia_id
            logger.info(f'[EMAIL] Pendência criada: {pendencia_id} ({tipo_problema})')

        resultado['anexos_processados'].append(item)

    return jsonify(resultado)


@app.route('/email/alerta-formato-invalido', methods=['POST', 'OPTIONS'])
def email_alerta_formato_invalido():
    """
    Dispara automaticamente um e-mail de resposta avisando que o(s) anexo(s)
    recebido(s) estão em formato não suportado (.RAR, .7z etc.) e pedindo
    reenvio em PDF solto ou .ZIP.

    Chamado pelo Google Apps Script (Magnata Email Listener) sempre que ele
    detecta, num e-mail de remetente monitorado, anexo(s) compactado(s) que
    não são PDF/ZIP. O envio em si usa a mesma infraestrutura de e-mail já
    configurada no Render (Resend API / SMTP — _smtp_enviar_email), e não o
    GmailApp do Apps Script, para manter toda a lógica de envio centralizada
    no app.py.

    Protegido por X-API-KEY (mesma EMAIL_WEBHOOK_KEY do /email/webhook).

    Body JSON:
      {
        "remetente": "dpessoal.contabilidade1@hotmail.com",  [obrigatório]
        "assunto_original": "MAGNATA - RESCISÕES",            [opcional]
        "colaboradores": ["Antonio Marcelino Valerio", ...],  [opcional]
        "arquivos_invalidos": ["01 - VITOR....rar", ...],     [opcional]
        "dry_run": false
      }
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401

    data = request.get_json(force=True, silent=True) or {}

    remetente = (data.get('remetente') or '').strip()
    if not remetente:
        return jsonify({'status': 'erro', 'erro': 'remetente é obrigatório'}), 400

    assunto_original = data.get('assunto_original', '') or ''
    colaboradores = data.get('colaboradores') or []
    arquivos_invalidos = data.get('arquivos_invalidos') or []
    dry_run = str(data.get('dry_run', False)).strip().lower() in ('1', 'true', 'yes', 'sim')

    assunto = 'Magnata - Documentos em formato não aceito (.RAR) - reenvio necessário'

    linhas_colab = '\n'.join(f'  - {nome}' for nome in colaboradores) or '  (lista de colaboradores não identificada)'
    linhas_arq = '\n'.join(f'  - {fn}' for fn in arquivos_invalidos) or '  (nomes de arquivo não identificados)'

    corpo = f"""Boa tarde, Jacqueline!

Nosso sistema de recebimento automático de documentos identificou que o e-mail{f' "{assunto_original}"' if assunto_original else ''} veio com anexo(s) em formato compactado (.RAR), que não conseguimos abrir automaticamente.

Poderia, por favor, reenviar o(s) documento(s) abaixo em PDF solto ou em arquivo .ZIP?

Colaborador(es) pendente(s):
{linhas_colab}

Arquivo(s) recebido(s) no formato não aceito:
{linhas_arq}

Pedimos que, se possível, os próximos envios já sigam em PDF ou .ZIP, para que o processamento continue automático.

Qualquer dúvida, estamos à disposição. Muito obrigado!

Att,
Magnata Portaria e Serviços
(mensagem enviada automaticamente pelo sistema de processamento de documentos)
"""

    logger.info(
        f'[ALERTA-FORMATO] remetente={remetente!r} | colaboradores={colaboradores} | '
        f'arquivos={arquivos_invalidos} | dry_run={dry_run}'
    )

    if dry_run:
        return jsonify({
            'status': 'ok',
            'dry_run': True,
            'acao': 'enviaria_alerta',
            'remetente': remetente,
            'assunto': assunto,
            'corpo': corpo,
        })

    try:
        _smtp_enviar_email(remetente, assunto, corpo, [])
    except Exception as exc:
        logger.exception('[ALERTA-FORMATO] Erro ao enviar e-mail de alerta')
        return jsonify({'status': 'erro', 'erro': str(exc)}), 500

    return jsonify({
        'status': 'ok',
        'dry_run': False,
        'acao': 'alerta_enviado',
        'remetente': remetente,
        'assunto': assunto,
    })


@app.route('/processar-fila', methods=['POST', 'OPTIONS'])
def processar_fila():
    """
    Fase 4 — Processador genérico da fila "Processar Arquivos".

    Estrutura preparada para todos os tipos de TIPO_DOC_REGRAS + 'Outro', mas
    nesta implementação apenas 'Holerite' tem handler (demais retornam 400).

    Body JSON:
      {
        "tipo_documento": "Holerite",   [obrigatório]
        "dry_run": true,
        "limit": 1,
        "record_id": "rec...",          [opcional — processa só este registro]
        "folha_mensal": "Fevereiro 2026",  [opcional — senão usa mês anterior]
        "mes_cont_id": "rec..."         [opcional]
      }
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401

    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    data = request.get_json(force=True, silent=True) or {}

    tipo_documento = data.get('tipo_documento')
    if not tipo_documento:
        return jsonify({'status': 'erro', 'erro': 'tipo_documento é obrigatório'}), 400

    if tipo_documento not in PROCESSADORES_DOCUMENTO:
        return jsonify({
            'status': 'erro',
            'erro': f'tipo_documento desconhecido: {tipo_documento}',
            'tipos_validos': list(PROCESSADORES_DOCUMENTO.keys()),
        }), 400

    handler = PROCESSADORES_DOCUMENTO[tipo_documento]
    if handler is None:
        return jsonify({
            'status': 'erro',
            'erro': f'Processamento de "{tipo_documento}" ainda não implementado',
        }), 400

    dry_run = str(data.get('dry_run', False)).strip().lower() in ('1', 'true', 'yes', 'sim')
    limit = data.get('limit', 1)
    try:
        limit = max(1, int(limit))
    except (TypeError, ValueError):
        limit = 1
    record_id = data.get('record_id')

    # Aprovação por exceção (v2.66): a trava antiga que exigia record_id+limit=1
    # para Contrato de Experiência/Trabalho foi removida — agora a fila de
    # contratos pode ser processada em lote. Cada admissão cria o cadastro
    # imediatamente com Status="Validação Pendente" (revisão manual posterior),
    # em vez de reter o arquivo. Cap de segurança para evitar lote gigante
    # acidental num único request.
    if tipo_documento in TIPOS_CONTRATO_5C and not dry_run and not record_id:
        limit = min(limit, 25)

    # Resolver folha_mensal / mes_cont_id / data_holerite
    folha_mensal = data.get('folha_mensal')
    mes_cont_id = data.get('mes_cont_id')
    if not folha_mensal:
        nome_mes, ano, mes_num = mes_anterior_info()
        folha_mensal = f'{nome_mes} {ano}'
        data_holerite = f'{ano}-{mes_num:02d}-01'
    else:
        data_holerite = datetime.now().strftime('%Y-%m-%d')

    if not mes_cont_id:
        mes_cont_id = _buscar_contabilidade_mensal_por_nome(folha_mensal)
        if not mes_cont_id:
            mes_cont_id, _ = buscar_mes_contabilidade_atual()

    logger.info(
        f'[FILA] tipo_documento={tipo_documento} | dry_run={dry_run} | limit={limit} | '
        f'record_id={record_id} | folha_mensal={folha_mensal!r} | mes_cont_id={mes_cont_id}'
    )

    # Montar filtro
    if record_id:
        formula = f'RECORD_ID()="{record_id}"'
    else:
        formula = (
            f'AND('
            f'OR({{Status}}="Processando", {{Status}}="Pendente"), '
            f'{{Tipo de Documento}}="{tipo_documento}"'
            f')'
        )

    _at_throttle()
    r = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_PROCESSAR}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        params={
            'filterByFormula': formula,
            'maxRecords': limit,
            'sort[0][field]': 'Data Processo',
            'sort[0][direction]': 'asc',
            'returnFieldsByFieldId': 'true',
        },
        timeout=30,
    )
    r.raise_for_status()
    registros = r.json().get('records', [])

    resultado = {
        'status': 'ok',
        'dry_run': dry_run,
        'tipo_documento': tipo_documento,
        'folha_mensal': folha_mensal,
        'mes_cont_id': mes_cont_id,
        'registros_encontrados': len(registros),
        'processados': [],
    }

    for rec in registros:
        proc_id = rec['id']
        fields = rec.get('fields', {})
        item = {'processar_id': proc_id}

        tmp_path = None
        try:
            # Localizar Arquivo vinculado
            arquivos_link = fields.get(F_PROC_ARQUIVOS2) or []
            if not arquivos_link:
                item['erro'] = 'Nenhum Arquivo vinculado em "Arquivos 2"'
                if not dry_run:
                    _atualizar_status_processar(proc_id, 'Erro')
                    _criar_pendencia(proc_id, 'Arquivo não vinculado', item['erro'])
                item['status_final'] = 'Erro'
                resultado['processados'].append(item)
                logger.error(f'[FILA] {proc_id}: {item["erro"]}')
                continue

            primeiro_link = arquivos_link[0]
            arquivo_id = primeiro_link['id'] if isinstance(primeiro_link, dict) else primeiro_link
            item['arquivo_id'] = arquivo_id

            _at_throttle()
            r_arq = requests.get(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ARQUIVOS}/{arquivo_id}',
                headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
                params={'returnFieldsByFieldId': 'true'},
                timeout=30,
            )
            r_arq.raise_for_status()
            arquivo_fields = r_arq.json().get('fields', {})

            attachments = arquivo_fields.get(F_ARQ_ATTACH) or []
            if not attachments:
                item['erro'] = 'Arquivo sem anexo PDF'
                if not dry_run:
                    _atualizar_status_processar(proc_id, 'Erro')
                    _criar_pendencia(arquivo_id, 'Arquivo sem anexo', item['erro'])
                item['status_final'] = 'Erro'
                resultado['processados'].append(item)
                logger.error(f'[FILA] {proc_id}: {item["erro"]}')
                continue

            nome_arquivo = attachments[0].get('filename', 'arquivo.pdf')
            pdf_url = attachments[0]['url']
            pdf_bytes = requests.get(pdf_url, timeout=60).content

            # Extrair texto do PDF
            tmp_path = os.path.join(tempfile.gettempdir(), f'fila_{_uuid.uuid4().hex}.pdf')
            with open(tmp_path, 'wb') as f:
                f.write(pdf_bytes)
            texto = ''
            with pdfplumber.open(tmp_path) as pdf_doc:
                for pg in pdf_doc.pages:
                    texto += (pg.extract_text() or '') + '\n'

            status_atual = fields.get(F_PROC_STATUS, 'Processando')

            # Vínculo determinístico com a mensagem de origem (Macro 6A —
            # correção de auditoria do agrupamento do Kit de Admissão):
            # já lido em arquivo_fields acima, sem chamada extra à API.
            email_links = arquivo_fields.get(F_ARQ_EMAILS) or []
            email_savian_id = (
                email_links[0]['id'] if email_links and isinstance(email_links[0], dict)
                else (email_links[0] if email_links else None)
            )

            ctx = {
                'proc_id': proc_id,
                'arquivo_id': arquivo_id,
                'pdf_bytes': pdf_bytes,
                'pdf_hash': hashlib.sha256(pdf_bytes).hexdigest(),
                'nome_arquivo': nome_arquivo,
                'texto': texto,
                'folha_mensal': folha_mensal,
                'mes_cont_id': mes_cont_id,
                'data_holerite': data_holerite,
                'tipo_documento': tipo_documento,
                'status_atual': status_atual,
                'email_savian_id': email_savian_id,
            }

            resultado_handler = handler(ctx, dry_run)
            item['acao'] = resultado_handler['acao']
            item['status_final'] = resultado_handler['status_final']
            item['detalhes'] = resultado_handler['detalhes']

            if not dry_run:
                _atualizar_status_processar(proc_id, resultado_handler['status_final'])
                pendencia = resultado_handler.get('pendencia')
                if pendencia:
                    pend_id = _criar_pendencia(arquivo_id, pendencia['tipo'], pendencia['observacao'])
                    item['pendencia_id'] = pend_id

            logger.info(f'[FILA] {proc_id}: acao={item["acao"]} status_final={item["status_final"]}')

        except Exception as exc:
            logger.exception(f'[FILA] Erro ao processar {proc_id}')
            item['erro'] = str(exc)
            item['status_final'] = 'Erro'
            if not dry_run:
                _atualizar_status_processar(proc_id, 'Erro')
                _criar_pendencia(proc_id, 'Erro de processamento', str(exc))
        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.unlink(tmp_path)

        resultado['processados'].append(item)

    return jsonify(resultado)


# ── Diagnóstico de Distribuição (read-only) ──────────────────────────────────

def _at_listar_todos(table_id, field_names=None, filter_formula=None):
    """Lista TODOS os registros de uma tabela (paginação via offset). Somente leitura."""
    registros = []
    params = {}
    if field_names:
        params['fields[]'] = field_names
    if filter_formula:
        params['filterByFormula'] = filter_formula
    offset = None
    while True:
        if offset:
            params['offset'] = offset
        _at_throttle()
        r = requests.get(
            f'https://api.airtable.com/v0/{BASE_ID}/{table_id}',
            headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
            params=params,
            timeout=30,
        )
        r.raise_for_status()
        data = r.json()
        registros.extend(data.get('records', []))
        offset = data.get('offset')
        if not offset:
            break
    return registros


def _at_obter_registro(table_id: str, record_id: str, field_names=None) -> dict:
    """Lê um único registro do Airtable por ID."""
    params = {}
    if field_names:
        params['fields[]'] = field_names
    _at_throttle()
    r = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{table_id}/{record_id}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        params=params,
        timeout=15,
    )
    r.raise_for_status()
    return r.json()


def _carregar_contexto_distribuicao():
    """Carrega as tabelas de apoio (Clientes, Locais, Funcionários) usadas
    para classificar Holerites na Distribuição Mensal de Documentos."""
    clientes = {
        rec['id']: {
            'nome': rec['fields'].get('Nome', ''),
            'email': rec['fields'].get('Email'),
        }
        for rec in _at_listar_todos(TABLE_CLIENTES, ['Nome', 'Email'])
    }

    locais = {
        rec['id']: {
            'nome': rec['fields'].get('Nome', ''),
            'cliente_ids': rec['fields'].get('Cliente', []),
        }
        for rec in _at_listar_todos(TABLE_LOCAIS, ['Nome', 'Cliente'])
    }

    funcionarios = {
        rec['id']: {
            'nome': rec['fields'].get('Nome Completo', ''),
            'whatsapp': rec['fields'].get('WhatsApp'),
            'locais_ids': rec['fields'].get('Locais de trabalho', []),
            'pdf_folha_ponto': rec['fields'].get('PDF Folha Ponto', []),
            'recibos_pagamento': rec['fields'].get('Recibos Pagamento', []),
            'resumo_ponto': rec['fields'].get('Resumo Cartão Ponto (extraído)', ''),
        }
        for rec in _at_listar_todos(TABLE_FUNC, [
            'Nome Completo', 'WhatsApp', 'Locais de trabalho',
            'PDF Folha Ponto', 'Recibos Pagamento', 'Resumo Cartão Ponto (extraído)',
        ])
    }
    return clientes, locais, funcionarios


def _classificar_holerite_distribuicao(rec, clientes, locais, funcionarios):
    """Classifica 1 registro de Holerites para a Distribuição Mensal de
    Documentos. Retorna um dict com func/cliente identificados, motivos de
    pendência e a 'situacao' final ('pronto_whatsapp_colaborador',
    'pronto_pacote_cliente', 'pronto_ambos' ou 'pendente')."""
    f = rec['fields']
    func_ids = f.get('Funcionário', []) or []
    func = funcionarios.get(func_ids[0]) if func_ids else None

    whatsapp = func.get('whatsapp') if func else None
    arquivo_ok = bool(f.get('PDF HOLERITE'))
    folha_ok = bool(f.get('Folha Mensal'))

    cliente_ids = set()
    if func:
        for local_id in func.get('locais_ids', []) or []:
            local = locais.get(local_id)
            if local:
                cliente_ids.update(local.get('cliente_ids', []) or [])

    cliente_id = next(iter(cliente_ids)) if cliente_ids else None
    cliente_email = None
    cliente_nome = None
    if cliente_id:
        primeiro = clientes.get(cliente_id)
        if primeiro:
            cliente_nome = primeiro.get('nome')
            cliente_email = primeiro.get('email')

    motivos = []
    if not func_ids:
        motivos.append('funcionario_nao_identificado')
    if not whatsapp:
        motivos.append('whatsapp_ausente')
    if not arquivo_ok:
        motivos.append('arquivo_holerite_ausente')
    if not folha_ok:
        motivos.append('folha_mensal_ausente')
    if not cliente_ids:
        motivos.append('cliente_local_nao_identificado')
    elif not cliente_email:
        motivos.append('email_cliente_ausente')

    ok_whatsapp = bool(func_ids and whatsapp and arquivo_ok and folha_ok)
    ok_cliente = bool(cliente_ids and cliente_email and arquivo_ok and folha_ok)

    if ok_whatsapp and ok_cliente:
        situacao = 'pronto_ambos'
    elif ok_whatsapp:
        situacao = 'pronto_whatsapp_colaborador'
    elif ok_cliente:
        situacao = 'pronto_pacote_cliente'
    else:
        situacao = 'pendente'

    return {
        'func_ids': func_ids,
        'func': func,
        'whatsapp': whatsapp,
        'arquivo_ok': arquivo_ok,
        'folha_ok': folha_ok,
        'cliente_id': cliente_id,
        'cliente_nome': cliente_nome,
        'cliente_email': cliente_email,
        'motivos': motivos,
        'ok_whatsapp': ok_whatsapp,
        'ok_cliente': ok_cliente,
        'situacao': situacao,
    }


def _classificar_folha_ponto_distribuicao(dados_func, clientes, locais):
    """Classifica 1 Funcionário quanto ao envio do Cartão Ponto/Folha de Ponto
    mensal — espelha _classificar_holerite_distribuicao, mas a partir do
    prontuário em Funcionários (não há 1 registro por mês como em Holerites).
    Retorna dict com pdf_anexo (último anexo de "PDF Folha Ponto"), whatsapp,
    cliente identificado, motivos de pendência e 'situacao'."""
    anexos = dados_func.get('pdf_folha_ponto') or []
    pdf_anexo = anexos[-1] if anexos else None
    whatsapp = dados_func.get('whatsapp')

    cliente_ids = set()
    for local_id in dados_func.get('locais_ids', []) or []:
        local = locais.get(local_id)
        if local:
            cliente_ids.update(local.get('cliente_ids', []) or [])

    cliente_id = next(iter(cliente_ids)) if cliente_ids else None
    cliente_email = None
    cliente_nome = None
    if cliente_id:
        cliente = clientes.get(cliente_id)
        if cliente:
            cliente_nome = cliente.get('nome')
            cliente_email = cliente.get('email')

    motivos = []
    if not pdf_anexo:
        motivos.append('pdf_folha_ponto_ausente')
    if not whatsapp:
        motivos.append('whatsapp_ausente')
    if not cliente_ids:
        motivos.append('cliente_local_nao_identificado')
    elif not cliente_email:
        motivos.append('email_cliente_ausente')

    ok_whatsapp = bool(whatsapp and pdf_anexo)
    ok_cliente = bool(cliente_email and pdf_anexo)

    if ok_whatsapp and ok_cliente:
        situacao = 'pronto_ambos'
    elif ok_whatsapp:
        situacao = 'pronto_whatsapp_colaborador'
    elif ok_cliente:
        situacao = 'pronto_pacote_cliente'
    else:
        situacao = 'pendente'

    return {
        'pdf_anexo': pdf_anexo,
        'whatsapp': whatsapp,
        'cliente_id': cliente_id,
        'cliente_nome': cliente_nome,
        'cliente_email': cliente_email,
        'motivos': motivos,
        'ok_whatsapp': ok_whatsapp,
        'ok_cliente': ok_cliente,
        'situacao': situacao,
    }


def _carregar_envios_pendentes(tipo):
    """Retorna (func_ids_pendentes, cliente_ids_pendentes): conjuntos de
    Funcionários/Clientes que já possuem um envio do `tipo` informado com
    Status em ('Preparando', 'Enviado') — usado para não duplicar a fila em
    execuções repetidas no mesmo ciclo."""
    func_ids_pendentes = set()
    cliente_ids_pendentes = set()
    campos = ['Tipo', 'Status', 'Canal', 'Funcionário(s) Vinculado(s)', 'Cliente']
    for rec in _at_listar_todos(TABLE_ENVIOS, campos):
        f = rec['fields']
        if f.get('Tipo') != tipo or f.get('Status') not in ('Preparando', 'Enviado'):
            continue
        func_ids_pendentes.update(f.get('Funcionário(s) Vinculado(s)', []) or [])
        if f.get('Canal') == 'E-mail':
            cliente_ids_pendentes.update(f.get('Cliente', []) or [])
    return func_ids_pendentes, cliente_ids_pendentes


def _diagnostico_holerites(folha_mensal=None, limit=None):
    """
    Diagnóstico read-only: classifica cada Holerite existente em
    'pronto_whatsapp_colaborador', 'pronto_pacote_cliente', 'pronto_ambos'
    ou 'pendente', conforme a diretriz de Distribuição Mensal de Documentos.

    Nenhuma escrita é feita — apenas leitura em Holerites, Funcionários,
    Locais e Clientes.
    """
    clientes, locais, funcionarios = _carregar_contexto_distribuicao()

    filtro = f'{{Folha Mensal}}="{folha_mensal}"' if folha_mensal else None
    holerite_fields = [
        'Holerite', 'Status', 'Funcionário', 'Data', 'Folha Mensal',
        'PDF HOLERITE', 'Envios de Documentos',
    ]
    holerites = _at_listar_todos(TABLE_HOL, holerite_fields, filtro)
    if limit:
        holerites = holerites[:limit]

    contagem = {
        'pronto_whatsapp_colaborador': 0,
        'pronto_pacote_cliente': 0,
        'pronto_ambos': 0,
        'pendente': 0,
    }
    motivos_pendencia = Counter()
    exemplos_prontos = []
    exemplos_pendentes = []
    com_envio_vinculado = 0
    sem_envio_vinculado = 0

    for rec in holerites:
        f = rec['fields']
        c = _classificar_holerite_distribuicao(rec, clientes, locais, funcionarios)
        situacao = c['situacao']
        if situacao == 'pendente':
            motivos_pendencia.update(c['motivos'])
        contagem[situacao] += 1

        envios = f.get('Envios de Documentos', []) or []
        if envios:
            com_envio_vinculado += 1
        else:
            sem_envio_vinculado += 1

        exemplo = {
            'record_id': rec['id'],
            'holerite': f.get('Holerite'),
            'funcionario': c['func'].get('nome') if c['func'] else None,
            'whatsapp': c['whatsapp'],
            'cliente': c['cliente_nome'],
            'email_cliente': c['cliente_email'],
            'folha_mensal': f.get('Folha Mensal'),
            'situacao': situacao,
            'motivos': c['motivos'] or None,
        }
        if situacao == 'pendente':
            if len(exemplos_pendentes) < 5:
                exemplos_pendentes.append(exemplo)
        else:
            if len(exemplos_prontos) < 5:
                exemplos_prontos.append(exemplo)

    return {
        'total_holerites_analisados': len(holerites),
        'contagem': contagem,
        'principais_motivos_pendencia': motivos_pendencia.most_common(),
        'exemplos_prontos': exemplos_prontos,
        'exemplos_pendentes': exemplos_pendentes,
        'campos_tabelas_usados': {
            'Holerites': holerite_fields,
            'Funcionários': ['Nome Completo', 'WhatsApp', 'Locais de trabalho'],
            'Locais': ['Nome', 'Cliente'],
            'Clientes': ['Nome', 'Email'],
        },
        'campos_faltando_para_destravar': [
            'WhatsApp ausente em Funcionários' if motivos_pendencia.get('whatsapp_ausente') else None,
            'Cliente/Local não identificado a partir de Funcionários->Locais de trabalho->Cliente'
                if motivos_pendencia.get('cliente_local_nao_identificado') else None,
            'Email ausente em Clientes' if motivos_pendencia.get('email_cliente_ausente') else None,
            'Holerite sem Funcionário vinculado' if motivos_pendencia.get('funcionario_nao_identificado') else None,
            'Holerite sem PDF anexado' if motivos_pendencia.get('arquivo_holerite_ausente') else None,
            'Holerite sem Folha Mensal preenchida' if motivos_pendencia.get('folha_mensal_ausente') else None,
        ],
        'investigacao_tela_enviar_holerites': {
            'tabela_provavel': (
                f'Envios de Documentos ({TABLE_ENVIOS}) — possui campos Status, Tipo, '
                'Canal, Enviar (checkbox), Holerites (link), Email, Destinatário, Erro, '
                'ID da Mensagem, Tentativa, consistentes com uma fila de despacho.'
            ),
            'holerites_com_envio_vinculado': com_envio_vinculado,
            'holerites_sem_envio_vinculado': sem_envio_vinculado,
            'hipotese': (
                'A tela "Enviar Holerites" provavelmente lista registros da tabela '
                '"Envios de Documentos" (filtrados por Tipo/Status), não os Holerites '
                'diretamente. Holerites já existentes que nunca tiveram um registro '
                'correspondente criado em "Envios de Documentos" não apareceriam na '
                'tela, mesmo estando prontos.'
            ),
            'limitacoes': (
                'Esta API não tem acesso a automações, botões, views/filtros do '
                'Airtable nem a configurações de Make/Apps Script/Gmail — a hipótese '
                'acima é inferida apenas a partir do schema de dados (somente leitura). '
                'Confirmação definitiva requer inspeção manual da view/automação na '
                'interface do Airtable.'
            ),
        },
        'observacao': (
            'Diagnóstico 100% somente leitura. Nenhum WhatsApp ou e-mail foi enviado, '
            'nenhum registro foi criado/alterado em Holerites, Funcionários, Clientes, '
            'Locais, Envios de Documentos ou Pendências/Revisar.'
        ),
    }


@app.route('/diagnostico-holerites', methods=['GET', 'OPTIONS'])
def diagnostico_holerites():
    """
    Diagnóstico read-only dos Holerites existentes (Distribuição Mensal de
    Documentos — menor próximo passo seguro). Não envia WhatsApp/e-mail, não
    altera nada no Airtable.

    Query params opcionais:
      - folha_mensal: filtra Holerites por "Folha Mensal" (ex.: "Maio 2026")
      - limit: limita a quantidade de Holerites analisados
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401

    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    folha_mensal = request.args.get('folha_mensal')
    limit = request.args.get('limit')
    try:
        limit = int(limit) if limit else None
    except ValueError:
        limit = None

    resultado = _diagnostico_holerites(folha_mensal=folha_mensal, limit=limit)
    return jsonify(resultado)


# ── Fila de Envios + Recibo Digital de Leitura (Fase 3) ──────────────────────

def _gerar_hash_recibo() -> str:
    """Token único (URL-safe) para o link de Recibo Digital de Leitura."""
    return secrets.token_urlsafe(24)


def _criar_envio_documento(campos: dict):
    """Cria um registro em Envios de Documentos com um Hash Recibo único.
    Retorna (envio_id, hash_recibo)."""
    campos = dict(campos)
    hash_recibo = _gerar_hash_recibo()
    campos[F_ENVIO_HASH] = hash_recibo
    campos.setdefault(F_ENVIO_STATUS, 'Preparando')
    envio_id = _criar_registro(TABLE_ENVIOS, campos)
    return envio_id, hash_recibo


def _gerar_fila_envios(folha_mensal=None, limit=None, dry_run=True):
    """
    Fase 3 — Fila de Envios (Distribuição Mensal de Documentos).

    Para cada Holerite "pronto" (ver _classificar_holerite_distribuicao) que
    ainda não possui nenhum registro vinculado em "Envios de Documentos",
    cria:
      - 1 envio Canal="WhatsApp" para o colaborador (Destinatário=WhatsApp)
      - 1 envio Canal="E-mail" para o cliente (Destinatário=Email do Cliente)

    Cada envio criado recebe um "Hash Recibo" único. A mensagem enviada pela
    automação externa (WhatsApp/E-mail) deve incluir o link /recibo/<hash> —
    ao ser clicado, o link marca o envio como "Lido" (Recibo Digital de
    Leitura) e redireciona para o PDF do Holerite.

    dry_run=True (padrão): não cria nada no Airtable, apenas retorna o que
    seria criado.
    """
    clientes, locais, funcionarios = _carregar_contexto_distribuicao()

    filtro = f'{{Folha Mensal}}="{folha_mensal}"' if folha_mensal else None
    holerite_fields = ['Holerite', 'Funcionário', 'Folha Mensal', 'PDF HOLERITE', 'Envios de Documentos']
    holerites = _at_listar_todos(TABLE_HOL, holerite_fields, filtro)
    if limit:
        holerites = holerites[:limit]

    envios = []
    ignorados = []

    for rec in holerites:
        f = rec['fields']
        if f.get('Envios de Documentos'):
            ignorados.append({
                'record_id': rec['id'], 'holerite': f.get('Holerite'),
                'motivo': 'ja_possui_envio_vinculado',
            })
            continue

        c = _classificar_holerite_distribuicao(rec, clientes, locais, funcionarios)

        candidatos = []
        if c['ok_whatsapp']:
            candidatos.append({
                'canal': 'WhatsApp', 'destinatario': c['whatsapp'],
                'funcionario_ids': c['func_ids'], 'cliente_ids': [],
            })
        if c['ok_cliente']:
            candidatos.append({
                'canal': 'E-mail', 'destinatario': c['cliente_email'],
                'funcionario_ids': c['func_ids'], 'cliente_ids': [c['cliente_id']],
            })

        if not candidatos:
            ignorados.append({
                'record_id': rec['id'], 'holerite': f.get('Holerite'),
                'motivo': 'nao_pronto', 'motivos': c['motivos'],
            })
            continue

        for candidato in candidatos:
            item = {
                'holerite': f.get('Holerite'),
                'holerite_id': rec['id'],
                'canal': candidato['canal'],
                'destinatario': candidato['destinatario'],
                'funcionario': c['func'].get('nome') if c['func'] else None,
                'cliente': c['cliente_nome'],
            }

            if dry_run:
                item['acao'] = 'criaria_envio'
                envios.append(item)
                continue

            campos = {
                F_ENVIO_TIPO: 'Relatório Mensal',
                F_ENVIO_CANAL: candidato['canal'],
                F_ENVIO_DEST: candidato['destinatario'],
                F_ENVIO_HOLERITES: [rec['id']],
            }
            if candidato['funcionario_ids']:
                campos[F_ENVIO_FUNC] = candidato['funcionario_ids']
            if candidato['cliente_ids']:
                campos[F_ENVIO_CLIENTE] = candidato['cliente_ids']
            if candidato['canal'] == 'E-mail':
                campos[F_ENVIO_EMAIL] = candidato['destinatario']

            try:
                envio_id, hash_recibo = _criar_envio_documento(campos)
                item['acao'] = 'envio_criado'
                item['envio_id'] = envio_id
                item['hash_recibo'] = hash_recibo
                item['link_recibo'] = f'/recibo/{hash_recibo}'
            except AirtableError as exc:
                logger.error(f'[FILA-ENVIOS] falha ao criar envio para holerite {rec["id"]}: {exc}')
                item['acao'] = 'falhou'
                item['erro'] = str(exc)

            envios.append(item)

    return {
        'total_holerites_analisados': len(holerites),
        'envios': envios,
        'holerites_ignorados': ignorados,
        'dry_run': dry_run,
    }


def _gerar_fila_envios_folha_ponto(limit=None, dry_run=True):
    """
    Fase 3 — Fila de Envios de Folha de Ponto/Cartão Ponto, espelhada a
    _gerar_fila_envios. Como a Folha de Ponto não gera 1 registro por mês
    (ao contrário de Holerites), a origem dos dados é o prontuário em
    Funcionários ("PDF Folha Ponto" + "Resumo Cartão Ponto (extraído)"):

      - 1 envio Canal="WhatsApp" por Colaborador, com o Cartão Ponto
        individual (último anexo de "PDF Folha Ponto").
      - 1 envio Canal="E-mail" por Cliente, em lote, com os Cartões Ponto de
        todos os colaboradores prontos daquele cliente.

    Cada envio recebe um "Hash Recibo" único e o(s) PDF(s) são anexados
    diretamente no campo "Arquivos" do envio (sem link para uma tabela de
    Folha de Ponto mensal, que não existe).

    Funcionários/Clientes que já possuem um envio "Cartão Ponto Mensal"
    pendente (Status="Preparando"/"Enviado") são ignorados, para não
    duplicar a fila em execuções repetidas. dry_run=True (padrão): não cria
    nada no Airtable, apenas simula.
    """
    clientes, locais, funcionarios = _carregar_contexto_distribuicao()
    func_pendentes, cliente_pendentes = _carregar_envios_pendentes(TIPO_ENVIO_PONTO)

    func_ids = list(funcionarios.keys())
    if limit:
        func_ids = func_ids[:limit]

    envios = []
    ignorados = []
    lotes_cliente = {}

    for func_id in func_ids:
        dados = funcionarios[func_id]

        if func_id in func_pendentes:
            ignorados.append({
                'funcionario_id': func_id, 'funcionario': dados.get('nome'),
                'motivo': 'ja_possui_envio_ponto_pendente',
            })
            continue

        c = _classificar_folha_ponto_distribuicao(dados, clientes, locais)

        if not c['pdf_anexo']:
            ignorados.append({
                'funcionario_id': func_id, 'funcionario': dados.get('nome'),
                'motivo': 'pdf_folha_ponto_ausente',
            })
            continue

        if c['ok_whatsapp']:
            item = {
                'funcionario': dados.get('nome'), 'funcionario_id': func_id,
                'canal': 'WhatsApp', 'destinatario': c['whatsapp'],
                'arquivo': c['pdf_anexo'].get('filename'),
            }
            if dry_run:
                item['acao'] = 'criaria_envio'
            else:
                campos = {
                    F_ENVIO_TIPO: TIPO_ENVIO_PONTO,
                    F_ENVIO_CANAL: 'WhatsApp',
                    F_ENVIO_DEST: c['whatsapp'],
                    F_ENVIO_FUNC: [func_id],
                    F_ENVIO_ARQUIVOS: [{'url': c['pdf_anexo']['url']}],
                }
                try:
                    envio_id, hash_recibo = _criar_envio_documento(campos)
                    item['acao'] = 'envio_criado'
                    item['envio_id'] = envio_id
                    item['hash_recibo'] = hash_recibo
                    item['link_recibo'] = f'/recibo/{hash_recibo}'
                except AirtableError as exc:
                    logger.error(f'[FILA-ENVIOS-PONTO] falha ao criar envio WhatsApp para {func_id}: {exc}')
                    item['acao'] = 'falhou'
                    item['erro'] = str(exc)
            envios.append(item)

        if c['ok_cliente']:
            lote = lotes_cliente.setdefault(c['cliente_id'], {
                'cliente_nome': c['cliente_nome'], 'cliente_email': c['cliente_email'],
                'func_ids': [], 'funcionarios': [], 'anexos': [],
            })
            lote['func_ids'].append(func_id)
            lote['funcionarios'].append(dados.get('nome'))
            lote['anexos'].append(c['pdf_anexo'])

        if not c['ok_whatsapp'] and not c['ok_cliente']:
            ignorados.append({
                'funcionario_id': func_id, 'funcionario': dados.get('nome'),
                'motivo': 'nao_pronto', 'motivos': c['motivos'],
            })

    for cliente_id, lote in lotes_cliente.items():
        if cliente_id in cliente_pendentes:
            ignorados.append({
                'cliente_id': cliente_id, 'cliente': lote['cliente_nome'],
                'motivo': 'ja_possui_envio_ponto_pendente',
            })
            continue

        item = {
            'cliente': lote['cliente_nome'], 'cliente_id': cliente_id,
            'canal': 'E-mail', 'destinatario': lote['cliente_email'],
            'funcionarios': lote['funcionarios'],
        }
        if dry_run:
            item['acao'] = 'criaria_envio'
        else:
            campos = {
                F_ENVIO_TIPO: TIPO_ENVIO_PONTO,
                F_ENVIO_CANAL: 'E-mail',
                F_ENVIO_DEST: lote['cliente_email'],
                F_ENVIO_EMAIL: lote['cliente_email'],
                F_ENVIO_CLIENTE: [cliente_id],
                F_ENVIO_FUNC: lote['func_ids'],
                F_ENVIO_ARQUIVOS: [{'url': a['url']} for a in lote['anexos']],
            }
            try:
                envio_id, hash_recibo = _criar_envio_documento(campos)
                item['acao'] = 'envio_criado'
                item['envio_id'] = envio_id
                item['hash_recibo'] = hash_recibo
                item['link_recibo'] = f'/recibo/{hash_recibo}'
            except AirtableError as exc:
                logger.error(f'[FILA-ENVIOS-PONTO] falha ao criar envio E-mail para cliente {cliente_id}: {exc}')
                item['acao'] = 'falhou'
                item['erro'] = str(exc)
        envios.append(item)

    return {
        'total_funcionarios_analisados': len(func_ids),
        'envios': envios,
        'ignorados': ignorados,
        'dry_run': dry_run,
    }


@app.route('/gerar-fila-envios', methods=['POST', 'OPTIONS'])
def gerar_fila_envios():
    """
    Fase 3 — Fila de Envios (Distribuição Mensal de Documentos).

    Body JSON opcional:
      - folha_mensal: filtra Holerites por "Folha Mensal" (ex.: "Maio 2026")
      - limit: limita a quantidade de Holerites analisados
      - dry_run: true (padrão) não grava nada, apenas simula
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401

    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    body = request.get_json(silent=True) or {}
    folha_mensal = body.get('folha_mensal')
    limit = body.get('limit')
    try:
        limit = int(limit) if limit else None
    except (ValueError, TypeError):
        limit = None
    dry_run = body.get('dry_run', True)

    resultado = _gerar_fila_envios(folha_mensal=folha_mensal, limit=limit, dry_run=dry_run)
    return jsonify(resultado)


@app.route('/gerar-fila-envios-ponto', methods=['POST', 'OPTIONS'])
def gerar_fila_envios_ponto():
    """
    Fase 3 — Fila de Envios de Folha de Ponto/Cartão Ponto (espelha
    /gerar-fila-envios).

    Body JSON opcional:
      - limit: limita a quantidade de Funcionários analisados
      - dry_run: true (padrão) não grava nada, apenas simula
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401

    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    body = request.get_json(silent=True) or {}
    limit = body.get('limit')
    try:
        limit = int(limit) if limit else None
    except (ValueError, TypeError):
        limit = None
    dry_run = body.get('dry_run', True)

    resultado = _gerar_fila_envios_folha_ponto(limit=limit, dry_run=dry_run)
    return jsonify(resultado)


def _gerar_fila_envios_combinado(folha_mensal=None, limit=None, dry_run=True):
    """
    v2.25 — Fila de Envio COMBINADO (Holerite + Cartão Ponto na MESMA
    mensagem de WhatsApp).

    Para cada Colaborador pronto em AMBOS os documentos, cria 1 único envio
    Canal="WhatsApp" anexando os 2 PDFs no campo "Arquivos", na ordem
    [Holerite, Ponto]. A automação da Evolution API envia os dois anexos em
    sequência numa só mensagem.

    Critério de "pronto": WhatsApp preenchido + Holerite do mês com PDF
    (filtrado por `folha_mensal`) + PDF Folha de Ponto no prontuário. Quem
    estiver pronto em só um dos dois é ignorado (e segue pelos endpoints
    individuais já existentes). Colaboradores com envio combinado pendente
    (Status="Preparando"/"Enviado") são ignorados p/ não duplicar a fila.

    dry_run=True (padrão): não grava nada, apenas simula.
    """
    clientes, locais, funcionarios = _carregar_contexto_distribuicao()
    func_pendentes, _ = _carregar_envios_pendentes(TIPO_ENVIO_COMBINADO)

    # Mapa funcionario_id -> Holerite (com PDF) da folha_mensal informada.
    filtro = f'{{Folha Mensal}}="{folha_mensal}"' if folha_mensal else None
    holerite_fields = ['Holerite', 'Funcionário', 'Folha Mensal', 'PDF HOLERITE']
    holerite_por_func = {}
    for rec in _at_listar_todos(TABLE_HOL, holerite_fields, filtro):
        f = rec['fields']
        anexos = f.get('PDF HOLERITE') or []
        if not anexos:
            continue
        for fid in (f.get('Funcionário', []) or []):
            holerite_por_func[fid] = {
                'holerite_id': rec['id'],
                'holerite': f.get('Holerite'),
                'pdf': anexos[-1],
            }

    func_ids = list(funcionarios.keys())
    if limit:
        func_ids = func_ids[:limit]

    envios = []
    ignorados = []

    for func_id in func_ids:
        dados = funcionarios[func_id]

        if func_id in func_pendentes:
            ignorados.append({
                'funcionario_id': func_id, 'funcionario': dados.get('nome'),
                'motivo': 'ja_possui_envio_combinado_pendente',
            })
            continue

        whatsapp = dados.get('whatsapp')
        anexos_ponto = dados.get('pdf_folha_ponto') or []
        pdf_ponto = anexos_ponto[-1] if anexos_ponto else None
        hol = holerite_por_func.get(func_id)

        motivos = []
        if not whatsapp:
            motivos.append('whatsapp_ausente')
        if not hol:
            motivos.append('holerite_do_mes_ausente')
        if not pdf_ponto:
            motivos.append('pdf_folha_ponto_ausente')

        if motivos:
            ignorados.append({
                'funcionario_id': func_id, 'funcionario': dados.get('nome'),
                'motivo': 'nao_pronto', 'motivos': motivos,
            })
            continue

        item = {
            'funcionario': dados.get('nome'), 'funcionario_id': func_id,
            'canal': 'WhatsApp', 'destinatario': whatsapp,
            'holerite': hol['holerite'],
            'arquivos': [hol['pdf'].get('filename'), pdf_ponto.get('filename')],
        }

        if dry_run:
            item['acao'] = 'criaria_envio'
        else:
            campos = {
                F_ENVIO_TIPO: TIPO_ENVIO_COMBINADO,
                F_ENVIO_CANAL: 'WhatsApp',
                F_ENVIO_DEST: whatsapp,
                F_ENVIO_FUNC: [func_id],
                F_ENVIO_HOLERITES: [hol['holerite_id']],
                # Ordem garantida: Holerite primeiro, Cartão Ponto depois.
                F_ENVIO_ARQUIVOS: [
                    {'url': hol['pdf']['url']},
                    {'url': pdf_ponto['url']},
                ],
            }
            try:
                envio_id, hash_recibo = _criar_envio_documento(campos)
                item['acao'] = 'envio_criado'
                item['envio_id'] = envio_id
                item['hash_recibo'] = hash_recibo
                item['link_recibo'] = f'/recibo/{hash_recibo}'
            except AirtableError as exc:
                logger.error(f'[FILA-COMBINADO] falha ao criar envio p/ {func_id}: {exc}')
                item['acao'] = 'falhou'
                item['erro'] = str(exc)

        envios.append(item)

    return {
        'total_funcionarios_analisados': len(func_ids),
        'envios': envios,
        'ignorados': ignorados,
        'dry_run': dry_run,
    }


@app.route('/gerar-fila-envios-combinado', methods=['POST', 'OPTIONS'])
def gerar_fila_envios_combinado():
    """
    v2.25 — Gera a fila de envio COMBINADO: 1 mensagem WhatsApp por
    colaborador com Holerite + Cartão Ponto anexados em sequência.

    Body JSON opcional:
      - folha_mensal: filtra o Holerite do mês (ex.: "Maio 2026")
      - limit: limita a quantidade de Funcionários analisados
      - dry_run: true (padrão) não grava nada, apenas simula
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401

    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    body = request.get_json(silent=True) or {}
    folha_mensal = body.get('folha_mensal')
    limit = body.get('limit')
    try:
        limit = int(limit) if limit else None
    except (ValueError, TypeError):
        limit = None
    dry_run = body.get('dry_run', True)

    resultado = _gerar_fila_envios_combinado(folha_mensal=folha_mensal, limit=limit, dry_run=dry_run)
    return jsonify(resultado)


def _normalizar_whatsapp_funcionarios(limit=None, dry_run=True):
    """Varre Funcionários e normaliza o campo "WhatsApp" para o formato
    DDI 55 + DDD + número (somente dígitos), exigido pela Evolution API.
    Registros já normalizados ou sem WhatsApp são ignorados."""
    atualizados = []
    ignorados = []

    for rec in _at_listar_todos(TABLE_FUNC, ['Nome Completo', 'WhatsApp']):
        if limit is not None and len(atualizados) >= limit:
            break

        f = rec['fields']
        numero_original = f.get('WhatsApp')
        numero_normalizado = _normalizar_telefone_br(numero_original)

        if not numero_normalizado:
            ignorados.append({
                'funcionario_id': rec['id'],
                'nome': f.get('Nome Completo', ''),
                'whatsapp_atual': numero_original,
                'motivo': 'whatsapp_ausente' if not numero_original else 'ja_normalizado_ou_invalido',
            })
            continue

        item = {
            'funcionario_id': rec['id'],
            'nome': f.get('Nome Completo', ''),
            'whatsapp_antes': numero_original,
            'whatsapp_depois': numero_normalizado,
        }

        if dry_run:
            item['acao'] = 'atualizaria'
        else:
            _at_throttle()
            r = requests.patch(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}/{rec["id"]}',
                headers=_at_headers(),
                json={'fields': {'WhatsApp': numero_normalizado}, 'typecast': True},
                timeout=15,
            )
            _raise_for_airtable(r, f'normalizar WhatsApp {rec["id"]}')
            item['acao'] = 'atualizado'

        atualizados.append(item)

    return {
        'total_funcionarios_atualizados': len(atualizados),
        'atualizados': atualizados,
        'ignorados': ignorados,
        'dry_run': dry_run,
    }


@app.route('/normalizar-whatsapp', methods=['POST', 'OPTIONS'])
def normalizar_whatsapp():
    """
    Normaliza o campo "WhatsApp" de Funcionários para o formato DDI 55 +
    DDD + número (somente dígitos), necessário para a Evolution API.

    Body JSON opcional:
      - limit: limita a quantidade de registros atualizados
      - dry_run: true (padrão) não grava nada, apenas simula
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401

    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    body = request.get_json(silent=True) or {}
    limit = body.get('limit')
    try:
        limit = int(limit) if limit else None
    except (ValueError, TypeError):
        limit = None
    dry_run = body.get('dry_run', True)

    resultado = _normalizar_whatsapp_funcionarios(limit=limit, dry_run=dry_run)
    return jsonify(resultado)


def _normalizar_numero_evolution(dest: str):
    """Garante número no formato da Evolution (DDI 55 + DDD + número, só dígitos)."""
    if not dest:
        return None
    digitos = re.sub(r'\D', '', str(dest))
    if not digitos:
        return None
    if digitos.startswith('0') and len(digitos) in (12, 13):
        digitos = digitos[1:]
    if not digitos.startswith('55') and len(digitos) in (10, 11):
        digitos = '55' + digitos
    return digitos if len(digitos) in (12, 13) else None


def _evolution_enviar_texto(numero: str, texto: str):
    """Envia 1 mensagem de texto puro via Evolution API v2 (sendText) — usado
    para o link de Assinatura Nativa (sem custo extra, mesma instância já
    configurada para o envio de holerites)."""
    endpoint = f'{EVOLUTION_API_URL}/message/sendText/{EVOLUTION_INSTANCE}'
    payload = {'number': numero, 'text': texto}
    r = requests.post(
        endpoint,
        headers={'apikey': EVOLUTION_API_KEY, 'Content-Type': 'application/json'},
        json=payload,
        timeout=90,
    )
    if not (200 <= r.status_code < 300):
        raise RuntimeError(f'Evolution HTTP {r.status_code}: {r.text[:300]}')
    return r.json()


@app.route('/evolution/status', methods=['GET', 'OPTIONS'])
def evolution_status():
    """Consulta o estado real da conexão da instância Evolution API (o celular
    vinculado está online/'open' ou caiu/'close')."""
    if request.method == 'OPTIONS':
        return '', 204
    endpoint = f'{EVOLUTION_API_URL}/instance/connectionState/{EVOLUTION_INSTANCE}'
    try:
        r = requests.get(endpoint, headers={'apikey': EVOLUTION_API_KEY}, timeout=30)
        return jsonify({'http_status': r.status_code, 'body': r.json() if r.content else None})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@app.route('/whatsapp/enviar-texto', methods=['POST', 'OPTIONS'])
def whatsapp_enviar_texto():
    """Envia 1 mensagem de texto customizada via WhatsApp (Evolution API) para
    um número específico — usado para comunicações pontuais (ex.: avisos de
    correção de incidente) que não seguem o fluxo automático de nenhum
    processador. Protegido por X-API-KEY (EMAIL_WEBHOOK_KEY)."""
    if request.method == 'OPTIONS':
        return '', 204
    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401

    body = request.get_json(silent=True) or {}
    numero = _normalizar_numero_evolution(body.get('numero'))
    texto = (body.get('mensagem') or '').strip()
    if not numero:
        return jsonify({'status': 'erro', 'erro': 'Número inválido.'}), 400
    if not texto:
        return jsonify({'status': 'erro', 'erro': 'Mensagem vazia.'}), 400

    try:
        resultado = _evolution_enviar_texto(numero, texto)
        return jsonify({'status': 'ok', 'numero': numero, 'resultado': resultado})
    except Exception as exc:
        return jsonify({'status': 'erro', 'erro': str(exc)}), 502


def _sanitizar_nome_video(nome_arquivo):
    """Retorna um nome MP4 seguro, sem caminho ou caracteres de controle."""
    if nome_arquivo is None:
        return 'video.mp4'
    if not isinstance(nome_arquivo, str):
        return None
    nome = nome_arquivo.strip()
    if not nome or '/' in nome or '\\' in nome or '\x00' in nome:
        return None
    nome = unicodedata.normalize('NFKD', nome).encode('ascii', 'ignore').decode('ascii')
    nome = re.sub(r'[^A-Za-z0-9._-]', '_', nome)
    nome = re.sub(r'_+', '_', nome).strip('._')
    if not nome or len(nome) > 120 or not nome.lower().endswith('.mp4'):
        return None
    return nome


def _decodificar_video_mp4(video_base64):
    """Valida base64 estrito, tamanho e assinatura ISO BMFF/MP4."""
    if not isinstance(video_base64, str) or not video_base64:
        raise ValueError('video_base64 é obrigatório.')
    if len(video_base64) > 4 * ((WHATSAPP_VIDEO_MAX_BYTES + 2) // 3):
        raise ValueError('Vídeo acima do limite permitido.')
    try:
        conteudo = base64.b64decode(video_base64, validate=True)
    except (ValueError, UnicodeEncodeError) as exc:
        raise ValueError('video_base64 inválido.') from exc
    if len(conteudo) > WHATSAPP_VIDEO_MAX_BYTES:
        raise ValueError('Vídeo acima do limite permitido.')
    if len(conteudo) < 12 or conteudo[4:8] != b'ftyp':
        raise ValueError('Conteúdo não é um MP4 válido.')
    tamanho_caixa = int.from_bytes(conteudo[:4], 'big')
    if tamanho_caixa < 8 or tamanho_caixa > len(conteudo):
        raise ValueError('Conteúdo não é um MP4 válido.')
    return conteudo


def _evolution_enviar_video(numero: str, video_base64: str, nome_arquivo: str):
    """Envia um MP4 já validado como base64 pela Evolution API."""
    endpoint = f'{EVOLUTION_API_URL}/message/sendMedia/{EVOLUTION_INSTANCE}'
    payload = {
        'number': numero,
        'mediatype': 'video',
        'mimetype': 'video/mp4',
        'media': video_base64,
        'fileName': nome_arquivo,
    }
    resposta = requests.post(
        endpoint,
        headers={'apikey': EVOLUTION_API_KEY, 'Content-Type': 'application/json'},
        json=payload,
        timeout=90,
    )
    if not (200 <= resposta.status_code < 300):
        raise RuntimeError(f'Evolution HTTP {resposta.status_code}')
    return resposta.json()


@app.route('/whatsapp/enviar-video', methods=['POST', 'OPTIONS'])
def whatsapp_enviar_video():
    """Envia exclusivamente um MP4 em base64; não aceita URLs externas."""
    if request.method == 'OPTIONS':
        return '', 204

    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or not secrets.compare_digest(api_key, EMAIL_WEBHOOK_KEY):
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401
    if not EVOLUTION_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'Integração indisponível.'}), 503

    body = request.get_json(silent=True)
    campos_permitidos = {'numero', 'video_base64', 'nome_arquivo'}
    if not isinstance(body, dict) or set(body) - campos_permitidos:
        return jsonify({'status': 'erro', 'erro': 'JSON ou campos inválidos.'}), 400

    numero = _normalizar_numero_evolution(body.get('numero'))
    if not numero:
        return jsonify({'status': 'erro', 'erro': 'Número inválido.'}), 400
    nome_arquivo = _sanitizar_nome_video(body.get('nome_arquivo'))
    if not nome_arquivo:
        return jsonify({'status': 'erro', 'erro': 'Nome de arquivo inválido.'}), 400
    try:
        _decodificar_video_mp4(body.get('video_base64'))
    except ValueError as exc:
        return jsonify({'status': 'erro', 'erro': str(exc)}), 400

    try:
        resultado = _evolution_enviar_video(numero, body['video_base64'], nome_arquivo)
    except (requests.RequestException, RuntimeError, ValueError):
        return jsonify({'status': 'erro', 'erro': 'Falha ao enviar vídeo.'}), 502

    identificador = resultado.get('key', {}).get('id') if isinstance(resultado, dict) else None
    if not identificador and isinstance(resultado, dict):
        identificador = resultado.get('id')
    resposta = {'status': 'ok'}
    if identificador:
        resposta['id'] = identificador
    return jsonify(resposta), 200


def _evolution_enviar_documento(numero: str, media_url: str, filename: str, caption=None,
                                 media_bytes: bytes = None):
    """
    Envia 1 documento (PDF) via Evolution API v2 (sendMedia).

    Por padrão manda 'media' como URL (Evolution busca o arquivo). Se
    media_bytes for informado, manda em base64 embutido no payload —
    mais confiável: elimina a etapa em que a própria Evolution API busca
    a URL (CloudFront/S3 do Airtable) por conta própria, etapa em que já
    foi observado em produção (08/07/2026) entrega de PDF corrompido/
    incompleto para uma minoria dos destinatários ("Erro de formato de
    arquivo" no WhatsApp), mesmo com o arquivo de origem íntegro
    (confirmado por download e validação direta do mesmo PDF).
    """
    endpoint = f'{EVOLUTION_API_URL}/message/sendMedia/{EVOLUTION_INSTANCE}'
    payload = {
        'number': numero,
        'mediatype': 'document',
        'mimetype': 'application/pdf',
        'media': base64.b64encode(media_bytes).decode('utf-8') if media_bytes else media_url,
        'fileName': filename,
    }
    if caption:
        payload['caption'] = caption
    r = requests.post(
        endpoint,
        headers={'apikey': EVOLUTION_API_KEY, 'Content-Type': 'application/json'},
        json=payload,
        timeout=90,
    )
    if not (200 <= r.status_code < 300):
        raise RuntimeError(f'Evolution HTTP {r.status_code}: {r.text[:300]}')
    return r.json()


def _sanitizar_nome_documento_pdf(nome_arquivo):
    """Retorna um nome PDF seguro, sem caminho ou caracteres de controle."""
    if not isinstance(nome_arquivo, str):
        return None
    nome = nome_arquivo.strip()
    if not nome or '/' in nome or '\\' in nome or '\x00' in nome:
        return None
    nome = unicodedata.normalize('NFKD', nome).encode('ascii', 'ignore').decode('ascii')
    nome = re.sub(r'[^A-Za-z0-9._-]', '_', nome)
    nome = re.sub(r'_+', '_', nome).strip('._')
    if not nome or len(nome) > 120 or not nome.lower().endswith('.pdf'):
        return None
    return nome


def _decodificar_documento_pdf(documento_base64):
    """Valida base64 estrito, limite binário e assinatura de um PDF."""
    if not isinstance(documento_base64, str) or not documento_base64:
        raise ValueError('documento_base64 é obrigatório.')
    limite_base64 = 4 * ((WHATSAPP_DOCUMENTO_MAX_BYTES + 2) // 3)
    if len(documento_base64) > limite_base64:
        raise ValueError('Documento acima do limite permitido.')
    try:
        conteudo = base64.b64decode(documento_base64, validate=True)
    except (ValueError, UnicodeEncodeError) as exc:
        raise ValueError('documento_base64 inválido.') from exc
    if len(conteudo) > WHATSAPP_DOCUMENTO_MAX_BYTES:
        raise ValueError('Documento acima do limite permitido.')
    if not conteudo.startswith(b'%PDF-'):
        raise ValueError('Conteúdo não é um PDF válido.')
    return conteudo


@app.route('/whatsapp/enviar-documento', methods=['POST', 'OPTIONS'])
def whatsapp_enviar_documento():
    """Envia diretamente um PDF em base64, sem Airtable ou assinatura."""
    if request.method == 'OPTIONS':
        return '', 204

    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or not secrets.compare_digest(api_key, EMAIL_WEBHOOK_KEY):
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401
    if not EVOLUTION_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'Integração indisponível.'}), 503

    body = request.get_json(silent=True)
    campos_permitidos = {'numero', 'documento_base64', 'nome_arquivo'}
    if not isinstance(body, dict) or set(body) - campos_permitidos:
        return jsonify({'status': 'erro', 'erro': 'JSON ou campos inválidos.'}), 400

    numero = _normalizar_numero_evolution(body.get('numero'))
    if not numero:
        return jsonify({'status': 'erro', 'erro': 'Número inválido.'}), 400
    nome_arquivo = _sanitizar_nome_documento_pdf(body.get('nome_arquivo'))
    if not nome_arquivo:
        return jsonify({'status': 'erro', 'erro': 'Nome de arquivo inválido.'}), 400
    try:
        documento = _decodificar_documento_pdf(body.get('documento_base64'))
    except ValueError as exc:
        return jsonify({'status': 'erro', 'erro': str(exc)}), 400

    try:
        resultado = _evolution_enviar_documento(
            numero,
            None,
            nome_arquivo,
            media_bytes=documento,
        )
    except (requests.RequestException, RuntimeError, ValueError):
        return jsonify({'status': 'erro', 'erro': 'Falha ao enviar documento.'}), 502

    identificador = resultado.get('key', {}).get('id') if isinstance(resultado, dict) else None
    if not identificador and isinstance(resultado, dict):
        identificador = resultado.get('id')
    resposta = {'status': 'ok'}
    if identificador:
        resposta['id'] = identificador
    return jsonify(resposta), 200


def _marcar_envio_status(envio_id: str, status: str, erro: str = None):
    _at_throttle()
    fields = {F_ENVIO_STATUS: status}
    if status == 'Enviado':
        agora_brt = datetime.now(timezone(timedelta(hours=-3)))   # horário de Brasília
        fields[F_ENVIO_DATA] = agora_brt.strftime('%Y-%m-%d')
        fields[F_ENVIO_ENVIADO_EM] = agora_brt.isoformat()        # timestamp exato p/ Protocolo
    if erro:
        fields[F_ENVIO_ERRO] = erro[:1000]
    r = requests.patch(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ENVIOS}/{envio_id}',
        headers=_at_headers(),
        json={'fields': fields, 'typecast': True},
        timeout=15,
    )
    _raise_for_airtable(r, f'marcar envio {envio_id}')


def _disparar_fila_combinado(limit=None, dry_run=True, numero_teste=None):
    """
    v2.28 — Lê os envios COMBINADOS (Holerite + Cartão Ponto) em Status
    "Preparando" e dispara cada um pela Evolution API: envia os 2 PDFs em
    sequência para o WhatsApp do colaborador e marca o envio como "Enviado".

    Segurança:
      - dry_run=True (padrão): não envia nada, só lista o que enviaria.
      - numero_teste: se informado, TODOS os envios vão para esse número
        (teste controlado), sem tocar os destinatários reais.
      - limit: limita quantos envios processa (ex.: limit=1 p/ teste).
    """
    if not dry_run and not EVOLUTION_API_KEY:
        return {'status': 'erro', 'erro': 'EVOLUTION_API_KEY não configurada no ambiente'}

    enviados, falhas, ignorados = [], [], []
    campos = ['Tipo', 'Status', 'Canal', 'Destinatário', 'Arquivos',
              'Funcionário(s) Vinculado(s)', 'Hash Recibo']

    for rec in _at_listar_todos(TABLE_ENVIOS, campos):
        f = rec['fields']
        if (f.get('Tipo') != TIPO_ENVIO_COMBINADO or f.get('Status') != 'Preparando'
                or f.get('Canal') != 'WhatsApp'):
            continue
        if limit is not None and (len(enviados) + len(falhas)) >= limit:
            break

        func = (f.get('Funcionário(s) Vinculado(s)') or [])
        nome = func[0]['name'] if func and isinstance(func[0], dict) else None
        arquivos = f.get('Arquivos') or []
        hash_recibo = f.get('Hash Recibo')
        numero = _normalizar_numero_evolution(numero_teste or f.get('Destinatário'))

        item = {
            'envio_id': rec['id'], 'funcionario': nome,
            'destinatario': numero, 'qtd_anexos': len(arquivos),
        }

        if not numero or len(arquivos) < 2:
            item['motivo'] = 'numero_invalido' if not numero else 'menos_de_2_anexos'
            ignorados.append(item)
            continue

        if dry_run:
            item['acao'] = 'enviaria'
            enviados.append(item)
            continue

        # Link de confirmação de recebimento (comprovação de leitura): ao abrir,
        # o colaborador marca "Recibo Lido em" -> entra no Protocolo de Entrega.
        link_recibo = f'{RECIBO_BASE_URL}/recibo/{hash_recibo}' if hash_recibo else None

        try:
            for i, a in enumerate(arquivos[:2]):
                caption = ((
                    f'Olá{(" " + nome) if nome else ""}! Seguem seus documentos do mês — '
                    f'Holerite e Cartão Ponto. Magnata Portaria e Serviços.'
                    + (f'\n\nConfirme o recebimento abrindo: {link_recibo}' if link_recibo else '')
                ) if i == 0 else None)
                _evolution_enviar_documento(numero, a['url'], a.get('filename', 'documento.pdf'), caption)
                _at_throttle()
            _marcar_envio_status(rec['id'], 'Enviado')
            item['acao'] = 'enviado'
            enviados.append(item)
        except Exception as exc:
            logger.error(f'[DISPARO] falha no envio {rec["id"]} ({nome}): {exc}')
            try:
                _marcar_envio_status(rec['id'], 'Erro', erro=str(exc))
            except Exception:
                pass
            item['acao'] = 'falhou'
            item['erro'] = str(exc)
            falhas.append(item)

    return {
        'status': 'concluido',
        'dry_run': dry_run,
        'numero_teste': numero_teste,
        'total_enviados': len(enviados),
        'total_falhas': len(falhas),
        'total_ignorados': len(ignorados),
        'enviados': enviados,
        'falhas': falhas,
        'ignorados': ignorados,
    }


@app.route('/disparar-fila-combinado', methods=['POST', 'OPTIONS'])
def disparar_fila_combinado():
    """
    v2.28 — Dispara pela Evolution API os envios combinados (Holerite + Ponto)
    em Status "Preparando". Body JSON:
      - dry_run: true (padrão) só simula
      - limit: limita a quantidade (ex.: 1 para teste)
      - numero_teste: redireciona TODOS os envios para 1 número (teste seguro)
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401
    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    body = request.get_json(silent=True) or {}
    limit = body.get('limit')
    try:
        limit = int(limit) if limit else None
    except (ValueError, TypeError):
        limit = None
    dry_run = body.get('dry_run', True)
    numero_teste = body.get('numero_teste')

    resultado = _disparar_fila_combinado(limit=limit, dry_run=dry_run, numero_teste=numero_teste)
    return jsonify(resultado)


# ═══════════════════════════════════════════════════════════════════════════
# v2.29 — DISTRIBUIÇÃO POR E-MAIL (CLIENTES/ESCRITÓRIOS)
#   1) Fatiador POR CLIENTE (Extrato Mensal / FGTS Digital) — espelha o
#      fatiador por CPF, mas casa cada página a um Cliente (por CNPJ; fallback
#      por Nome). Resolve de vez o histórico "SEM CLIENTE".
#   2) Gerador da fila de e-mail: 1 pacote por Cliente Ativo (holerites do mês
#      + extrato + FGTS exclusivos + guias comuns).
#   3) Motor SMTP: dispara os pacotes (dry_run/limit/email_teste — testar antes).
# ═══════════════════════════════════════════════════════════════════════════

def extrair_cnpj(texto: str):
    """Extrai o 1º CNPJ (14 dígitos) do texto, normalizado só com dígitos."""
    if not texto:
        return None
    m = re.search(r'\d{2}[.\s]?\d{3}[.\s]?\d{3}[/\s]?\d{4}[-\s]?\d{2}', texto)
    if not m:
        return None
    dig = re.sub(r'\D', '', m.group(0))
    return dig if len(dig) == 14 else None


def _normalizar_texto_busca(s: str) -> str:
    s = unicodedata.normalize('NFKD', s or '')
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', s).strip().upper()


def _carregar_indice_clientes():
    """Índice p/ casar páginas → Cliente. Retorna (por_cnpj, nomes):
      - por_cnpj: {cnpj14: (cliente_id, nome)}
      - nomes: [(nome_norm, cliente_id, nome)] ordenado do nome mais longo p/ o
        mais curto (evita casar 'SKY' antes de 'EDIFICIO SKY TATUI').
    """
    por_cnpj = {}
    nomes = []
    for rec in _at_listar_todos(TABLE_CLIENTES, ['Nome', 'Cliente', 'CNPJ']):
        f = rec['fields']
        nome = f.get('Nome') or f.get('Cliente') or ''
        nome_norm = _normalizar_texto_busca(nome)
        # Exclui a PRÓPRIA Magnata (empregador): o nome/CNPJ dela aparece no
        # cabeçalho de TODA página do extrato/FGTS — não é condomínio-destino.
        # Sem isso, páginas que não casam o condomínio "vazam" para a Magnata.
        if 'MAGNATA' in nome_norm:
            continue
        cnpj = re.sub(r'\D', '', f.get('CNPJ') or '')
        if len(cnpj) == 14 and cnpj != CNPJ_MAGNATA:
            por_cnpj[cnpj] = (rec['id'], nome)
        if len(nome_norm) >= 4:
            nomes.append((nome_norm, rec['id'], nome))
    nomes.sort(key=lambda x: len(x[0]), reverse=True)
    return por_cnpj, nomes


def construir_mapa_cliente(caminho_pdf: str):
    """
    Pass 1 do fatiador POR CLIENTE — modelo SECCIONADO POR TOMADOR.

    O Extrato Mensal/FGTS é organizado em seções: a 1ª página de cada tomador
    traz o CNPJ (ou nome) do condomínio; as páginas seguintes (só com o CNPJ da
    Magnata, o empregador) são o DETALHE daquele tomador. Por isso:
      1) Página com CNPJ/nome de cliente conhecido → identifica a seção (vira o
         "tomador atual") e a página entra para ele.
      2) Página SEM identificação E sem outro CNPJ tomador → DETALHE: herda o
         tomador atual (carry-forward).
      3) Página com um CNPJ tomador DESCONHECIDO (≠ Magnata, não cadastrado) →
         é uma nova seção de cliente não cadastrado: QUEBRA o carry-forward
         (zera o tomador atual) e vai para sem_cliente — evita grudar no anterior.

    Retorna ({cliente_id: {'nome':.., 'paginas':[int]}}, total, paginas_sem_cliente).
    """
    por_cnpj, nomes = _carregar_indice_clientes()
    mapa, sem_cliente = {}, []
    cliente_atual = None  # (id, nome) da seção/tomador corrente
    with pdfplumber.open(caminho_pdf) as pdf:
        total = len(pdf.pages)
        logger.info(f'[MAPA-CLI] Total de páginas: {total}')
        for i in range(total):
            try:
                texto = pdf.pages[i].extract_text() or ''
            except Exception as exc:
                logger.warning(f'[MAPA-CLI] pág {i+1}: erro extração → {exc}')
                texto = ''

            # CNPJs da página, excluindo o da Magnata (empregador onipresente)
            cnpjs_pagina = [
                re.sub(r'\D', '', c) for c in
                re.findall(r'\d{2}[.\s]?\d{3}[.\s]?\d{3}[/\s]?\d{4}[-\s]?\d{2}', texto)
            ]
            tomador_cnpjs = [c for c in cnpjs_pagina if len(c) == 14 and c != CNPJ_MAGNATA]

            cliente_id = cliente_nome = None
            for c in tomador_cnpjs:                       # 1a) por CNPJ do tomador
                if c in por_cnpj:
                    cliente_id, cliente_nome = por_cnpj[c]
                    break
            if not cliente_id:                            # 1b) por Nome do cliente
                texto_norm = _normalizar_texto_busca(texto)
                for nome_norm, cid, cnome in nomes:
                    if nome_norm in texto_norm:
                        cliente_id, cliente_nome = cid, cnome
                        break

            if cliente_id:
                cliente_atual = (cliente_id, cliente_nome)   # abre/atualiza a seção
            elif tomador_cnpjs:
                # nova seção de tomador NÃO cadastrado → quebra o carry-forward
                cliente_atual = None
                sem_cliente.append(i)
                logger.info(f'[MAPA-CLI] pág {i+1:03d}/{total}: SEM CLIENTE (tomador desconhecido {tomador_cnpjs})')
                del texto
                continue
            elif cliente_atual:
                cliente_id, cliente_nome = cliente_atual     # detalhe → herda a seção

            if cliente_id:
                if cliente_id not in mapa:
                    mapa[cliente_id] = {'nome': cliente_nome, 'paginas': []}
                mapa[cliente_id]['paginas'].append(i)
                logger.info(f'[MAPA-CLI] pág {i+1:03d}/{total}: → {cliente_nome}')
            else:
                sem_cliente.append(i)
                logger.info(f'[MAPA-CLI] pág {i+1:03d}/{total}: SEM CLIENTE')

            del texto
            if (i + 1) % 15 == 0:
                gc.collect()
    gc.collect()
    logger.info(f'[MAPA-CLI] {len(mapa)} clientes | {len(sem_cliente)} págs sem cliente')
    return mapa, total, sem_cliente


def _processar_doc_cliente_master(caminho_pdf: str, tipo: str, folha_mensal: str):
    """Fatia um mestre (Extrato Mensal OU FGTS Digital) por cliente e cria 1
    registro por cliente na tabela respectiva, anexando só as páginas dele.
    tipo: 'extrato' | 'fgts'."""
    if tipo == 'extrato':
        table, f_cli, f_folha, f_pdf, f_data, f_status = (
            TABLE_EXTRATO, F_EXT_CLIENTE, F_EXT_FOLHA, F_EXT_PDF, F_EXT_DATA, F_EXT_STATUS)
        rotulo = 'Extrato Mensal'
    elif tipo == 'fgts':
        table, f_cli, f_folha, f_pdf, f_data, f_status = (
            TABLE_FGTS, F_FGTS_CLIENTE, F_FGTS_FOLHA, F_FGTS_PDF, F_FGTS_DATA, F_FGTS_STATUS)
        rotulo = 'FGTS Digital'
    else:
        return {'status': 'erro', 'erro': f'tipo inválido: {tipo} (use extrato|fgts)'}

    mapa, total, sem_cliente = construir_mapa_cliente(caminho_pdf)
    hoje = datetime.now().strftime('%Y-%m-%d')
    criados, falhas = [], []
    for cliente_id, dados in mapa.items():
        nome_cli = dados['nome'] or cliente_id
        try:
            pdf_bytes = extrair_pdf_colaborador(caminho_pdf, dados['paginas'])
            campos = {
                f_cli: [cliente_id],
                f_folha: folha_mensal or '',
                f_status: 'Recebido',
                f_data: hoje,
            }
            rec_id = _criar_registro(table, campos)
            fname = f'{rotulo} - {folha_mensal or ""} - {nome_cli}.pdf'.strip()
            _anexar_attachment(table, rec_id, f_pdf, pdf_bytes, fname)
            criados.append({
                'cliente': nome_cli, 'cliente_id': cliente_id,
                'paginas': [p + 1 for p in dados['paginas']], 'record_id': rec_id,
            })
        except Exception as exc:
            logger.error(f'[DOC-CLI] falha cliente {nome_cli}: {exc}')
            falhas.append({'cliente': nome_cli, 'cliente_id': cliente_id, 'erro': str(exc)})

    return {
        'status': 'concluido', 'tipo': rotulo, 'folha_mensal': folha_mensal,
        'total_paginas': total, 'clientes_identificados': len(mapa),
        'registros_criados': len(criados),
        'paginas_sem_cliente': [p + 1 for p in sem_cliente],
        'criados': criados, 'falhas': falhas,
    }


@app.route('/processar-doc-cliente', methods=['POST', 'OPTIONS'])
def processar_doc_cliente():
    """
    v2.29 — Fatia um PDF mestre (Extrato Mensal ou FGTS Digital) por CLIENTE.
    multipart/form-data:
      - pdf: arquivo mestre
      - tipo: 'extrato' | 'fgts'
      - folha_mensal: ex. "Maio 2026"
    Protegido por X-API-KEY (EMAIL_WEBHOOK_KEY) — Macro 6A, correção de
    auditoria (achado #4): esta rota escrevia em Extrato Mensal/FGTS de
    CLIENTES aceitando qualquer chamador que soubesse a URL, sem nenhuma
    credencial própria (só a checagem de que o SERVIDOR tem AIRTABLE_API_KEY
    configurada — isso não autentica o chamador). Reaproveita o mesmo
    mecanismo e a mesma variável já usados por /email/webhook e
    /processar-fila — sem segredo novo. Cria 1 registro por cliente na aba
    respectiva.
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401
    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    tipo = (request.form.get('tipo') or '').strip().lower()
    folha_mensal = (request.form.get('folha_mensal') or '').strip()
    if tipo not in ('extrato', 'fgts'):
        return jsonify({'status': 'erro', 'erro': "tipo deve ser 'extrato' ou 'fgts'"}), 400

    caminho = extrair_pdf_do_request()
    if not caminho:
        return jsonify({'status': 'erro', 'erro': 'PDF não enviado (campo "pdf")'}), 400
    try:
        resultado = _processar_doc_cliente_master(caminho, tipo, folha_mensal)
        return jsonify(resultado)
    finally:
        try:
            os.remove(caminho)
        except OSError:
            pass


# ── Benefícios: Horas Extras, Assiduidade, VR/VA (v3.12) ──────────────────────
# Bloco ISOLADO: só escreve nos 3 campos novos de Clientes (Horas Extras,
# Assiduidade, VR/VA Benefícios) — não toca em Holerite, Ponto (Regra de
# Prevalência intacta) nem no campo "Documento PDF"/imutabilidade das
# Assinaturas Digitais.

_RECIBO_TERCEIROS_VALOR_RE = re.compile(r'quantia de\s*(?:R\$\s*)+([\d.,]+)\s*,\s*referente', re.IGNORECASE)
_RECIBO_TERCEIROS_EMPREGADO_RE = re.compile(r'Empregado\(a\):\s*(.+)')


def _extrair_recibo_terceiros(texto: str):
    """Extrai (nome, cpf, valor_float) de 1 página do 'RECIBO DE PAGAMENTO DE
    TERCEIROS' (Sicoob) — formato compartilhado por Horas Extras e
    Assiduidade. cpf reaproveita extrair_cpf (a linha 'CPF/CNPJ' do
    empregador não bate no formato XXX.XXX.XXX-XX, então não dá falso
    positivo antes de chegar na linha 'CPF:' do empregado)."""
    if not texto:
        return None, None, None
    m_nome = _RECIBO_TERCEIROS_EMPREGADO_RE.search(texto)
    nome = m_nome.group(1).strip() if m_nome else None
    cpf = extrair_cpf(texto)
    m_valor = _RECIBO_TERCEIROS_VALOR_RE.search(texto)
    valor = None
    if m_valor:
        try:
            valor = float(m_valor.group(1).replace('.', '').replace(',', '.'))
        except ValueError:
            valor = None
    return nome, cpf, valor


def _buscar_cliente_do_funcionario(func_id: str):
    """Resolve o(s) Cliente(s) de um Funcionário via Locais de trabalho →
    Local.Cliente (2 saltos de multipleRecordLinks). O lookup 'Cliente'
    direto do Funcionário não expõe o record id pela API padrão (só o nome
    já resolvido), por isso a travessia manual em vez de ler o lookup."""
    _at_throttle()
    r = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}/{func_id}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        timeout=15,
    )
    if not r.ok:
        return []
    locais_ids = r.json().get('fields', {}).get('Locais de trabalho') or []
    if not locais_ids:
        return []
    clientes, vistos = [], set()
    for local_id in locais_ids:
        _at_throttle()
        rl = requests.get(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_LOCAIS}/{local_id}',
            headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
            timeout=15,
        )
        if not rl.ok:
            continue
        for cid in (rl.json().get('fields', {}).get('Cliente') or []):
            if cid in vistos:
                continue
            vistos.add(cid)
            _at_throttle()
            rc = requests.get(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_CLIENTES}/{cid}',
                headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
                timeout=15,
            )
            # Sem params 'fields[]' de propósito: GET de registro único do
            # Airtable ignora esse filtro (mesma causa raiz do bug de CPF do
            # Venceslau/v3.01 hoje mais cedo) — sempre retorna todos os
            # campos, então busca-se Nome com fallback para Cliente
            # (fórmula), igual a _carregar_indice_clientes.
            campos_cli = rc.json().get('fields', {}) if rc.ok else {}
            nome_cli = campos_cli.get('Nome') or campos_cli.get('Cliente') or ''
            clientes.append((cid, nome_cli))
    return clientes


def _processar_lote_sicoob(caminho_pdf: str, tipo: str, folha_mensal: str, dry_run: bool = False) -> dict:
    """v3.12 — Fatia um mestre de recibos Sicoob (1 página = 1 colaborador)
    de Horas Extras ou Assiduidade, agrupa por Cliente (via Locais de
    trabalho do Funcionário) e funde num único PDF por cliente.
    tipo: 'horas_extras' | 'assiduidade'.

    Assiduidade tem trava de segurança: só entra quem tem valor == R$110,00
    exatos; quem diverge vira Pendência para revisão humana, nunca é
    descartado silenciosamente."""
    if tipo == 'horas_extras':
        campo_cliente, rotulo = F_CLI_HORAS_EXTRAS, 'Horas Extras'
    elif tipo == 'assiduidade':
        campo_cliente, rotulo = F_CLI_ASSIDUIDADE, 'Assiduidade'
    else:
        return {'status': 'erro', 'erro': f"tipo inválido: {tipo} (use horas_extras|assiduidade)"}

    paginas_por_cliente = {}
    erros = []
    with pdfplumber.open(caminho_pdf) as pdf:
        total = len(pdf.pages)
        for i in range(total):
            texto = pdf.pages[i].extract_text() or ''
            nome, cpf, valor = _extrair_recibo_terceiros(texto)
            if not cpf:
                erros.append({'pagina': i + 1, 'motivo': 'CPF não extraído da página', 'nome': nome})
                continue

            if tipo == 'assiduidade' and (valor is None or round(valor, 2) != 110.00):
                if not dry_run:
                    _criar_registro(TABLE_PENDENCIAS, {
                        F_PEND_NOME: f'Assiduidade com valor divergente de R$ 110,00: {nome or cpf}',
                        F_PEND_STATUS: 'Pendente',
                        F_PEND_TIPO: 'Assiduidade — valor divergente (Formato B)',
                        F_PEND_OBS: f'Página {i + 1}. Valor extraído: {valor}. CPF: {cpf}.',
                        F_PEND_DATA: datetime.now().isoformat(),
                    })
                erros.append({'pagina': i + 1, 'motivo': 'valor diferente de R$ 110,00', 'nome': nome, 'valor': valor})
                continue

            func_id, nome_at = buscar_funcionario_por_cpf(cpf)
            if not func_id:
                if not dry_run:
                    _criar_registro(TABLE_PENDENCIAS, {
                        F_PEND_NOME: f'{rotulo}: funcionário não encontrado — {nome or cpf}',
                        F_PEND_STATUS: 'Pendente',
                        F_PEND_TIPO: f'{rotulo} — funcionário não encontrado',
                        F_PEND_OBS: f'Página {i + 1}. CPF: {cpf}.',
                        F_PEND_DATA: datetime.now().isoformat(),
                    })
                erros.append({'pagina': i + 1, 'motivo': 'funcionário não encontrado', 'nome': nome, 'cpf': cpf})
                continue

            clientes = _buscar_cliente_do_funcionario(func_id)
            if not clientes:
                if not dry_run:
                    _criar_registro(TABLE_PENDENCIAS, {
                        F_PEND_NOME: f'{rotulo}: sem Cliente/posto vinculado — {nome_at or nome}',
                        F_PEND_STATUS: 'Pendente',
                        F_PEND_TIPO: f'{rotulo} — sem cliente vinculado',
                        F_PEND_OBS: f'Página {i + 1}. Funcionário {func_id} sem Locais de trabalho/Cliente.',
                        F_PEND_DATA: datetime.now().isoformat(),
                    })
                erros.append({'pagina': i + 1, 'motivo': 'funcionário sem cliente vinculado', 'nome': nome_at or nome})
                continue

            for cliente_id, cliente_nome in clientes:
                if cliente_id not in paginas_por_cliente:
                    paginas_por_cliente[cliente_id] = {'nome': cliente_nome, 'paginas': [], 'pessoas': []}
                paginas_por_cliente[cliente_id]['paginas'].append(i)
                paginas_por_cliente[cliente_id]['pessoas'].append(nome_at or nome)

            if (i + 1) % 15 == 0:
                gc.collect()

    if dry_run:
        return {
            'status': 'concluido', 'tipo': rotulo, 'dry_run': True, 'folha_mensal': folha_mensal,
            'total_paginas': total, 'clientes_identificados': len(paginas_por_cliente),
            'clientes_a_anexar': [
                {'cliente': d['nome'], 'cliente_id': cid, 'colaboradores': d['pessoas'],
                 'paginas': [p + 1 for p in d['paginas']]}
                for cid, d in paginas_por_cliente.items()
            ],
            'erros': erros,
        }

    anexados = []
    for cliente_id, dados in paginas_por_cliente.items():
        try:
            pdf_bytes = extrair_pdf_colaborador(caminho_pdf, dados['paginas'])
            fname = f'{rotulo} {folha_mensal or ""} - {dados["nome"] or cliente_id}.pdf'.strip()
            _anexar_attachment(TABLE_CLIENTES, cliente_id, campo_cliente, pdf_bytes, fname)
            anexados.append({
                'cliente': dados['nome'], 'cliente_id': cliente_id,
                'colaboradores': dados['pessoas'], 'paginas': [p + 1 for p in dados['paginas']],
            })
        except Exception as exc:
            logger.error(f'[BENEFICIOS] falha ao anexar {rotulo} cliente {dados["nome"]}: {exc}')
            erros.append({'cliente': dados['nome'], 'cliente_id': cliente_id, 'motivo': str(exc)})

    return {
        'status': 'concluido', 'tipo': rotulo, 'dry_run': False, 'folha_mensal': folha_mensal,
        'total_paginas': total, 'clientes_identificados': len(paginas_por_cliente),
        'clientes_anexados': anexados, 'erros': erros,
    }


def _processar_vr_va_pix(caminho_pdf: str, folha_mensal: str, dry_run: bool = False) -> dict:
    """v3.12 — PIX individual de VR/VA para colaborador novo sem cartão de
    benefício ainda. NÃO TESTADO contra um exemplo real (só a descrição
    fornecida) — layout de comprovante bancário comum, com 'VR' ou 'VA' na
    descrição/detalhes do PIX. 1 página = 1 colaborador."""
    paginas_por_cliente = {}
    erros = []
    with pdfplumber.open(caminho_pdf) as pdf:
        total = len(pdf.pages)
        for i in range(total):
            texto = pdf.pages[i].extract_text() or ''
            if not re.search(r'\bVR\b|\bVA\b', texto):
                erros.append({'pagina': i + 1, 'motivo': "sem 'VR'/'VA' na descrição — não é PIX de benefício"})
                continue
            cpf = extrair_cpf(texto)
            if not cpf:
                erros.append({'pagina': i + 1, 'motivo': 'CPF não extraído'})
                continue
            func_id, nome_at = buscar_funcionario_por_cpf(cpf)
            if not func_id:
                if not dry_run:
                    _criar_registro(TABLE_PENDENCIAS, {
                        F_PEND_NOME: f'VR/VA PIX individual: funcionário não encontrado — CPF {cpf}',
                        F_PEND_STATUS: 'Pendente',
                        F_PEND_TIPO: 'VR/VA PIX — funcionário não encontrado',
                        F_PEND_OBS: f'Página {i + 1}.',
                        F_PEND_DATA: datetime.now().isoformat(),
                    })
                erros.append({'pagina': i + 1, 'motivo': 'funcionário não encontrado', 'cpf': cpf})
                continue
            clientes = _buscar_cliente_do_funcionario(func_id)
            if not clientes:
                if not dry_run:
                    _criar_registro(TABLE_PENDENCIAS, {
                        F_PEND_NOME: f'VR/VA PIX individual: sem Cliente/posto vinculado — {nome_at}',
                        F_PEND_STATUS: 'Pendente',
                        F_PEND_TIPO: 'VR/VA PIX — sem cliente vinculado',
                        F_PEND_OBS: f'Página {i + 1}. Funcionário {func_id}.',
                        F_PEND_DATA: datetime.now().isoformat(),
                    })
                erros.append({'pagina': i + 1, 'motivo': 'funcionário sem cliente vinculado', 'nome': nome_at})
                continue
            for cliente_id, cliente_nome in clientes:
                if cliente_id not in paginas_por_cliente:
                    paginas_por_cliente[cliente_id] = {'nome': cliente_nome, 'paginas': [], 'pessoas': []}
                paginas_por_cliente[cliente_id]['paginas'].append(i)
                paginas_por_cliente[cliente_id]['pessoas'].append(nome_at)

    if dry_run:
        return {
            'status': 'concluido', 'dry_run': True, 'folha_mensal': folha_mensal,
            'total_paginas': total,
            'clientes_a_anexar': [
                {'cliente': d['nome'], 'cliente_id': cid, 'colaboradores': d['pessoas'],
                 'paginas': [p + 1 for p in d['paginas']]}
                for cid, d in paginas_por_cliente.items()
            ],
            'erros': erros,
        }

    anexados = []
    for cliente_id, dados in paginas_por_cliente.items():
        try:
            pdf_bytes = extrair_pdf_colaborador(caminho_pdf, dados['paginas'])
            fname = f'VR-VA PIX Individual {folha_mensal or ""} - {dados["nome"] or cliente_id}.pdf'.strip()
            _anexar_attachment(TABLE_CLIENTES, cliente_id, F_CLI_VRVA, pdf_bytes, fname)
            anexados.append({
                'cliente': dados['nome'], 'cliente_id': cliente_id,
                'colaboradores': dados['pessoas'], 'paginas': [p + 1 for p in dados['paginas']],
            })
        except Exception as exc:
            logger.error(f'[BENEFICIOS] falha ao anexar VR/VA PIX cliente {dados["nome"]}: {exc}')
            erros.append({'cliente': dados['nome'], 'cliente_id': cliente_id, 'motivo': str(exc)})

    return {
        'status': 'concluido', 'dry_run': False, 'folha_mensal': folha_mensal,
        'total_paginas': total, 'clientes_anexados': anexados, 'erros': erros,
    }


# ── Extrato Bancário Consolidador (v3.15) ──────────────────────────────────
# O banco manda 1 PDF único de extrato PIX com centenas de comprovantes
# concatenados (achado real: 321 num arquivo de 122 páginas), misturando
# Horas Extras/Almoço-Jantar/VR-VA com pagamentos sem relação nenhuma
# (diárias, fornecedores, salário — a maioria real do arquivo). CPF vem
# MASCARADO ("***.018.388-**"), então a identificação é por NOME (fuzzy,
# reaproveitando os helpers da Folha de Ponto Formato B). Transações podem
# atravessar página — por isso, em vez de fatiar páginas originais, gera-se
# um PDF-relatório novo por Cliente+Categoria (mesmo padrão de
# _gerar_pdf_vr_va_cliente), evitando depender de fronteira de página.

_EXTRATO_BLOCO_RE = re.compile(r'COMPROVANTE DE EFETIVA[ÇC][ÃA]O DE PAGAMENTO PIX')
_EXTRATO_RUIDO_RES = [
    re.compile(r'^\d{2}/\d{2}/\d{4},?\s*\d{2}:\d{2}\s*Sicoob\s*\|\s*Internet Banking\s*$', re.MULTILINE),
    re.compile(r'^https://ib\.sicoob\.com\.br.*$', re.MULTILINE),
    re.compile(r'^OUVIDORIA SICOOB.*$', re.MULTILINE),
]


def _extrair_transacoes_extrato_pix(caminho_pdf: str) -> list:
    """Extrai todas as transações PIX de um extrato consolidado. Retorna
    (transacoes, ilegiveis) — transacoes é lista de {id_transacao, nome,
    cpf_mascarado, valor, descricao, data}; ilegiveis é a contagem de blocos
    onde Nome/Descrição não puderam ser lidos (ex.: campo cortado ao meio
    por quebra de página) — nunca descartados sem rastro, sempre contados.

    Dedup por 'ID Transação' (único por PIX, sempre presente no bloco): em
    lotes com sobreposição de página (v3.16, para não perder transação
    cortada na borda), a mesma transação completa pode aparecer 2x — aqui
    dentro de UM arquivo. Dedup entre chamadas/lotes diferentes é
    responsabilidade de quem agrega os resultados dos lotes."""
    with pdfplumber.open(caminho_pdf) as pdf:
        texto_total = '\n'.join((pg.extract_text() or '') for pg in pdf.pages)

    texto_limpo = texto_total
    for ruido_re in _EXTRATO_RUIDO_RES:
        texto_limpo = ruido_re.sub('', texto_limpo)

    blocos = _EXTRATO_BLOCO_RE.split(texto_limpo)
    transacoes, ilegiveis = [], 0
    ids_vistos = set()
    for bloco in blocos[1:]:  # blocos[0] é preâmbulo antes do 1º comprovante
        if 'Destinat' not in bloco:
            ilegiveis += 1
            continue
        sub_destinatario = bloco.split('Destinat', 1)[1]
        m_nome = re.search(r'Nome:\s*(.+)', sub_destinatario)
        m_cpf = re.search(r'CPF/CNPJ:\s*([\d*./-]+)', sub_destinatario)
        m_valor = re.search(r'Valor:\s*R\$\s*([\d.,]+)', bloco)
        m_desc = re.search(r'Descri[çc][ãa]o:\s*(.+)', bloco)
        m_data = re.search(r'Data do pagamento:\s*([\d/]+)', bloco)
        m_id = re.search(r'ID Transa[çc][ãa]o:\s*(\S+)', bloco)
        if not m_nome or not m_desc:
            ilegiveis += 1
            continue
        id_transacao = m_id.group(1).strip() if m_id else None
        if id_transacao and id_transacao in ids_vistos:
            continue
        if id_transacao:
            ids_vistos.add(id_transacao)
        valor = None
        if m_valor:
            try:
                valor = float(m_valor.group(1).replace('.', '').replace(',', '.'))
            except ValueError:
                pass
        transacoes.append({
            'id_transacao': id_transacao,
            'nome': m_nome.group(1).strip(),
            'cpf_mascarado': m_cpf.group(1).strip() if m_cpf else '',
            'valor': valor,
            'descricao': m_desc.group(1).strip(),
            'data': m_data.group(1).strip() if m_data else '',
        })
    return transacoes, ilegiveis


def _classificar_transacao_extrato(descricao: str):
    """Ordem = prioridade das regras (v3.15/v3.16): VR/VA > Almoço/Jantar >
    Extra/Horas > Diárias — cobre os casos reais onde uma descrição cita
    mais de uma categoria junto (ex.: 'DIARIAS 750 VR 130 HORA ALMOÇO 56
    45' vira VR/VA, não Diárias; 'VR ref 3 diarias Carlos' também vira
    VR/VA). Diárias por último pega só quem não bateu em nenhuma categoria
    mais específica. Retorna None só para o que sobra sem nenhuma palavra-
    chave (fornecedor, salário, etc.)."""
    d = descricao or ''
    if re.search(r'\bvr\b|\bva\b|alimenta[çc][ãa]o|refei[çc][ãa]o', d, re.IGNORECASE):
        return 'vrva'
    if re.search(r'almo[çc]?o?|jant', d, re.IGNORECASE):
        return 'almoco_jantar'
    if re.search(r'extra|\bhora', d, re.IGNORECASE):
        return 'horas_extras'
    if re.search(r'di[áa]rias?', d, re.IGNORECASE):
        return 'diarias'
    return None


_CATEGORIAS_EXTRATO = {
    'vrva':          {'rotulo': 'VR/VA', 'campo': F_CLI_VRVA},
    'almoco_jantar': {'rotulo': 'Hora Almoço e Jantar', 'campo': F_CLI_ALMOCO_JANTA},
    'horas_extras':  {'rotulo': 'Horas Extras', 'campo': F_CLI_HORAS_EXTRAS},
    'diarias':       {'rotulo': 'Diárias', 'campo': F_CLI_DIARIAS},
}


def _gerar_pdf_extrato_categoria(nome_cliente: str, rotulo_categoria: str,
                                  folha_mensal: str, transacoes: list) -> bytes:
    """Relatório consolidado (mesmo estilo visual de _gerar_pdf_vr_va_cliente)
    de PIX extraídos do extrato bancário, 1 tabela por Cliente+Categoria."""
    from fpdf import FPDF

    def fmt_brl(v):
        if v is None:
            return '-'
        return f'{v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

    _TZ = timezone(timedelta(hours=-3))
    pdf = FPDF()
    pdf.add_page()

    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 8, 'Magnata Portaria e Servicos Ltda', ln=True, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, f'{rotulo_categoria} - {folha_mensal}', ln=True, align='C')
    pdf.cell(0, 6, f'Cliente: {nome_cliente}', ln=True, align='C')
    pdf.ln(4)

    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_fill_color(220, 220, 220)
    pdf.cell(55, 7, 'Nome', border=1, fill=True)
    pdf.cell(30, 7, 'CPF (parcial)', border=1, fill=True)
    pdf.cell(22, 7, 'Data', border=1, fill=True)
    pdf.cell(66, 7, 'Descricao (PIX)', border=1, fill=True)
    pdf.cell(0, 7, 'Valor (R$)', border=1, fill=True, align='R', ln=True)

    pdf.set_font('Helvetica', '', 7)
    total_geral = 0.0
    for t in sorted(transacoes, key=lambda x: x['nome']):
        v = t.get('valor') or 0.0
        total_geral += v
        pdf.cell(55, 6, t['nome'][:32], border=1)
        pdf.cell(30, 6, t.get('cpf_mascarado', ''), border=1)
        pdf.cell(22, 6, t.get('data', ''), border=1)
        pdf.cell(66, 6, (t.get('descricao') or '')[:38], border=1)
        pdf.cell(0, 6, fmt_brl(t.get('valor')), border=1, align='R', ln=True)

    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_fill_color(200, 230, 200)
    pdf.cell(173, 7, f'TOTAL - {len(transacoes)} lançamento(s)', border=1, fill=True, align='R')
    pdf.cell(0, 7, f'R$ {fmt_brl(total_geral)}', border=1, fill=True, align='R', ln=True)

    pdf.ln(4)
    pdf.set_font('Helvetica', 'I', 7)
    agora = datetime.now(tz=_TZ).strftime('%d/%m/%Y %H:%M')
    pdf.cell(0, 4, f'Gerado em {agora} a partir do Extrato Bancario consolidado (CPF mascarado na origem).', ln=True)

    return bytes(pdf.output())


def _carregar_indice_funcionario_cliente():
    """v3.16 — Carrega Funcionários + Locais + Clientes em LOTE (3 chamadas
    ao Airtable, no total) e devolve uma função resolver(nome) -> (status,
    dado), tudo em memória a partir daí.

    Causa raiz real do 502 na 1ª tentativa do Extrato Bancário Consolidado:
    _buscar_funcionario_por_nome recarregava a tabela INTEIRA de
    Funcionários a cada nome diferente da transação (~200 nomes únicos =
    ~200 recargas completas da tabela) — não era o tamanho do PDF. Mesma
    estratégia de eficiência já usada em /processar-vr-va (v2.34)."""
    locais_map = {
        rec['id']: [c['id'] if isinstance(c, dict) else c for c in (rec['fields'].get('Cliente') or [])]
        for rec in _at_listar_todos(TABLE_LOCAIS, ['Nome', 'Cliente'])
    }
    clientes_nome = {
        rec['id']: (rec['fields'].get('Nome') or rec['fields'].get('Cliente') or rec['id'])
        for rec in _at_listar_todos(TABLE_CLIENTES, ['Nome', 'Cliente'])
    }

    exatos = {}       # nome_norm -> (func_id, nome_completo, [locais_ids])
    homonimos_exatos = {}  # nome_norm -> [{'id':.., 'nome':..}, ...]
    parcial_idx = {}  # (primeiro, ultimo) -> [(func_id, nome_completo, [locais_ids]), ...]

    for rec in _at_listar_todos(TABLE_FUNC, ['Nome Completo', 'Locais de trabalho']):
        nome_completo = rec['fields'].get('Nome Completo', '')
        nome_norm = _normalizar_nome_busca(nome_completo)
        if not nome_norm:
            continue
        locais_ids = rec['fields'].get('Locais de trabalho') or []
        if nome_norm in exatos:
            homonimos_exatos.setdefault(nome_norm, [
                {'id': exatos[nome_norm][0], 'nome': exatos[nome_norm][1]}
            ]).append({'id': rec['id'], 'nome': nome_completo})
            continue
        exatos[nome_norm] = (rec['id'], nome_completo, locais_ids)
        primeiro, ultimo = _primeiro_ultimo_nome(nome_norm)
        if primeiro and ultimo:
            parcial_idx.setdefault((primeiro, ultimo), []).append((rec['id'], nome_completo, locais_ids))

    def _clientes_do_funcionario(locais_ids):
        clientes, vistos = [], set()
        for lid in locais_ids:
            for cid in locais_map.get(lid, []):
                if cid in vistos:
                    continue
                vistos.add(cid)
                clientes.append((cid, clientes_nome.get(cid, cid)))
        return clientes

    def resolver(nome_bruto):
        nome_norm = _normalizar_nome_busca(nome_bruto)
        if nome_norm in homonimos_exatos:
            return ('homonimos', homonimos_exatos[nome_norm])
        if nome_norm in exatos:
            func_id, nome_completo, locais_ids = exatos[nome_norm]
        else:
            primeiro, ultimo = _primeiro_ultimo_nome(nome_norm)
            candidatos = parcial_idx.get((primeiro, ultimo), []) if primeiro and ultimo else []
            if len(candidatos) > 1:
                return ('homonimos', [{'id': c[0], 'nome': c[1]} for c in candidatos])
            if len(candidatos) == 1:
                func_id, nome_completo, locais_ids = candidatos[0]
            else:
                return ('nao_encontrado', None)
        clientes = _clientes_do_funcionario(locais_ids)
        return ('ok', clientes) if clientes else ('sem_cliente', nome_completo)

    return resolver


def _processar_lista_transacoes_extrato(transacoes: list, folha_mensal: str, dry_run: bool = False) -> dict:
    """v3.16 — Metade "depois da extração" do fluxo: classifica, resolve
    Funcionário→Cliente (índice em memória) e agrupa/gera/anexa. Separado de
    _processar_extrato_beneficios para poder ser chamado direto com uma
    lista de transações JÁ EXTRAÍDA E DEDUPLICADA (por 'id_transacao') vinda
    da agregação de vários lotes de PDF — usado pelo passo de finalização
    depois que o extrato grande é fatiado com sobreposição de página."""
    resolver = _carregar_indice_funcionario_cliente()

    grupos = {}  # (cliente_id, categoria) -> {'cliente_nome':.., 'transacoes': []}
    erros = []
    ignoradas_sem_categoria = 0
    ja_resolvidos = {}

    for t in transacoes:
        categoria = _classificar_transacao_extrato(t['descricao'])
        if not categoria:
            ignoradas_sem_categoria += 1
            continue

        nome_norm = _normalizar_nome_busca(t['nome'])
        if nome_norm not in ja_resolvidos:
            ja_resolvidos[nome_norm] = resolver(t['nome'])

        status, dado = ja_resolvidos[nome_norm]
        if status == 'homonimos':
            erros.append({'nome': t['nome'], 'descricao': t['descricao'],
                           'motivo': 'homônimos — não resolvido automaticamente', 'candidatos': dado})
            continue
        if status == 'nao_encontrado':
            erros.append({'nome': t['nome'], 'descricao': t['descricao'], 'motivo': 'funcionário não encontrado'})
            continue
        if status == 'sem_cliente':
            erros.append({'nome': dado, 'descricao': t['descricao'], 'motivo': 'funcionário sem cliente vinculado'})
            continue

        for cliente_id, cliente_nome in dado:
            chave = (cliente_id, categoria)
            if chave not in grupos:
                grupos[chave] = {'cliente_nome': cliente_nome, 'transacoes': []}
            grupos[chave]['transacoes'].append(t)

    if dry_run:
        return {
            'status': 'concluido', 'dry_run': True, 'folha_mensal': folha_mensal,
            'total_transacoes_no_extrato': len(transacoes),
            'ignoradas_sem_categoria': ignoradas_sem_categoria,
            'grupos_a_gerar': [
                {'cliente': v['cliente_nome'], 'cliente_id': cid, 'categoria': _CATEGORIAS_EXTRATO[cat]['rotulo'],
                 'lancamentos': len(v['transacoes']),
                 'total_R$': round(sum(t.get('valor') or 0 for t in v['transacoes']), 2),
                 'transacoes': v['transacoes']}
                for (cid, cat), v in grupos.items()
            ],
            'erros': erros,
        }

    anexados = []
    for (cliente_id, categoria), dados in grupos.items():
        cat_info = _CATEGORIAS_EXTRATO[categoria]
        try:
            pdf_bytes = _gerar_pdf_extrato_categoria(
                dados['cliente_nome'], cat_info['rotulo'], folha_mensal, dados['transacoes'])
            fname = f'{cat_info["rotulo"]} {folha_mensal or ""} - {dados["cliente_nome"] or cliente_id}.pdf'.strip()
            _anexar_attachment(TABLE_CLIENTES, cliente_id, cat_info['campo'], pdf_bytes, fname)
            anexados.append({
                'cliente': dados['cliente_nome'], 'cliente_id': cliente_id, 'categoria': cat_info['rotulo'],
                'lancamentos': len(dados['transacoes']),
                'total_R$': round(sum(t.get('valor') or 0 for t in dados['transacoes']), 2),
            })
        except Exception as exc:
            logger.error(f'[BENEFICIOS] falha ao gerar/anexar extrato {cat_info["rotulo"]} cliente {dados["cliente_nome"]}: {exc}')
            erros.append({'cliente': dados['cliente_nome'], 'cliente_id': cliente_id,
                           'categoria': cat_info['rotulo'], 'motivo': str(exc)})

    return {
        'status': 'concluido', 'dry_run': False, 'folha_mensal': folha_mensal,
        'total_transacoes_no_extrato': len(transacoes),
        'ignoradas_sem_categoria': ignoradas_sem_categoria,
        'clientes_anexados': anexados, 'erros': erros,
    }


def _processar_extrato_beneficios(caminho_pdf: str, folha_mensal: str, dry_run: bool = False) -> dict:
    """v3.15/v3.16 — Lê o extrato bancário (1 PDF) e processa por completo
    (extração + classificação + resolução + agrupamento + geração/anexo).
    Para arquivos grandes que estouram o tempo de request do Render, ver
    /processar-beneficios-extrato-lista (v3.16) — recebe direto uma lista
    de transações já extraídas de vários lotes menores, sem reabrir PDF."""
    transacoes, ilegiveis = _extrair_transacoes_extrato_pix(caminho_pdf)
    resultado = _processar_lista_transacoes_extrato(transacoes, folha_mensal, dry_run=dry_run)
    resultado['transacoes_ilegiveis'] = ilegiveis
    return resultado


@app.route('/processar-beneficios', methods=['POST', 'OPTIONS'])
def processar_beneficios():
    """
    v3.15 — Endpoint único para os fluxos de Benefícios (o VR/VA em lote via
    planilha continua em /processar-vr-va — v2.34, estendido em v3.12 para
    aceitar .xlsx moderno além do .xls legado):
      - tipo=extrato_pix  : multipart 'pdf' (RECOMENDADO — extrato bancário
        consolidado do mês, centenas de PIX concatenados; classifica
        VR/VA, Almoço/Jantar e Horas Extras automaticamente por palavra-
        chave na descrição e ignora o resto — v3.15)
      - tipo=horas_extras : multipart 'pdf' (mestre Sicoob avulso, 1 pág/colaborador — pré-v3.15)
      - tipo=assiduidade  : multipart 'pdf' (mestre Sicoob avulso, valida R$110,00)
      - tipo=vrva_pix      : multipart 'pdf' (PIX individual avulso, fora do
        extrato — mantido para caso raro de comprovante solto; o fluxo real
        de colaborador novo sem cartão passou a vir dentro do extrato_pix)
    Body comum: folha_mensal, dry_run (default TRUE — nunca disparar direto
    sem revisar o dry_run primeiro).
    Não exige X-API-KEY (só AIRTABLE_API_KEY do servidor), igual aos outros
    fatiadores por cliente (/processar-doc-cliente, /processar-folha-ponto).
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    tipo = (request.form.get('tipo') or '').strip().lower()
    folha_mensal = (request.form.get('folha_mensal') or '').strip()
    dry_run = str(request.form.get('dry_run', 'true')).strip().lower() in ('1', 'true', 'yes', 'sim')

    if tipo not in ('horas_extras', 'assiduidade', 'vrva_pix', 'extrato_pix'):
        return jsonify({
            'status': 'erro',
            'erro': "tipo deve ser 'extrato_pix', 'horas_extras', 'assiduidade' ou 'vrva_pix' "
                    "(VR/VA em lote via planilha é em /processar-vr-va)",
        }), 400

    caminho_pdf = extrair_pdf_do_request('pdf')

    try:
        if not caminho_pdf:
            return jsonify({'status': 'erro', 'erro': 'PDF não enviado (campo "pdf")'}), 400
        if tipo == 'extrato_pix':
            resultado = _processar_extrato_beneficios(caminho_pdf, folha_mensal, dry_run=dry_run)
        elif tipo in ('horas_extras', 'assiduidade'):
            resultado = _processar_lote_sicoob(caminho_pdf, tipo, folha_mensal, dry_run=dry_run)
        else:
            resultado = _processar_vr_va_pix(caminho_pdf, folha_mensal, dry_run=dry_run)
        return jsonify(resultado)
    finally:
        try:
            os.remove(caminho_pdf)
        except OSError:
            pass


@app.route('/processar-beneficios-extrato-lista', methods=['POST', 'OPTIONS'])
def processar_beneficios_extrato_lista():
    """
    v3.16 — Passo de FINALIZAÇÃO do Extrato Bancário Consolidado para
    arquivos grandes demais pra processar em 1 request (achado real: 122
    páginas / 4,6 MB estoura o tempo de request do Render mesmo já com a
    resolução de Funcionário/Cliente em lote — o gargalo é ler/processar o
    texto do PDF inteiro, não o Airtable).

    Fluxo: fatiar o PDF grande em pedaços menores COM SOBREPOSIÇÃO de
    página (pra nenhuma transação ficar cortada na borda), chamar
    /processar-beneficios (tipo=extrato_pix, dry_run=true) em cada pedaço,
    juntar o campo 'transacoes' de cada grupo de todas as respostas,
    deduplicar por 'id_transacao' (a mesma transação completa aparece nos
    2 lotes vizinhos por causa da sobreposição) e mandar a lista final
    aqui — 1 chamada só, leve (só classifica/resolve/agrupa/gera, sem abrir
    PDF nenhum).

    Body JSON:
      {
        "transacoes": [{"id_transacao":.., "nome":.., "cpf_mascarado":..,
                         "valor":.., "descricao":.., "data":..}, ...],
        "folha_mensal": "Junho 2026",
        "dry_run": true
      }
    Não exige X-API-KEY (só AIRTABLE_API_KEY do servidor), igual ao resto
    dos fatiadores de Benefícios.
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    data = request.get_json(force=True, silent=True) or {}
    transacoes = data.get('transacoes')
    if not isinstance(transacoes, list) or not transacoes:
        return jsonify({'status': 'erro', 'erro': "'transacoes' deve ser uma lista não vazia"}), 400
    folha_mensal = (data.get('folha_mensal') or '').strip()
    dry_run = str(data.get('dry_run', True)).strip().lower() in ('1', 'true', 'yes', 'sim')

    # Dedup defensivo por id_transacao (o chamador já deve deduplicar, mas
    # não custa garantir aqui também — transação sem id_transacao passa
    # direto, sem dedup, pois não tem chave confiável pra isso).
    vistos = set()
    lista_final = []
    for t in transacoes:
        tid = t.get('id_transacao')
        if tid:
            if tid in vistos:
                continue
            vistos.add(tid)
        lista_final.append(t)

    resultado = _processar_lista_transacoes_extrato(lista_final, folha_mensal, dry_run=dry_run)
    resultado['transacoes_recebidas'] = len(transacoes)
    resultado['transacoes_apos_dedup'] = len(lista_final)
    return jsonify(resultado)


def _gerar_fila_envios_email(folha_mensal=None, folha_mensal_d2=None,
                             guias_ids=None, certidoes_ids=None, limit=None, dry_run=True,
                             cliente_id_filtro=None, force=False):
    """
    v2.29/v2.92 — Gera a fila de e-mail por CLIENTE. Para cada Cliente Ativo com
    e-mail (campo Email; fallback Email Contador do escritório), consolida num
    único envio Canal="E-mail":
      - Holerites do mês dos funcionários daquele cliente (link)
      - Extrato Mensal do mês do cliente (link)
      - FGTS Digital do mês do cliente (link)
      - Guias comuns (broadcast) — apenas se `guias_ids` for informado

    Flag "Exige D-2" (Clientes): clientes marcados recebem os documentos de
    `folha_mensal_d2` (2 meses atrás, ex.: Abril); os demais recebem
    `folha_mensal` (mês passado, ex.: Maio). Por isso buscamos os documentos de
    AMBAS as folhas e escolhemos a certa por cliente.

    cliente_id_filtro (v2.92): se informado, processa apenas esse cliente.
    force (v2.92): se True, ignora a dedup de envio pendente (útil em webhooks).
    Ignora clientes já com envio do tipo pendente (não duplica), a menos que
    force=True. Os PDFs em si são resolvidos na hora do disparo via os campos
    lookup da Envios.

    dry_run=True (padrão): não cria nada, só simula.
    """
    clientes_ctx, locais, funcionarios = _carregar_contexto_distribuicao()
    _, cliente_pendentes = _carregar_envios_pendentes(TIPO_ENVIO_EMAIL_CLIENTE)
    guias_ids = guias_ids or []
    certidoes_ids = certidoes_ids or []
    folhas_alvo = [f for f in (folha_mensal, folha_mensal_d2) if f]

    # Recibos de Pagamento e Folha de Ponto por cliente (via funcionários)
    recibos_por_cli = _recibos_pdfs_por_cliente((clientes_ctx, locais, funcionarios))
    ponto_por_cli   = _ponto_pdfs_por_cliente((clientes_ctx, locais, funcionarios))

    # Holerites por folha → por cliente: {folha: {cliente_id: [holerite_ids]}}
    holerites_idx = {}
    for folha in folhas_alvo:
        dest = holerites_idx.setdefault(folha, {})
        filtro = f'{{Folha Mensal}}="{folha}"'
        for rec in _at_listar_todos(TABLE_HOL, ['Holerite', 'Funcionário', 'Folha Mensal', 'PDF HOLERITE'], filtro):
            if not rec['fields'].get('PDF HOLERITE'):
                continue
            c = _classificar_holerite_distribuicao(rec, clientes_ctx, locais, funcionarios)
            if c['cliente_id']:
                dest.setdefault(c['cliente_id'], []).append(rec['id'])

    # Extratos / FGTS por folha → por cliente (mesma estrutura)
    def _docs_idx(table):
        idx = {}
        for folha in folhas_alvo:
            dest = idx.setdefault(folha, {})
            filtro = f'{{Folha Mensal}}="{folha}"'
            for rec in _at_listar_todos(table, ['Folha Mensal', 'Cliente', 'PDF ARQUIVO'], filtro):
                f = rec['fields']
                if not f.get('PDF ARQUIVO'):
                    continue
                for cid in (f.get('Cliente') or []):
                    cid = cid['id'] if isinstance(cid, dict) else cid
                    dest.setdefault(cid, []).append(rec['id'])
        return idx

    extratos_idx = _docs_idx(TABLE_EXTRATO)
    fgts_idx = _docs_idx(TABLE_FGTS)

    # Clientes Ativos + e-mail (com fallback p/ escritório) + flag Exige D-2
    envios, ignorados = [], []
    clientes = _at_listar_todos(
        TABLE_CLIENTES, ['Nome', 'Cliente', 'Email', 'Email Contador', 'Status', 'Exige D-2'])
    if limit:
        clientes = clientes[:limit]

    for rec in clientes:
        cid = rec['id']
        f = rec['fields']
        nome = f.get('Nome') or f.get('Cliente') or cid
        status = f.get('Status')
        status_nome = status['name'] if isinstance(status, dict) else status
        if status_nome != 'Ativo':
            continue
        if cliente_id_filtro and cid != cliente_id_filtro:
            continue
        if not force and cid in cliente_pendentes:
            ignorados.append({'cliente': nome, 'motivo': 'ja_possui_envio_email_pendente'})
            continue

        exige_d2 = bool(f.get('Exige D-2'))
        if exige_d2 and not folha_mensal_d2:
            ignorados.append({'cliente': nome, 'motivo': 'exige_d2_sem_folha_d2_informada'})
            continue
        folha_cli = folha_mensal_d2 if exige_d2 else folha_mensal

        email = f.get('Email')
        if not email:
            contador = f.get('Email Contador') or []
            email = contador[0] if contador else None
        hol = holerites_idx.get(folha_cli, {}).get(cid, [])
        ext = extratos_idx.get(folha_cli, {}).get(cid, [])
        fgts = fgts_idx.get(folha_cli, {}).get(cid, [])

        motivos = []
        if not email:
            motivos.append('email_ausente')
        if not (hol or ext or fgts or guias_ids or certidoes_ids):
            motivos.append('sem_documentos_do_mes')
        if motivos:
            ignorados.append({'cliente': nome, 'motivo': 'nao_pronto', 'motivos': motivos})
            continue

        item = {
            'cliente': nome, 'cliente_id': cid, 'email': email,
            'folha': folha_cli, 'd2': exige_d2,
            'qtd_holerites': len(hol), 'qtd_extratos': len(ext),
            'qtd_fgts': len(fgts), 'qtd_guias': len(guias_ids),
        }
        if dry_run:
            item['acao'] = 'criaria_envio'
        else:
            campos = {
                F_ENVIO_TIPO: TIPO_ENVIO_EMAIL_CLIENTE,
                F_ENVIO_CANAL: 'E-mail',
                F_ENVIO_DEST: email,
                F_ENVIO_EMAIL: email,
                F_ENVIO_CLIENTE: [cid],
                F_ENVIO_TEXTO: _corpo_maggie(nome, folha_cli),   # saudação (Bia) + competência
            }
            if hol:
                campos[F_ENVIO_HOLERITES] = hol
            if ext:
                campos[F_ENVIO_EXTRATO] = ext
            if fgts:
                campos[F_ENVIO_FGTS] = fgts
            if guias_ids:
                campos[F_ENVIO_GUIAS] = guias_ids
            if certidoes_ids:
                campos[F_ENVIO_CERTIDOES] = certidoes_ids
            # Recibos de Pagamento + Folha de Ponto (URLs diretas via F_ENVIO_ARQUIVOS)
            arquivos_extras = (
                [{'url': u, 'filename': fn} for fn, u in recibos_por_cli.get(cid, [])] +
                [{'url': u, 'filename': fn} for fn, u in ponto_por_cli.get(cid, [])]
            )
            if arquivos_extras:
                campos[F_ENVIO_ARQUIVOS] = arquivos_extras
            envio_id, _ = _criar_envio_documento(campos)
            item['acao'] = 'envio_criado'
            item['envio_id'] = envio_id
        envios.append(item)

    return {
        'status': 'concluido', 'dry_run': dry_run,
        'folha_mensal': folha_mensal, 'folha_mensal_d2': folha_mensal_d2,
        'total_envios': len(envios), 'total_ignorados': len(ignorados),
        'envios': envios, 'ignorados': ignorados,
    }


@app.route('/gerar-fila-envios-email', methods=['POST', 'OPTIONS'])
def gerar_fila_envios_email():
    """
    v2.29 — Gera a fila de e-mail (1 pacote por Cliente Ativo). Body JSON:
      - folha_mensal: mês normal (D-1), ex. "Maio 2026"
      - folha_mensal_d2: mês de 2 atrás (D-2), ex. "Abril 2026" — usado só pelos
        clientes com a flag "Exige D-2" marcada no Airtable.
      - guias_ids: lista de record ids da aba Guias e Comprovantes a anexar em
        TODOS os clientes (broadcast). Opcional.
      - limit / dry_run (padrão true).
    Protegido por X-API-KEY (EMAIL_WEBHOOK_KEY).
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401
    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    body = request.get_json(silent=True) or {}
    folha_mensal = body.get('folha_mensal')
    folha_mensal_d2 = body.get('folha_mensal_d2')
    guias_ids = body.get('guias_ids') or []
    certidoes_ids = body.get('certidoes_ids') or []
    limit = body.get('limit')
    try:
        limit = int(limit) if limit else None
    except (ValueError, TypeError):
        limit = None
    dry_run = body.get('dry_run', True)

    resultado = _gerar_fila_envios_email(
        folha_mensal=folha_mensal, folha_mensal_d2=folha_mensal_d2,
        guias_ids=guias_ids, certidoes_ids=certidoes_ids, limit=limit, dry_run=dry_run)
    return jsonify(resultado)


def _baixar_attachment_bytes(url: str) -> bytes:
    r = requests.get(url, timeout=90)
    r.raise_for_status()
    return r.content


def _smtp_enviar_email(destinatario: str, assunto: str, corpo_texto: str, anexos: list):
    """Envia 1 e-mail via Resend API (HTTPS/443) com fallback para SMTP.
    anexos: lista de (filename, bytes).
    Prioridade: RESEND_API_KEY presente → Resend; senão → SMTP clássico."""

    RESEND_KEY = os.environ.get('RESEND_API_KEY', '')

    if RESEND_KEY:
        # ── Resend API (funciona no Render free tier — usa HTTPS porta 443) ──
        destinatarios = [d.strip() for d in destinatario.split(',') if d.strip()]
        payload = {
            'from':    f'{EMAIL_SENDER_NOME} <{EMAIL_SENDER or "noreply@magnataservicos.com.br"}>',
            'to':      destinatarios,
            'subject': assunto,
            'text':    corpo_texto,
        }
        if anexos:
            payload['attachments'] = [
                {'filename': fname, 'content': list(conteudo)}
                for fname, conteudo in anexos
            ]
        r = requests.post(
            'https://api.resend.com/emails',
            headers={'Authorization': f'Bearer {RESEND_KEY}',
                     'Content-Type': 'application/json'},
            json=payload,
            timeout=120,
        )
        if not r.ok:
            raise RuntimeError(f'Resend erro {r.status_code}: {r.text[:300]}')
        return  # sucesso

    # ── Fallback SMTP ─────────────────────────────────────────────────────────
    import smtplib
    import mimetypes
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    from email.utils import formataddr

    if not EMAIL_SENDER or not EMAIL_SENDER_PASSWORD:
        raise RuntimeError('Configure RESEND_API_KEY (recomendado) ou EMAIL_SENDER+EMAIL_SENDER_PASSWORD')

    msg = MIMEMultipart()
    msg['From'] = formataddr((EMAIL_SENDER_NOME, EMAIL_SENDER))
    msg['To'] = destinatario
    msg['Subject'] = assunto
    msg.attach(MIMEText(corpo_texto, 'plain', 'utf-8'))
    for fname, conteudo in anexos:
        ctype, _enc = mimetypes.guess_type(fname)
        maintype, subtype = (ctype.split('/', 1) if ctype else ('application', 'octet-stream'))
        part = MIMEBase(maintype, subtype)
        part.set_payload(conteudo)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=fname)
        msg.attach(part)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=90) as server:
        server.ehlo()
        server.starttls()
        server.login(EMAIL_SENDER, EMAIL_SENDER_PASSWORD)
        server.send_message(msg)


def _fmt_quando(valor):
    """Formata um ISO datetime/date -> 'dd/mm/aaaa às HH:MM' (ou só a data),
    sempre no horário de Brasília.

    O Airtable NORMALIZA campo de data/hora para UTC e devolve com 'Z'
    (ex.: '2026-08-19T02:06:00.000Z'), mesmo quando a gravação foi feita com
    offset -03:00 — é o que o app faz em F_ASS_DATA_ASSINATURA e
    F_ENVIO_ENVIADO_EM. Sem converter na leitura, a tela mostrava o relógio
    UTC como se fosse local: 3 horas adiantado e, perto da virada, no dia
    seguinte. Achado de 20/08/2026 — para a MESMA assinatura, o carimbo do
    PDF dizia "18/08/2026 23:06" e a tela de sucesso dizia "19/08/2026 às
    02:06".

    Valor COM fuso (venha 'Z' ou '-03:00') é convertido para BRT — o que
    torna esta função correta tanto para o que vem do Airtable quanto para o
    `agora_brt` passado direto, que já está em -03:00 e sai inalterado.
    Valor SEM fuso é formatado como está: não há como saber a que zona
    pertence, e chutar seria pior do que preservar o comportamento antigo.
    """
    if not valor:
        return 'data não registrada'
    try:
        if 'T' in valor:
            dt = datetime.fromisoformat(valor)
            if dt.tzinfo is not None:
                dt = dt.astimezone(timezone(timedelta(hours=-3)))
            return dt.strftime('%d/%m/%Y às %H:%M')
        return datetime.strptime(valor[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
    except Exception:
        return str(valor)


def _mapa_func_cliente(locais, funcionarios):
    """{funcionario_id: cliente_id} via Funcionário->Locais de trabalho->Cliente."""
    func_cliente = {}
    for fid, dados in funcionarios.items():
        for lid in (dados.get('locais_ids') or []):
            loc = locais.get(lid)
            if loc and (loc.get('cliente_ids')):
                func_cliente[fid] = loc['cliente_ids'][0]
                break
    return func_cliente


def _protocolos_entrega_por_cliente(contexto=None):
    """Protocolo de Entrega (segurança jurídica): {cliente_id: [(nome, enviado,
    lido)]} a partir dos envios COMBINADOS (Holerite+Ponto) já ENVIADOS por
    WhatsApp. `enviado` usa o timestamp exato ("Enviado em"); `lido` vem do
    "Recibo Lido em" (quando o colaborador abriu o link de recibo) — '' se ainda
    não lido."""
    _clientes, locais, funcionarios = contexto or _carregar_contexto_distribuicao()
    func_cliente = _mapa_func_cliente(locais, funcionarios)
    out = {}
    campos = ['Tipo', 'Status', 'Canal', 'Funcionário(s) Vinculado(s)',
              'Data envio', 'Enviado em', 'Recibo Lido em']
    for rec in _at_listar_todos(TABLE_ENVIOS, campos):
        f = rec['fields']
        if (f.get('Tipo') != TIPO_ENVIO_COMBINADO or f.get('Status') != 'Enviado'
                or f.get('Canal') != 'WhatsApp'):
            continue
        enviado = _fmt_quando(f.get('Enviado em') or f.get('Data envio'))
        lido = _fmt_quando(f.get('Recibo Lido em')) if f.get('Recibo Lido em') else ''
        for fl in (f.get('Funcionário(s) Vinculado(s)') or []):
            fid = fl['id'] if isinstance(fl, dict) else fl
            nome = fl['name'] if isinstance(fl, dict) else funcionarios.get(fid, {}).get('nome', fid)
            cid = func_cliente.get(fid)
            if cid:
                out.setdefault(cid, []).append((nome, enviado, lido))
    return out


def _ponto_pdfs_por_cliente(contexto=None):
    """{cliente_id: [(filename, url)]} dos Cartões de Ponto (último PDF Folha
    Ponto de cada colaborador), agrupados pelo cliente do colaborador."""
    _clientes, locais, funcionarios = contexto or _carregar_contexto_distribuicao()
    func_cliente = _mapa_func_cliente(locais, funcionarios)
    out = {}
    for fid, dados in funcionarios.items():
        anexos = dados.get('pdf_folha_ponto') or []
        if not anexos:
            continue
        pdf = anexos[-1]
        cid = func_cliente.get(fid)
        if cid and isinstance(pdf, dict) and pdf.get('url'):
            out.setdefault(cid, []).append((pdf.get('filename', 'cartao_ponto.pdf'), pdf['url']))
    return out


def _recibos_pdfs_por_cliente(contexto=None):
    """{cliente_id: [(filename, url)]} dos Recibos de Pagamento (salário,
    assiduidade, almoço/janta) fatiados por colaborador, agrupados pelo cliente
    do colaborador. v2.33."""
    _clientes, locais, funcionarios = contexto or _carregar_contexto_distribuicao()
    func_cliente = _mapa_func_cliente(locais, funcionarios)
    out = {}
    for fid, dados in funcionarios.items():
        cid = func_cliente.get(fid)
        if not cid:
            continue
        for a in (dados.get('recibos_pagamento') or []):
            if isinstance(a, dict) and a.get('url'):
                out.setdefault(cid, []).append((a.get('filename', 'recibo.pdf'), a['url']))
    return out


# ── Fix v3.01 — Recibos de Assiduidade via "Outros documentos" ──────────────
# Caminho alternativo a _recibos_pdfs_por_cliente(): em vez de depender de
# Funcionário->Locais de trabalho->Cliente (sujeito a Local atual em vez do
# Local da competência, e sem filtro de mês — ver achados da auditoria de
# 2026-07-21), lê diretamente a tabela "Outros documentos", que já guarda o
# vínculo correto e definitivo com o Envio de Documentos (setado no momento
# em que o recibo foi fatiado e classificado), sem nenhuma resolução de
# Funcionário/Local/Cliente em tempo de leitura.

TABLE_OUTROS_DOCS = 'tblanxELlj11HjJEV'
F_OD_DOCUMENTO = 'Documento'
F_OD_PDF = 'PDF ARQUIVO'
F_OD_ENVIO = 'Envios de Documentos 2'


def _outros_documentos_por_envio(competencia_substr):
    """Somente leitura. Lê 'Outros documentos' e agrupa por Envio de
    Documentos vinculado, exigindo PDF anexado + vínculo de Envio, e
    filtrando pela substring de competência no nome do 'Documento' (ex.:
    'Junho 2026'). Não usa Funcionário/Local/Cliente em nenhum momento —
    o Envio já vem pronto no próprio registro.

    Retorna (por_envio, rejeitados):
      por_envio:  {envio_id: [{'filename','url','size','documento','source_record_id'}]}
      rejeitados: [{'record_id','documento','motivo'}]
    """
    registros = _at_listar_todos(TABLE_OUTROS_DOCS, [
        F_OD_DOCUMENTO, F_OD_PDF, F_OD_ENVIO,
    ])
    por_envio = {}
    rejeitados = []
    for rec in registros:
        f = rec['fields']
        doc_nome = f.get(F_OD_DOCUMENTO, '') or ''
        if competencia_substr not in doc_nome:
            continue  # fora do escopo desta competência — não é rejeição, é filtro normal
        pdf = f.get(F_OD_PDF) or []
        envio_links = f.get(F_OD_ENVIO) or []
        motivo = None
        if not pdf:
            motivo = 'sem_pdf_anexado'
        elif not envio_links:
            motivo = 'sem_envio_vinculado'
        if motivo:
            rejeitados.append({'record_id': rec['id'], 'documento': doc_nome, 'motivo': motivo})
            continue
        envio_id = envio_links[0]['id'] if isinstance(envio_links[0], dict) else envio_links[0]
        att = pdf[0]
        por_envio.setdefault(envio_id, []).append({
            'filename': att.get('filename'),
            'url': att.get('url'),
            'size': att.get('size'),
            'documento': doc_nome,
            'source_record_id': rec['id'],
        })
    return por_envio, rejeitados


def _normalizar_filename(nome):
    """Normaliza um nome de arquivo para comparação (espaços nas pontas +
    caixa). Não mexe em acentuação — os nomes já vêm consistentes de quem
    gera o PDF."""
    return (nome or '').strip().casefold()


def _classificar_novos_vs_existentes(arquivos_existentes, candidatos):
    """Decide o que fazer com cada candidato comparado aos anexos já
    presentes em 'Arquivos' — SEM usar a URL como parte da chave (URLs de
    anexo do Airtable são assinadas e mudam a cada leitura; usá-las em
    dedup falha silenciosamente e causa duplicação real — foi o que
    aconteceu com Teodolino/João Batista na Unimed em 2026-07-21).

    Chave de dedup: filename normalizado. Quando o nome já existe,
    confirma que é de fato o MESMO arquivo comparando o tamanho (bytes) —
    campo estável, presente tanto no anexo já hospedado quanto no
    candidato. Nome igual + tamanho igual = duplicata real (não
    adiciona). Nome igual + tamanho diferente = conflito real (não
    adiciona, registra para revisão manual).

    Retorna (a_adicionar, ja_existentes, conflitos) — três listas de
    candidatos.
    """
    existentes_por_nome = {}
    for a in arquivos_existentes:
        chave = _normalizar_filename(a.get('filename'))
        existentes_por_nome.setdefault(chave, []).append(a)

    a_adicionar, ja_existentes, conflitos = [], [], []
    vistos_nesta_rodada = set()
    for c in candidatos:
        chave = _normalizar_filename(c['filename'])
        if chave in vistos_nesta_rodada:
            conflitos.append({**c, 'motivo': 'duplicado_dentro_dos_proprios_candidatos'})
            continue
        vistos_nesta_rodada.add(chave)
        existentes_mesmo_nome = existentes_por_nome.get(chave)
        if not existentes_mesmo_nome:
            a_adicionar.append(c)
            continue
        tamanho_bate = any(
            e.get('size') is not None and e.get('size') == c.get('size')
            for e in existentes_mesmo_nome
        )
        if tamanho_bate:
            ja_existentes.append(c)
        else:
            conflitos.append({**c, 'motivo': 'mesmo_nome_tamanho_diferente'})
    return a_adicionar, ja_existentes, conflitos


def _dry_run_outros_documentos_para_envios(competencia_substr):
    """Somente leitura. Simula o que _aplicar_outros_documentos_no_envio()
    faria para CADA Envio afetado, sem escrever nada. Dedup por filename
    normalizado + tamanho (ver _classificar_novos_vs_existentes) — não usa
    URL como chave."""
    por_envio, rejeitados = _outros_documentos_por_envio(competencia_substr)
    if not por_envio:
        return {'envios': [], 'rejeitados': rejeitados}

    envio_ids = list(por_envio.keys())
    envios_atuais = {}
    for envio_id in envio_ids:
        _at_throttle()
        r = requests.get(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ENVIOS}/{envio_id}',
            headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
            timeout=30,
        )
        r.raise_for_status()
        envios_atuais[envio_id] = r.json().get('fields', {})

    resultado = []
    for envio_id, novos in por_envio.items():
        atuais = envios_atuais.get(envio_id, {})
        arquivos_existentes = atuais.get('Arquivos') or []
        a_adicionar, ja_existentes, conflitos = _classificar_novos_vs_existentes(
            arquivos_existentes, novos)
        cliente_nomes = [c.get('name') if isinstance(c, dict) else c
                         for c in (atuais.get('Cliente') or [])]
        resultado.append({
            'envio_id': envio_id,
            'envio_nome': atuais.get('Name'),
            'cliente': cliente_nomes,
            'status_atual': atuais.get('Status'),
            'qtd_recibos_novos': len(novos),
            'qtd_a_adicionar': len(a_adicionar),
            'filenames_a_adicionar': [n['filename'] for n in a_adicionar],
            'qtd_ja_existentes_no_envio': len(arquivos_existentes),
            'duplicados_evitados': [n['filename'] for n in ja_existentes],
            'conflitos': conflitos,
        })
    return {'envios': resultado, 'rejeitados': rejeitados}


def _aplicar_outros_documentos_no_envio(envio_id, competencia_substr):
    """ESCREVE. Aplica, para UM único envio_id, os recibos de
    'Outros documentos' filtrados por competência — preserva os anexos já
    existentes em 'Arquivos' e só adiciona os que não estão lá. Dedup por
    filename normalizado + tamanho (ver _classificar_novos_vs_existentes),
    NÃO por URL. Se houver conflito real (mesmo nome, tamanho diferente),
    NÃO escreve nada neste Envio e retorna o conflito para revisão manual."""
    por_envio, rejeitados = _outros_documentos_por_envio(competencia_substr)
    novos = por_envio.get(envio_id, [])
    if not novos:
        return {'envio_id': envio_id, 'status': 'nada_a_fazer', 'rejeitados': rejeitados}

    _at_throttle()
    r = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ENVIOS}/{envio_id}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        timeout=30,
    )
    r.raise_for_status()
    arquivos_existentes = r.json().get('fields', {}).get('Arquivos') or []

    a_adicionar, ja_existentes, conflitos = _classificar_novos_vs_existentes(
        arquivos_existentes, novos)

    if conflitos:
        return {
            'envio_id': envio_id, 'status': 'conflito_nao_aplicado',
            'conflitos': conflitos, 'rejeitados': rejeitados,
        }
    if not a_adicionar:
        return {
            'envio_id': envio_id, 'status': 'sem_novidade_apos_dedup',
            'duplicados_evitados': [n['filename'] for n in ja_existentes],
            'rejeitados': rejeitados,
        }

    novo_valor = (
        [{'id': a['id']} for a in arquivos_existentes]
        + [{'url': n['url'], 'filename': n['filename']} for n in a_adicionar]
    )
    _at_throttle()
    r2 = requests.patch(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ENVIOS}/{envio_id}',
        headers=_at_headers(),
        json={'fields': {'Arquivos': novo_valor}, 'typecast': True},
        timeout=30,
    )
    _raise_for_airtable(r2, f'aplicar outros documentos no envio {envio_id}')
    return {
        'envio_id': envio_id,
        'status': 'aplicado',
        'adicionados': [a['filename'] for a in a_adicionar],
        'duplicados_evitados': [n['filename'] for n in ja_existentes],
        'total_apos': len(novo_valor),
        'rejeitados': rejeitados,
    }


def _formatar_protocolo(linhas):
    """Monta o texto do Protocolo de Entrega para o corpo do e-mail."""
    if not linhas:
        return ''
    corpo = [
        '', '', '====== PROTOCOLO DE ENTREGA ======',
        'Comprovação de envio dos documentos individuais (Holerite e Folha de '
        'Ponto) aos colaboradores deste condomínio, via WhatsApp:', '',
    ]
    # "LIDO em" só aparece quando há leitura registrada (a partir do mês em que o
    # link de recibo passa a acompanhar o WhatsApp). Sem isso, fica só o ENVIO.
    for nome, enviado, lido in sorted(set(linhas)):
        leitura = f' · LIDO em {lido} ✓' if lido else ''
        corpo.append(f'  • {nome} — enviado em {enviado}{leitura}')
    corpo += ['', 'Protocolo gerado automaticamente por Magnata Portaria e Serviços.']
    return '\n'.join(corpo)


def _corpo_maggie(cli_nome, competencia):
    """Corpo do e-mail — estilo humanizado assinado pela Bia. v2.34."""
    comp = f' referentes à competência {competencia}' if competencia else ' do mês'
    cond = f' referente ao condomínio {cli_nome}' if cli_nome else ''
    return (
        'Olá, tudo bem?\n\n'
        f'Segue em anexo a documentação{comp}{cond}.\n\n'
        'Qualquer dúvida, é só chamar — estou à disposição!'
    )


def _manifesto_anexos(nomes):
    """Lista numerada dos anexos incluídos (transparência p/ o síndico conferir)."""
    nomes = [n for n in nomes if n]
    if not nomes:
        return ''
    linhas = ['', '', '📎 DOCUMENTOS ANEXADOS NESTE E-MAIL:']
    for i, n in enumerate(sorted(set(nomes)), 1):
        linhas.append(f'  {i}. {n}')
    return '\n'.join(linhas)


def _rodape_email():
    """Assinatura + aviso de confidencialidade/LGPD."""
    return (
        '\n\n'
        'Um abraço,\n'
        'Bia\n'
        'Magnata Portaria e Serviços\n'
        '(15) 98161-2263  ·  contato@magnataservicos.com.br\n\n'
        '─────────────────────────────────────────────\n'
        'AVISO DE CONFIDENCIALIDADE (LGPD – Lei nº 13.709/2018): Este e-mail e seus '
        'anexos contêm informações confidenciais e dados pessoais destinados '
        'EXCLUSIVAMENTE ao condomínio/empresa identificado. É vedada a cópia, o uso '
        'ou a divulgação a terceiros não autorizados. Caso tenha recebido esta '
        'mensagem por engano, por favor exclua-a e notifique o remetente.'
    )


def _disparar_fila_email(limit=None, dry_run=True, email_teste=None):
    """
    v2.29/v2.30 — Lê os envios Canal="E-mail" / Tipo="Pacote Mensal Cliente" em
    Status "Preparando", monta o e-mail consolidado (todos os PDFs dos campos
    lookup) + o PROTOCOLO DE ENTREGA (colaboradores + data/hora do WhatsApp),
    envia via SMTP e marca "Enviado".

    Segurança:
      - dry_run=True (padrão): não envia nada, só lista o que enviaria.
      - email_teste: redireciona TODOS os envios p/ esse e-mail (teste seguro).
      - limit: limita quantos envios processa (ex.: 1 p/ teste).
    """
    if not dry_run and (not EMAIL_SENDER or not EMAIL_SENDER_PASSWORD):
        return {'status': 'erro', 'erro': 'EMAIL_SENDER/EMAIL_SENDER_PASSWORD não configurados no ambiente'}

    contexto = _carregar_contexto_distribuicao()
    protocolos = _protocolos_entrega_por_cliente(contexto)     # {cliente_id: [(nome, enviado, lido)]}
    ponto_por_cliente = _ponto_pdfs_por_cliente(contexto)      # {cliente_id: [(filename, url)]}
    recibos_por_cliente = _recibos_pdfs_por_cliente(contexto)  # {cliente_id: [(filename, url)]}
    campos = ['Tipo', 'Status', 'Canal', 'Email', 'Destinatário', 'Cliente',
              'Texto do Email'] + ENVIO_LOOKUP_PDFS
    enviados, falhas, ignorados = [], [], []

    for rec in _at_listar_todos(TABLE_ENVIOS, campos):
        f = rec['fields']
        if (f.get('Tipo') != TIPO_ENVIO_EMAIL_CLIENTE or f.get('Status') != 'Preparando'
                or f.get('Canal') != 'E-mail'):
            continue
        if limit is not None and (len(enviados) + len(falhas)) >= limit:
            break

        destino = email_teste or f.get('Email') or f.get('Destinatário')
        cli = f.get('Cliente') or []
        cli_nome = cli[0]['name'] if cli and isinstance(cli[0], dict) else None
        cli_id = cli[0]['id'] if cli and isinstance(cli[0], dict) else None
        protocolo_linhas = protocolos.get(cli_id, [])

        anexos_meta = []
        for campo in ENVIO_LOOKUP_PDFS:
            for a in (f.get(campo) or []):
                if isinstance(a, dict) and a.get('url'):
                    anexos_meta.append((a.get('filename', 'documento.pdf'), a['url']))
        # Cartões de Ponto + Recibos dos colaboradores deste cliente (Funcionários)
        anexos_meta += ponto_por_cliente.get(cli_id, [])
        anexos_meta += recibos_por_cliente.get(cli_id, [])

        item = {'envio_id': rec['id'], 'cliente': cli_nome,
                'destinatario': destino, 'qtd_anexos': len(anexos_meta),
                'qtd_colaboradores_protocolo': len(protocolo_linhas)}

        if not destino or '@' not in destino:
            item['motivo'] = 'email_invalido'
            ignorados.append(item)
            continue
        if not anexos_meta:
            item['motivo'] = 'sem_anexos'
            ignorados.append(item)
            continue
        if dry_run:
            item['acao'] = 'enviaria'
            enviados.append(item)
            continue

        try:
            anexos = [(fn, _baixar_attachment_bytes(u)) for fn, u in anexos_meta]
            assunto = (f'Documentos da Folha de Pagamento'
                       f'{(" — " + cli_nome) if cli_nome else ""} — Grupo Magnata')
            # Corpo = saudação (Bia, vem do gerador c/ competência) + manifesto
            # dos anexos + Protocolo de Entrega + rodapé/LGPD.
            corpo_base = f.get('Texto do Email') or _corpo_maggie(cli_nome, '')
            corpo = (corpo_base
                     + _manifesto_anexos([fn for fn, _ in anexos_meta])
                     + _formatar_protocolo(protocolo_linhas)
                     + _rodape_email())
            _smtp_enviar_email(destino, assunto, corpo, anexos)
            _marcar_envio_status(rec['id'], 'Enviado')
            item['acao'] = 'enviado'
            enviados.append(item)
        except Exception as exc:
            logger.error(f'[EMAIL] falha no envio {rec["id"]} ({cli_nome}): {exc}')
            try:
                _marcar_envio_status(rec['id'], 'Erro', erro=str(exc))
            except Exception:
                pass
            item['acao'] = 'falhou'
            item['erro'] = str(exc)
            falhas.append(item)

    return {
        'status': 'concluido', 'dry_run': dry_run, 'email_teste': email_teste,
        'total_enviados': len(enviados), 'total_falhas': len(falhas),
        'total_ignorados': len(ignorados),
        'enviados': enviados, 'falhas': falhas, 'ignorados': ignorados,
    }


@app.route('/disparar-fila-email', methods=['POST', 'OPTIONS'])
def disparar_fila_email():
    """
    v2.29 — Dispara via SMTP os pacotes de e-mail (Tipo="Pacote Mensal Cliente",
    Status="Preparando"). Body JSON:
      - dry_run: true (padrão) só simula
      - limit: limita a quantidade (ex.: 1 para teste)
      - email_teste: redireciona TODOS os envios p/ 1 e-mail (teste seguro)
    Protegido por X-API-KEY (EMAIL_WEBHOOK_KEY).
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401

    body = request.get_json(silent=True) or {}
    limit = body.get('limit')
    try:
        limit = int(limit) if limit else None
    except (ValueError, TypeError):
        limit = None
    dry_run = body.get('dry_run', True)
    email_teste = body.get('email_teste')

    resultado = _disparar_fila_email(limit=limit, dry_run=dry_run, email_teste=email_teste)
    return jsonify(resultado)


def _processar_guia_comum(caminho_pdf: str, tipo_guia: str, nome_doc: str, folha_mensal: str):
    """Arquiva 1 Guia/Comprovante (documento COMUM — INSS, DCTFWeb, PIS/COFINS)
    como 1 registro na aba Guias e Comprovantes, anexando o PDF. Não fatia nem
    casa cliente (é broadcast: o mesmo doc vai p/ todos via guias_ids)."""
    campos = {F_GUIA_STATUS: 'Recebido'}
    if tipo_guia:
        campos[F_GUIA_TIPO] = tipo_guia
    if nome_doc:
        campos[F_GUIA_NOME] = nome_doc
    rec_id = _criar_registro(TABLE_GUIAS, campos)
    fname = nome_doc or f'Guia {tipo_guia or ""} {folha_mensal or ""}'.strip() or 'guia.pdf'
    if not fname.lower().endswith('.pdf'):
        fname += '.pdf'
    with open(caminho_pdf, 'rb') as fh:
        _anexar_attachment(TABLE_GUIAS, rec_id, F_GUIA_PDF, fh.read(), fname)
    return {'status': 'concluido', 'record_id': rec_id, 'tipo': tipo_guia,
            'nome_documento': nome_doc, 'folha_mensal': folha_mensal}


@app.route('/processar-guia', methods=['POST', 'OPTIONS'])
def processar_guia():
    """
    v2.29/v2.31 — Arquiva 1 Guia/Comprovante COMUM (INSS, DCTFWeb, FGTS, VR/VA…)
    na aba Guias e Comprovantes (documento broadcast). multipart/form-data:
      - 1+ arquivos (qualquer tipo: PDF, XLS…) — qualquer nome de campo
      - tipo: opcional (ex.: "DCTFWeb", "FGTS", "VR Benefícios")
      - nome: opcional (Nome documento)
      - campo: 'guia' (padrão) ou 'comprovante' — onde anexar (PDF GUIA vs PDF COMPROVANTE)
    Protegido por X-API-KEY (EMAIL_WEBHOOK_KEY) — Macro 6A, correção de
    auditoria (achado #4): mesma lacuna e mesma correção de
    /processar-doc-cliente (ver docstring lá). Retorna record_id — use em
    guias_ids no broadcast.
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401
    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    tipo_guia = (request.form.get('tipo') or '').strip()
    nome_doc = (request.form.get('nome') or '').strip()
    campo = (request.form.get('campo') or 'guia').strip().lower()
    field_id = F_GUIA_COMPROV if campo == 'comprovante' else F_GUIA_PDF

    # Coleta TODOS os arquivos enviados (qualquer tipo), em memória (são leves).
    arquivos = []
    for fs in request.files.values():
        dados = fs.read()
        if dados:
            arquivos.append((fs.filename or 'arquivo', dados))
    if not arquivos:
        return jsonify({'status': 'erro', 'erro': 'nenhum arquivo enviado'}), 400

    campos = {F_GUIA_STATUS: 'Recebido'}
    if tipo_guia:
        campos[F_GUIA_TIPO] = tipo_guia
    if nome_doc:
        campos[F_GUIA_NOME] = nome_doc
    rec_id = _criar_registro(TABLE_GUIAS, campos)

    anexados, falhas = [], []
    for fname, dados in arquivos:
        ct = (mimetypes.guess_type(fname)[0] or 'application/octet-stream')
        try:
            _anexar_attachment(TABLE_GUIAS, rec_id, field_id, dados, fname, ct)
            anexados.append(fname)
        except Exception as exc:
            logger.error(f'[GUIA] falha ao anexar {fname}: {exc}')
            falhas.append({'arquivo': fname, 'erro': str(exc)})

    return jsonify({'status': 'concluido', 'record_id': rec_id, 'campo': campo,
                    'tipo': tipo_guia, 'anexos': anexados, 'falhas': falhas})


def _processar_recibos_master(caminho_pdf, tipo, folha_mensal):
    """v2.33 — Fatia um mestre de RECIBOS (1 colaborador por página) por CPF e
    anexa o recibo de cada colaborador no campo 'Recibos Pagamento' do
    Funcionário. `tipo`: rótulo (ex.: 'Salário', 'Assiduidade', 'Almoço e Janta')."""
    mapa, total = construir_mapa_cpf(caminho_pdf)
    anexados, nao_vinculados = [], []
    for cpf, dados in mapa.items():
        if str(cpf).startswith('sem_cpf_'):
            nao_vinculados.append({'motivo': 'cpf_nao_extraido',
                                   'paginas': [p + 1 for p in dados['paginas']]})
            continue
        func_id, func_nome = buscar_funcionario_por_cpf(cpf)
        if not func_id:
            nao_vinculados.append({'cpf': cpf, 'nome': dados.get('nome'),
                                   'motivo': 'funcionario_nao_encontrado'})
            continue
        try:
            pdf_bytes = extrair_pdf_colaborador(caminho_pdf, dados['paginas'])
            nome_final = dados.get('nome') or func_nome or cpf
            fname = (f'Recibo {tipo} {folha_mensal or ""} - '
                     f'{nome_final}.pdf').replace('  ', ' ').strip()
            _anexar_attachment(TABLE_FUNC, func_id, F_FUNC_RECIBOS, pdf_bytes, fname)
            anexados.append({'cpf': cpf, 'nome': nome_final, 'func_id': func_id,
                             'paginas': [p + 1 for p in dados['paginas']]})
        except Exception as exc:
            logger.error(f'[RECIBOS] falha {_mascarar_cpf(cpf)}: {exc}')
            nao_vinculados.append({'cpf': cpf, 'erro': str(exc)})
    return {
        'status': 'concluido', 'tipo': tipo, 'folha_mensal': folha_mensal,
        'total_paginas': total, 'recibos_anexados': len(anexados),
        'anexados': anexados, 'nao_vinculados': nao_vinculados,
    }


@app.route('/processar-recibos', methods=['POST', 'OPTIONS'])
def processar_recibos():
    """
    v2.33 — Fatia um PDF mestre de RECIBOS por colaborador (CPF) e anexa cada
    recibo ao Funcionário (campo 'Recibos Pagamento') → entra no pacote de
    e-mail do cliente daquele colaborador. multipart/form-data:
      - pdf: arquivo mestre (1 colaborador por página)
      - tipo: rótulo (ex.: "Salário", "Assiduidade", "Almoço e Janta")
      - folha_mensal: ex. "Abril 2026"
    Protegido por X-API-KEY (EMAIL_WEBHOOK_KEY) — Macro 6A, correção de
    auditoria (achado #4): mesma lacuna e mesma correção de
    /processar-doc-cliente (ver docstring lá).
    ATENÇÃO: rode 1x por tipo (re-rodar duplica os anexos no colaborador) —
    risco de idempotência pré-existente, registrado separadamente no
    relatório da Macro 6A; não corrigido nesta rodada (fora do achado
    específico de autenticação desta tarefa).
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401
    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    tipo = (request.form.get('tipo') or 'Recibo').strip()
    folha_mensal = (request.form.get('folha_mensal') or '').strip()
    caminho = extrair_pdf_do_request()
    if not caminho:
        return jsonify({'status': 'erro', 'erro': 'PDF não enviado (campo "pdf")'}), 400
    try:
        return jsonify(_processar_recibos_master(caminho, tipo, folha_mensal))
    finally:
        try:
            os.remove(caminho)
        except OSError:
            pass


@app.route('/recibo/<hash_recibo>', methods=['GET'])
def recibo_leitura(hash_recibo):
    """
    Recibo Digital de Leitura: marca o envio (Envios de Documentos) com a
    data/hora do clique em "Recibo Lido em" (Status="Lido") e redireciona
    para o PDF do Holerite associado. Endpoint público — o hash funciona
    como token de acesso ao documento.
    """
    if not AIRTABLE_API_KEY:
        return 'Serviço indisponível.', 503

    registro = _buscar_por_campo(TABLE_ENVIOS, 'Hash Recibo', hash_recibo)
    if not registro:
        return 'Link inválido ou expirado.', 404

    fields = registro.get('fields', {})

    if not fields.get('Recibo Lido em'):
        try:
            _at_throttle()
            r = requests.patch(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ENVIOS}/{registro["id"]}',
                headers=_at_headers(),
                json={
                    'fields': {
                        F_ENVIO_LIDO_EM: datetime.now().isoformat(),
                        F_ENVIO_STATUS: 'Lido',
                    },
                    'typecast': True,
                },
                timeout=15,
            )
            _raise_for_airtable(r, f'marcar recibo lido {registro["id"]}')
        except AirtableError as exc:
            logger.warning(f'[RECIBO] falha ao marcar leitura do envio {registro["id"]}: {exc}')

    holerite_ids = fields.get('Holerites') or []
    if holerite_ids:
        _at_throttle()
        r = requests.get(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_HOL}/{holerite_ids[0]}',
            headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
            timeout=30,
        )
        if r.ok:
            anexos = r.json().get('fields', {}).get('PDF HOLERITE') or []
            if anexos:
                return redirect(anexos[0]['url'])

    # Envios de Folha de Ponto/Cartão Ponto (Fase 3): PDF(s) anexados
    # diretamente no envio, sem registro vinculado em outra tabela.
    anexos_diretos = fields.get('Arquivos') or []
    if anexos_diretos:
        return redirect(anexos_diretos[0]['url'])

    return 'Documento não encontrado.', 404


# ─── ASSINATURA NATIVA (IP + CPF) ─────────────────────────────────────────────
#
# Fluxo: 1) /assinatura/gerar cria o registro em "Assinaturas Digitais" com um
# Hash Token único e devolve o link. 2) O link é enviado ao colaborador (ex.:
# via WhatsApp/Evolution API). 3) Colaborador abre /assinatura/<hash> (GET) —
# página pede o CPF. 4) Ao confirmar (POST), o CPF informado é validado contra
# o CPF cadastrado em Funcionários; se bater, grava IP real, User-Agent e
# timestamp, marca a Assinatura como "Assinado" e também atualiza o Status do
# registro de origem em Processar Arquivos para "Assinado".

def _gerar_hash_assinatura() -> str:
    """Token único (URL-safe) para o link de Assinatura Nativa."""
    return secrets.token_urlsafe(24)


def _ip_real_da_requisicao() -> str:
    """Render fica atrás de proxy — o IP real do cliente vem em X-Forwarded-For
    (primeiro IP da lista); fallback para remote_addr se o header não vier."""
    xff = request.headers.get('X-Forwarded-For', '')
    if xff:
        return xff.split(',')[0].strip()
    return request.remote_addr or 'desconhecido'


def _pagina_assinatura_html(hash_token: str, nome_doc: str, erro: str = None, mostrar_formulario: bool = True,
                             documentos: list = None) -> str:
    """
    documentos (opcional, só usado hoje pelo pacote HOLERITE_FOLHA_PONTO —
    correção desta Macro de fechamento, auditoria de 18-19/08 encontrou que
    o colaborador nunca via o PDF em lugar nenhum, nem aqui nem no WhatsApp):
    lista de {'nome': 'Holerite', 'url': 'https://...'} a exibir embutido
    ANTES do formulário de CPF. Quando None (todo outro tipo de documento —
    Kit Admissão, Rescisão, EPI, Contratos, Folha de Ponto isolada — e as
    páginas de erro/status terminal do próprio pacote), o comportamento é
    EXATAMENTE o de antes desta correção: formulário liberado de cara, sem
    visualizador. Nenhuma mudança de comportamento fora do pacote.
    """
    erro_html = f'<p style="color:#c0392b;font-weight:bold;">{erro}</p>' if erro else ''
    # mostrar_formulario=False: usado para status terminais em que reenviar o
    # CPF nunca teria efeito (ex.: pacote Cancelado por vínculo não mais
    # ativo) -- evita sugerir uma ação que a idempotência já bloqueia por
    # dentro, sem alterar o comportamento existente de Expirado/erro genérico.
    tem_visualizador = bool(documentos) and mostrar_formulario
    botao_attrs = 'disabled' if tem_visualizador else ''
    botao_texto = 'Visualize os documentos acima para liberar a assinatura' if tem_visualizador else 'Confirmar e Assinar'
    aviso_leitura_html = (
        '<p id="avisoLeitura" class="aviso-leitura">Role/visualize os documentos acima até o fim '
        'para liberar o botão de assinatura.</p>'
    ) if tem_visualizador else ''
    form_html = f"""
    <form method="POST" action="/assinatura/{hash_token}">
      <input type="text" name="cpf" inputmode="numeric" pattern="\\d{{4}}" placeholder="Últimos 4 números do CPF" maxlength="4" required>
      <p class="termo">Ao clicar abaixo, concordo eletronicamente com o recebimento e os termos do documento acima, sob as penas da lei.</p>
      {aviso_leitura_html}
      <button type="submit" id="btnAssinar" {botao_attrs}>{botao_texto}</button>
    </form>""" if mostrar_formulario else ''

    visualizador_html = ''
    visualizador_script = ''
    if tem_visualizador:
        blocos = []
        chamadas_render = []
        for idx, doc in enumerate(documentos):
            container_id = f'pdfContainer{idx}'
            url_js = json.dumps(doc['url'])
            if doc.get('paginas'):
                # Caminho normal: páginas como imagem, marcadas como
                # pré-visualização. Nenhum PDF chega ao navegador aqui.
                # Cada página fica dentro de um bloco com proporção de folha
                # A4 e o aviso "Carregando documento…". Dois motivos, ambos
                # do incidente de 20/08/2026:
                #   1. sem isso a área da página fica vazia (branca no
                #      navegador comum, PRETA no navegador interno do
                #      WhatsApp em modo escuro) enquanto a imagem não chega,
                #      e o colaborador acha que travou e fecha;
                #   2. o bloco já ocupa a altura final antes da imagem
                #      carregar, então a rede de segurança de rolagem não
                #      dispara numa página "curta" e o botão não destrava
                #      antes de haver o que ver.
                # onload/onerror são inline de propósito: não dependem de o
                # <script> do fim da página já ter sido interpretado, o que
                # falharia para imagem servida do cache do navegador.
                imgs = '\n'.join(
                    f'        <div class="pagina-wrap">'
                    f'<img class="pdf-page" loading="lazy" alt="{doc["nome"]} — página {p}" '
                    f'src="{doc["url_pagina"]}{p}" '
                    f'onload="this.parentNode.classList.add(\'carregada\')" '
                    f'onerror="this.parentNode.classList.add(\'falhou\')">'
                    f'<div class="pagina-aviso"><span class="girador"></span>'
                    f'<span class="rotulo">Carregando documento…</span></div>'
                    f'</div>'
                    for p in range(1, doc['paginas'] + 1)
                )
                blocos.append(f"""
    <div class="doc-bloco">
      <h2>{doc['nome']}</h2>
      <div id="{container_id}" class="pdf-container">
{imgs}
      </div>
    </div>""")
                chamadas_render.append(f"observarUltimaImagem({idx}, {json.dumps(container_id)});")
            else:
                # Renderização em imagem indisponível para este documento —
                # cai no PDF embutido, comportamento anterior.
                blocos.append(f"""
    <div class="doc-bloco">
      <h2>{doc['nome']}</h2>
      <div id="{container_id}" class="pdf-container"></div>
    </div>""")
                chamadas_render.append(
                    f"ativarFallback({idx}, {url_js}, {json.dumps(container_id)});"
                )
        visualizador_html = '\n'.join(blocos)
        visualizador_script = f"""
<script>
  var TOTAL_DOCS = {len(documentos)};
  var docsVistos = {{}};

  function liberarBotao() {{
    var btn = document.getElementById('btnAssinar');
    if (btn) {{ btn.disabled = false; btn.textContent = 'Confirmar e Assinar'; }}
    var aviso = document.getElementById('avisoLeitura');
    if (aviso) {{ aviso.style.display = 'none'; }}
  }}
  function marcarVisto(idx) {{
    if (!docsVistos[idx]) {{
      docsVistos[idx] = true;
      if (Object.keys(docsVistos).length >= TOTAL_DOCS) {{ liberarBotao(); }}
    }}
  }}

  function observarUltimaImagem(idx, containerId) {{
    // As páginas já vêm no HTML como <img> (servidas pelo próprio
    // domínio, com marca d'água). Aqui só marcamos o documento como
    // "visto" quando a ÚLTIMA página dele aparecer na tela.
    var container = document.getElementById(containerId);
    if (!container) {{ return; }}
    var imagens = container.getElementsByTagName('img');
    if (!imagens.length) {{ return; }}
    var ultima = imagens[imagens.length - 1];
    if (!('IntersectionObserver' in window)) {{
      // Navegador sem suporte: a rede de segurança de rolagem da página
      // (mais abaixo) continua valendo, então nunca trava a assinatura.
      return;
    }}
    var obs = new IntersectionObserver(function(entries) {{
      entries.forEach(function(entry) {{ if (entry.isIntersecting) {{ marcarVisto(idx); }} }});
    }}, {{threshold: 0.3}});
    obs.observe(ultima);
  }}

  function ativarFallback(idx, url, containerId) {{
    // PDF.js indisponível (CDN bloqueado por política de rede/corporativa,
    // navegador antigo, etc.) — fallback para o visualizador nativo do
    // navegador via iframe; como não dá pra detectar rolagem dentro de um
    // iframe, libera por confirmação explícita do colaborador em vez de
    // travar a assinatura indefinidamente.
    //
    // #toolbar=0 (feedback do teste de homologação, 19/08/2026): esconde a
    // barra de ferramentas do visualizador nativo do Chrome/Edge, junto com
    // o botão de download que ela traz -- alinhado à intenção de que o
    // colaborador assine antes de guardar o documento. Não é garantia (é
    // um parâmetro não-padrão, ignorado por outros navegadores, e print de
    // tela sempre existe) -- é redução de atrito, não controle de acesso.
    // O link "Abrir/baixar em nova aba" foi removido pelo mesmo motivo: era
    // um caminho explícito de download antes da assinatura. O documento
    // continua visível no iframe; quem precisar de cópia recebe o PDF
    // combinado carimbado logo após assinar.
    var container = document.getElementById(containerId);
    container.innerHTML =
      '<iframe src="' + url + '#toolbar=0" class="pdf-fallback-iframe"></iframe>' +
      '<label class="ack"><input type="checkbox" onchange="if(this.checked){{marcarVisto(' + idx + ');}}"> Já visualizei este documento</label>';
  }}

  window.addEventListener('DOMContentLoaded', function() {{
    {' '.join(chamadas_render)}
  }});

  // Correção adicional (feedback do teste de homologação, 19/08/2026):
  // a detecção por documento (IntersectionObserver na última página de
  // cada PDF) pode falhar de forma silenciosa dependendo de como o
  // navegador mede visibilidade dentro do container do PDF -- confirmado
  // no teste real: os documentos renderizaram certinho, mas o botão
  // continuou travado mesmo depois do colaborador rolar a página inteira
  // até o campo de CPF. Esta checagem independente serve de rede de
  // segurança: se a pessoa rolou até perto do fim da PÁGINA (não de cada
  // caixa de PDF individualmente), libera o botão de qualquer forma --
  // não depende de canvas, PDF.js ou de nenhum dos containers.
  function pertoDoFimDaPagina() {{
    return (window.innerHeight + window.scrollY) >= (document.body.scrollHeight - 60);
  }}
  window.addEventListener('scroll', function() {{
    if (pertoDoFimDaPagina()) {{ liberarBotao(); }}
  }});
  window.addEventListener('load', function() {{
    if (pertoDoFimDaPagina()) {{ liberarBotao(); }}
  }});
</script>"""

    return f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Assinatura — Magnata Portaria e Serviços</title>
<style>
  body {{ font-family: Arial, sans-serif; background:#f4f6f8; margin:0; padding:24px; }}
  .card {{ max-width:640px; margin:0 auto; background:#fff; border-radius:10px; padding:28px;
           box-shadow:0 2px 10px rgba(0,0,0,0.08); }}
  h1 {{ font-size:18px; color:#1a1a1a; }}
  h2 {{ font-size:15px; color:#2c5f8a; margin:20px 0 8px; }}
  p {{ color:#444; font-size:14px; line-height:1.5; }}
  input[type=text] {{ width:100%; padding:12px; font-size:16px; border:1px solid #ccc; border-radius:6px;
           margin:12px 0; box-sizing:border-box; }}
  button {{ width:100%; padding:14px; font-size:16px; background:#2c5f8a; color:#fff;
            border:none; border-radius:6px; cursor:pointer; font-weight:bold; }}
  button:hover {{ background:#234b6e; }}
  button:disabled {{ background:#9db3c4; cursor:not-allowed; }}
  .aviso {{ font-size:12px; color:#888; margin-top:16px; }}
  .aviso-leitura {{ font-size:13px; color:#a05a00; background:#fff6e6; border-radius:6px; padding:10px; }}
  .termo {{ font-size:12px; color:#555; font-style:italic; margin:14px 0 8px; line-height:1.4; }}
  .doc-bloco {{ border:1px solid #e2e2e2; border-radius:8px; padding:12px; margin-bottom:16px; background:#fafafa; }}
  .pdf-container {{ border:1px solid #ddd; border-radius:6px; background:#fff; }}
  .pdf-page {{ display:block; width:100%; height:auto; margin:0 auto; }}
  /* Bloco de página com aviso de carregamento — ver comentário em
     _pagina_assinatura_html. A proporção A4 segura a altura final antes da
     imagem chegar. */
  .pagina-wrap {{ position:relative; width:100%; aspect-ratio:1/1.414;
                  background:#eef1f4; border-radius:4px; margin-bottom:10px; }}
  .pagina-wrap.carregada {{ aspect-ratio:auto; background:none; }}
  .pagina-wrap.falhou {{ aspect-ratio:auto; min-height:96px; background:#fdf1f1; }}
  .pagina-aviso {{ position:absolute; top:0; left:0; right:0; bottom:0;
                   display:flex; align-items:center; justify-content:center;
                   color:#5b6b7a; font-size:14px; padding:12px; text-align:center; }}
  .pagina-wrap.carregada .pagina-aviso {{ display:none; }}
  .girador {{ width:15px; height:15px; margin-right:8px; flex:none;
              border:2px solid #c8d2dc; border-top-color:#2c5f8a;
              border-radius:50%; animation:girar 0.8s linear infinite; }}
  @keyframes girar {{ to {{ transform:rotate(360deg); }} }}
  .pagina-wrap.falhou .girador {{ display:none; }}
  .pagina-wrap.falhou .rotulo {{ display:none; }}
  .pagina-wrap.falhou .pagina-aviso {{ color:#a12d2d; }}
  .pagina-wrap.falhou .pagina-aviso::after {{
      content:'Não foi possível carregar esta página. Recarregue a tela.'; }}
  @media (prefers-reduced-motion: reduce) {{ .girador {{ animation:none; }} }}
  .pdf-fallback-iframe {{ width:100%; height:480px; border:1px solid #ddd; border-radius:6px; }}
  .ack {{ display:block; font-size:13px; margin-top:8px; }}
</style></head>
<body>
  <div class="card">
    <h1>Confirmação de Recebimento — {nome_doc}</h1>
    <p>Para confirmar que você é o(a) destinatário(a) deste documento, digite os <strong>4 últimos números</strong> do seu CPF abaixo.</p>
    {visualizador_html}
    {erro_html}{form_html}
    <p class="aviso">Ao confirmar, você declara ter recebido e tomado conhecimento do documento. Data, hora e IP de acesso serão registrados como comprovação.</p>
  </div>
  {visualizador_script}
</body></html>"""


def _downloads_do_registro_assinado(fields: dict):
    """Downloads a oferecer na tela de "já assinado" do pacote Holerite +
    Folha de Ponto. Devolve lista de {'rotulo', 'url'} — vazia se não houver
    nada entregável.

    Preferência: o PDF único "Combinado Carimbado", gerado no momento da
    assinatura desde 5bca9e8 (19/08/2026 13:58).

    Registros assinados ANTES desse commit não têm esse arquivo. Auditoria de
    20/08/2026: 65 dos 71 registros assinados do lote Julho/2026 estavam nessa
    situação, e a tela mostrava só "Assinatura confirmada", sem documento
    nenhum — embora os 3 PDFs individuais já estivessem anexados ao registro.
    Para esses, devolve os 3 individuais em vez de deixar a tela vazia.

    Só entrega arquivo já carimbado (" - ASSINADO") ou o comprovante. O PDF
    original, não carimbado, nunca é oferecido por aqui — quem ainda não
    assinou não passa por esta função (o status não é "Assinado"), e um
    registro que por qualquer motivo ainda guarde o original não deve
    entregá-lo como se fosse o documento assinado.
    """
    anexos = fields.get('Documento PDF') or []

    for att in anexos:
        if (att.get('filename') or '').startswith('Combinado Carimbado'):
            return [{'rotulo': '⬇ Baixar PDF Combinado Carimbado', 'url': att['url']}]

    ROTULOS = {
        'Holerite': '⬇ Baixar Holerite assinado',
        'Folha de Ponto': '⬇ Baixar Folha de Ponto assinada',
        'Comprovante Assinatura': '⬇ Baixar Comprovante de Assinatura',
    }
    ORDEM = ('Holerite', 'Folha de Ponto', 'Comprovante Assinatura')

    achados = {}
    for att in anexos:
        nome = att.get('filename') or ''
        if ' - ASSINADO' not in nome and not nome.startswith('Comprovante Assinatura'):
            continue
        for prefixo in ORDEM:
            if nome.startswith(prefixo) and prefixo not in achados:
                achados[prefixo] = {'rotulo': ROTULOS[prefixo], 'url': att['url']}
                break

    return [achados[p] for p in ORDEM if p in achados]


def _pagina_assinatura_sucesso_html(nome_doc: str, quando: str, downloads=None) -> str:
    # downloads (opcional): lista de {'rotulo', 'url'} mostrada como botões na
    # tela de sucesso. None/vazio preserva o comportamento anterior (nenhum
    # botão) para todo tipo de documento que não passa esse parâmetro.
    # Substitui o antigo parâmetro download_url (1 botão só), que não dava
    # conta dos registros sem o PDF combinado — ver
    # _downloads_do_registro_assinado.
    botao_download_html = ''
    if downloads:
        botoes = ''.join(
            f'<a class="botao-download" href="{(d["url"] or "").replace(chr(34), "&quot;")}" '
            f'target="_blank" rel="noopener" download>{d["rotulo"]}</a>'
            for d in downloads if d.get('url')
        )
        if botoes:
            legenda = ('<p class="legenda-downloads">Seus documentos assinados:</p>'
                       if len(downloads) > 1 else '')
            botao_download_html = legenda + botoes
    return f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Assinatura confirmada</title>
<style>
  body {{ font-family: Arial, sans-serif; background:#f4f6f8; margin:0; padding:24px; }}
  .card {{ max-width:480px; margin:0 auto; background:#fff; border-radius:10px; padding:28px;
           box-shadow:0 2px 10px rgba(0,0,0,0.08); text-align:center; }}
  h1 {{ color:#1e7a34; font-size:20px; }}
  p {{ color:#444; font-size:14px; line-height:1.5; }}
  .botao-download {{ display:block; margin-top:12px; padding:14px; font-size:15px; font-weight:bold;
                      background:#2c5f8a; color:#fff; text-decoration:none; border-radius:6px; }}
  .botao-download:hover {{ background:#234b6e; }}
  .legenda-downloads {{ margin-top:22px; margin-bottom:0; font-weight:bold; color:#333; }}
</style></head>
<body>
  <div class="card">
    <h1>✅ Assinatura confirmada!</h1>
    <p><strong>{nome_doc}</strong></p>
    <p>Recebimento confirmado em {quando}.</p>
    <p>Obrigado!</p>
    {botao_download_html}
  </div>
</body></html>"""


def _buscar_funcionario_nome_whatsapp(func_id: str):
    """Retorna (nome, whatsapp_normalizado) de um Funcionário. whatsapp_normalizado
    é None se o campo estiver vazio ou em formato inválido."""
    _at_throttle()
    r = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}/{func_id}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        timeout=30,
    )
    if not r.ok:
        return None, None
    fields = r.json().get('fields', {})
    nome = fields.get('Nome Completo')
    whatsapp = _normalizar_numero_evolution(fields.get('WhatsApp'))
    return nome, whatsapp


# Único valor de Status que autoriza assinatura -- comparação exata, nunca
# por lista de nomes/telefones. Qualquer outro valor real (Inativo, Empresa,
# Pessoal, Outro) ou ausência de valor bloqueia. "Status desconhecido" (uma
# string que não é nem "Ativo" nem nenhum dos valores conhecidos -- ex.: erro
# de digitação futuro no Airtable) é tratado como indeterminado, nunca como
# aprovado por omissão.
STATUS_FUNCIONARIO_ATIVO = 'Ativo'
STATUS_FUNCIONARIO_CONHECIDOS = {'Ativo', 'Inativo', 'Empresa', 'Pessoal', 'Outro'}


def _status_funcionario_elegivel(funcionario_id: str):
    """Retorna (elegivel: bool, motivo: Optional[str]).

    Quando elegivel=False, `motivo` é SEMPRE um destes 2 valores fixos --
    nunca contém dado pessoal, nome de campo do Airtable ou texto de erro:
      'vinculo_nao_ativo'      status conhecido, mas != Ativo;
      'vinculo_indeterminado'  id ausente, falha de rede/HTTP, exceção, ou
                               status ausente/desconhecido.

    Duas razões para não devolver o valor bruto, ambas reais:
    1. LGPD (`CLAUDE.md` §6) -- a versão anterior interpolava
       `r.text[:80]` e `list(fields.keys())` no motivo, e esse motivo vai
       para log. O corpo de um registro de Funcionários contém nome, CPF
       e WhatsApp.
    2. Nunca aprovar por omissão -- status desconhecido é tratado como
       indeterminado (bloqueia), nunca como "provavelmente ativo".

    `returnFieldsByFieldId` é obrigatório, não cosmético: F_FUNC_STATUS é
    um Field ID (`fld...`). Sem esse parâmetro o Airtable devolve as
    chaves por NOME, `fields.get(F_FUNC_STATUS)` seria sempre None, e era
    exatamente isso que as heurísticas de fallback removidas aqui
    tentavam compensar -- varrendo todos os valores do registro atrás de
    algo parecido com um status, o que podia casar com o campo errado.
    Mesmo padrão já usado em outros 18 pontos deste arquivo.
    """
    if not funcionario_id:
        return False, 'vinculo_indeterminado'
    _at_throttle()
    try:
        r = requests.get(
            f"https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}/{funcionario_id}",
            headers={"Authorization": f"Bearer {AIRTABLE_API_KEY}"},
            params={'returnFieldsByFieldId': 'true'},
            timeout=30,
        )
        if not r.ok:
            return False, 'vinculo_indeterminado'
        fields = r.json().get("fields", {}) or {}
        raw = fields.get(F_FUNC_STATUS)
        if isinstance(raw, (list, tuple)) and raw:
            raw = raw[0]
        st = str(raw).strip() if raw is not None else ''
        if st == STATUS_FUNCIONARIO_ATIVO:
            return True, None
        if st in STATUS_FUNCIONARIO_CONHECIDOS:
            return False, 'vinculo_nao_ativo'
        return False, 'vinculo_indeterminado'
    except Exception:
        return False, 'vinculo_indeterminado'


def _montar_mensagem_assinatura(nome: str, tipo_documento: str, link: str) -> str:
    return (
        f'Olá{(" " + nome) if nome else ""}! Você tem um documento pendente de confirmação '
        f'({tipo_documento}).\n\nPara confirmar o recebimento, acesse o link abaixo e informe '
        f'os 4 últimos números do seu CPF:\n{link}\n\nMagnata Portaria e Serviços.'
    )


def _carimbar_pdf_assinado(pdf_bytes: bytes, nome: str, cpf4: str, dt_str: str) -> bytes:
    """Sobrepõe um rodapé discreto ("Assinado digitalmente por...") em todas
    as páginas do PDF, via pypdf merge_page + uma página de overlay gerada
    com reportlab (mesmo tamanho de cada página original). Usa reportlab (e
    não fpdf2) especificamente aqui porque overlays fpdf2+pypdf sofrem de
    colisão de nome de recurso de fonte entre PDFs de origens diferentes,
    fazendo o texto do carimbo "sumir" silenciosamente (bug conhecido do
    merge_page). Não usa marca d'água diagonal nem cobre o conteúdo
    original — só uma faixa fina no rodapé.
    """
    from reportlab.pdfgen import canvas

    texto = f'Assinado digitalmente por {nome} | CPF final {cpf4} | {dt_str} | Magnata Portaria e Serviços'

    reader = PdfReader(io.BytesIO(pdf_bytes))
    writer = PdfWriter()
    carimbo_cache = {}

    for page in reader.pages:
        box = page.mediabox
        largura_pt, altura_pt = float(box.width), float(box.height)
        chave = (round(largura_pt), round(altura_pt))
        if chave not in carimbo_cache:
            buf_stamp = io.BytesIO()
            c = canvas.Canvas(buf_stamp, pagesize=(largura_pt, altura_pt))
            c.setFillColorRGB(0.92, 0.92, 0.92)
            c.rect(0, 0, largura_pt, 28, fill=1, stroke=0)
            c.setFillColorRGB(0, 0, 0)
            c.setFont('Helvetica', 7)
            c.drawString(10, 10, texto[:180])
            c.save()
            buf_stamp.seek(0)
            carimbo_cache[chave] = PdfReader(buf_stamp).pages[0]
        page.merge_page(carimbo_cache[chave])
        writer.add_page(page)

    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue()


def _gerar_comprovante_assinatura_pdf(nome_documento: str, func_nome: str, ip: str, user_agent: str,
                                       dt_str: str, cpf4: str, doc_bytes: bytes, doc_nome: str) -> bytes:
    """Gera o PDF 'Comprovante de Assinatura Eletrônica' — documento autônomo
    (não é merge de PDFs, então fpdf2 é seguro aqui) consolidando as
    evidências capturadas na confirmação da assinatura, incluindo hash
    SHA-256 do documento original para prova de integridade."""
    from fpdf import FPDF

    sha256 = hashlib.sha256(doc_bytes).hexdigest() if doc_bytes else '(indisponível)'

    def s(t):
        return str(t).encode('latin-1', 'replace').decode('latin-1')

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    larg = pdf.w - pdf.l_margin - pdf.r_margin

    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(larg, 8, s('COMPROVANTE DE ASSINATURA ELETRÔNICA'), new_x='LMARGIN', new_y='NEXT', align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(larg, 6, s('MAGNATA PORTARIA E SERVIÇOS LTDA - CNPJ 17.987.187/0001-61'), new_x='LMARGIN', new_y='NEXT', align='C')
    pdf.ln(6)

    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(larg, 7, s('1. Identificação'), new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 10)
    for ln in [f'Colaborador: {func_nome}', f'Documento assinado: {nome_documento}',
               f'Nome do arquivo: {doc_nome}', 'Status: Assinado']:
        pdf.cell(larg, 6, s(ln), new_x='LMARGIN', new_y='NEXT')
    pdf.ln(3)

    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(larg, 7, s('2. Evidências da Assinatura'), new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 10)
    for ln in [f'Data/Hora da confirmação: {dt_str} (horário de Brasília)', f'Endereço IP de origem: {ip}',
               f'CPF confirmado pelo colaborador (4 últimos dígitos): {cpf4}',
               f'Dispositivo/Navegador (User-Agent): {user_agent}']:
        pdf.multi_cell(larg, 6, s(ln))
    pdf.ln(2)
    pdf.set_font('Helvetica', 'I', 9)
    pdf.multi_cell(larg, 5, s(
        'Declaração aceita eletronicamente pelo colaborador no momento da confirmação: '
        '"Ao clicar, concordo eletronicamente com o recebimento e os termos do documento, '
        'sob as penas da lei."'
    ))
    pdf.ln(3)

    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(larg, 7, s('3. Integridade do Documento'), new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 9)
    pdf.multi_cell(larg, 5, s(f'Hash SHA-256 do arquivo assinado (identificador único de integridade): {sha256}'))
    pdf.ln(3)

    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(larg, 7, s('4. Base Legal'), new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 9)
    pdf.multi_cell(larg, 5, s(
        'Este comprovante registra a manifestação de vontade eletrônica do colaborador, '
        'nos termos do art. 10, §2º da Medida Provisória 2.200-2/2001 e da Lei 14.063/2020, '
        'mediante confirmação de identidade por CPF e captura de metadados de sessão '
        '(IP, timestamp e dispositivo), constituindo prova de aceite do documento anexo.'
    ))
    pdf.ln(8)
    pdf.set_font('Helvetica', 'I', 8)
    pdf.cell(larg, 5, s(f'Comprovante gerado automaticamente em {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}.'), new_x='LMARGIN', new_y='NEXT', align='C')

    return bytes(pdf.output())


def _gerar_comprovante_assinatura_pacote_pdf(competencia: str, func_nome: str, ip: str, user_agent: str,
                                              dt_str: str, cpf4: str,
                                              doc1_bytes: bytes, doc1_nome: str,
                                              doc2_bytes: bytes, doc2_nome: str) -> bytes:
    """Variante do Comprovante de Assinatura Eletrônica para o PACOTE
    Holerite + Folha de Ponto — mesmo modelo e mesma base legal do
    comprovante de documento único (`_gerar_comprovante_assinatura_pdf`,
    não alterado), mas listando os 2 documentos e os 2 hashes SHA-256 na
    mesma seção de integridade, cobrindo explicitamente ambos — exigência
    desta Macro (item 1.7). Terminologia: "assinatura eletrônica com
    evidências" — não usa nem implica certificação ICP-Brasil."""
    from fpdf import FPDF

    sha1 = hashlib.sha256(doc1_bytes).hexdigest() if doc1_bytes else '(indisponível)'
    sha2 = hashlib.sha256(doc2_bytes).hexdigest() if doc2_bytes else '(indisponível)'

    def s(t):
        return str(t).encode('latin-1', 'replace').decode('latin-1')

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    larg = pdf.w - pdf.l_margin - pdf.r_margin

    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(larg, 8, s('COMPROVANTE DE ASSINATURA ELETRÔNICA'), new_x='LMARGIN', new_y='NEXT', align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(larg, 6, s('MAGNATA PORTARIA E SERVIÇOS LTDA - CNPJ 17.987.187/0001-61'), new_x='LMARGIN', new_y='NEXT', align='C')
    pdf.ln(6)

    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(larg, 7, s('1. Identificação'), new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 10)
    for ln in [f'Colaborador: {func_nome}', f'Competência: {competencia}',
               f'Documento 1 (Holerite): {doc1_nome}', f'Documento 2 (Folha de Ponto): {doc2_nome}',
               'Status: Assinado (pacote - os 2 documentos confirmados numa única transação)']:
        pdf.multi_cell(larg, 6, s(ln))
    pdf.ln(3)

    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(larg, 7, s('2. Evidências da Assinatura'), new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 10)
    for ln in [f'Data/Hora da confirmação: {dt_str} (horário de Brasília)', f'Endereço IP de origem: {ip}',
               f'CPF confirmado pelo colaborador (4 últimos dígitos): {cpf4}',
               f'Dispositivo/Navegador (User-Agent): {user_agent}']:
        pdf.multi_cell(larg, 6, s(ln))
    pdf.ln(2)
    pdf.set_font('Helvetica', 'I', 9)
    pdf.multi_cell(larg, 5, s(
        'Declaração aceita eletronicamente pelo colaborador no momento da confirmação: '
        '"Ao clicar, concordo eletronicamente com o recebimento e os termos dos dois '
        'documentos acima (Holerite e Folha de Ponto), sob as penas da lei."'
    ))
    pdf.ln(3)

    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(larg, 7, s('3. Integridade dos Documentos'), new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 9)
    pdf.multi_cell(larg, 5, s(f'Hash SHA-256 do Holerite assinado (identificador único de integridade): {sha1}'))
    pdf.multi_cell(larg, 5, s(f'Hash SHA-256 da Folha de Ponto assinada (identificador único de integridade): {sha2}'))
    pdf.ln(3)

    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(larg, 7, s('4. Base Legal'), new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 9)
    pdf.multi_cell(larg, 5, s(
        'Este comprovante registra a manifestação de vontade eletrônica do colaborador, '
        'nos termos do art. 10, §2º da Medida Provisória 2.200-2/2001 e da Lei 14.063/2020, '
        'mediante confirmação de identidade por CPF e captura de metadados de sessão '
        '(IP, timestamp e dispositivo), constituindo prova de aceite dos dois documentos '
        'anexos (Holerite e Folha de Ponto). Trata-se de assinatura eletrônica com '
        'evidências - não de assinatura digital certificada ICP-Brasil.'
    ))
    pdf.ln(8)
    pdf.set_font('Helvetica', 'I', 8)
    pdf.cell(larg, 5, s(f'Comprovante gerado automaticamente em {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}.'), new_x='LMARGIN', new_y='NEXT', align='C')

    return bytes(pdf.output())


@app.route('/assinatura/gerar', methods=['POST', 'OPTIONS'])
def assinatura_gerar():
    """
    ✅ v3.6: Cria um novo registro de Assinatura Nativa com identidade documental obrigatória.

    Mudanças v3.6:
    - arquivo_record_id é OBRIGATÓRIO (sem fallback automático para Kit)
    - tipo_documento contra whitelist TIPOS_DOCUMENTO_VALIDOS
    - documento_url removido (URL vem do arquivo)
    - disparar_whatsapp default FALSE (mais seguro)
    - Idempotência por SHA-256 do PDF
    - Estados: PREPARADO → AGUARDANDO_ENVIO → ENVIADO_AGUARDANDO_ASSINATURA

    Protegido por X-API-KEY (EMAIL_WEBHOOK_KEY).

    Body JSON:
      {
        "funcionario_id": "rec...",              [OBRIGATÓRIO]
        "tipo_documento": "KIT_ADMISSAO",        [OBRIGATÓRIO — valores: KIT_ADMISSAO, FOLHA_PONTO, FICHA_EPI, RESCISAO, CONTRATO_EXPERIENCIA, CONTRATO_TRABALHO, HOLERITE_FOLHA_PONTO]
        "arquivo_record_id": "rec...",           [OBRIGATÓRIO — exceto para HOLERITE_FOLHA_PONTO, ver abaixo]
        "kit_arquivo_record_ids": ["rec...", ...],  [OPCIONAL — para Kit multi-arquivo]
        "processar_id": "rec...",                [opcional — Processar Arquivos a marcar como concluído]
        "nome_documento": "Rescisão - João Silva",  [opcional — se ausente, usa nome canônico]
        "mensagem_extra": "texto",               [opcional — texto adicional no WhatsApp]
        "disparar_whatsapp": false,              [opcional — default FALSE]
        "dry_run": false                         [opcional]
      }

    tipo_documento="HOLERITE_FOLHA_PONTO" (pacote atômico, decisão desta
    Macro — Holerite nunca é assinável isolado): usa 2 campos no lugar de
    "arquivo_record_id":
      {
        "funcionario_id": "rec...",                    [OBRIGATÓRIO]
        "tipo_documento": "HOLERITE_FOLHA_PONTO",      [OBRIGATÓRIO]
        "arquivo_holerite_record_id": "rec...",        [OBRIGATÓRIO]
        "arquivo_folha_ponto_record_id": "rec...",     [OBRIGATÓRIO]
        "mensagem_extra": "texto", "disparar_whatsapp": false, "dry_run": false  [opcionais]
      }
    Bloqueia se as competências extraídas dos 2 PDFs não coincidirem.
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    request_id = f"req{secrets.token_hex(8)}"

    # ✅ Validar X-API-KEY
    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        logger.warning(f'[ASSINATURA] Acesso negado | Request: {request_id}')
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente', 'request_id': request_id}), 401

    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    # ✅ v3.6: Validar configuração de Field IDs
    config_ok, config_msg = _validar_configuracao_assinatura_v36()
    if not config_ok:
        logger.error(f'[ASSINATURA] Configuracao invalida: {config_msg}')
        return jsonify({
            'status': 'erro',
            'erro': 'Configuracao v3.6 incompleta. Substituir placeholders fldXXXX pelos Field IDs reais do Airtable.',
            'request_id': request_id,
        }), 500

    data = request.get_json(force=True, silent=True) or {}

    funcionario_id = data.get('funcionario_id')
    tipo_documento = data.get('tipo_documento')
    arquivo_record_id = data.get('arquivo_record_id')
    kit_arquivo_record_ids = data.get('kit_arquivo_record_ids')

    # ✅ Validações obrigatórias
    if not funcionario_id or not tipo_documento:
        return jsonify({
            'status': 'erro',
            'erro': 'funcionario_id e tipo_documento são obrigatórios',
            'request_id': request_id,
        }), 400

    # Pacote Holerite + Folha de Ponto (decisão desta Macro): 2 documentos,
    # 2 campos próprios — nunca usa arquivo_record_id nem entra no fluxo
    # de documento único abaixo.
    if tipo_documento == TIPO_PACOTE_HOLERITE_PONTO:
        arquivo_holerite_id = data.get('arquivo_holerite_record_id')
        arquivo_ponto_id = data.get('arquivo_folha_ponto_record_id')
        if not arquivo_holerite_id or not arquivo_ponto_id:
            return jsonify({
                'status': 'erro',
                'erro': 'arquivo_holerite_record_id e arquivo_folha_ponto_record_id são obrigatórios para HOLERITE_FOLHA_PONTO',
                'request_id': request_id,
            }), 400

        resultado, status_code = _gerar_pacote_assinatura_holerite_ponto(
            funcionario_id=funcionario_id,
            arquivo_holerite_id=arquivo_holerite_id,
            arquivo_ponto_id=arquivo_ponto_id,
            mensagem_extra=data.get('mensagem_extra') or '',
            disparar_whatsapp=str(data.get('disparar_whatsapp', False)).strip().lower() in ('1', 'true', 'yes', 'sim'),
            dry_run=str(data.get('dry_run', False)).strip().lower() in ('1', 'true', 'yes', 'sim'),
            request_id=request_id,
        )
        return jsonify(resultado), status_code

    if tipo_documento != 'KIT_ADMISSAO' and not arquivo_record_id:
        return jsonify({
            'status': 'erro',
            'erro': 'arquivo_record_id obrigatório (exceto para Kit que pode usar kit_arquivo_record_ids)',
            'tipo_documento': tipo_documento,
            'request_id': request_id,
        }), 400

    # Chamar versão corrigida
    resultado, status_code = _gerar_assinatura_core(
        funcionario_id=funcionario_id,
        tipo_documento=tipo_documento,
        arquivo_record_id=arquivo_record_id,
        kit_arquivo_record_ids=kit_arquivo_record_ids,
        processar_id=data.get('processar_id', '') or '',
        nome_documento=data.get('nome_documento'),  # None = usar canônico
        mensagem_extra=data.get('mensagem_extra') or '',
        disparar_whatsapp=str(data.get('disparar_whatsapp', False)).strip().lower() in ('1', 'true', 'yes', 'sim'),
        dry_run=str(data.get('dry_run', False)).strip().lower() in ('1', 'true', 'yes', 'sim'),
        request_id=request_id,
    )
    return jsonify(resultado), status_code


def _carregar_documento_url(url: str) -> bytes:
    """Carrega os bytes de documento_url — lê direto do disco quando a URL
    aponta para o /static/ deste MESMO servidor, em vez de HTTP GET.

    Necessário porque em Render free tier (1 worker síncrono) o processo
    chamar a si mesmo via HTTPS público causa deadlock: o worker fica
    esperando uma resposta que só um worker livre poderia servir, e não há
    outro — trava até estourar o timeout (visto em produção: ReadTimeout
    de 60s em TODO item de /assinatura/gerar-lote que usava documento_url
    do próprio /static/).
    """
    prefixo_static = f'{RECIBO_BASE_URL}/static/'
    if url.startswith(prefixo_static):
        nome_arquivo = url[len(prefixo_static):].split('?')[0]
        caminho = os.path.join(app.static_folder, nome_arquivo)
        with open(caminho, 'rb') as f:
            return f.read()

    r = requests.get(url, timeout=60)
    # Achado real em produção (08/07/2026): URLs assinadas do Airtable
    # (v5.airtableusercontent.com) EXPIRAM — num lote grande (99 itens,
    # processado em ~13 chamadas sequenciais ao longo de vários minutos),
    # a URL de alguns itens do FIM do lote já tinha expirado quando
    # chegou a vez de buscá-la. O código antigo aceitava r.content sem
    # checar nada — a resposta de erro (corpo pequeno, ex.: 21 bytes) foi
    # gravada como se fosse o PDF real, tanto no Airtable (anexo de 21
    # bytes) quanto no envio por WhatsApp ("documento inválido" ao abrir).
    # Agora falha alto e claro em vez de silenciosamente salvar lixo.
    if not r.ok:
        raise RuntimeError(f'documento_url retornou HTTP {r.status_code} (possível URL expirada): {url[:120]}')
    if not r.content.startswith(b'%PDF'):
        raise RuntimeError(
            f'documento_url não retornou um PDF válido ({len(r.content)} bytes recebidos, '
            f'possível URL expirada ou resposta de erro): {url[:120]}'
        )
    return r.content


def _gerar_assinatura_core(funcionario_id, tipo_documento, arquivo_record_id=None, processar_id='',
                            nome_documento=None, mensagem_extra='', disparar_whatsapp=False, dry_run=False,
                            request_id=None, kit_arquivo_record_ids=None):
    """
    ✅ v3.6: Núcleo de /assinatura/gerar com identidade documental obrigatória.

    Mudanças:
    - arquivo_record_id obrigatório para todos (exceto Kit que pode vir com kit_arquivo_record_ids)
    - tipo_documento contra whitelist (TIPOS_DOCUMENTO_VALIDOS)
    - Busca arquivo pelo Record ID (não por URL)
    - Calcula SHA-256 do PDF real
    - Idempotência por SHA256(func + arquivo + pdf + tipo)
    - SEM fallback automático para Kit
    - disparar_whatsapp=False padrão (seguro)
    - Estados de assinatura (PREPARADO → AGUARDANDO_ENVIO → ENVIADO_AGUARDANDO_ASSINATURA)
    """

    request_id = request_id or f"req{secrets.token_hex(8)}"

    # ✅ DEFESA PROFUNDA: Validar configuração de Field IDs (aplicável a TODOS os caminhos)
    config_ok, config_msg = _validar_configuracao_assinatura_v36()
    if not config_ok:
        logger.error(f'[ASSINATURA CORE] Configuracao invalida: {config_msg}')
        return {
            'status': 'erro',
            'erro': 'Configuracao v3.6 incompleta. Nenhuma assinatura pode ser criada.',
            'request_id': request_id,
        }, 500

    # ✅ CORREÇÃO 1: Validar tipo contra whitelist
    if tipo_documento not in TIPOS_DOCUMENTO_VALIDOS:
        logger.warning(f'[ASSINATURA] Tipo inválido: {tipo_documento} | Func: {funcionario_id}')
        return {
            'status': 'erro',
            'erro': f'tipo_documento inválido: "{tipo_documento}". Aceitos: {", ".join(sorted(TIPOS_DOCUMENTO_VALIDOS))}',
            'request_id': request_id,
        }, 400

    # ✅ CORREÇÃO 2: arquivo_record_id obrigatório (exceto Kit que pode vir com lista)
    if tipo_documento != 'KIT_ADMISSAO' and not arquivo_record_id:
        logger.warning(f'[ASSINATURA] Arquivo ausente: {tipo_documento} | Func: {funcionario_id}')
        return {
            'status': 'erro',
            'erro': 'arquivo_record_id obrigatório para documentos individuais',
            'tipo_documento': tipo_documento,
            'request_id': request_id,
        }, 400

    # Kit de Admissão exige lista de arquivos OU arquivo único consolidado
    if tipo_documento == 'KIT_ADMISSAO' and not arquivo_record_id and not kit_arquivo_record_ids:
        logger.warning(f'[ASSINATURA] Kit sem identidade: {funcionario_id}')
        return {
            'status': 'erro',
            'erro': 'Kit de Admissão exige arquivo_record_id (consolidado) ou kit_arquivo_record_ids (lista)',
            'request_id': request_id,
        }, 400

    nome_documento = nome_documento or NOMES_DOCUMENTOS.get(tipo_documento, tipo_documento)
    nome_func, whatsapp = _buscar_funcionario_nome_whatsapp(funcionario_id)

    # Bloqueio preventivo: disparar_whatsapp=True exige WhatsApp
    if disparar_whatsapp and not whatsapp:
        logger.warning(f'[ASSINATURA] WhatsApp ausente: {funcionario_id}')
        return {
            'status': 'erro', 'erro': 'whatsapp_ausente',
            'mensagem': f'Funcionário "{nome_func or funcionario_id}" não tem WhatsApp cadastrado.',
            'funcionario_id': funcionario_id,
            'request_id': request_id,
        }, 400

    # ✅ CORREÇÃO 3: Buscar arquivo e calcular SHA-256
    pdf_bytes = None
    pdf_sha256 = None
    pdf_filename = 'Documento.pdf'

    if arquivo_record_id:
        try:
            # Buscar arquivo no Airtable
            _at_throttle()
            r_arquivo = requests.get(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ARQUIVOS}/{arquivo_record_id}',  # ✅ corrigido: tblbgJqXh0u1Cjq1s não existe na base
                headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
                params={'returnFieldsByFieldId': 'true'},
                timeout=15,
            )

            if not r_arquivo.ok:
                logger.error(f'[ASSINATURA] Arquivo não encontrado: {arquivo_record_id}')
                return {
                    'status': 'erro',
                    'erro': f'Arquivo {arquivo_record_id} não encontrado',
                    'request_id': request_id,
                }, 404

            # Extrair attachment (primeiro anexo)
            campos_arquivo = r_arquivo.json().get('fields', {})
            anexos = campos_arquivo.get(F_ARQ_ATTACH, [])  # F_ARQ_ATTACH

            if not anexos:
                logger.error(f'[ASSINATURA] Arquivo sem attachment: {arquivo_record_id}')
                return {
                    'status': 'erro',
                    'erro': f'Arquivo {arquivo_record_id} sem attachment',
                    'request_id': request_id,
                }, 400

            # Validar funcionário do arquivo
            func_arquivo = campos_arquivo.get('fldxbZwVNa01pchqF', [])  # F_ARQ_FUNC (link)
            if func_arquivo and func_arquivo[0] != funcionario_id:
                logger.error(f'[ASSINATURA] Arquivo de outro funcionário: {arquivo_record_id} != {funcionario_id}')
                return {
                    'status': 'erro',
                    'erro': f'Arquivo pertence a outro funcionário',
                    'arquivo_func': func_arquivo[0] if func_arquivo else None,
                    'request_id': request_id,
                }, 403

            # Baixar PDF
            attachment = anexos[0]
            pdf_url = attachment.get('url')
            pdf_filename = attachment.get('filename', 'Documento.pdf')

            pdf_bytes = _carregar_documento_url(pdf_url)

            # ✅ CORREÇÃO 4: Calcular SHA-256 do PDF real
            pdf_sha256 = hashlib.sha256(pdf_bytes).hexdigest()
            logger.info(f'[ASSINATURA] PDF carregado: {arquivo_record_id} | SHA256: {pdf_sha256[:16]}...')

        except Exception as exc:
            logger.error(f'[ASSINATURA] Erro ao carregar arquivo: {exc}')
            return {
                'status': 'erro',
                'erro': f'Erro ao carregar arquivo: {exc}',
                'arquivo_record_id': arquivo_record_id,
                'request_id': request_id,
            }, 400

    # ✅ CORREÇÃO 5: Implementar idempotência com SHA-256 físico COMPLETO (sem truncamento)
    if arquivo_record_id and pdf_sha256:
        idempotency_key = hashlib.sha256(
            f"{funcionario_id}|{arquivo_record_id}|{pdf_sha256}|{tipo_documento}".encode()
        ).hexdigest()  # ✅ SHA-256 completo (64 chars), não truncado

        # Verificar se já existe com mesma chave
        _at_throttle()
        try:
            r_existente = requests.get(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}',
                headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
                params={
                    'filterByFormula': f'{{{F_ASS_CHAVE_IDEMPOTENCIA}}}="{idempotency_key}"',
                    'maxRecords': 1,
                    'returnFieldsByFieldId': 'true',
                },
                timeout=30,
            )

            if r_existente.ok and r_existente.json().get('records'):
                rec_existente = r_existente.json()['records'][0]
                status_existente = rec_existente['fields'].get(F_ASS_STATUS)
                existente_id = rec_existente['id']

                logger.info(f'[ASSINATURA] Idempotência: {existente_id} (status: {status_existente})')

                # Comportamento por estado
                if status_existente in ('PREPARADO', 'AGUARDANDO_ENVIO'):
                    return {
                        'status': 'duplicado',
                        'assinatura_id': existente_id,
                        'motivo': f'idempotência: já criada ({status_existente})',
                        'idempotency_key': idempotency_key,
                        'request_id': request_id,
                    }, 409
                elif status_existente in ('ENVIADO_AGUARDANDO_ASSINATURA', 'ENVIANDO'):
                    return {
                        'status': 'ok',
                        'assinatura_id': existente_id,
                        'link': f'{RECIBO_BASE_URL}/assinatura/{rec_existente["fields"].get(F_ASS_HASH)}',
                        'motivo': 'retornando link existente',
                        'request_id': request_id,
                    }, 200
                elif status_existente == 'ASSINADO':
                    return {
                        'status': 'erro',
                        'erro': 'Documento já foi assinado',
                        'assinatura_id': existente_id,
                        'request_id': request_id,
                    }, 409
        except Exception as exc:
            logger.warning(f'[ASSINATURA] Erro ao verificar idempotência: {exc}')
    else:
        idempotency_key = None

    hash_token = _gerar_hash_assinatura()
    link = f'{RECIBO_BASE_URL}/assinatura/{hash_token}'

    if dry_run:
        return {
            'status': 'ok', 'dry_run': True,
            'acao': 'criaria_assinatura_e_dispararia' if disparar_whatsapp else 'criaria_assinatura',
            'link': link, 'funcionario_id': funcionario_id, 'nome_funcionario': nome_func,
            'tipo_documento': tipo_documento,
            'arquivo_record_id': arquivo_record_id,
            'pdf_sha256': pdf_sha256[:16] + '...' if pdf_sha256 else None,
            'request_id': request_id,
        }, 200

    # ✅ CORREÇÃO 6: Criar registro com status intermediário PREPARADO
    campos = {
        F_ASS_NOME:             nome_documento,
        F_ASS_TIPO_DOC:         tipo_documento,
        F_ASS_STATUS:           'PREPARADO',  # Status intermediário (antes de enviar)
        F_ASS_HASH:             hash_token,
        F_ASS_FUNCIONARIO:      [funcionario_id],
        F_ASS_PROCESSAR_ID:     processar_id,
        F_ASS_DATA_GERACAO:     datetime.now().isoformat(),
        F_ASS_TENTATIVAS:       0,
        F_ASS_ARQUIVO_RECORD_ID: arquivo_record_id,  # ✅ Novo: rastrear arquivo
        F_ASS_REQUEST_ID:       request_id,  # ✅ Novo: logging
    }

    if pdf_sha256:
        campos[F_ASS_PDF_SHA256] = pdf_sha256  # ✅ Novo: rastrear integridade
    if idempotency_key:
        campos[F_ASS_CHAVE_IDEMPOTENCIA] = idempotency_key  # ✅ Novo: proteção duplicidade

    assinatura_id = _criar_registro(TABLE_ASSINATURAS, campos)
    logger.info(f'[ASSINATURA] Criada: {assinatura_id} | Tipo: {tipo_documento} | Arquivo: {arquivo_record_id}')

    # ✅ Status já criado como PREPARADO; será AGUARDANDO_ENVIO após validação persistente

    # ✅ CORREÇÃO 7: Anexar PDF se foi carregado
    if pdf_bytes:
        _anexar_attachment(TABLE_ASSINATURAS, assinatura_id, F_ASS_DOCUMENTO_PDF, pdf_bytes, pdf_filename)

    # ✅ CORREÇÃO 8: Disparar WhatsApp apenas se explicitamente solicitado
    disparo_resultado = None

    if disparar_whatsapp:
        try:
            # Status será atualizado via reconciliação persistente (BLOQUEADOR 2)
            mensagem = _montar_mensagem_assinatura(nome_func, tipo_documento, link)
            if mensagem_extra:
                mensagem = f'{mensagem}\n\n{mensagem_extra}'

            if pdf_bytes:
                _evolution_enviar_documento(whatsapp, None, pdf_filename, caption=mensagem, media_bytes=pdf_bytes)
            else:
                _evolution_enviar_texto(whatsapp, mensagem)

            disparo_resultado = 'enviado'
            logger.info(f'[ASSINATURA] WhatsApp enviado: {assinatura_id} | WhatsApp: {whatsapp}')

        except Exception as exc:
            logger.error(f'[ASSINATURA] Falha ao disparar WhatsApp: {assinatura_id} | Erro: {exc}')
            disparo_resultado = f'falha: {exc}'
    else:
        # Não vai disparar, deixar em AGUARDANDO_ENVIO (será disparado por reenvio manual)
        logger.info(f'[ASSINATURA] Criada sem disparo automático: {assinatura_id}')

    return {
        'status': 'ok', 'dry_run': False, 'acao': 'assinatura_criada',
        'assinatura_id': assinatura_id, 'link': link, 'hash_token': hash_token,
        'nome_funcionario': nome_func, 'whatsapp': whatsapp,
        'whatsapp_disparo': disparo_resultado,
        'tipo_documento': tipo_documento,
        'arquivo_record_id': arquivo_record_id,
        'pdf_sha256': pdf_sha256[:16] + '...' if pdf_sha256 else None,
        'idempotency_key': idempotency_key,
        'request_id': request_id,
    }, 200


def _gerar_pacote_assinatura_holerite_ponto(funcionario_id, arquivo_holerite_id, arquivo_ponto_id,
                                             mensagem_extra='', disparar_whatsapp=False, dry_run=False,
                                             request_id=None):
    """
    Pacote atômico de assinatura: Holerite + Folha de Ponto da MESMA
    competência, em UMA solicitação, UM token, UM comprovante.

    Decisão arquitetural (documentada, não presumida): Holerite nunca é
    assinável isolado — só dentro deste pacote, sempre pareado com a
    Folha de Ponto da mesma competência. Reaproveita 100% a tabela
    Assinaturas Digitais e os 4 campos v3.6 já reais no Airtable — ver
    comentário da constante TIPO_PACOTE_HOLERITE_PONTO, acima. Nenhum
    campo novo no Airtable, nenhuma rota nova (usa /assinatura/gerar e
    /assinatura/<hash> já existentes, com branch por tipo_documento).

    Falhas parciais — todas explícitas, nenhuma silenciosamente
    "concluída" (ver auditoria prévia desta Macro, §8):
      - vínculo do colaborador não Ativo -> erro 403, nada criado (checado
        AQUI na preparação/dry-run, de novo imediatamente antes do
        disparo, e de novo em _confirmar_assinatura_pacote_holerite_ponto
        imediatamente antes de concluir a assinatura — 3 checkpoints,
        nenhum depende de lista fixa de nome/telefone)
      - arquivo ausente / não-PDF / de outro funcionário -> erro 4xx, nada criado
      - competência ausente em qualquer um dos 2 docs -> erro 422, nada criado
      - competências divergentes entre os 2 docs -> erro 409, nada criado
      - idempotência cobre TODOS os estados alcançáveis de verdade (corrige
        a lacuna encontrada na auditoria — a versão anterior comparava
        contra nomes ALL-CAPS que a própria gravação nunca escreve; ver
        revisão adversarial desta Macro de fechamento):
          PREPARADO         -> 409 duplicado (criado, ainda sem envio confirmado)
          Pendente          -> 200 link existente (WhatsApp já enviado, aguardando assinatura)
          FALHA_ENVIO       -> 200 link existente p/ reenviar (não duplica)
          Assinado          -> 409 já assinado
          Expirado/Cancelado -> 409 bloqueia recriação automática, exige o
                                 fluxo humano de reenvio já existente
                                 (/assinatura/processar-reenvios, agora
                                 também estendido a este tipo)

    Risco residual declarado (não escondido): Airtable não tem transação
    nem constraint único — duas chamadas concorrentes com o MESMO par de
    documentos podem, em tese, passar a checagem de idempotência ao mesmo
    tempo antes de qualquer uma criar o registro, e cada uma criar seu
    próprio registro com a mesma chave (duplicidade de registro, nunca de
    EFEITO sobre o colaborador — nenhuma delas reenvia duas vezes se a
    aplicação chamadora serializar as duas competências antes de disparar
    o WhatsApp). Mitigação real exigiria um lock externo (Redis, já usado
    pelo Celery) — fora do escopo desta fase; registrado como risco
    residual no relatório final, não como bug corrigido.
    """
    request_id = request_id or f"req{secrets.token_hex(8)}"

    config_ok, config_msg = _validar_configuracao_assinatura_v36()
    if not config_ok:
        logger.error(f'[PACOTE HOL+PONTO] Configuracao invalida: {config_msg}')
        return {
            'status': 'erro',
            'erro': 'Configuracao v3.6 incompleta. Nenhum pacote pode ser criado.',
            'request_id': request_id,
        }, 500

    if not funcionario_id or not arquivo_holerite_id or not arquivo_ponto_id:
        return {
            'status': 'erro',
            'erro': 'funcionario_id, arquivo_holerite_record_id e arquivo_folha_ponto_record_id são obrigatórios',
            'request_id': request_id,
        }, 400

    if arquivo_holerite_id == arquivo_ponto_id:
        return {
            'status': 'erro',
            'erro': 'arquivo_holerite_record_id e arquivo_folha_ponto_record_id não podem ser o mesmo arquivo',
            'request_id': request_id,
        }, 400

    # Checkpoint 1/3 de vínculo ativo: na preparação/dry-run. Bloqueia antes
    # de tocar em qualquer documento — nunca depende de nome/telefone.
    elegivel, motivo_vinculo = _status_funcionario_elegivel(funcionario_id)
    if not elegivel:
        logger.warning(f'[PACOTE HOL+PONTO] Bloqueado na preparação — motivo={motivo_vinculo}')
        return {
            'status': 'erro', 'erro': motivo_vinculo,
            'mensagem': 'Colaborador não está com vínculo ativo confirmado — pacote não pode ser preparado.',
            'request_id': request_id,
        }, 403

    nome_func, whatsapp = _buscar_funcionario_nome_whatsapp(funcionario_id)

    if disparar_whatsapp and not whatsapp:
        logger.warning(f'[PACOTE HOL+PONTO] WhatsApp ausente: {funcionario_id}')
        return {
            'status': 'erro', 'erro': 'whatsapp_ausente',
            'mensagem': f'Funcionário "{nome_func or funcionario_id}" não tem WhatsApp cadastrado.',
            'funcionario_id': funcionario_id,
            'request_id': request_id,
        }, 400

    def _carregar_e_validar(arquivo_id, rotulo):
        _at_throttle()
        r_arq = requests.get(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ARQUIVOS}/{arquivo_id}',
            headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
            params={'returnFieldsByFieldId': 'true'},
            timeout=15,
        )
        if not r_arq.ok:
            return None, ({'status': 'erro', 'erro': f'{rotulo} {arquivo_id} não encontrado',
                            'request_id': request_id}, 404)

        campos_arquivo = r_arq.json().get('fields', {})
        anexos = campos_arquivo.get(F_ARQ_ATTACH, [])
        if not anexos:
            return None, ({'status': 'erro', 'erro': f'{rotulo} {arquivo_id} sem attachment',
                            'request_id': request_id}, 400)

        func_arquivo = campos_arquivo.get('fldxbZwVNa01pchqF', [])  # F_ARQ_FUNC (link)
        if func_arquivo and func_arquivo[0] != funcionario_id:
            return None, ({'status': 'erro', 'erro': f'{rotulo} pertence a outro funcionário',
                            'arquivo_func': func_arquivo[0], 'request_id': request_id}, 403)

        attachment = anexos[0]
        try:
            pdf_bytes = _carregar_documento_url(attachment.get('url'))
        except Exception as exc:
            return None, ({'status': 'erro', 'erro': f'Erro ao carregar {rotulo}: {exc}',
                            'request_id': request_id}, 400)

        return {
            'bytes': pdf_bytes,
            'filename': attachment.get('filename', 'Documento.pdf'),
            'sha256': hashlib.sha256(pdf_bytes).hexdigest(),
        }, None

    holerite, erro_hol = _carregar_e_validar(arquivo_holerite_id, 'Holerite')
    if erro_hol:
        return erro_hol
    ponto, erro_ponto = _carregar_e_validar(arquivo_ponto_id, 'Folha de Ponto')
    if erro_ponto:
        return erro_ponto

    texto_hol = _extrair_texto_pdf_bytes(holerite['bytes'])
    competencia_hol, _ = extrair_competencia_holerite(texto_hol)
    texto_ponto = _extrair_texto_pdf_bytes(ponto['bytes'])
    competencia_ponto, _ = _extrair_competencia_folha_ponto(texto_ponto)

    if not competencia_hol:
        return {
            'status': 'erro', 'erro': 'competencia_holerite_indeterminavel',
            'mensagem': 'Não foi possível ler a competência impressa no Holerite.',
            'request_id': request_id,
        }, 422
    if not competencia_ponto:
        return {
            'status': 'erro', 'erro': 'competencia_folha_ponto_indeterminavel',
            'mensagem': 'Não foi possível ler a competência impressa na Folha de Ponto.',
            'request_id': request_id,
        }, 422
    if competencia_hol != competencia_ponto:
        return {
            'status': 'erro', 'erro': 'competencia_divergente',
            'competencia_holerite': competencia_hol, 'competencia_folha_ponto': competencia_ponto,
            'mensagem': 'Holerite e Folha de Ponto não são da mesma competência — pacote bloqueado.',
            'request_id': request_id,
        }, 409

    competencia = competencia_hol

    idempotency_key = hashlib.sha256(
        f"{funcionario_id}|{competencia}|HOLERITE|{holerite['sha256']}|FOLHA_PONTO|{ponto['sha256']}|v1".encode()
    ).hexdigest()

    _at_throttle()
    try:
        r_existente = requests.get(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}',
            headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
            params={
                'filterByFormula': f'{{{F_ASS_CHAVE_IDEMPOTENCIA}}}="{idempotency_key}"',
                'maxRecords': 1,
                'returnFieldsByFieldId': 'true',
            },
            timeout=30,
        )
        if r_existente.ok and r_existente.json().get('records'):
            rec_existente = r_existente.json()['records'][0]
            status_existente = rec_existente['fields'].get(F_ASS_STATUS)
            existente_id = rec_existente['id']
            logger.info(f'[PACOTE HOL+PONTO] Idempotência: {existente_id} (status: {status_existente})')

            # Casing real (não o nome ALL-CAPS aspiracional de ESTADOS_ASSINATURA):
            # 'PREPARADO' é o único all-caps que a criação de fato escreve;
            # 'Pendente'/'Assinado'/'Expirado' são Title-Case porque é isso
            # que o handler de confirmação (compartilhado, não desta Macro)
            # e o reenvio genérico de fato gravam — comparar contra o
            # all-caps aspiracional aqui faria essas 3 ramificações nunca
            # baterem contra um registro real (achado da revisão adversarial
            # desta macro de fechamento).
            if status_existente == 'PREPARADO':
                return {
                    'status': 'duplicado', 'assinatura_id': existente_id,
                    'motivo': 'idempotência: pacote já criado, aguardando confirmação de envio',
                    'idempotency_key': idempotency_key, 'request_id': request_id,
                }, 409
            if status_existente == 'Pendente':
                return {
                    'status': 'ok', 'assinatura_id': existente_id,
                    'link': f'{RECIBO_BASE_URL}/assinatura/{rec_existente["fields"].get(F_ASS_HASH)}',
                    'motivo': 'retornando link existente', 'request_id': request_id,
                }, 200
            if status_existente == 'Assinado':
                return {
                    'status': 'erro', 'erro': 'pacote_ja_assinado', 'assinatura_id': existente_id,
                    'request_id': request_id,
                }, 409
            if status_existente == 'FALHA_ENVIO':
                return {
                    'status': 'ok', 'assinatura_id': existente_id,
                    'link': f'{RECIBO_BASE_URL}/assinatura/{rec_existente["fields"].get(F_ASS_HASH)}',
                    'motivo': 'falha_envio_anterior — reenviar o mesmo link, não recriar pacote',
                    'request_id': request_id,
                }, 200
            if status_existente in ('Expirado', 'Cancelado'):
                return {
                    'status': 'erro', 'erro': 'pacote_expirado_ou_cancelado', 'assinatura_id': existente_id,
                    'mensagem': (
                        'Pacote com estes 2 documentos já expirou ou foi cancelado. '
                        'Reenvio automático bloqueado — use o fluxo humano de reenvio '
                        '(Status="Reenviar" em Assinaturas Digitais, agora também disponível '
                        'para este tipo) antes de tentar novamente.'
                    ),
                    'request_id': request_id,
                }, 409
    except Exception as exc:
        logger.warning(f'[PACOTE HOL+PONTO] Erro ao verificar idempotência: {exc}')

    hash_token = _gerar_hash_assinatura()
    link = f'{RECIBO_BASE_URL}/assinatura/{hash_token}'

    if dry_run:
        return {
            'status': 'ok', 'dry_run': True,
            'acao': 'criaria_pacote_e_dispararia' if disparar_whatsapp else 'criaria_pacote',
            'link': link, 'funcionario_id': funcionario_id, 'nome_funcionario': nome_func,
            'tipo_documento': TIPO_PACOTE_HOLERITE_PONTO, 'competencia': competencia,
            'holerite': {'arquivo_record_id': arquivo_holerite_id, 'filename': holerite['filename'],
                         'sha256': holerite['sha256'][:16] + '...'},
            'folha_ponto': {'arquivo_record_id': arquivo_ponto_id, 'filename': ponto['filename'],
                            'sha256': ponto['sha256'][:16] + '...'},
            'idempotency_key': idempotency_key, 'request_id': request_id,
        }, 200

    nome_documento = f'Holerite + Folha de Ponto - {competencia}'
    campos = {
        F_ASS_NOME: nome_documento,
        F_ASS_TIPO_DOC: TIPO_PACOTE_HOLERITE_PONTO,
        F_ASS_STATUS: 'PREPARADO',
        F_ASS_HASH: hash_token,
        F_ASS_FUNCIONARIO: [funcionario_id],
        F_ASS_DATA_GERACAO: datetime.now().isoformat(),
        F_ASS_TENTATIVAS: 0,
        F_ASS_ARQUIVO_RECORD_ID: f'{arquivo_holerite_id}{_SEPARADOR_PACOTE}{arquivo_ponto_id}',
        F_ASS_PDF_SHA256: f'{holerite["sha256"]}{_SEPARADOR_PACOTE}{ponto["sha256"]}',
        F_ASS_CHAVE_IDEMPOTENCIA: idempotency_key,
        F_ASS_REQUEST_ID: request_id,
    }

    assinatura_id = _criar_registro(TABLE_ASSINATURAS, campos)
    logger.info(f'[PACOTE HOL+PONTO] Criado: {assinatura_id} | Competência: {competencia} | '
                f'Holerite: {arquivo_holerite_id} | Ponto: {arquivo_ponto_id}')

    _anexar_attachment(TABLE_ASSINATURAS, assinatura_id, F_ASS_DOCUMENTO_PDF, holerite['bytes'], holerite['filename'])
    _anexar_attachment(TABLE_ASSINATURAS, assinatura_id, F_ASS_DOCUMENTO_PDF, ponto['bytes'], ponto['filename'])

    disparo_resultado = None
    if disparar_whatsapp:
        # Checkpoint 2/3 de vínculo ativo: imediatamente antes do disparo,
        # não reaproveita a checagem do início da função — é uma nova
        # consulta, cobrindo o caso real de desligamento no intervalo
        # (mesmo que hoje esse intervalo seja curto dentro da mesma
        # chamada; o reenvio, que pode ocorrer dias depois, reusa este
        # mesmo padrão de checkpoint — ver assinatura_processar_reenvios).
        elegivel_no_disparo, motivo_no_disparo = _status_funcionario_elegivel(funcionario_id)
        if not elegivel_no_disparo:
            logger.warning(f'[PACOTE HOL+PONTO] Disparo abortado — vínculo mudou antes do envio: '
                            f'{assinatura_id} | motivo={motivo_no_disparo}')
            _at_throttle()
            requests.patch(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{assinatura_id}',
                headers=_at_headers(),
                json={'fields': {F_ASS_STATUS: 'Cancelado'}, 'typecast': True}, timeout=15,
            )
            return {
                'status': 'ok', 'dry_run': False, 'acao': 'pacote_criado_e_cancelado',
                'assinatura_id': assinatura_id, 'erro': motivo_no_disparo,
                'mensagem': 'Pacote criado, mas cancelado antes do disparo — vínculo não mais ativo.',
                'tipo_documento': TIPO_PACOTE_HOLERITE_PONTO, 'competencia': competencia,
                'idempotency_key': idempotency_key, 'request_id': request_id,
            }, 200

        try:
            # Correção desta Macro (auditoria de 18-19/08 encontrou a falha):
            # a versão original só mandava TEXTO com o link — o Holerite e a
            # Folha de Ponto nunca chegavam ao colaborador em lugar nenhum
            # (nem WhatsApp, nem a própria página do link, que só pedia CPF
            # sem exibir documento nenhum). Corrigido em 2 passos, nesta
            # ordem, dentro da mesma Macro:
            #   1) passou a mandar os 2 PDFs soltos por WhatsApp + o link;
            #   2) decisão do usuário (19/08/2026, após teste de homologação
            #      real confirmar que a página do link já exibe os 2
            #      documentos embutidos via PDF.js): manda SÓ o link — os
            #      PDFs soltos ficaram redundantes, e um clique único (ver +
            #      assinar no mesmo lugar) é mais simples que 3 mensagens
            #      separadas. Risco aceito conscientemente: exige o Render
            #      sempre acessível (dependência de plano pago, sem
            #      hibernação por inatividade — decisão de infra do
            #      usuário, fora deste código).
            saudacao = f'Olá{(" " + nome_func) if nome_func else ""}!'
            mensagem = (
                f'{saudacao} Seu Holerite e sua Folha de Ponto de {competencia} estão '
                f'prontos. Para visualizar os 2 documentos e concluir a assinatura '
                f'digital, acesse o link abaixo e informe os 4 últimos números do seu '
                f'CPF:\n{link}\n\nMagnata Portaria e Serviços.'
            )
            if mensagem_extra:
                mensagem = f'{mensagem}\n\n{mensagem_extra}'
            _evolution_enviar_texto(whatsapp, mensagem)
            disparo_resultado = 'enviado'
            logger.info(f'[PACOTE HOL+PONTO] WhatsApp enviado (só link, documentos embutidos na página): {assinatura_id} | WhatsApp: {whatsapp}')
            _at_throttle()
            requests.patch(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{assinatura_id}',
                headers=_at_headers(),
                json={'fields': {F_ASS_STATUS: 'Pendente'}, 'typecast': True}, timeout=15,
            )
        except Exception as exc:
            # Log do servidor pode conter o detalhe completo (erro da
            # Evolution API às vezes inclui o próprio número no texto) --
            # o campo persistido no Airtable, exigido "sem PII" por esta
            # Macro, guarda só uma categoria sanitizada.
            logger.error(f'[PACOTE HOL+PONTO] Falha ao disparar WhatsApp: {assinatura_id} | Erro: {exc}')
            disparo_resultado = 'falha'
            _at_throttle()
            requests.patch(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{assinatura_id}',
                headers=_at_headers(),
                json={'fields': {
                    F_ASS_STATUS: 'FALHA_ENVIO',
                    F_ASS_EVIDENCIAS: 'Falha no envio WhatsApp (Evolution API) — ver logs do servidor para detalhe técnico.',
                }, 'typecast': True}, timeout=15,
            )
    else:
        logger.info(f'[PACOTE HOL+PONTO] Criado sem disparo automático: {assinatura_id}')

    return {
        'status': 'ok', 'dry_run': False, 'acao': 'pacote_criado',
        'assinatura_id': assinatura_id, 'link': link, 'hash_token': hash_token,
        'nome_funcionario': nome_func, 'whatsapp': whatsapp,
        'whatsapp_disparo': disparo_resultado,
        'tipo_documento': TIPO_PACOTE_HOLERITE_PONTO, 'competencia': competencia,
        'holerite_sha256': holerite['sha256'][:16] + '...',
        'folha_ponto_sha256': ponto['sha256'][:16] + '...',
        'idempotency_key': idempotency_key, 'request_id': request_id,
    }, 200


@app.route('/assinatura/gerar-lote', methods=['POST', 'OPTIONS'])
def assinatura_gerar_lote():
    """
    ✅ v3.6: Disparo em lote com identidade documental obrigatória.

    Mudanças v3.6:
    - X-API-KEY OBRIGATÓRIA (proteção contra acesso não autorizado)
    - arquivo_record_id OBRIGATÓRIO por item (sem fallback para Kit)
    - Idempotência por SHA-256 do PDF
    - disparar_whatsapp default FALSE

    Protegido por X-API-KEY (EMAIL_WEBHOOK_KEY).

    Body JSON:
      {
        "tipo_documento": "FOLHA_PONTO",          [OBRIGATÓRIO — tipos canônicos]
        "dry_run": false,
        "itens": [
          {
            "funcionario_id": "rec...",           [OBRIGATÓRIO]
            "arquivo_record_id": "rec...",        [OBRIGATÓRIO — Record ID em Arquivos]
            "mensagem_extra": "texto..."          [opcional]
          },
          ...
        ]
      }
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    request_id = f"req{secrets.token_hex(8)}"

    # ✅ NOVO: Validar X-API-KEY
    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        logger.warning(f'[ASSINATURA LOTE] Acesso negado | Request: {request_id}')
        return jsonify({
            'status': 'erro',
            'erro': 'X-API-KEY inválida ou ausente',
            'request_id': request_id,
        }), 401

    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    # ✅ v3.6: Validar configuração de Field IDs
    config_ok, config_msg = _validar_configuracao_assinatura_v36()
    if not config_ok:
        logger.error(f'[ASSINATURA LOTE] Configuracao invalida: {config_msg}')
        return jsonify({
            'status': 'erro',
            'erro': 'Configuracao v3.6 incompleta. Substituir placeholders fldXXXX pelos Field IDs reais do Airtable.',
            'request_id': request_id,
        }), 500

    data = request.get_json(force=True, silent=True) or {}
    tipo_documento = data.get('tipo_documento')
    dry_run = str(data.get('dry_run', False)).strip().lower() in ('1', 'true', 'yes', 'sim')
    itens = data.get('itens') or []

    if not tipo_documento or not itens:
        return jsonify({
            'status': 'erro',
            'erro': 'tipo_documento e itens são obrigatórios',
            'request_id': request_id,
        }), 400

    resultados = []

    for item in itens:
        funcionario_id = item.get('funcionario_id')
        arquivo_record_id = item.get('arquivo_record_id')
        item_request_id = f"{request_id}_{len(resultados)}"

        # ✅ Validar item
        if not funcionario_id:
            resultados.append({
                'status': 'erro',
                'erro': 'funcionario_id ausente no item',
                'request_id': item_request_id,
            })
            continue

        if not arquivo_record_id:
            resultados.append({
                'status': 'erro',
                'funcionario_id': funcionario_id,
                'erro': 'arquivo_record_id ausente no item',
                'request_id': item_request_id,
            })
            continue

        try:
            resultado, _ = _gerar_assinatura_core(
                funcionario_id=funcionario_id,
                tipo_documento=tipo_documento,
                arquivo_record_id=arquivo_record_id,
                mensagem_extra=item.get('mensagem_extra') or '',
                disparar_whatsapp=False,  # ✅ Default FALSE
                dry_run=dry_run,
                request_id=item_request_id,
            )
            resultado['funcionario_id'] = funcionario_id
        except Exception as exc:
            logger.exception(f'[ASSINATURA LOTE] Erro no item {funcionario_id} | Request: {item_request_id}')
            resultado = {
                'status': 'erro',
                'funcionario_id': funcionario_id,
                'erro': f'{type(exc).__name__}: {exc}',
                'request_id': item_request_id,
            }

        resultados.append(resultado)

    logger.info(f'[ASSINATURA LOTE] Processado | Total: {len(itens)} | '
                f'Sucesso: {sum(1 for r in resultados if r.get("status") == "ok")} | '
                f'Request: {request_id}')

    return jsonify({
        'status': 'ok',
        'dry_run': dry_run,
        'total': len(itens),
        'resultados': resultados,
        'request_id': request_id,
    })


@app.route('/admissao/consolidar-kit', methods=['POST', 'OPTIONS'])
def admissao_consolidar_kit():
    """
    Funde manualmente um Kit de Admissão a partir de anexos já recebidos por
    e-mail (tabela Arquivos), quando o agrupamento automático por janela de
    tempo (_buscar_documentos_irmaos_kit_admissao) não rodou — caso real
    encontrado em auditoria (05/07/2026): 3 admissões (08/06) tiveram os 3
    documentos capturados em "Arquivos" (status Recebido) mas nunca
    passaram por "Processar Arquivos", então o agrupamento automático nunca
    disparou e o Kit ficou parado sem sequer virar Pendência.

    Sem X-API-KEY (uso administrativo manual, mesmo padrão de
    /processar-holerites).

    Body JSON:
      {
        "funcionario_id": "rec...",
        "ficha_url": "https://...",      [opcional]
        "contrato_url": "https://...",   [opcional]
        "vt_url": "https://...",         [opcional]
        "nome_arquivo": "Kit Admissao - Fulano.pdf"  [opcional]
      }
    Anexa o PDF fundido em Kit Admissão - PDF Consolidado (F_FUNC_KIT_CONSOLIDADO)
    e em Documentos Admissionais (F_FUNC_DOCS_ADMISSAO) do Funcionário.
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    data = request.get_json(force=True, silent=True) or {}
    funcionario_id = data.get('funcionario_id')
    if not funcionario_id:
        return jsonify({'status': 'erro', 'erro': 'funcionario_id é obrigatório'}), 400

    ficha_url = data.get('ficha_url') or ''
    contrato_url = data.get('contrato_url') or ''
    vt_url = data.get('vt_url') or ''
    if not (ficha_url or contrato_url or vt_url):
        return jsonify({'status': 'erro', 'erro': 'informe ao menos uma URL (ficha_url/contrato_url/vt_url)'}), 400

    pdf_ficha = requests.get(ficha_url, timeout=60).content if ficha_url else None
    pdf_contrato = requests.get(contrato_url, timeout=60).content if contrato_url else None
    pdf_vt = requests.get(vt_url, timeout=60).content if vt_url else None

    pdf_consolidado = _montar_pdf_consolidado_kit_admissao(pdf_ficha, pdf_contrato, pdf_vt)

    _at_throttle()
    r_func = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}/{funcionario_id}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        params={'returnFieldsByFieldId': 'true'}, timeout=15,
    )
    nome_func = (r_func.json().get('fields', {}).get(F_FUNC_NOME) if r_func.ok else None) or funcionario_id
    nome_arquivo = data.get('nome_arquivo') or f'Kit Admissao - {nome_func}.pdf'

    _anexar_attachment(TABLE_FUNC, funcionario_id, F_FUNC_KIT_CONSOLIDADO, pdf_consolidado, nome_arquivo)
    _anexar_attachment(TABLE_FUNC, funcionario_id, F_FUNC_DOCS_ADMISSAO, pdf_consolidado, nome_arquivo)

    return jsonify({
        'status': 'ok', 'funcionario_id': funcionario_id, 'nome_funcionario': nome_func,
        'paginas_incluidas': sum(1 for p in (pdf_ficha, pdf_contrato, pdf_vt) if p),
        'tamanho_bytes': len(pdf_consolidado),
    })


MAX_REENVIOS_PACOTE = 5  # mesmo teto de MAX_TENTATIVAS_ASSINATURA, por consistência

_REENVIO_PACOTE_REGEX = re.compile(r'Reenvios do pacote:\s*(\d+)')


def _contar_reenvios_pacote(evidencias_texto: str) -> int:
    m = _REENVIO_PACOTE_REGEX.search(evidencias_texto or '')
    return int(m.group(1)) if m else 0


def _reenviar_pacote_holerite_ponto(rec: dict, fields: dict, func_id: str, dry_run: bool, disparar_whatsapp: bool) -> dict:
    """
    Extensão mínima do mecanismo de reenvio já existente
    (/assinatura/processar-reenvios) para o tipo HOLERITE_FOLHA_PONTO.

    Reaproveita a MESMA transação e chave de idempotência (nunca cria um
    registro novo — só faz PATCH no registro existente, então nunca gera
    uma segunda assinatura para o mesmo pacote). Isolado do reenvio
    genérico: chamado a partir de um branch por tipo, sem alterar
    nenhuma linha do reenvio de Kit Admissão/Rescisão/EPI/Contratos/
    Folha de Ponto isolada.
    """
    item = {'assinatura_id': rec['id'], 'nome': fields.get('Nome')}

    # Revalida vínculo ativo antes de reenviar -- nunca ressuscita pacote
    # de colaborador que deixou de ser elegível desde a preparação.
    elegivel, motivo_vinculo = _status_funcionario_elegivel(func_id)
    if not elegivel:
        item['erro'] = motivo_vinculo
        item['acao'] = 'cancelado_vinculo_nao_ativo'
        if not dry_run:
            _at_throttle()
            requests.patch(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{rec["id"]}',
                headers=_at_headers(), json={'fields': {F_ASS_STATUS: 'Cancelado'}, 'typecast': True}, timeout=15,
            )
        return item

    nome_func, whatsapp = _buscar_funcionario_nome_whatsapp(func_id)
    if not whatsapp:
        item['erro'] = 'whatsapp_ausente'
        if not dry_run:
            _at_throttle()
            requests.patch(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{rec["id"]}',
                headers=_at_headers(), json={'fields': {F_ASS_STATUS: 'FALHA_ENVIO'}, 'typecast': True}, timeout=15,
            )
        return item

    # Limite de tentativas de reenvio (contador embutido em Evidencias_
    # Assinatura -- sem campo novo no Airtable, ver decisão registrada em
    # docs/decisoes/pacote-holerite-folha-ponto.md). Checado ANTES de
    # revalidar documentos -- um pacote já esgotado nunca precisa baixar
    # nada de novo para ser bloqueado.
    reenvios_ate_agora = _contar_reenvios_pacote(fields.get('Evidencias_Assinatura'))
    if reenvios_ate_agora >= MAX_REENVIOS_PACOTE:
        item['erro'] = 'limite_de_reenvios_excedido'
        item['acao'] = 'bloqueado_para_revisao'
        if not dry_run:
            _at_throttle()
            requests.patch(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{rec["id"]}',
                headers=_at_headers(), json={'fields': {F_ASS_STATUS: 'Expirado'}, 'typecast': True}, timeout=15,
            )
        return item

    # Revalida os 2 Record IDs e hashes: os anexos ORIGINAIS ainda
    # precisam corresponder ao par gravado na criação. Um pacote em
    # reenvio nunca deveria ter sido assinado (chegaria aqui só se
    # Status="Reenviar", nunca "Assinado"), então os originais devem
    # continuar lá -- mas se algo os alterou por fora (documento
    # corrigido depois da preparação), o reenvio é bloqueado, não
    # silenciosamente aceito.
    doc_atual = fields.get('Documento PDF') or []
    anexos_originais = [a for a in doc_atual if not a['filename'].startswith('Comprovante Assinatura')
                         and ' - ASSINADO' not in a['filename']]
    hashes_gravados = (fields.get('PDF SHA-256') or '').split(_SEPARADOR_PACOTE)
    if len(anexos_originais) != 2 or len(hashes_gravados) != 2:
        item['erro'] = 'estrutura_inesperada'
        item['acao'] = 'bloqueado_para_revisao'
        return item

    if not dry_run:
        try:
            hashes_atuais = set()
            for att in anexos_originais:
                r_doc = requests.get(att['url'], timeout=60)
                if not r_doc.ok:
                    raise RuntimeError('download falhou')
                hashes_atuais.add(hashlib.sha256(r_doc.content).hexdigest())
            if hashes_atuais != set(hashes_gravados):
                item['erro'] = 'documento_alterado_apos_preparacao'
                item['acao'] = 'bloqueado_documento_trocado'
                return item
        except Exception as exc:
            logger.error(f'[PACOTE HOL+PONTO] Reenvio: falha ao revalidar documentos de {rec["id"]}: {exc}')
            item['erro'] = 'falha_ao_revalidar_documentos'
            return item

    novo_hash = _gerar_hash_assinatura()
    link = f'{RECIBO_BASE_URL}/assinatura/{novo_hash}'
    item['novo_hash'] = novo_hash
    item['link'] = link
    item['reenvios_apos_este'] = reenvios_ate_agora + 1

    if dry_run:
        item['acao'] = 'reenviaria_pacote'
        item['disparar_whatsapp'] = disparar_whatsapp
        return item

    competencia_pacote = (fields.get('Nome') or '').replace('Holerite + Folha de Ponto - ', '')
    mensagem = (
        f'Olá{(" " + nome_func) if nome_func else ""}! Você tem 2 documentos pendentes de '
        f'confirmação — Holerite e Folha de Ponto de {competencia_pacote}.\n\nPara confirmar o '
        f'recebimento dos dois, acesse o link abaixo e informe os 4 últimos números do seu '
        f'CPF:\n{link}\n\nMagnata Portaria e Serviços.'
    )

    # Concorrência best-effort: relê o Status imediatamente antes de
    # gravar, reduzindo (não eliminando -- mesma limitação já declarada
    # do resto do pacote) a janela de corrida com outra execução do
    # mesmo reenvio.
    _at_throttle()
    r_recheck = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{rec["id"]}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'}, timeout=15,
    )
    if r_recheck.ok and r_recheck.json().get('fields', {}).get('Status') != 'Reenviar':
        item['acao'] = 'ja_processado_outra_instancia'
        item['status_atual'] = r_recheck.json().get('fields', {}).get('Status')
        return item

    whatsapp_enviado = False
    try:
        if disparar_whatsapp:
            _evolution_enviar_texto(whatsapp, mensagem)
            whatsapp_enviado = True
        evidencias_novas = f'Reenvios do pacote: {reenvios_ate_agora + 1}'
        _at_throttle()
        requests.patch(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{rec["id"]}',
            headers=_at_headers(),
            json={'fields': {
                F_ASS_HASH: novo_hash,
                F_ASS_STATUS: 'Pendente' if whatsapp_enviado or not disparar_whatsapp else 'FALHA_ENVIO',
                F_ASS_TENTATIVAS: 0,
                F_ASS_EVIDENCIAS: evidencias_novas,
            }, 'typecast': True}, timeout=15,
        )
        item['acao'] = 'reenviado'
        item['disparar_whatsapp'] = disparar_whatsapp
        item['whatsapp_enviado'] = whatsapp_enviado
        logger.info(f'[PACOTE HOL+PONTO] Reenvio OK — {rec["id"]} novo_hash={novo_hash[:8]}...')
    except Exception as exc:
        logger.error(f'[PACOTE HOL+PONTO] falha no reenvio {rec["id"]}: {exc}')
        item['erro'] = 'falha_ao_reenviar'
        _at_throttle()
        requests.patch(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{rec["id"]}',
            headers=_at_headers(),
            json={'fields': {
                F_ASS_STATUS: 'FALHA_ENVIO',
                F_ASS_EVIDENCIAS: f'Reenvios do pacote: {reenvios_ate_agora + 1}\nÚltimo erro (reenvio): falha_ao_reenviar',
            }, 'typecast': True}, timeout=15,
        )

    return item


@app.route('/assinatura/processar-reenvios', methods=['POST', 'OPTIONS'])
def assinatura_processar_reenvios():
    """
    Cenário preventivo 2 (Lógica de Reenvio): varre "Assinaturas Digitais"
    procurando registros com Status="Reenviar" (definido manualmente por um
    humano direto no Airtable) e, para cada um: zera o contador de
    tentativas, gera um NOVO Hash Token (o anterior pode ter sido o que
    esgotou as tentativas — por segurança não se reaproveita), volta o
    Status para "Pendente" e dispara um novo link via WhatsApp (controlável).

    Reaplica a mesma validação de telefone do cenário 1 — se o WhatsApp
    estiver vazio, marca erro e não tenta enviar.

    Protegido por X-API-KEY (EMAIL_WEBHOOK_KEY). Suporta dry_run, limit e disparar_whatsapp.

    v3.5 (correções de segurança):
    - Parametrização de disparar_whatsapp (default=True, backward compatible)
    - Proteção contra duplicação: verifica Status antes de processar (idempotente)
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401

    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY não configurada'}), 500

    data = request.get_json(force=True, silent=True) or {}
    dry_run = str(data.get('dry_run', True)).strip().lower() in ('1', 'true', 'yes', 'sim')
    disparar_whatsapp = str(data.get('disparar_whatsapp', True)).strip().lower() in ('1', 'true', 'yes', 'sim')
    limit = data.get('limit')

    _at_throttle()
    r = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        params={'filterByFormula': '{Status}="Reenviar"'},
        timeout=30,
    )
    r.raise_for_status()
    registros = r.json().get('records', [])
    if limit:
        registros = registros[:int(limit)]

    resultado = {'status': 'ok', 'dry_run': dry_run, 'disparar_whatsapp': disparar_whatsapp, 'encontrados': len(registros), 'processados': []}

    for rec in registros:
        fields = rec.get('fields', {})
        item = {'assinatura_id': rec['id'], 'nome': fields.get('Nome')}

        # PROTEÇÃO CONTRA DUPLICAÇÃO (idempotência v3.5):
        # Se Status já não é "Reenviar", alguém/algo já processou este registro
        status_atual = fields.get('Status', 'Pendente')
        if status_atual != 'Reenviar':
            item['acao'] = 'ja_processado_outra_instancia'
            item['status_atual'] = status_atual
            logger.warning(f'[ASSINATURA] Reenvio saltado — {rec["id"]} já não está com Status=Reenviar (Status atual={status_atual})')
            resultado['processados'].append(item)
            continue

        func_links = fields.get('Funcionário') or []
        func_id = func_links[0]['id'] if func_links and isinstance(func_links[0], dict) else (func_links[0] if func_links else None)
        if not func_id:
            item['erro'] = 'sem_funcionario_vinculado'
            resultado['processados'].append(item)
            continue

        # Pacote Holerite + Folha de Ponto: branch isolado, não altera
        # nenhuma linha do reenvio genérico abaixo (usado por Kit
        # Admissão/Rescisão/EPI/Contratos/Folha de Ponto isolada).
        if fields.get('Tipo de Documento') == TIPO_PACOTE_HOLERITE_PONTO:
            resultado['processados'].append(
                _reenviar_pacote_holerite_ponto(rec, fields, func_id, dry_run, disparar_whatsapp)
            )
            continue

        nome_func, whatsapp = _buscar_funcionario_nome_whatsapp(func_id)
        if not whatsapp:
            item['erro'] = 'whatsapp_ausente'
            if not dry_run:
                requests.patch(
                    f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{rec["id"]}',
                    headers=_at_headers(), json={'fields': {F_ASS_STATUS: 'Pendente'}, 'typecast': True}, timeout=15,
                )
            resultado['processados'].append(item)
            continue

        novo_hash = _gerar_hash_assinatura()
        link = f'{RECIBO_BASE_URL}/assinatura/{novo_hash}'
        item['novo_hash'] = novo_hash
        item['link'] = link

        if dry_run:
            item['acao'] = 'reenviaria'
            item['disparar_whatsapp'] = disparar_whatsapp
            resultado['processados'].append(item)
            continue

        try:
            mensagem = _montar_mensagem_assinatura(nome_func, fields.get('Tipo de Documento', 'documento'), link)

            # PARAMETRIZAÇÃO v3.5: respeita disparar_whatsapp
            whatsapp_enviado = False
            if disparar_whatsapp:
                _evolution_enviar_texto(whatsapp, mensagem)
                whatsapp_enviado = True
                logger.info(f'[ASSINATURA] Reenvio WhatsApp disparado — {rec["id"]} para {whatsapp}')
            else:
                logger.info(f'[ASSINATURA] Reenvio gerado mas WhatsApp NÃO disparado — {rec["id"]} (disparar_whatsapp=False)')

            requests.patch(
                f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{rec["id"]}',
                headers=_at_headers(),
                json={'fields': {
                    F_ASS_HASH: novo_hash, F_ASS_STATUS: 'Pendente', F_ASS_TENTATIVAS: 0,
                }, 'typecast': True}, timeout=15,
            )
            item['acao'] = 'reenviado'
            item['disparar_whatsapp'] = disparar_whatsapp
            item['whatsapp_enviado'] = whatsapp_enviado
            logger.info(f'[ASSINATURA] Reenvio OK — {rec["id"]} novo_hash={novo_hash[:8]}... whatsapp_enviado={whatsapp_enviado}')
        except Exception as exc:
            item['erro'] = str(exc)
            logger.error(f'[ASSINATURA] falha no reenvio {rec["id"]}: {exc}')

        resultado['processados'].append(item)

    return jsonify(resultado)


def _confirmar_assinatura_pacote_holerite_ponto(registro, fields, hash_token, nome_doc, func_id,
                                                 cpf_informado, ip_captura, user_agent, agora_brt_dt,
                                                 nome_func=''):
    """
    Confirmação de assinatura do PACOTE Holerite + Folha de Ponto — ordem de
    escrita DIFERENTE do fluxo de documento único (esse fluxo, compartilhado
    por Kit Admissão/Rescisão/EPI/Contratos/Folha de Ponto isolada,
    permanece inalterado): só grava Status="Assinado" DEPOIS que os 2
    documentos foram revalidados, carimbados e o comprovante único foi
    gerado e persistido — nunca antes.

    Correção desta Macro de fechamento (revisão adversarial encontrou o
    problema): a versão anterior gravava "Assinado" incondicionalmente
    ANTES de tentar carimbar (código compartilhado, no topo da rota) e só
    depois tentava o carimbo/comprovante em bloco isolado — uma falha
    parcial ali (documento trocado, exceção no carimbo, falha ao persistir
    o comprovante) deixava o pacote marcado como concluído mesmo sem ter
    concluído nada. Agora a escrita de "Assinado" é uma ÚNICA chamada,
    feita só no fim, junto com a lista final de anexos e as evidências —
    se essa chamada não for alcançada ou falhar, nada fica marcado como
    assinado (fica no status anterior, seguro para nova tentativa).
    """
    agora_brt = agora_brt_dt.isoformat()
    dt_fmt = agora_brt_dt.strftime('%d/%m/%Y %H:%M')

    erro_generico = _pagina_assinatura_html(
        hash_token, nome_doc,
        erro='Não foi possível concluir a confirmação agora. Tente novamente em alguns minutos ou contate o RH.'
    )

    # Checkpoint 3/3 de vínculo ativo: imediatamente antes de concluir a
    # assinatura. Mudança de status entre preparação/envio e este clique
    # invalida o pacote em vez de assinar — a mesma idempotência que já
    # bloqueia Expirado/Cancelado impede que uma reexecução ressuscite ou
    # reenvie automaticamente este pacote depois disso.
    elegivel, motivo_vinculo = _status_funcionario_elegivel(func_id)
    if not elegivel:
        logger.warning(f'[PACOTE HOL+PONTO] Assinatura bloqueada — vínculo não ativo: '
                        f'{registro["id"]} | motivo={motivo_vinculo}')
        _at_throttle()
        requests.patch(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{registro["id"]}',
            headers=_at_headers(), json={'fields': {F_ASS_STATUS: 'Cancelado'}, 'typecast': True}, timeout=15,
        )
        return _pagina_assinatura_html(hash_token, nome_doc, erro='Não foi possível concluir a confirmação. Contate o RH.',
                                        mostrar_formulario=False)

    doc_atual = fields.get('Documento PDF') or []
    anexos_originais = [a for a in doc_atual if not a['filename'].startswith('Comprovante Assinatura')
                         and ' - ASSINADO' not in a['filename']]
    hashes_gravados = (fields.get('PDF SHA-256') or '').split(_SEPARADOR_PACOTE)

    if len(anexos_originais) != 2 or len(hashes_gravados) != 2:
        logger.error(
            f'[PACOTE HOL+PONTO] Estrutura inesperada no registro {registro["id"]}: '
            f'{len(anexos_originais)} anexo(s) original(is), {len(hashes_gravados)} hash(es) gravado(s) '
            f'— esperado 2 e 2. Assinatura NÃO registrada — caso isolado, requer revisão manual, '
            f'não bloqueia outros colaboradores.'
        )
        return erro_generico

    # nome_func vem do próprio r_func já buscado pela rota que chamou esta
    # função (validação de CPF) — evita uma 2ª chamada redundante à mesma
    # Funcionário para o mesmo request.

    try:
        baixados = {}
        for att in anexos_originais:
            r_doc_pacote = requests.get(att['url'], timeout=60)
            if not r_doc_pacote.ok:
                raise RuntimeError(f'download falhou para anexo do pacote {registro["id"]}')
            conteudo = r_doc_pacote.content
            sha_real = hashlib.sha256(conteudo).hexdigest()
            baixados[sha_real] = {'bytes': conteudo, 'filename': att['filename'], 'id': att['id']}

        # Bloqueio contra documento trocado: os 2 hashes baixados AGORA
        # precisam ser exatamente os 2 gravados na criação do pacote.
        if set(baixados.keys()) != set(hashes_gravados):
            logger.error(
                f'[PACOTE HOL+PONTO] BLOQUEIO — hash dos anexos não corresponde ao par gravado '
                f'na criação do pacote {registro["id"]}. Assinatura NÃO registrada (documento '
                f'possivelmente trocado após a preparação).'
            )
            return erro_generico

        holerite_dl = baixados[hashes_gravados[0]]
        ponto_dl = baixados[hashes_gravados[1]]

        def _nome_carimbado(nome_original):
            return (nome_original[:-4] + ' - ASSINADO.pdf'
                    if nome_original.lower().endswith('.pdf') else nome_original + ' - ASSINADO')

        pdf_hol_carimbado = _carimbar_pdf_assinado(holerite_dl['bytes'], nome_func, cpf_informado, dt_fmt)
        pdf_ponto_carimbado = _carimbar_pdf_assinado(ponto_dl['bytes'], nome_func, cpf_informado, dt_fmt)
        nome_hol_carimbado = _nome_carimbado(holerite_dl['filename'])
        nome_ponto_carimbado = _nome_carimbado(ponto_dl['filename'])

        competencia_pacote = nome_doc.replace('Holerite + Folha de Ponto - ', '') if nome_doc else ''
        pdf_comprovante_pacote = _gerar_comprovante_assinatura_pacote_pdf(
            competencia=competencia_pacote, func_nome=nome_func, ip=ip_captura, user_agent=user_agent,
            dt_str=dt_fmt, cpf4=cpf_informado,
            doc1_bytes=holerite_dl['bytes'], doc1_nome=holerite_dl['filename'],
            doc2_bytes=ponto_dl['bytes'], doc2_nome=ponto_dl['filename'],
        )
        nome_comprovante_pacote = f'Comprovante Assinatura - {(nome_func or "Colaborador").title()} - {competencia_pacote}.pdf'

        upload_hol = _anexar_attachment(TABLE_ASSINATURAS, registro['id'], F_ASS_DOCUMENTO_PDF, pdf_hol_carimbado, nome_hol_carimbado)
        upload_ponto = _anexar_attachment(TABLE_ASSINATURAS, registro['id'], F_ASS_DOCUMENTO_PDF, pdf_ponto_carimbado, nome_ponto_carimbado)
        upload_comp = _anexar_attachment(TABLE_ASSINATURAS, registro['id'], F_ASS_DOCUMENTO_PDF, pdf_comprovante_pacote, nome_comprovante_pacote)

        lista_pos_upload = (upload_comp.get('fields', {}) or {}).get(F_ASS_DOCUMENTO_PDF) or []
        nomes_manter = (nome_hol_carimbado, nome_ponto_carimbado, nome_comprovante_pacote)
        ids_novos = [a['id'] for a in lista_pos_upload if a['filename'] in nomes_manter]
        if len(ids_novos) != 3:
            logger.error(f'[PACOTE HOL+PONTO] upload dos 3 anexos finais incompleto: {registro["id"]} '
                         f'({len(ids_novos)}/3) — Assinatura NÃO registrada.')
            return erro_generico

        ids_originais = {holerite_dl['id'], ponto_dl['id']}
        lista_final = ids_novos + [a['id'] for a in doc_atual if a['id'] not in ids_originais]

        evidencias = (
            f'IP: {ip_captura}\n'
            f'Timestamp: {dt_fmt} (horário de Brasília)\n'
            f'User-Agent: {user_agent}\n'
            f'CPF confirmado (4 últimos dígitos): {cpf_informado}\n'
            f'Declaração: "Ao clicar, concordo eletronicamente com o recebimento e os termos '
            f'dos dois documentos (Holerite e Folha de Ponto), sob as penas da lei."'
        )

        # Escrita ÚNICA e atômica do lado da aplicação: Status="Assinado"
        # só é gravado JUNTO com a lista final de anexos e as evidências —
        # nunca em passo separado (isso é o que corrige o achado de §4).
        _at_throttle()
        r_patch_final = requests.patch(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{registro["id"]}',
            headers=_at_headers(),
            json={'fields': {
                F_ASS_STATUS: 'Assinado',
                F_ASS_CPF_INFORMADO: cpf_informado,
                F_ASS_IP: ip_captura,
                F_ASS_USER_AGENT: user_agent,
                F_ASS_DATA_ASSINATURA: agora_brt,
                F_ASS_EVIDENCIAS: evidencias,
                F_ASS_DOCUMENTO_PDF: [{'id': i} for i in lista_final],
            }, 'typecast': True}, timeout=30,
        )
        if not r_patch_final.ok:
            logger.error(f'[PACOTE HOL+PONTO] falha ao gravar assinatura final {registro["id"]}: {r_patch_final.text[:300]}')
            return erro_generico

        # Cópia para o Funcionário — melhor esforço; não afeta o que já foi
        # gravado como fonte da verdade (Assinaturas Digitais).
        try:
            _anexar_attachment(TABLE_FUNC, func_id, F_FUNC_DOCUMENTOS, pdf_hol_carimbado, nome_hol_carimbado)
            _anexar_attachment(TABLE_FUNC, func_id, F_FUNC_DOCUMENTOS, pdf_ponto_carimbado, nome_ponto_carimbado)
            _anexar_attachment(TABLE_FUNC, func_id, F_FUNC_DOCUMENTOS, pdf_comprovante_pacote, nome_comprovante_pacote)
        except Exception as exc:
            logger.warning(f'[PACOTE HOL+PONTO] assinado, mas cópia para o Funcionário falhou (não bloqueante): '
                            f'{registro["id"]} | {exc}')

        # PDF único combinado (Holerite + Folha de Ponto + Comprovante), para
        # o botão de download na tela de sucesso — correção desta Macro.
        # Melhor esforço: se falhar, a assinatura já está gravada e correta
        # (bloco acima), só não aparece o botão de download imediato; os 3
        # arquivos separados continuam disponíveis normalmente no registro.
        url_combinado = None
        try:
            merged_writer = PdfWriter()
            for pdf_parte in (pdf_hol_carimbado, pdf_ponto_carimbado, pdf_comprovante_pacote):
                merged_writer.append(io.BytesIO(pdf_parte))
            buf_combinado = io.BytesIO()
            merged_writer.write(buf_combinado)
            pdf_combinado_bytes = buf_combinado.getvalue()
            merged_writer.close()

            nome_combinado = f'Combinado Carimbado - {(nome_func or "Colaborador").title()} - {competencia_pacote}.pdf'
            upload_combinado = _anexar_attachment(TABLE_ASSINATURAS, registro['id'], F_ASS_DOCUMENTO_PDF,
                                                   pdf_combinado_bytes, nome_combinado)
            for att in (upload_combinado.get('fields', {}) or {}).get(F_ASS_DOCUMENTO_PDF) or []:
                if att['filename'] == nome_combinado:
                    url_combinado = att['url']
                    break
        except Exception as exc:
            logger.warning(f'[PACOTE HOL+PONTO] assinado, mas PDF combinado p/ download falhou (não bloqueante): '
                            f'{registro["id"]} | {exc}')

        logger.info(f'[PACOTE HOL+PONTO] Assinado com sucesso: {registro["id"]}')
        downloads = ([{'rotulo': '⬇ Baixar PDF Combinado Carimbado', 'url': url_combinado}]
                     if url_combinado else None)
        return _pagina_assinatura_sucesso_html(nome_doc, _fmt_quando(agora_brt), downloads=downloads)

    except Exception as exc:
        logger.warning(f'[PACOTE HOL+PONTO] falha ao processar assinatura do pacote {registro["id"]}: {exc}')
        return erro_generico


# Marca d'água da pré-visualização — texto fixo, nunca configurável por
# request (não é conteúdo do documento, é rótulo de estado).
_MARCA_DAGUA_PREVIEW = 'PRÉ-VISUALIZAÇÃO — NÃO ASSINADO'


def _contar_paginas_pdf(pdf_bytes: bytes) -> int:
    """Só conta páginas — não renderiza nada. Usado pelo GET da página de
    assinatura, que precisa saber quantas <img> emitir mas não precisa das
    imagens em si (cada uma é pedida depois, sob demanda, pelo navegador)."""
    import pypdfium2 as _pdfium
    documento = _pdfium.PdfDocument(pdf_bytes)
    try:
        return len(documento)
    finally:
        documento.close()


def _pdf_para_paginas_png(pdf_bytes: bytes, escala: float = 1.4, apenas: int = None):
    """
    Converte cada página de um PDF em PNG, com marca d'água diagonal de
    pré-visualização. Retorna lista de bytes PNG (1 por página).

    Decisão desta Macro (pedido do usuário, 19-20/08/2026): a
    pré-visualização da página de assinatura passa a ser IMAGEM, não o PDF
    original. Motivo: enquanto o navegador recebe o PDF, ele sempre pode
    salvá-lo (trocando a URL na barra de endereço, pelo visualizador
    nativo, etc.) ANTES de assinar — o que contraria a intenção de que o
    colaborador só fique com o documento depois de confirmar. Com imagens
    não existe arquivo PDF do lado do colaborador; o PDF só aparece depois
    de assinar, no "Baixar PDF Combinado Carimbado".

    Efeito colateral positivo: elimina a dependência do PDF.js via CDN
    (bloqueado por política corporativa em pelo menos um ambiente real,
    ver histórico desta Macro) e, com ela, o visualizador nativo do
    navegador que aparecia no fallback.

    Limite declarado, não escondido: captura de tela continua possível —
    isso vale para qualquer site e não tem solução técnica. A marca
    d'água serve justamente para que uma cópia obtida assim fique
    visivelmente distinta do documento oficial carimbado.

    Import tardio e proposital: pypdfium2/Pillow chegam como dependência
    transitiva do pdfplumber (confirmado instalado no build do Render),
    não estão fixados em requirements.txt. Se algum dia sumirem, esta
    função levanta e o chamador cai no caminho antigo (servir o PDF) em
    vez de derrubar a rota inteira.
    """
    import pypdfium2 as _pdfium
    from PIL import Image as _Image, ImageDraw as _ImageDraw, ImageFont as _ImageFont

    documento = _pdfium.PdfDocument(pdf_bytes)
    paginas = []
    try:
        # `apenas` (1-based) renderiza uma única página. Sem isso, cada
        # requisição de imagem re-renderizaria o documento inteiro só para
        # devolver uma página — num documento de 5 páginas seriam 25
        # renderizações por colaborador (medido: ~0,7s cada).
        if apenas is not None:
            if not (1 <= apenas <= len(documento)):
                return []
            indices = [apenas - 1]
        else:
            indices = range(len(documento))
        for indice in indices:
            imagem = documento[indice].render(scale=escala).to_pil().convert('RGBA')

            # A camada da marca é desenhada MAIOR que a página e depois
            # recortada no centro: girar uma camada do tamanho exato corta
            # as frases nas bordas e deixa palavras pela metade
            # ("LIZAÇÃO —"), o que fica visualmente desleixado.
            lado = int((imagem.width ** 2 + imagem.height ** 2) ** 0.5) + 40
            camada = _Image.new('RGBA', (lado, lado), (255, 255, 255, 0))
            desenho = _ImageDraw.Draw(camada)
            try:
                fonte = _ImageFont.truetype('DejaVuSans-Bold.ttf', max(18, imagem.width // 28))
            except Exception:
                fonte = _ImageFont.load_default()

            largura_texto = desenho.textlength(_MARCA_DAGUA_PREVIEW, font=fonte)
            # Marca repetida cobrindo a camada inteira: uma única faixa
            # central seria fácil de recortar de uma captura de tela.
            passo_y = max(140, imagem.height // 6)
            passo_x = int(largura_texto) + max(60, imagem.width // 12)
            for linha, pos_y in enumerate(range(0, lado, passo_y)):
                deslocamento = -(passo_x // 2) if linha % 2 else 0
                for pos_x in range(deslocamento, lado, passo_x):
                    desenho.text((pos_x, pos_y), _MARCA_DAGUA_PREVIEW,
                                 font=fonte, fill=(190, 30, 30, 70))

            camada = camada.rotate(28, resample=_Image.BICUBIC, expand=False)
            esq = (lado - imagem.width) // 2
            topo = (lado - imagem.height) // 2
            camada = camada.crop((esq, topo, esq + imagem.width, topo + imagem.height))
            composta = _Image.alpha_composite(imagem, camada).convert('RGB')

            saida = io.BytesIO()
            composta.save(saida, format='PNG', optimize=True)
            paginas.append(saida.getvalue())
    finally:
        documento.close()
    return paginas


# Cache em memória dos PDFs originais do pacote, por URL de anexo.
#
# Motivo (incidente de 20/08/2026): cada PÁGINA que o colaborador vê é uma
# requisição própria, e cada uma baixava o PDF INTEIRO do Airtable de novo.
# Um pacote de 5 páginas baixava os mesmos 2 arquivos 5 vezes. Com o lote de
# 102 pessoas clicando junto num servidor de 0,5 vCPU, isso vira fila e a
# tela fica em branco esperando.
#
# Escopo deliberadamente pequeno: guarda só os bytes já baixados, por URL
# assinada do Airtable (que já é de curta duração), com teto de itens e TTL
# curto. NÃO guarda o registro do Airtable — a checagem de Status/Tipo
# continua sendo feita a cada requisição, então assinar continua refletindo
# na hora. Cada worker tem o seu; não há estado compartilhado.
_CACHE_PDF_MAX = 8
_CACHE_PDF_TTL_S = 300
_cache_pdf = OrderedDict()   # url -> (expira_em_epoch, bytes)


def _pdf_de_anexo_com_cache(url: str) -> bytes:
    agora = time.time()
    item = _cache_pdf.get(url)
    if item and item[0] > agora:
        _cache_pdf.move_to_end(url)
        return item[1]

    conteudo = _carregar_documento_url(url)

    # Limpa expirados antes de inserir, para o teto não guardar lixo.
    for chave in [k for k, v in _cache_pdf.items() if v[0] <= agora]:
        _cache_pdf.pop(chave, None)
    _cache_pdf[url] = (agora + _CACHE_PDF_TTL_S, conteudo)
    _cache_pdf.move_to_end(url)
    while len(_cache_pdf) > _CACHE_PDF_MAX:
        _cache_pdf.popitem(last=False)
    return conteudo


def _pdf_original_do_pacote(fields, idx):
    """
    Devolve (pdf_bytes, erro) do documento `idx` (0=Holerite, 1=Folha de
    Ponto) de um registro de pacote — mesma seleção de anexos originais
    usada pela rota GET e pela confirmação. Nunca usa os arquivos
    carimbados nem o comprovante.
    """
    doc_atual = fields.get('Documento PDF') or []
    anexos_originais = [
        a for a in doc_atual
        if not a['filename'].startswith('Comprovante Assinatura')
        and not a['filename'].startswith('Combinado Carimbado')
        and ' - ASSINADO' not in a['filename']
    ]
    if len(anexos_originais) != 2 or idx not in (0, 1):
        return None, 'estrutura_inesperada'
    try:
        return _pdf_de_anexo_com_cache(anexos_originais[idx]['url']), None
    except Exception as exc:
        return None, str(exc)


@app.route('/assinatura/<hash_token>/doc/<int:idx>/pagina/<int:num>', methods=['GET'])
def assinatura_documento_pagina_png(hash_token, idx, num):
    """
    Uma página do documento `idx` do pacote, como PNG com marca d'água de
    pré-visualização — ver _pdf_para_paginas_png. `num` é 1-based.
    Mesmo controle de acesso do resto do fluxo: só o hash_token.
    """
    if not AIRTABLE_API_KEY:
        return 'Serviço indisponível.', 503

    registro = _buscar_por_campo(TABLE_ASSINATURAS, 'Hash Token', hash_token)
    if not registro:
        return 'Link inválido ou expirado.', 404

    fields = registro.get('fields', {})
    if fields.get('Tipo de Documento') != TIPO_PACOTE_HOLERITE_PONTO:
        return 'Documento inválido.', 404

    pdf_bytes, erro = _pdf_original_do_pacote(fields, idx)
    if erro or not pdf_bytes:
        logger.warning(f'[PACOTE HOL+PONTO] preview PNG indisponível: {registro["id"]} | idx={idx} | {erro}')
        return 'Documento indisponível.', 404

    try:
        paginas = _pdf_para_paginas_png(pdf_bytes, apenas=num)
    except Exception as exc:
        logger.warning(f'[PACOTE HOL+PONTO] falha ao renderizar página: {registro["id"]} | idx={idx} | pg={num} | {exc}')
        return 'Documento indisponível.', 404
    if not paginas:
        return 'Página inexistente.', 404

    resposta = app.response_class(paginas[0], mimetype='image/png')
    resposta.headers['Cache-Control'] = 'private, max-age=300'
    return resposta


@app.route('/assinatura/<hash_token>/doc/<int:idx>', methods=['GET'])
def assinatura_documento_proxy(hash_token, idx):
    """
    Repassa (proxy) um dos 2 PDFs originais do pacote HOLERITE_FOLHA_PONTO
    pelo NOSSO domínio, em vez do navegador buscar direto na URL do
    Airtable — ver comentário completo na chamada (GET de
    assinatura_pagina). Existe só pra resolver o problema de CORS que
    fazia o PDF.js cair no visualizador nativo do navegador; não é uma
    rota de download de uso geral, nem substitui nenhuma checagem de
    acesso já existente (o hash_token continua sendo o único controle,
    igual ao resto do fluxo de assinatura).
    idx: 0 = Holerite, 1 = Folha de Ponto — mesma ordem de anexação usada
    em toda a Macro (holerite sempre primeiro, ponto sempre depois).
    """
    if not AIRTABLE_API_KEY:
        return 'Serviço indisponível.', 503
    if idx not in (0, 1):
        return 'Documento inválido.', 404

    registro = _buscar_por_campo(TABLE_ASSINATURAS, 'Hash Token', hash_token)
    if not registro:
        return 'Link inválido ou expirado.', 404

    fields = registro.get('fields', {})
    if fields.get('Tipo de Documento') != TIPO_PACOTE_HOLERITE_PONTO:
        return 'Documento inválido.', 404

    doc_atual = fields.get('Documento PDF') or []
    anexos_originais = [
        a for a in doc_atual
        if not a['filename'].startswith('Comprovante Assinatura')
        and not a['filename'].startswith('Combinado Carimbado')
        and ' - ASSINADO' not in a['filename']
    ]
    if len(anexos_originais) != 2:
        return 'Documento indisponível.', 404

    anexo = anexos_originais[idx]
    try:
        pdf_bytes = _carregar_documento_url(anexo['url'])
    except Exception as exc:
        logger.warning(f'[PACOTE HOL+PONTO] proxy de documento falhou: {registro["id"]} | idx={idx} | {exc}')
        return 'Erro ao carregar documento.', 502

    resposta = app.response_class(pdf_bytes, mimetype='application/pdf')
    resposta.headers['Content-Disposition'] = f'inline; filename="{anexo["filename"]}"'
    resposta.headers['Cache-Control'] = 'private, max-age=300'
    return resposta


@app.route('/assinatura/<hash_token>', methods=['GET', 'POST'])
def assinatura_pagina(hash_token):
    """Página pública de assinatura nativa — GET mostra o formulário de CPF,
    POST valida o CPF contra Funcionários e, se bater, registra IP/User-Agent/
    timestamp e marca a Assinatura (e o Processar Arquivos de origem) como
    "Assinado". O hash funciona como token de acesso ao documento."""
    if not AIRTABLE_API_KEY:
        return 'Serviço indisponível.', 503

    registro = _buscar_por_campo(TABLE_ASSINATURAS, 'Hash Token', hash_token)
    if not registro:
        return 'Link inválido ou expirado.', 404

    fields = registro.get('fields', {})
    nome_doc = fields.get('Nome', 'Documento')
    status_atual = fields.get('Status', 'Pendente')

    if status_atual == 'Assinado':
        quando = _fmt_quando(fields.get('Data/Hora Assinatura'))
        downloads = None
        if fields.get('Tipo de Documento') == TIPO_PACOTE_HOLERITE_PONTO:
            downloads = _downloads_do_registro_assinado(fields)
            if not downloads:
                logger.warning(
                    f'[PACOTE HOL+PONTO] registro assinado sem nenhum PDF entregável '
                    f'na tela de sucesso: {registro["id"]}'
                )
        return _pagina_assinatura_sucesso_html(nome_doc, quando, downloads=downloads)

    if status_atual == 'Expirado':
        return _pagina_assinatura_html(
            hash_token, nome_doc,
            erro='Link expirado por excesso de tentativas incorretas. Contate o RH para gerar um novo link.'
        )

    if status_atual == 'Cancelado':
        return _pagina_assinatura_html(
            hash_token, nome_doc,
            erro='Não foi possível concluir a confirmação. Contate o RH.',
            mostrar_formulario=False,
        )

    if request.method == 'GET':
        documentos_preview = None
        if fields.get('Tipo de Documento') == TIPO_PACOTE_HOLERITE_PONTO:
            # Correção desta Macro: busca os 2 PDFs originais (ainda não
            # carimbados) já anexados ao registro na criação do pacote, para
            # o colaborador ver o conteúdo real antes de assinar — mesma
            # ordem de anexação usada em _confirmar_assinatura_pacote_
            # holerite_ponto (holerite primeiro, ponto depois).
            doc_atual = fields.get('Documento PDF') or []
            anexos_originais = [
                a for a in doc_atual
                if not a['filename'].startswith('Comprovante Assinatura')
                and not a['filename'].startswith('Combinado Carimbado')
                and ' - ASSINADO' not in a['filename']
            ]
            if len(anexos_originais) == 2:
                # Evolução em 3 passos dentro desta mesma Macro, cada um
                # motivado por um teste real (ver histórico e autorizações
                # em .magnata/app-py-authorizations/):
                #   1) URL direta do Airtable + PDF.js -> falhava por CORS e
                #      caía no visualizador nativo do navegador;
                #   2) proxy same-origin do PDF -> não resolveu: o PDF.js
                #      nem chegava a carregar (CDN bloqueado por política
                #      corporativa no ambiente real testado);
                #   3) agora: PÁGINAS COMO IMAGEM (PNG com marca d'água de
                #      pré-visualização), servidas pelo próprio domínio.
                #      Sem PDF do lado do colaborador antes de assinar, sem
                #      dependência de CDN, sem visualizador nativo.
                # O PDF só chega ao colaborador depois da assinatura, no
                # "Baixar PDF Combinado Carimbado".
                #
                # A contagem de páginas é feita aqui (renderizando uma vez)
                # para o HTML já sair com as <img> certas. Se a renderização
                # falhar por qualquer motivo, cai no caminho anterior
                # (iframe com o PDF) em vez de deixar o colaborador sem
                # documento nenhum.
                documentos_preview = []
                for indice, rotulo in ((0, 'Holerite'), (1, 'Folha de Ponto')):
                    total_paginas = None
                    pdf_bytes, erro_preview = _pdf_original_do_pacote(fields, indice)
                    if pdf_bytes and not erro_preview:
                        try:
                            total_paginas = _contar_paginas_pdf(pdf_bytes)
                        except Exception as exc:
                            erro_preview = str(exc)
                    if not total_paginas:
                        logger.warning(
                            f'[PACOTE HOL+PONTO] preview em imagem indisponível ({erro_preview}) — '
                            f'usando PDF embutido para "{rotulo}": {registro["id"]}'
                        )
                        documentos_preview.append(
                            {'nome': rotulo, 'url': f'/assinatura/{hash_token}/doc/{indice}'}
                        )
                    else:
                        documentos_preview.append({
                            'nome': rotulo,
                            'url': f'/assinatura/{hash_token}/doc/{indice}',
                            'paginas': total_paginas,
                            'url_pagina': f'/assinatura/{hash_token}/doc/{indice}/pagina/',
                        })
            else:
                logger.warning(
                    f'[PACOTE HOL+PONTO] GET com estrutura inesperada de anexos ({len(anexos_originais)}, '
                    f'esperado 2) — mostrando formulário sem visualizador embutido: {registro["id"]}'
                )
        return _pagina_assinatura_html(hash_token, nome_doc, documentos=documentos_preview)

    # POST — validar os 4 últimos dígitos do CPF informado
    cpf_informado = re.sub(r'\D', '', request.form.get('cpf', ''))
    if len(cpf_informado) != 4:
        return _pagina_assinatura_html(hash_token, nome_doc, erro='Digite os 4 últimos números do seu CPF.')

    func_links = fields.get('Funcionário') or []
    func_id = func_links[0]['id'] if func_links and isinstance(func_links[0], dict) else (func_links[0] if func_links else None)
    if not func_id:
        return _pagina_assinatura_html(hash_token, nome_doc, erro='Funcionário não vinculado a esta assinatura. Contate o RH.')

    _at_throttle()
    r_func = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_FUNC}/{func_id}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'},
        timeout=30,
    )
    if not r_func.ok:
        return _pagina_assinatura_html(hash_token, nome_doc, erro='Erro ao validar dados. Tente novamente.')

    cpf_cadastrado = re.sub(r'\D', '', r_func.json().get('fields', {}).get('CPF', '') or '')
    ultimos4_cadastrado = cpf_cadastrado[-4:] if len(cpf_cadastrado) >= 4 else cpf_cadastrado

    if cpf_informado != ultimos4_cadastrado:
        tentativas = (fields.get('Tentativas') or 0) + 1
        logger.warning(f'[ASSINATURA] CPF não confere (tentativa {tentativas}) — hash={hash_token}, func_id={func_id}')

        campos_falha = {F_ASS_TENTATIVAS: tentativas}
        if tentativas >= MAX_TENTATIVAS_ASSINATURA:
            campos_falha[F_ASS_STATUS] = 'Expirado'

        _at_throttle()
        requests.patch(
            f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{registro["id"]}',
            headers=_at_headers(), json={'fields': campos_falha, 'typecast': True}, timeout=15,
        )

        if tentativas >= MAX_TENTATIVAS_ASSINATURA:
            return _pagina_assinatura_html(
                hash_token, nome_doc,
                erro='Link expirado por excesso de tentativas incorretas. Contate o RH para gerar um novo link.'
            )
        return _pagina_assinatura_html(hash_token, nome_doc, erro='Os números não correspondem ao cadastro. Verifique e tente novamente.')

    # CPF confere — registrar assinatura
    ip_captura = _ip_real_da_requisicao()
    user_agent = request.headers.get('User-Agent', 'desconhecido')[:500]
    agora_brt_dt = datetime.now(timezone(timedelta(hours=-3)))
    agora_brt = agora_brt_dt.isoformat()

    # Checagem de concorrência (best-effort — Airtable não tem transação
    # nem constraint único, então isto reduz mas não elimina a janela de
    # corrida): relê o Status imediatamente antes de gravar "Assinado".
    # Se outra requisição já confirmou entre o GET inicial e agora, aborta
    # em vez de carimbar/gerar comprovante de novo (nunca reprocessa um
    # pacote/documento já assinado).
    _at_throttle()
    r_recheck = requests.get(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{registro["id"]}',
        headers={'Authorization': f'Bearer {AIRTABLE_API_KEY}'}, timeout=15,
    )
    if r_recheck.ok and r_recheck.json().get('fields', {}).get('Status') == 'Assinado':
        quando = _fmt_quando(r_recheck.json().get('fields', {}).get('Data/Hora Assinatura'))
        return _pagina_assinatura_sucesso_html(nome_doc, quando)

    # Pacote Holerite + Folha de Ponto: ordem de escrita própria (nunca
    # grava "Assinado" antes de concluir carimbo + comprovante dos 2
    # documentos — ver docstring de _confirmar_assinatura_pacote_holerite_
    # ponto). Sai daqui ANTES do bloco compartilhado abaixo, que grava
    # "Assinado" incondicionalmente — esse bloco compartilhado continua
    # exatamente como estava para os demais tipos.
    if fields.get('Tipo de Documento') == TIPO_PACOTE_HOLERITE_PONTO:
        return _confirmar_assinatura_pacote_holerite_ponto(
            registro=registro, fields=fields, hash_token=hash_token, nome_doc=nome_doc,
            func_id=func_id, cpf_informado=cpf_informado, ip_captura=ip_captura,
            user_agent=user_agent, agora_brt_dt=agora_brt_dt,
            nome_func=r_func.json().get('fields', {}).get('Nome Completo', ''),
        )

    evidencias = (
        f'IP: {ip_captura}\n'
        f'Timestamp: {agora_brt_dt.strftime("%d/%m/%Y %H:%M:%S")} (horário de Brasília)\n'
        f'User-Agent: {user_agent}\n'
        f'CPF confirmado (4 últimos dígitos): {cpf_informado}\n'
        f'Declaração: "Ao clicar, concordo eletronicamente com o recebimento e os termos '
        f'do documento, sob as penas da lei."'
    )

    _at_throttle()
    r_patch = requests.patch(
        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{registro["id"]}',
        headers=_at_headers(),
        json={
            'fields': {
                F_ASS_STATUS:         'Assinado',
                F_ASS_CPF_INFORMADO:  cpf_informado,
                F_ASS_IP:             ip_captura,
                F_ASS_USER_AGENT:     user_agent,
                F_ASS_DATA_ASSINATURA: agora_brt,
                F_ASS_EVIDENCIAS:     evidencias,
            },
            'typecast': True,
        },
        timeout=15,
    )
    if not r_patch.ok:
        logger.error(f'[ASSINATURA] falha ao gravar assinatura {registro["id"]}: {r_patch.text[:300]}')
        return _pagina_assinatura_html(hash_token, nome_doc, erro='Erro ao registrar assinatura. Tente novamente em alguns instantes.')

    # Regra universal (v2.68): para QUALQUER documento assinado pelo sistema —
    # (1) carimbar todas as páginas do original ("Assinado digitalmente por..."),
    # (2) gerar o Comprovante de Assinatura Eletrônica em separado,
    # (3) substituir o anexo original pelos dois em Documento PDF (Assinaturas
    #     Digitais), e (4) copiar ambos para o campo "Documentos" do
    #     Funcionário — automático, sem depender de script manual.
    #
    # Tipo HOLERITE_FOLHA_PONTO nunca chega até aqui: retorna antes, mais
    # acima nesta mesma rota, via _confirmar_assinatura_pacote_holerite_
    # ponto (esse tipo tem 2 documentos no mesmo campo "Documento PDF", e
    # este bloco universal usa next(...) para pegar só o PRIMEIRO — se
    # rodasse aqui, carimbaria e geraria comprovante de só 1 dos 2).
    try:
        doc_atual = fields.get('Documento PDF') or []
        original_att = next(
            (a for a in doc_atual if not a['filename'].startswith('Comprovante Assinatura')), None
        )
        if original_att:
            nome_func = r_func.json().get('fields', {}).get('Nome Completo', '')
            r_doc = requests.get(original_att['url'], timeout=60)
            if r_doc.ok:
                doc_bytes_original = r_doc.content
                dt_fmt = agora_brt_dt.strftime('%d/%m/%Y %H:%M')

                pdf_carimbado = _carimbar_pdf_assinado(doc_bytes_original, nome_func, cpf_informado, dt_fmt)
                nome_carimbado = original_att['filename']
                nome_carimbado = (
                    nome_carimbado[:-4] + ' - ASSINADO.pdf'
                    if nome_carimbado.lower().endswith('.pdf') else nome_carimbado + ' - ASSINADO'
                )

                pdf_comprovante = _gerar_comprovante_assinatura_pdf(
                    nome_documento=nome_doc, func_nome=nome_func, ip=ip_captura, user_agent=user_agent,
                    dt_str=dt_fmt, cpf4=cpf_informado, doc_bytes=doc_bytes_original,
                    doc_nome=original_att['filename'],
                )
                nome_comprovante = f'Comprovante Assinatura - {(nome_func or "Colaborador").title()}.pdf'

                _anexar_attachment(TABLE_ASSINATURAS, registro['id'], F_ASS_DOCUMENTO_PDF, pdf_carimbado, nome_carimbado)
                upload_comprovante = _anexar_attachment(
                    TABLE_ASSINATURAS, registro['id'], F_ASS_DOCUMENTO_PDF, pdf_comprovante, nome_comprovante,
                )
                lista_pos_upload = (upload_comprovante.get('fields', {}) or {}).get(F_ASS_DOCUMENTO_PDF) or []
                ids_novos = [a['id'] for a in lista_pos_upload if a['filename'] in (nome_carimbado, nome_comprovante)]
                if ids_novos:
                    manter = ids_novos + [a['id'] for a in doc_atual if a['id'] != original_att['id']]
                    _at_throttle()
                    requests.patch(
                        f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ASSINATURAS}/{registro["id"]}',
                        headers=_at_headers(),
                        json={'fields': {F_ASS_DOCUMENTO_PDF: [{'id': i} for i in manter]}, 'typecast': True},
                        timeout=30,
                    )

                # Copiar Kit carimbado + Comprovante para "Documentos" do Funcionário (visível/fácil acesso)
                _anexar_attachment(TABLE_FUNC, func_id, F_FUNC_DOCUMENTOS, pdf_carimbado, nome_carimbado)
                _anexar_attachment(TABLE_FUNC, func_id, F_FUNC_DOCUMENTOS, pdf_comprovante, nome_comprovante)
    except Exception as exc:
        logger.warning(f'[ASSINATURA] falha ao carimbar/gerar comprovante/salvar no funcionário {registro["id"]}: {exc}')

    # ── Fase Definitiva da Folha de Ponto (v3.00) ────────────────────────────
    # Bloco ISOLADO: só executa quando Tipo de Documento == 'Folha de Ponto'.
    # Não depende de nenhuma variável do bloco universal acima (busca e
    # carimba por conta própria) e não altera Rescisão/Admissão/outros tipos
    # nem o campo "Documentos" — mexe exclusivamente em "PDF Folha Ponto".
    # Substitui o PDF provisório (não assinado) que ficou lá pela versão
    # carimbada, via _substituir_attachment — nunca acumula os dois no mesmo
    # campo (Regra "Sem Duplicidade").
    if fields.get('Tipo de Documento') == 'Folha de Ponto':
        try:
            # Nunca reler o campo compartilhado "PDF Folha Ponto" do Funcionário
            # e pegar "o último anexo" — esse campo pode ter sido alterado por
            # outro processamento (ex.: reprocesso, disparo duplicado) entre a
            # geração deste link e o momento da assinatura, fazendo o código
            # carimbar o documento errado. Em vez disso, reaproveita o mesmo
            # `doc_atual`/`original_att` já obtidos acima a partir do PRÓPRIO
            # registro de Assinatura (cópia imutável, gravada no momento da
            # geração deste link específico — sempre correta).
            if original_att and r_doc.ok:
                nome_func_ponto = r_func.json().get('fields', {}).get('Nome Completo', '')
                dt_fmt_ponto = agora_brt_dt.strftime('%d/%m/%Y %H:%M')
                pdf_ponto_assinado = _carimbar_pdf_assinado(
                    doc_bytes_original, nome_func_ponto, cpf_informado, dt_fmt_ponto)
                nome_ponto_assinado = original_att['filename']
                nome_ponto_assinado = (
                    nome_ponto_assinado[:-4] + ' - ASSINADO.pdf'
                    if nome_ponto_assinado.lower().endswith('.pdf') else nome_ponto_assinado + ' - ASSINADO'
                )
                _substituir_attachment(TABLE_FUNC, func_id, F_FUNC_PDF_FOLHA,
                                        pdf_ponto_assinado, nome_ponto_assinado)
                logger.info(f'[ASSINATURA][PONTO] PDF Folha Ponto substituído pelo assinado — func_id={func_id}')
        except Exception as exc:
            logger.warning(f'[ASSINATURA][PONTO] falha ao substituir PDF Folha Ponto assinado (func_id={func_id}): {exc}')

    # Atualizar Status do documento de origem em Processar Arquivos, se vinculado
    processar_id = fields.get('Processar Arquivos ID')
    if processar_id:
        try:
            _atualizar_status_processar(processar_id, 'Assinado')
        except Exception as exc:
            logger.warning(f'[ASSINATURA] não foi possível atualizar Processar Arquivos {processar_id}: {exc}')

    logger.info(f'[ASSINATURA] Assinado com sucesso — hash={hash_token}, func_id={func_id}, ip={ip_captura}')
    return _pagina_assinatura_sucesso_html(nome_doc, _fmt_quando(agora_brt))


# ─── VR/VA FATIAMENTO ─────────────────────────────────────────────────────────

def _ler_xls_vr_va(dados_xls: bytes) -> dict:
    """XLS relatorio VR/VA -> dict CPF_NUM -> {nome, cpf_fmt, va, vr}."""
    import xlrd  # xlrd==1.2.0 suporta .xls legado
    wb = xlrd.open_workbook(file_contents=dados_xls)
    sh = wb.sheets()[0]
    colaboradores = {}
    # Header linha 23 (0-based); dados a partir da linha 24 (0-based)
    # cols: 1=Nome, 2=Valor, 3=Produto, 4=CPF
    for row in range(24, sh.nrows):
        nome = str(sh.cell_value(row, 1)).strip()
        if not nome or nome == 'Nome':
            continue
        try:
            valor = float(sh.cell_value(row, 2))
        except (ValueError, TypeError):
            continue
        produto = str(sh.cell_value(row, 3)).strip().lower()
        cpf_raw = str(sh.cell_value(row, 4)).strip()
        cpf = re.sub(r'\D', '', cpf_raw)
        if len(cpf) < 11:
            continue
        if cpf not in colaboradores:
            colaboradores[cpf] = {'nome': nome, 'cpf_fmt': cpf_raw, 'va': 0.0, 'vr': 0.0}
        if 'alimenta' in produto:
            colaboradores[cpf]['va'] += valor
        elif 'refei' in produto:
            colaboradores[cpf]['vr'] += valor
    return colaboradores


def _ler_xlsx_vr_va(dados_xlsx: bytes) -> dict:
    """v3.12 — XLSX moderno 'Pedido de Benefícios' (mesmo relatório de
    _ler_xls_vr_va, mas formato .xlsx) -> dict CPF_NUM -> {nome, cpf_fmt, va, vr}.
    O layout novo não tem a linha de cabeçalho num número fixo (varia por
    quantos produtos/taxas aparecem no resumo antes da tabela), então
    localiza dinamicamente a linha cuja coluna B (índice 1) é 'Nome'."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(dados_xlsx), data_only=True)
    ws = wb.worksheets[0]
    linhas = list(ws.iter_rows(values_only=True))

    idx_header = None
    for i, row in enumerate(linhas):
        if row and len(row) > 1 and str(row[1] or '').strip() == 'Nome':
            idx_header = i
            break
    if idx_header is None:
        return {}

    colaboradores = {}
    for row in linhas[idx_header + 1:]:
        if not row or len(row) < 5 or not row[1]:
            continue
        nome = str(row[1]).strip()
        try:
            valor = float(row[2])
        except (ValueError, TypeError):
            continue
        produto = str(row[3] or '').strip().lower()
        cpf_raw = str(row[4] or '').strip()
        cpf = re.sub(r'\D', '', cpf_raw)
        if len(cpf) < 11:
            continue
        if cpf not in colaboradores:
            colaboradores[cpf] = {'nome': nome, 'cpf_fmt': cpf_raw, 'va': 0.0, 'vr': 0.0}
        if 'alimenta' in produto:
            colaboradores[cpf]['va'] += valor
        elif 'refei' in produto:
            colaboradores[cpf]['vr'] += valor
    return colaboradores


def _ler_planilha_vr_va(dados: bytes, nome_arquivo: str) -> dict:
    """v3.12 — Detecta .xls legado vs .xlsx moderno pelo nome do arquivo
    (fallback: assinatura de bytes — ZIP/'PK' = xlsx, OLE2 = xls) e despacha
    pro parser certo. Mesmo dict de saída dos dois lados, resto do fluxo de
    /processar-vr-va não muda."""
    nome_lower = (nome_arquivo or '').lower()
    eh_xlsx = nome_lower.endswith('.xlsx') or dados[:2] == b'PK'
    return _ler_xlsx_vr_va(dados) if eh_xlsx else _ler_xls_vr_va(dados)


def _gerar_pdf_vr_va_cliente(nome_cliente: str, folha_mensal: str,
                              colaboradores: list, pedido_num: str = '',
                              data_pedido: str = '') -> bytes:
    """PDF VR/VA por cliente. colaboradores: list de dicts {nome, cpf_fmt, va, vr}."""
    from fpdf import FPDF

    def fmt_brl(v: float) -> str:
        return f'{v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

    _TZ = timezone(timedelta(hours=-3))
    pdf = FPDF()
    pdf.add_page()

    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 8, 'Magnata Portaria e Servicos Ltda', ln=True, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, f'Relatorio VR e VA - {folha_mensal}', ln=True, align='C')
    pdf.cell(0, 6, f'Cliente: {nome_cliente}', ln=True, align='C')
    if pedido_num:
        pdf.cell(0, 5, f'Pedido N {pedido_num}  |  Pago em {data_pedido}', ln=True, align='C')
    pdf.ln(4)

    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_fill_color(220, 220, 220)
    pdf.cell(75, 7, 'Nome', border=1, fill=True)
    pdf.cell(38, 7, 'CPF', border=1, fill=True)
    pdf.cell(25, 7, 'VA (R$)', border=1, fill=True, align='R')
    pdf.cell(25, 7, 'VR (R$)', border=1, fill=True, align='R')
    pdf.cell(0, 7, 'Total (R$)', border=1, fill=True, align='R', ln=True)

    pdf.set_font('Helvetica', '', 7)
    total_geral = 0.0
    for col in sorted(colaboradores, key=lambda x: x['nome']):
        total = col['va'] + col['vr']
        total_geral += total
        pdf.cell(75, 6, col['nome'][:43], border=1)
        pdf.cell(38, 6, col['cpf_fmt'], border=1)
        pdf.cell(25, 6, fmt_brl(col['va']), border=1, align='R')
        pdf.cell(25, 6, fmt_brl(col['vr']), border=1, align='R')
        pdf.cell(0, 6, fmt_brl(total), border=1, align='R', ln=True)

    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_fill_color(200, 230, 200)
    pdf.cell(163, 7, f'TOTAL - {len(colaboradores)} colaborador(es)', border=1, fill=True, align='R')
    pdf.cell(0, 7, f'R$ {fmt_brl(total_geral)}', border=1, fill=True, align='R', ln=True)

    pdf.ln(4)
    pdf.set_font('Helvetica', 'I', 7)
    agora = datetime.now(tz=_TZ).strftime('%d/%m/%Y %H:%M')
    pdf.cell(0, 4, f'Gerado em {agora} | VA = Vale Alimentacao | VR = Vale Refeicao', ln=True)

    return bytes(pdf.output())


@app.route('/processar-vr-va', methods=['POST', 'OPTIONS'])
def processar_vr_va():
    """
    v2.34 - Fatia relatorio VR/VA (XLS) por cliente via CPF->Local->Cliente,
    gera PDF de prestacao de contas por condominio e indexa em Extratos Mensais
    (ja incluso automaticamente no pacote de e-mail).

    v3.12 - Aceita também .xlsx moderno (campo 'xls' OU 'xlsx' — mesmo
    parâmetro, detecção automática por nome de arquivo/assinatura de bytes
    em _ler_planilha_vr_va), além do .xls legado original.

    multipart/form-data:
      xls / xlsx   - arquivo do relatorio VR/VA, .xls legado ou .xlsx moderno (obrigatorio)
      comprovante  - PDF comprovante consolidado (opcional - arquiva em Guias)
      folha_mensal - ex. 'Maio 2026' (obrigatorio)
      pedido_num   - numero do pedido (opcional, ex. '767390003')
      data_pedido  - data do pagamento (opcional, ex. '01/05/2026')
      dry_run      - 'true'/'false' (padrao: 'true')
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    if not AIRTABLE_API_KEY:
        return jsonify({'status': 'erro', 'erro': 'AIRTABLE_API_KEY nao configurada'}), 500

    folha_mensal = (request.form.get('folha_mensal') or '').strip()
    dry_run      = request.form.get('dry_run', 'true').lower() != 'false'
    pedido_num   = (request.form.get('pedido_num') or '767390003').strip()
    data_pedido  = (request.form.get('data_pedido') or '01/05/2026').strip()

    if not folha_mensal:
        return jsonify({'status': 'erro', 'erro': 'folha_mensal obrigatorio'}), 400

    xls_file = request.files.get('xls') or request.files.get('xlsx')
    if not xls_file:
        return jsonify({'status': 'erro', 'erro': 'arquivo xls/xlsx obrigatorio'}), 400
    xls_bytes = xls_file.read()

    comp_file  = request.files.get('comprovante')
    comp_bytes = comp_file.read() if comp_file else None
    comp_nome  = comp_file.filename if comp_file else None

    try:
        # 1. Parse XLS/XLSX
        colaboradores_map = _ler_planilha_vr_va(xls_bytes, xls_file.filename)
        logger.info(f'[VR/VA] {len(colaboradores_map)} colaboradores no XLS')

        # 2. Batch: todos os Locais (id -> lista de cliente_ids)
        locais_map = {
            rec['id']: rec['fields'].get('Cliente', [])
            for rec in _at_listar_todos(TABLE_LOCAIS, ['Nome', 'Cliente'])
        }

        # 3. Batch: todos os Funcionarios (CPF -> locais_ids)
        cpf_para_locais = {}
        for rec in _at_listar_todos(TABLE_FUNC, ['CPF', 'Locais de trabalho']):
            cpf = re.sub(r'\D', '', str(rec['fields'].get('CPF') or ''))
            if cpf:
                cpf_para_locais[cpf] = rec['fields'].get('Locais de trabalho') or []

        # 4. Batch: todos os Clientes (id -> nome)
        clientes_nome = {
            rec['id']: rec['fields'].get('Nome', rec['id'])
            for rec in _at_listar_todos(TABLE_CLIENTES, ['Nome'])
        }

        # 5. CPF -> primeiro cliente encontrado via Local
        cpf_para_cliente_id = {}
        for cpf, local_ids in cpf_para_locais.items():
            for lid in local_ids:
                client_ids = locais_map.get(lid, [])
                if client_ids:
                    cid = client_ids[0]
                    # client_ids pode ser lista de strings ou de dicts
                    if isinstance(cid, dict):
                        cid = cid.get('id', cid)
                    cpf_para_cliente_id[cpf] = cid
                    break

        # 6. Agrupar colaboradores do XLS por cliente
        clientes_grupos = {}  # cliente_id -> {nome, colaboradores:[]}
        sem_cliente = []

        for cpf, dados in colaboradores_map.items():
            cid = cpf_para_cliente_id.get(cpf)
            if not cid:
                sem_cliente.append({'cpf': dados['cpf_fmt'], 'nome': dados['nome']})
                continue
            if cid not in clientes_grupos:
                clientes_grupos[cid] = {
                    'nome': clientes_nome.get(cid, cid),
                    'colaboradores': [],
                }
            clientes_grupos[cid]['colaboradores'].append({
                'nome':    dados['nome'],
                'cpf_fmt': dados['cpf_fmt'],
                'va':      dados['va'],
                'vr':      dados['vr'],
            })

        # 7. Gerar PDF e indexar por cliente em Extratos Mensais
        resultados = []
        _TZ = timezone(timedelta(hours=-3))
        for cid, grupo in clientes_grupos.items():
            nome_cli = grupo['nome']
            cols     = grupo['colaboradores']
            total    = sum(c['va'] + c['vr'] for c in cols)

            if dry_run:
                resultados.append({
                    'cliente_id': cid, 'cliente': nome_cli,
                    'colaboradores': len(cols), 'total_R$': round(total, 2),
                    'acao': 'dry_run - PDF nao gerado',
                })
                continue

            try:
                pdf_bytes = _gerar_pdf_vr_va_cliente(
                    nome_cliente=nome_cli,
                    folha_mensal=folha_mensal,
                    colaboradores=cols,
                    pedido_num=pedido_num,
                    data_pedido=data_pedido,
                )
            except Exception as exc:
                resultados.append({'cliente_id': cid, 'cliente': nome_cli, 'erro': str(exc)})
                continue

            fname  = f'VR e VA {folha_mensal} - {nome_cli}.pdf'
            rec_id = _criar_registro(TABLE_EXTRATO, {
                F_EXT_STATUS:  'Concluido',
                F_EXT_FOLHA:   folha_mensal,
                F_EXT_CLIENTE: [cid],
                F_EXT_DATA:    datetime.now(tz=_TZ).strftime('%Y-%m-%d'),
            })
            _anexar_attachment(TABLE_EXTRATO, rec_id, F_EXT_PDF, pdf_bytes, fname)

            resultados.append({
                'cliente_id': cid, 'cliente': nome_cli,
                'colaboradores': len(cols), 'total_R$': round(total, 2),
                'record_id': rec_id, 'arquivo': fname,
            })

        # 8. Comprovante consolidado (broadcast -> TABLE_GUIAS)
        comp_record_id = None
        if comp_bytes and not dry_run:
            comp_record_id = _criar_registro(TABLE_GUIAS, {
                F_GUIA_STATUS: 'Recebido',
                F_GUIA_TIPO:   'VR/VA',
                F_GUIA_NOME:   f'Comprovante VR e VA {folha_mensal}',
            })
            _anexar_attachment(
                TABLE_GUIAS, comp_record_id, F_GUIA_COMPROV,
                comp_bytes, comp_nome or f'Comprovante VR VA {folha_mensal}.pdf',
            )

        return jsonify({
            'status':                   'dry_run' if dry_run else 'concluido',
            'folha_mensal':             folha_mensal,
            'total_colaboradores_xls':  len(colaboradores_map),
            'clientes_processados':     len(clientes_grupos),
            'sem_cliente':              len(sem_cliente),
            'comprovante_record_id':    comp_record_id,
            'resultados':               resultados,
            'sem_cliente_detalhe':      sem_cliente,
        })

    except Exception as exc:
        logger.exception(f'[VR/VA] erro: {exc}')
        return jsonify({'status': 'erro', 'erro': str(exc)}), 500


@app.route('/webhook/enviar-whatsapp', methods=['POST', 'OPTIONS'])
def webhook_enviar_whatsapp():
    """
    v2.92 — Webhook Airtable: botão 'Enviar WhatsApp' em qualquer view de Funcionários.

    Body JSON (configurado na Automação Airtable):
      {
        "funcionario_id": "recXXX",              // ID do Funcionário (obrigatório)
        "documentos": ["Holerite", "Folha Ponto"],  // opcional; omitir = todos disponíveis
        "folha_mensal": "Junho 2026",            // opcional; omitir = holerite mais recente
        "dry_run": false
      }

    Envia cada PDF em base64 via Evolution API (mesmo fix da v2.89: evita
    PDF corrompido quando a Evolution busca a URL por conta própria).
    Cria registro em Envios de Documentos para rastreio.
    Protegido por X-API-KEY (EMAIL_WEBHOOK_KEY).
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401

    body = request.get_json(silent=True) or {}
    func_id = body.get('funcionario_id') or body.get('record_id')
    documentos = body.get('documentos')     # None = todos disponíveis
    folha_mensal = body.get('folha_mensal')
    dry_run = bool(body.get('dry_run', False))

    if not func_id:
        return jsonify({'status': 'erro', 'erro': 'funcionario_id ausente no body'}), 400

    try:
        func_rec = _at_obter_registro(TABLE_FUNC, func_id, [
            'Nome Completo', 'WhatsApp', 'Holerites', 'PDF Folha Ponto',
        ])
        f = func_rec.get('fields', {})
        nome = f.get('Nome Completo') or func_id
        whatsapp = f.get('WhatsApp')
        if not whatsapp:
            return jsonify({'status': 'nenhum_envio', 'funcionario': nome,
                            'motivo': 'whatsapp_ausente'}), 200
        numero = _normalizar_numero_evolution(whatsapp)
        if not numero:
            return jsonify({'status': 'nenhum_envio', 'funcionario': nome,
                            'motivo': f'numero_invalido ({whatsapp})'}), 200

        enviar_holerite = documentos is None or 'Holerite' in documentos
        enviar_ponto    = documentos is None or 'Folha Ponto' in documentos

        docs_enviados, erros = [], []

        # --- Holerite ---
        if enviar_holerite:
            hol_links = [
                (h['id'] if isinstance(h, dict) else h)
                for h in (f.get('Holerites') or [])
            ]
            if hol_links:
                candidatos = []
                for hid in hol_links[-5:]:   # últimos 5: evita scan desnecessário
                    try:
                        hr = _at_obter_registro(TABLE_HOL, hid, ['Folha Mensal', 'PDF HOLERITE'])
                        hf = hr.get('fields', {})
                        pdfs = hf.get('PDF HOLERITE') or []
                        if pdfs:
                            candidatos.append((hf.get('Folha Mensal', ''), pdfs[-1]))
                    except Exception:
                        pass
                if candidatos:
                    if folha_mensal:
                        match = [c for c in candidatos if c[0] == folha_mensal]
                        candidatos = match or candidatos
                    candidatos.sort(key=lambda x: x[0], reverse=True)
                    hol_folha, pdf_meta = candidatos[0]
                    try:
                        pdf_bytes = _carregar_documento_url(pdf_meta['url'])
                        fname = pdf_meta.get('filename') or f'Holerite {hol_folha} - {nome}.pdf'
                        if not dry_run:
                            _evolution_enviar_documento(
                                numero, pdf_meta['url'], fname,
                                caption=f'Holerite {hol_folha} — Grupo Magnata',
                                media_bytes=pdf_bytes,
                            )
                        docs_enviados.append({'tipo': 'Holerite', 'folha': hol_folha, 'arquivo': fname})
                    except Exception as exc:
                        erros.append({'tipo': 'Holerite', 'erro': str(exc)})

        # --- Folha de Ponto ---
        if enviar_ponto:
            ponto_anexos = f.get('PDF Folha Ponto') or []
            if ponto_anexos:
                ponto_meta = ponto_anexos[-1]
                try:
                    pdf_bytes = _carregar_documento_url(ponto_meta['url'])
                    fname = ponto_meta.get('filename') or f'Folha Ponto - {nome}.pdf'
                    if not dry_run:
                        _evolution_enviar_documento(
                            numero, ponto_meta['url'], fname,
                            caption='Cartão Ponto — Grupo Magnata',
                            media_bytes=pdf_bytes,
                        )
                    docs_enviados.append({'tipo': 'Folha Ponto', 'arquivo': fname})
                except Exception as exc:
                    erros.append({'tipo': 'Folha Ponto', 'erro': str(exc)})

        if not docs_enviados and not erros:
            return jsonify({'status': 'nenhum_documento', 'funcionario': nome,
                            'motivo': 'nenhum_pdf_disponivel'}), 200

        if not dry_run and docs_enviados:
            _criar_registro(TABLE_ENVIOS, {
                F_ENVIO_TIPO:   'Envio via Webhook',
                F_ENVIO_CANAL:  'WhatsApp',
                F_ENVIO_DEST:   numero,
                F_ENVIO_FUNC:   [func_id],
                F_ENVIO_STATUS: 'Enviado' if not erros else 'Erro',
                F_ENVIO_DATA:   datetime.now(timezone(timedelta(hours=-3))).strftime('%Y-%m-%d'),
            })

        status = 'dry_run' if dry_run else ('enviado' if not erros else 'parcial')
        return jsonify({
            'status': status, 'funcionario': nome, 'numero': numero,
            'docs_enviados': docs_enviados, 'erros': erros,
        })

    except Exception as exc:
        logger.exception(f'[WEBHOOK WA] {exc}')
        return jsonify({'status': 'erro', 'erro': str(exc)}), 500


@app.route('/webhook/enviar-email-cliente', methods=['POST', 'OPTIONS'])
def webhook_enviar_email_cliente():
    """
    v2.92 — Webhook Airtable: botão 'Enviar E-mail' na view de Clientes.

    Body JSON (configurado na Automação Airtable):
      {
        "cliente_id": "recXXX",          // ID do Cliente (obrigatório)
        "folha_mensal": "Junho 2026",    // competência dos documentos (obrigatório)
        "dry_run": false
      }

    Gera e dispara imediatamente 1 pacote de e-mail para o cliente especificado,
    LGPD-compliant: apenas documentos dos colaboradores lotados nos Locais desse
    cliente. Usa force=True (ignora dedup) pois o clique no botão é intencional.
    Protegido por X-API-KEY (EMAIL_WEBHOOK_KEY).
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    api_key = request.headers.get('X-API-KEY', '')
    if not EMAIL_WEBHOOK_KEY or api_key != EMAIL_WEBHOOK_KEY:
        return jsonify({'status': 'erro', 'erro': 'X-API-KEY inválida ou ausente'}), 401

    body = request.get_json(silent=True) or {}
    cliente_id   = body.get('cliente_id') or body.get('record_id')
    folha_mensal = body.get('folha_mensal')
    dry_run      = bool(body.get('dry_run', False))

    if not cliente_id:
        return jsonify({'status': 'erro', 'erro': 'cliente_id ausente no body'}), 400
    if not folha_mensal:
        return jsonify({'status': 'erro', 'erro': 'folha_mensal ausente no body'}), 400

    try:
        resultado_fila = _gerar_fila_envios_email(
            folha_mensal=folha_mensal,
            dry_run=dry_run,
            cliente_id_filtro=cliente_id,
            force=True,
        )

        if dry_run or not resultado_fila.get('envios'):
            return jsonify({
                'status': 'dry_run' if dry_run else 'nenhum_documento',
                'cliente_id': cliente_id,
                'folha_mensal': folha_mensal,
                'fila': resultado_fila,
            })

        resultado_disparo = _disparar_fila_email(dry_run=False)
        return jsonify({
            'status': 'enviado',
            'cliente_id': cliente_id,
            'folha_mensal': folha_mensal,
            'fila': resultado_fila,
            'disparo': resultado_disparo,
        })

    except Exception as exc:
        logger.exception(f'[WEBHOOK EMAIL] {exc}')
        return jsonify({'status': 'erro', 'erro': str(exc)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
